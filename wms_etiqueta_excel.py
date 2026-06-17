"""
Gera etiquetas longarina em Excel (layout ETIQUETA LONGARINA 102×73 mm).

Usa o template do usuário (Planilha2 + área A1:H23) e substitui a imagem
por uma renderização com os dados do WMS — mesma aparência do Excel manual.
"""
from __future__ import annotations

import io
import os
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont
from barcode import Code128
from barcode.writer import ImageWriter

from wms_etiqueta_zebra import ETIQUETA_ZEBRA_ZD220

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_BASE_DIR, 'data', 'wms', 'etiqueta_longarina_template.xlsx')

_DPI = 203


def _font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = []
    if bold:
        candidates.extend([
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            'C:/Windows/Fonts/arialbd.ttf',
            'C:/Windows/Fonts/Arialbd.ttf',
        ])
    else:
        candidates.extend([
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/Arial.ttf',
        ])
    for path in candidates:
        if os.path.isfile(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def _text_size(draw: ImageDraw.ImageDraw, text: str, font) -> tuple[int, int]:
    if hasattr(draw, 'textbbox'):
        box = draw.textbbox((0, 0), text, font=font)
        return box[2] - box[0], box[3] - box[1]
    return draw.textsize(text, font=font)


def _draw_centered(draw, box, text, font, fill='black'):
    x0, y0, x1, y1 = box
    tw, th = _text_size(draw, text, font)
    x = x0 + max(0, (x1 - x0 - tw) // 2)
    y = y0 + max(0, (y1 - y0 - th) // 2)
    draw.text((x, y), text, fill=fill, font=font)


def _barcode_png(data: str, target_w: int, target_h: int) -> Image.Image:
    buf = io.BytesIO()
    writer = ImageWriter()
    Code128(str(data), writer=writer).write(buf, options={
        'module_width': 0.28,
        'module_height': 12.0,
        'quiet_zone': 1.0,
        'write_text': False,
        'dpi': _DPI,
    })
    buf.seek(0)
    img = Image.open(buf).convert('RGB')
    ratio = min(target_w / img.width, target_h / img.height, 1.0)
    if ratio < 1.0:
        img = img.resize((max(1, int(img.width * ratio)), max(1, int(img.height * ratio))), Image.Resampling.LANCZOS)
    return img


def _draw_arrow_down(draw: ImageDraw.ImageDraw, cx: int, cy: int, size: int):
    half = size // 2
    draw.polygon([
        (cx, cy + half),
        (cx - half, cy - half),
        (cx + half, cy - half),
    ], fill='black')


def render_longarina_png(etiqueta: dict) -> bytes:
    """Renderiza etiqueta longarina em PNG (102×73 mm @ 203 dpi)."""
    z = ETIQUETA_ZEBRA_ZD220
    w = int(round(z['largura_mm'] * _DPI / 25.4))
    h = int(round(z['altura_mm'] * _DPI / 25.4))
    top_h = int(round(h * z['grid_top_pct'] / 100))
    mid_h = int(round(h * z['grid_mid_pct'] / 100))
    bot_h = h - top_h - mid_h

    img = Image.new('RGB', (w, h), 'white')
    draw = ImageDraw.Draw(img)
    lw = max(2, w // 250)
    draw.rectangle([0, 0, w - 1, h - 1], outline='black', width=lw)
    draw.line([(0, top_h), (w, top_h)], fill='black', width=lw)
    draw.line([(0, top_h + mid_h), (w, top_h + mid_h)], fill='black', width=lw)

    col_w = w // 4
    for i in range(1, 4):
        x = i * col_w
        draw.line([(x, 0), (x, top_h)], fill='black', width=lw)

    lbl_font = _font(max(14, h // 28), bold=True)
    val_font = _font(max(36, h // 8), bold=True)
    cols = [
        ('CÂMARA', str(etiqueta.get('camara') or etiqueta.get('rua_num') or '')),
        ('RUA', str(etiqueta.get('rua_letra') or '-')),
        ('COLUNA', str(etiqueta.get('predio') or '')),
        ('NÍVEL', str(etiqueta.get('nivel') or '')),
    ]
    for i, (lbl, val) in enumerate(cols):
        x0 = i * col_w
        x1 = x0 + col_w
        lbl_h = top_h // 3
        _draw_centered(draw, (x0, 2, x1, lbl_h), lbl, lbl_font)
        val_fs = max(28, h // 10) if len(val) > 2 else max(40, h // 7)
        _draw_centered(draw, (x0, lbl_h, x1, top_h - 2), val, _font(val_fs, bold=True))

    bc = str(etiqueta.get('barcode') or '')
    cod = str(etiqueta.get('codigo_wms') or etiqueta.get('codigo') or bc)
    bc_h = int(mid_h * 0.62)
    bc_w = int(w * 0.88)
    bc_img = _barcode_png(bc, bc_w, bc_h)
    bx = (w - bc_img.width) // 2
    by = top_h + max(4, (mid_h - bc_img.height) // 2 - h // 30)
    img.paste(bc_img, (bx, by))

    cod_font = _font(max(16, h // 22), bold=True)
    cod_y0 = by + bc_img.height + max(2, h // 80)
    _draw_centered(draw, (0, cod_y0, w, top_h + mid_h - 2), cod, cod_font)

    zona = str(etiqueta.get('zona') or ('PICKING' if etiqueta.get('picking') else 'PULMÃO')).upper()
    if etiqueta.get('destino_label'):
        zona = str(etiqueta['destino_label']).upper()
    zona_font = _font(max(28, h // 9), bold=True)
    _draw_centered(draw, (0, top_h + mid_h, int(w * 0.72), h), zona, zona_font)
    _draw_arrow_down(draw, int(w * 0.88), top_h + mid_h + bot_h // 2, max(16, bot_h // 3))

    out = io.BytesIO()
    img.save(out, format='PNG', dpi=(_DPI, _DPI))
    return out.getvalue()


def _sheet_title(etiqueta: dict, index: int) -> str:
    base = str(etiqueta.get('codigo_wms') or etiqueta.get('codigo') or etiqueta.get('barcode') or f'Etq{index + 1}')
    safe = re.sub(r'[\[\]\:\*\?\/\\]', '-', base)[:28]
    return safe or f'Etq{index + 1}'


def _apply_label_image(ws, png_bytes: bytes):
    ws._images = []
    xl_img = XLImage(io.BytesIO(png_bytes))
    z = ETIQUETA_ZEBRA_ZD220
    xl_img.width = int(z['largura_mm'] * 96 / 25.4)
    xl_img.height = int(z['altura_mm'] * 96 / 25.4)
    ws.add_image(xl_img, 'A1')


def gerar_workbook_longarina(etiquetas: list[dict]) -> bytes:
    """Monta .xlsx com uma aba por etiqueta (template ETIQUETA LONGARINA)."""
    if not etiquetas:
        raise ValueError('Nenhuma etiqueta para gerar.')
    if not os.path.isfile(TEMPLATE_PATH):
        raise FileNotFoundError(f'Template Excel não encontrado: {TEMPLATE_PATH}')

    wb = load_workbook(TEMPLATE_PATH)
    if 'Planilha1' in wb.sheetnames:
        del wb['Planilha1']

    base_name = 'Planilha2' if 'Planilha2' in wb.sheetnames else wb.sheetnames[0]
    base_ws = wb[base_name]

    for i, etiqueta in enumerate(etiquetas):
        if i == 0:
            ws = base_ws
        else:
            ws = wb.copy_worksheet(base_ws)
            ws.page_setup = copy(base_ws.page_setup)
            ws.page_margins = copy(base_ws.page_margins)
            ws.print_options = copy(base_ws.print_options)
        ws.title = _sheet_title(etiqueta, i)
        _apply_label_image(ws, render_longarina_png(etiqueta))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
