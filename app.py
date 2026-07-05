from flask import Flask, render_template, request, jsonify, send_file, Response, redirect, url_for, session, g, current_app
from datetime import datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import HTTPException
import sqlite3
import os
import sys
import time
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import uuid
import importlib.util
import re
import base64
import smtplib
import xml.etree.ElementTree as ET
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
from werkzeug.utils import secure_filename

try:
    from ravex_client import (
        get_token as ravex_get_token,
        obter_roteiro_por_id,
        obter_roteiro_por_periodo,
        obter_itens_roteiro,
        obter_itens_pedido,
        obter_viagem_por_id,
        obter_notas_fiscais_viagem,
        obter_itens_nota_fiscal,
        obter_canhotos_viagem,
        obter_ponto_referencia,
        obter_pedido_por_id,
        obter_veiculo_por_id,
        viagens_finalizadas_por_periodo,
    )
except ImportError:
    ravex_get_token = None
    obter_roteiro_por_id = None
    obter_roteiro_por_periodo = None
    obter_itens_roteiro = None
    obter_itens_pedido = None
    obter_viagem_por_id = None
    obter_notas_fiscais_viagem = None
    obter_itens_nota_fiscal = None
    obter_canhotos_viagem = None
    obter_ponto_referencia = None
    obter_pedido_por_id = None
    obter_veiculo_por_id = None
    viagens_finalizadas_por_periodo = None

try:
    from dotenv import load_dotenv  # type: ignore
except Exception:
    load_dotenv = None

try:
    import psycopg  # type: ignore
    from psycopg.rows import dict_row  # type: ignore
except Exception:
    psycopg = None
    dict_row = None

try:
    import requests as http_requests  # type: ignore
except Exception:
    http_requests = None

# Quando instalado (PyInstaller): dados em AppData; planilha pode ficar na pasta do exe
if getattr(sys, 'frozen', False):
    _BASE_DIR = os.path.join(os.environ.get('APPDATA', ''), 'ControleCarregamentoUltrapao')
    _EXE_DIR = os.path.dirname(sys.executable)
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    _EXE_DIR = _BASE_DIR
os.makedirs(_BASE_DIR, exist_ok=True)

if load_dotenv:
    try:
        load_dotenv()
    except Exception:
        pass

app = Flask(__name__)
app.config['SECRET_KEY'] = (os.environ.get('SECRET_KEY') or 'ultrapao-secret-key-2024')

SYSTEM_NAME = 'Stock System'
SYSTEM_TAGLINE = 'WMS · Gestão de estoque e armazenagem'


def _app_env():
    """producao | homologacao — controlado por APP_ENV ou nome do serviço no Render."""
    raw = (os.environ.get('APP_ENV') or '').strip().lower()
    if raw in ('homolog', 'homologacao', 'hml', 'staging', 'stg'):
        return 'homologacao'
    svc = (os.environ.get('RENDER_SERVICE_NAME') or '').strip().lower()
    if 'homolog' in svc or svc.endswith('-hml'):
        return 'homologacao'
    return 'producao'


def _app_env_url():
    return (os.environ.get('RENDER_EXTERNAL_URL') or '').strip() or None


@app.context_processor
def _inject_stock_system_branding():
    env = _app_env()
    return {
        'system_name': SYSTEM_NAME,
        'system_tagline': SYSTEM_TAGLINE,
        'app_env': env,
        'app_env_url': _app_env_url() if env == 'homologacao' else None,
        'is_homologacao': env == 'homologacao',
    }

# Atualização em tempo real: quando alguém bipa, todos os clientes (127.0.0.1 e 192.168.x.x) recebem e atualizam
_sse_queues = []
_sse_lock = threading.Lock()

def _broadcast_atualizar():
    """Notifica todos os clientes e envia totais para atualização imediata (sem esperar novo GET)."""
    def _emit():
        try:
            conn = get_db()
            row = conn.execute(
                'SELECT COUNT(*) as c, COALESCE(SUM(quantidade), 0) as s FROM produtos_bipados'
            ).fetchone()
            conn.close()
            payload = json.dumps({
                't': 'atualizar',
                'total_bipados': row['c'] if row else 0,
                'soma_quantidades': int(row['s']) if row else 0
            })
        except Exception:
            payload = 'atualizar'
        with _sse_lock:
            for q in _sse_queues:
                try:
                    q.put_nowait(payload)
                except Exception:
                    pass

    try:
        threading.Thread(target=_emit, daemon=True).start()
    except Exception:
        _emit()

# Configuração do banco de dados e upload (pasta gravável)
DB_NAME = os.path.join(_BASE_DIR, 'controle_carregamento.db')
UPLOAD_FOLDER = os.path.join(_BASE_DIR, 'uploads')
ALLOWED_SPREADSHEET_EXTENSIONS = {'xlsx', 'xls'}
ALLOWED_XML_EXTENSIONS = {'xml'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def init_db():
    """Inicializa o banco de dados"""
    conn = get_db()
    try:
        if getattr(conn, 'kind', 'sqlite') == 'pg':
            # Postgres (Supabase) — schema simplificado (completo está em supabase/schema.sql)
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.usuarios (
                    id BIGSERIAL PRIMARY KEY,
                    usuario TEXT NOT NULL UNIQUE,
                    senha_hash TEXT NOT NULL,
                    ativo BOOLEAN NOT NULL DEFAULT TRUE,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.colaboradores (
                    id BIGSERIAL PRIMARY KEY,
                    nome TEXT NOT NULL,
                    funcao TEXT,
                    centro_custo TEXT,
                    tipo TEXT,
                    cpf TEXT,
                    telefone TEXT,
                    email TEXT,
                    ativo BOOLEAN NOT NULL DEFAULT TRUE,
                    observacoes TEXT,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    criado_por TEXT,
                    atualizado_por TEXT
                )'''
            )
            conn.execute('CREATE INDEX IF NOT EXISTS idx_colaboradores_nome ON public.colaboradores (nome) WHERE ativo = TRUE')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_colaboradores_tipo ON public.colaboradores (tipo) WHERE ativo = TRUE AND tipo IS NOT NULL')
            conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_colaboradores_cpf ON public.colaboradores (cpf) WHERE cpf IS NOT NULL AND cpf != \'\'')
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.motoristas (
                    id BIGSERIAL PRIMARY KEY,
                    nome TEXT NOT NULL UNIQUE,
                    cpf TEXT,
                    cnh TEXT,
                    categoria_cnh TEXT,
                    validade_cnh DATE,
                    telefone TEXT,
                    email TEXT,
                    centro_custo TEXT,
                    ativo BOOLEAN NOT NULL DEFAULT TRUE,
                    observacoes TEXT,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    criado_por TEXT,
                    atualizado_por TEXT
                )'''
            )
            conn.execute('CREATE INDEX IF NOT EXISTS idx_motoristas_nome ON public.motoristas (nome) WHERE ativo = TRUE')
            conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_motoristas_cpf ON public.motoristas (cpf) WHERE cpf IS NOT NULL AND cpf != \'\'')
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.placas (
                    id BIGSERIAL PRIMARY KEY,
                    placa TEXT NOT NULL UNIQUE,
                    descricao TEXT,
                    tipo_veiculo TEXT,
                    capacidade_kg NUMERIC(10,2),
                    ano INTEGER,
                    marca TEXT,
                    modelo TEXT,
                    ativo BOOLEAN NOT NULL DEFAULT TRUE,
                    observacoes TEXT,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    criado_por TEXT,
                    atualizado_por TEXT
                )'''
            )
            conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_placas_placa ON public.placas (placa)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_placas_ativo ON public.placas (ativo) WHERE ativo = TRUE')

            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.produtos_bipados (
                    id BIGSERIAL PRIMARY KEY,
        codigo_barras TEXT NOT NULL,
                    codigo_interno TEXT,
                    codigo_dun TEXT,
        produto TEXT NOT NULL,
                    quantidade INTEGER NOT NULL CHECK (quantidade >= 1),
                    unidade TEXT,
                    peso TEXT,
                    id_viagem TEXT NOT NULL,
                    doca TEXT CHECK (doca IN ('1', '2', '3', '4')),
        veiculo TEXT,
                    status TEXT NOT NULL DEFAULT 'PENDENTE' CHECK (status IN ('PENDENTE', 'CARREGADO', 'CANCELADO')),
                    data_hora TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    usuario_bipagem TEXT,
                    fluxo TEXT DEFAULT 'carregamento',
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.viagem_periodo_bipagem (
                    id_viagem TEXT NOT NULL,
                    fluxo TEXT NOT NULL DEFAULT 'carregamento',
                    inicio_em TIMESTAMPTZ NOT NULL,
                    fim_em TIMESTAMPTZ NOT NULL,
                    PRIMARY KEY (id_viagem, fluxo)
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.viagem_placa (
                    id_viagem TEXT PRIMARY KEY,
                    placa TEXT NOT NULL,
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_por TEXT
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.viagem_motorista (
                    id_viagem TEXT PRIMARY KEY,
                    motorista TEXT NOT NULL,
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_por TEXT
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.romaneio (
                    id BIGSERIAL PRIMARY KEY,
                    codigo_barras TEXT NOT NULL UNIQUE,
                    quantidade_romaneio INTEGER NOT NULL DEFAULT 0,
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.viagem_responsaveis (
                    id_viagem TEXT PRIMARY KEY,
                    coordenador TEXT NOT NULL DEFAULT 'ASTROGILDO RODRIGUES DOS SANTOS',
                    conferente TEXT,
                    ajudante1 TEXT,
                    ajudante2 TEXT,
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_por TEXT
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.divergencia_motivo (
                    id_viagem TEXT NOT NULL,
                    codigo_produto TEXT NOT NULL,
                    motivo TEXT,
                    registrado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    registrado_por TEXT,
                    PRIMARY KEY (id_viagem, codigo_produto)
                )'''
            )
            conn.execute('CREATE INDEX IF NOT EXISTS idx_divergencia_motivo_id_viagem ON public.divergencia_motivo (id_viagem)')
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.terceiros_documentos (
                    id BIGSERIAL PRIMARY KEY,
                    area TEXT NOT NULL CHECK (area IN ('recebimento', 'expedicao', 'carreta')),
                    chave_nfe TEXT,
                    numero_nf TEXT,
                    serie_nf TEXT,
                    data_emissao TEXT,
                    remetente_nome TEXT,
                    remetente_cnpj TEXT,
                    destinatario_nome TEXT,
                    destinatario_cnpj TEXT,
                    destinatario_uf TEXT,
                    previsao_chegada TEXT,
                    arquivo_nome TEXT,
                    xml_conteudo TEXT,
                    recebimento_concluido BOOLEAN NOT NULL DEFAULT FALSE,
                    recebimento_concluido_em TIMESTAMPTZ,
                    recebimento_concluido_por TEXT,
                    nota_lancada TEXT,
                    nota_lancada_em TIMESTAMPTZ,
                    nota_lancada_por TEXT,
                    enviar_para_mg TEXT,
                    enviar_para_mg_em TIMESTAMPTZ,
                    enviar_para_mg_por TEXT,
                    motorista_carreta TEXT,
                    motorista_carreta_em TIMESTAMPTZ,
                    placa_carreta TEXT,
                    motorista_saida_mg TEXT,
                    motorista_saida_mg_em TIMESTAMPTZ,
                    placa_saida_mg TEXT,
                    carga_recebida_mg TEXT,
                    carga_recebida_mg_em TIMESTAMPTZ,
                    carga_recebida_mg_por TEXT,
                    recebedor_mg TEXT,
                    consumivel_sp TEXT,
                    recebedor_consumivel_sp TEXT,
                    consumivel_sp_historico TEXT,
                    consumivel_sp_historico_em TIMESTAMPTZ,
                    consumivel_sp_historico_por TEXT,
                    motivo_nao_lancada TEXT,
                    motivo_nao_enviar_mg TEXT,
                    motivo_nao_recebida_mg TEXT,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    criado_por TEXT,
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_por TEXT
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.terceiros_documento_itens (
                    id BIGSERIAL PRIMARY KEY,
                    documento_id BIGINT NOT NULL REFERENCES public.terceiros_documentos(id) ON DELETE CASCADE,
                    n_item INTEGER,
                    codigo_ean TEXT,
                    codigo_produto_xml TEXT,
                    descricao_xml TEXT,
                    unidade_xml TEXT,
                    quantidade_xml NUMERIC(14,3) NOT NULL DEFAULT 0,
                    codigo_produto_base TEXT,
                    codigo_barras_base TEXT,
                    descricao_base TEXT,
                    quantidade_bipada NUMERIC(14,3) NOT NULL DEFAULT 0,
                    status_bipagem TEXT NOT NULL DEFAULT 'PENDENTE',
                    ultimo_ean_bipado TEXT,
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_por TEXT
                )'''
            )
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.terceiros_documento_eventos (
                    id BIGSERIAL PRIMARY KEY,
                    documento_id BIGINT NOT NULL REFERENCES public.terceiros_documentos(id) ON DELETE CASCADE,
                    evento TEXT NOT NULL,
                    valor_anterior TEXT,
                    valor_novo TEXT,
                    usuario TEXT,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    detalhes TEXT
                )'''
            )
            # Índices
            conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_id_viagem ON public.produtos_bipados(id_viagem)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_codigo ON public.produtos_bipados(codigo_barras)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_viagem_codigo ON public.produtos_bipados(id_viagem, codigo_barras)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_viagem_fluxo ON public.produtos_bipados(id_viagem, fluxo)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_area ON public.terceiros_documentos(area, criado_em DESC)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_chave ON public.terceiros_documentos(chave_nfe)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_documento ON public.terceiros_documento_itens(documento_id)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_ean ON public.terceiros_documento_itens(codigo_ean)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_eventos_documento ON public.terceiros_documento_eventos(documento_id, criado_em DESC)')
            try:
                conn.execute("ALTER TABLE public.produtos_bipados ADD COLUMN IF NOT EXISTS fluxo TEXT DEFAULT 'carregamento'")
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
            try:
                conn.execute('ALTER TABLE public.romaneio_por_item ADD COLUMN identificador_rota TEXT')
            except Exception:
                conn.rollback()
            # Coluna importado_em no romaneio_por_item (para auditoria de importação Ravex)
            try:
                conn.execute('ALTER TABLE public.romaneio_por_item ADD COLUMN IF NOT EXISTS importado_em TIMESTAMPTZ')
            except Exception:
                conn.rollback()
            # Histórico de importações Ravex
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.ravex_importacoes (
                    id BIGSERIAL PRIMARY KEY,
                    dataset_id uuid REFERENCES public.excel_datasets(dataset_id) ON DELETE SET NULL,
                    tipo TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'OK',
                    parametros JSONB,
                    viagens_processadas INTEGER NOT NULL DEFAULT 0,
                    total_itens INTEGER NOT NULL DEFAULT 0,
                    usuario TEXT,
                    erros JSONB,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )'''
            )
            conn.execute('CREATE INDEX IF NOT EXISTS idx_ravex_importacoes_dataset ON public.ravex_importacoes (dataset_id)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_ravex_importacoes_criado_em ON public.ravex_importacoes (criado_em DESC)')
            # Tabela id_roteiros (registro dos roteiros vindos da API por período)
            conn.execute(
                '''CREATE TABLE IF NOT EXISTS public.id_roteiros (
                    dataset_id uuid NOT NULL REFERENCES public.excel_datasets(dataset_id) ON DELETE CASCADE,
                    id_roteiro TEXT NOT NULL,
                    id_viagem TEXT NOT NULL,
                    identificador_rota TEXT,
                    data_viagem TIMESTAMPTZ,
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    PRIMARY KEY (dataset_id, id_roteiro)
                )'''
            )
            conn.execute('CREATE INDEX IF NOT EXISTS idx_id_roteiros_dataset ON public.id_roteiros (dataset_id)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_id_roteiros_id_viagem ON public.id_roteiros (id_viagem) WHERE id_viagem IS NOT NULL')
            try:
                conn.execute('CREATE INDEX IF NOT EXISTS idx_romaneio_por_item_dataset_viagem ON public.romaneio_por_item (dataset_id, id_viagem)')
                conn.execute('CREATE INDEX IF NOT EXISTS idx_romaneio_por_item_dataset_roteiro ON public.romaneio_por_item (dataset_id, id_roteiro)')
                conn.execute('CREATE INDEX IF NOT EXISTS idx_romaneio_por_item_row_index ON public.romaneio_por_item (dataset_id, row_index)')
            except Exception:
                conn.rollback()
            try:
                _ensure_pg_tabela_geral_dados(conn)
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
            try:
                _ensure_terceiros_schema(conn, rodar_backfill=False)
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
            # Postgres: sem commit o DDL era revertido ao fechar a conexão (coluna fluxo sumia).
            try:
                conn.commit()
            except Exception:
                pass
            return

        # SQLite (local)
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS produtos_bipados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_barras TEXT NOT NULL,
        produto TEXT NOT NULL,
        quantidade INTEGER NOT NULL,
        data_hora TEXT NOT NULL,
        veiculo TEXT,
        status TEXT DEFAULT 'PENDENTE',
        id_viagem TEXT
            )'''
        )
        # Migrações SQLite (ADD COLUMN se não existir)
        for col_sql in (
            'ALTER TABLE produtos_bipados ADD COLUMN doca TEXT',
            'ALTER TABLE produtos_bipados ADD COLUMN codigo_interno TEXT',
            'ALTER TABLE produtos_bipados ADD COLUMN codigo_dun TEXT',
            'ALTER TABLE produtos_bipados ADD COLUMN unidade TEXT',
            'ALTER TABLE produtos_bipados ADD COLUMN peso TEXT',
            'ALTER TABLE produtos_bipados ADD COLUMN usuario_bipagem TEXT',
            "ALTER TABLE produtos_bipados ADD COLUMN fluxo TEXT DEFAULT 'carregamento'",
        ):
            try:
                conn.execute(col_sql)
                conn.commit()
            except Exception:
                pass

        conn.execute(
            '''CREATE TABLE IF NOT EXISTS romaneio (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_barras TEXT NOT NULL UNIQUE,
        quantidade_romaneio INTEGER NOT NULL DEFAULT 0
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS viagem_placa (
        id_viagem TEXT PRIMARY KEY,
                placa TEXT NOT NULL,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS viagem_motorista (
        id_viagem TEXT PRIMARY KEY,
                motorista TEXT NOT NULL,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS viagem_responsaveis (
        id_viagem TEXT PRIMARY KEY,
        coordenador TEXT NOT NULL,
        conferente TEXT NOT NULL,
        ajudante1 TEXT NOT NULL,
        ajudante2 TEXT NOT NULL
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS divergencia_motivo (
        id_viagem TEXT NOT NULL,
        codigo_produto TEXT NOT NULL,
        motivo TEXT,
        PRIMARY KEY (id_viagem, codigo_produto)
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS viagem_periodo_bipagem (
                id_viagem TEXT NOT NULL,
                fluxo TEXT NOT NULL DEFAULT 'carregamento',
                inicio_em TEXT NOT NULL,
                fim_em TEXT NOT NULL,
                PRIMARY KEY (id_viagem, fluxo)
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT NOT NULL UNIQUE,
        senha_hash TEXT NOT NULL,
        criado_em TEXT
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS colaboradores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                funcao TEXT,
                centro_custo TEXT,
                tipo TEXT,
                cpf TEXT UNIQUE,
                telefone TEXT,
                email TEXT,
                ativo INTEGER NOT NULL DEFAULT 1,
                observacoes TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS motoristas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE,
                cpf TEXT UNIQUE,
                cnh TEXT,
                categoria_cnh TEXT,
                validade_cnh TEXT,
                telefone TEXT,
                email TEXT,
                centro_custo TEXT,
                ativo INTEGER NOT NULL DEFAULT 1,
                observacoes TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS placas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                placa TEXT NOT NULL UNIQUE,
                descricao TEXT,
                tipo_veiculo TEXT,
                capacidade_kg REAL,
                ano INTEGER,
                marca TEXT,
                modelo TEXT,
                ativo INTEGER NOT NULL DEFAULT 1,
                observacoes TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS terceiros_documentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                area TEXT NOT NULL,
                chave_nfe TEXT,
                numero_nf TEXT,
                serie_nf TEXT,
                data_emissao TEXT,
                remetente_nome TEXT,
                remetente_cnpj TEXT,
                destinatario_nome TEXT,
                destinatario_cnpj TEXT,
                destinatario_uf TEXT,
                previsao_chegada TEXT,
                arquivo_nome TEXT,
                xml_conteudo TEXT,
                recebimento_concluido INTEGER NOT NULL DEFAULT 0,
                recebimento_concluido_em TEXT,
                recebimento_concluido_por TEXT,
                nota_lancada TEXT,
                nota_lancada_em TEXT,
                nota_lancada_por TEXT,
                enviar_para_mg TEXT,
                enviar_para_mg_em TEXT,
                enviar_para_mg_por TEXT,
                motorista_carreta TEXT,
                motorista_carreta_em TEXT,
                placa_carreta TEXT,
                motorista_saida_mg TEXT,
                motorista_saida_mg_em TEXT,
                placa_saida_mg TEXT,
                carga_recebida_mg TEXT,
                carga_recebida_mg_em TEXT,
                carga_recebida_mg_por TEXT,
                recebedor_mg TEXT,
                consumivel_sp TEXT,
                recebedor_consumivel_sp TEXT,
                consumivel_sp_historico TEXT,
                consumivel_sp_historico_em TEXT,
                consumivel_sp_historico_por TEXT,
                motivo_nao_lancada TEXT,
                motivo_nao_enviar_mg TEXT,
                motivo_nao_recebida_mg TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                criado_por TEXT,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_por TEXT
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS terceiros_documento_itens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                documento_id INTEGER NOT NULL,
                n_item INTEGER,
                codigo_ean TEXT,
                codigo_produto_xml TEXT,
                descricao_xml TEXT,
                unidade_xml TEXT,
                quantidade_xml REAL NOT NULL DEFAULT 0,
                codigo_produto_base TEXT,
                codigo_barras_base TEXT,
                descricao_base TEXT,
                quantidade_bipada REAL NOT NULL DEFAULT 0,
                status_bipagem TEXT NOT NULL DEFAULT 'PENDENTE',
                ultimo_ean_bipado TEXT,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_por TEXT,
                FOREIGN KEY (documento_id) REFERENCES terceiros_documentos(id) ON DELETE CASCADE
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS terceiros_documento_eventos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                documento_id INTEGER NOT NULL,
                evento TEXT NOT NULL,
                valor_anterior TEXT,
                valor_novo TEXT,
                usuario TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                detalhes TEXT,
                FOREIGN KEY (documento_id) REFERENCES terceiros_documentos(id) ON DELETE CASCADE
            )'''
        )
        conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_id_viagem ON produtos_bipados(id_viagem)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_colaboradores_nome ON colaboradores(nome) WHERE ativo = 1')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_colaboradores_tipo ON colaboradores(tipo) WHERE ativo = 1 AND tipo IS NOT NULL')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_motoristas_nome ON motoristas(nome) WHERE ativo = 1')
        conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_placas_placa ON placas(placa)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_codigo ON produtos_bipados(codigo_barras)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_produtos_bipados_viagem_codigo ON produtos_bipados(id_viagem, codigo_barras)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_area ON terceiros_documentos(area, criado_em DESC)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_chave ON terceiros_documentos(chave_nfe)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_documento ON terceiros_documento_itens(documento_id)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_ean ON terceiros_documento_itens(codigo_ean)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_eventos_documento ON terceiros_documento_eventos(documento_id, criado_em DESC)')
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass


def sync_usuarios_from_config():
    """
    Sincroniza a tabela 'usuarios' com o arquivo config_usuarios.py.
    - Usuários listados no arquivo: inseridos ou atualizados (senha em hash).
    - Usuários que não estão no arquivo: removidos do banco.
    Se o arquivo não existir, é criado com o usuário admin/admin.
    """
    config_path = os.path.join(_EXE_DIR, 'config_usuarios.py')
    default_content = '''# -*- coding: utf-8 -*-
# Controle de usuários. Edite a lista USUARIOS e reinicie o app.
# Formato: {"usuario": "nome", "senha": "senha_em_texto"}
USUARIOS = [
    {"usuario": "admin", "senha": "admin"},
]
'''
    if not os.path.isfile(config_path):
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                f.write(default_content)
        except Exception:
            pass
    lista = []
    if os.path.isfile(config_path):
        try:
            spec = importlib.util.spec_from_file_location('config_usuarios', config_path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            lista = getattr(mod, 'USUARIOS', [])
            if not isinstance(lista, list):
                lista = []
        except Exception:
            lista = []
    if not lista:
        lista = [{'usuario': 'admin', 'senha': 'admin'}]
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) == 'pg':
            try:
                _ensure_pg_tabela_geral_dados(conn)
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
        usuarios_do_arquivo = set()
        for item in lista:
            if not isinstance(item, dict):
                continue
            usuario = (item.get('usuario') or '').strip()
            senha = item.get('senha') or ''
            if not usuario:
                continue
            usuarios_do_arquivo.add(usuario)
            senha_hash = generate_password_hash(senha, method='pbkdf2:sha256')
            conn.execute(
                '''INSERT INTO usuarios (usuario, senha_hash, criado_em) VALUES (?, ?, ?)
                   ON CONFLICT(usuario) DO UPDATE SET senha_hash = excluded.senha_hash''',
                (usuario, senha_hash, datetime.now().isoformat())
            )
        conn.commit()
        # Remove do banco usuários que não estão mais no arquivo (excluídos no config = não podem mais logar)
        remover = [row['usuario'] for row in conn.execute('SELECT usuario FROM usuarios') if row['usuario'] not in usuarios_do_arquivo]
        for u in remover:
            conn.execute('DELETE FROM usuarios WHERE usuario = ?', (u,))
        conn.commit()
    finally:
        conn.close()


def adicionar_usuario_ao_config(usuario, senha):
    """
    Adiciona um novo usuário/senha ao arquivo config_usuarios.py (atualiza o arquivo
    quando alguém se cadastra pela tela de Cadastrar).
    """
    config_path = os.path.join(_EXE_DIR, 'config_usuarios.py')
    if not os.path.isfile(config_path):
        return
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        # Evitar duplicata: se o usuário já está no arquivo, não adicionar de novo
        for line in lines:
            if '"usuario"' in line and json.dumps(usuario) in line:
                return
        insert_at = -1
        in_list = False
        for i, line in enumerate(lines):
            if 'USUARIOS' in line and '=' in line and '[' in line:
                in_list = True
                continue
            if in_list and line.strip() == ']':
                insert_at = i
                break
        if insert_at < 0:
            return
        # Garantir vírgula na linha anterior
        prev = insert_at - 1
        while prev >= 0 and not lines[prev].strip():
            prev -= 1
        if prev >= 0 and lines[prev].rstrip() and not lines[prev].rstrip().endswith(','):
            lines[prev] = lines[prev].rstrip()
            if not lines[prev].endswith('\n'):
                lines[prev] += '\n'
            lines[prev] = lines[prev].rstrip().rstrip(',') + ',\n'
        nova_linha = '    {"usuario": %s, "senha": %s},\n' % (json.dumps(usuario), json.dumps(senha))
        lines.insert(insert_at, nova_linha)
        with open(config_path, 'w', encoding='utf-8') as f:
            f.writelines(lines)
    except Exception:
        pass


def get_db():
    """Retorna conexão com o banco de dados. Timeout para suportar 4 docas bipando ao mesmo tempo."""
    db_url = (os.environ.get('DATABASE_URL') or '').strip()
    if db_url:
        if psycopg is None:
            raise RuntimeError('psycopg não instalado. Instale as dependências do requirements.txt.')
        # Supabase Postgres normalmente exige SSL
        return CompatConn(
            psycopg.connect(
                db_url,
                sslmode=os.environ.get('PGSSLMODE', 'require'),
                connect_timeout=15,
                row_factory=dict_row,
                options='-c statement_timeout=60000 -c lock_timeout=15000',
            ),
            kind='pg',
        )
    conn = sqlite3.connect(DB_NAME, timeout=15.0)
    conn.row_factory = sqlite3.Row
    return CompatConn(conn, kind='sqlite')


class CompatConn:
    """
    Compatibilidade SQLite/Postgres:
    - SQLite usa '?' como placeholder
    - Postgres (psycopg) usa '%s'
    Este wrapper permite manter as queries atuais.
    """

    def __init__(self, conn, *, kind: str):
        self._conn = conn
        self._kind = kind
    
    @property
    def kind(self) -> str:
        return self._kind

    def execute(self, sql, params=()):
        if self._kind == 'pg' and sql and '?' in sql:
            sql = sql.replace('?', '%s')
        return self._conn.execute(sql, params or ())

    def executemany(self, sql, params_seq):
        if self._kind == 'pg' and sql and '?' in sql:
            sql = sql.replace('?', '%s')
        params_seq = list(params_seq or [])
        if not params_seq:
            return None
        if self._kind == 'pg':
            with self._conn.cursor() as cur:
                cur.executemany(sql, params_seq)
            return None
        return self._conn.executemany(sql, params_seq)

    def commit(self):
        return self._conn.commit()

    def rollback(self):
        if getattr(self._conn, 'rollback', None):
            self._conn.rollback()

    def close(self):
        return self._conn.close()


try:
    from wms_enderecamento import bp as wms_enderecamento_bp, init_wms_enderecamento, register_wms_db
    from wms_etiqueta_zebra import ctx_etiqueta_zebra
    register_wms_db(get_db)
    app.register_blueprint(wms_enderecamento_bp, url_prefix='/api/wms')

    @app.context_processor
    def _inject_etiqueta_zebra_config():
        return ctx_etiqueta_zebra()

    # Schema WMS é preparado sob demanda nas rotas /api/wms/* (_ensure_wms_schema_safe).
    # Init em background competia com init_db no Postgres e podia travar o worker no Render.
    if os.environ.get('WMS_INIT_BG', '').strip() == '1':
        def _init_wms_em_background():
            try:
                init_wms_enderecamento(get_db)
            except Exception as e:
                import traceback
                try:
                    print('[controle-carregamento] init_wms_enderecamento falhou:', e, flush=True)
                    traceback.print_exc()
                except Exception:
                    pass

        threading.Thread(target=_init_wms_em_background, daemon=True, name='init_wms').start()
except Exception as _wms_import_err:
    print('Aviso: módulo WMS endereçamento não carregado:', _wms_import_err)


def _ensure_pg_tabela_geral_dados(conn):
    """
    Triggers fn_sync_tabela_geral_dados no Supabase inserem nesta tabela em cada DML.
    Sem ela, sync_usuarios e outras gravações falham no boot do Render.
    """
    if getattr(conn, 'kind', None) != 'pg':
        return
    conn.execute(
        '''CREATE TABLE IF NOT EXISTS public.tabela_geral_dados (
            id BIGSERIAL PRIMARY KEY,
            fonte_tabela TEXT NOT NULL,
            row_id TEXT,
            acao TEXT NOT NULL,
            dados JSONB,
            criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )'''
    )
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_tabela_geral_dados_fonte '
        'ON public.tabela_geral_dados (fonte_tabela, criado_em DESC)'
    )
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_tabela_geral_dados_row '
        'ON public.tabela_geral_dados (fonte_tabela, row_id)'
    )
    conn.execute(
        '''CREATE TABLE IF NOT EXISTS public.tabela_geral_snapshot (
            fonte_tabela TEXT NOT NULL,
            row_id TEXT NOT NULL,
            dados JSONB,
            atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (fonte_tabela, row_id)
        )'''
    )


def _pg_auditoria_sync_desligar(conn):
    """Evita triggers fn_sync_tabela_geral_* bloquearem UPDATE/INSERT (ex.: concluir recebimento)."""
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute("SET LOCAL session_replication_role = replica")


_SCHEMA_ENSURE_DONE = set()


def _ensure_pg_produtos_bipados_fluxo(conn):
    """Garante coluna fluxo em produtos_bipados (Postgres). Migrações antigas sem commit perdiam o ALTER."""
    if getattr(conn, 'kind', None) != 'pg':
        return
    if 'pg_produtos_bipados_fluxo' in _SCHEMA_ENSURE_DONE:
        return
    try:
        conn.execute(
            "ALTER TABLE public.produtos_bipados ADD COLUMN IF NOT EXISTS fluxo TEXT DEFAULT 'carregamento'"
        )
        conn.commit()
        _SCHEMA_ENSURE_DONE.add('pg_produtos_bipados_fluxo')
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


DISPOSICOES_ESTOQUE_SP = frozenset({'reentrega', 'avaria', 'descarte_perdas', 'palete_bloqueado'})


def _ensure_produtos_bipados_disposicao(conn):
    """Coluna disposicao_estoque: reentrega, avaria, descarte_perdas, palete_bloqueado."""
    is_pg = getattr(conn, 'kind', None) == 'pg'
    cache_key = 'pb_disposicao_pg' if is_pg else 'pb_disposicao_sqlite'
    if cache_key in _SCHEMA_ENSURE_DONE:
        return
    tbl = 'public.produtos_bipados' if is_pg else 'produtos_bipados'
    try:
        if is_pg:
            conn.execute(f'ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS disposicao_estoque TEXT')
            conn.commit()
        else:
            cols = [r[1] for r in conn.execute('PRAGMA table_info(produtos_bipados)').fetchall()]
            if 'disposicao_estoque' not in cols:
                conn.execute('ALTER TABLE produtos_bipados ADD COLUMN disposicao_estoque TEXT')
                conn.commit()
        _SCHEMA_ENSURE_DONE.add(cache_key)
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


def _normalizar_disposicao_estoque(val, fluxo=None, devolucao_nf_id=None, conn=None):
    raw = (val or '').strip().lower()
    if raw in ('', 'normal', 'none', 'null'):
        disp = None
    elif raw in DISPOSICOES_ESTOQUE_SP:
        disp = raw
    else:
        disp = None
    if disp is None and fluxo == 'devolucao' and devolucao_nf_id and conn:
        try:
            row = conn.execute(
                'SELECT motivo FROM devolucao_nota_fiscal WHERE id = ?',
                (int(devolucao_nf_id),),
            ).fetchone()
            rd = _row_dict(row) if row else {}
            if str(rd.get('motivo') or '').lower() == 'reentrega':
                disp = 'reentrega'
        except Exception:
            pass
    return disp


def _estoque_sp_sql_apenas_disposicao_normal(prefix=''):
    col = f'{prefix}disposicao_estoque' if prefix else 'disposicao_estoque'
    return (
        f" AND (COALESCE(NULLIF(TRIM({col}), ''), 'normal') = 'normal' "
        f"OR {col} IS NULL) "
    )


def _parse_hora_hhmm(val, default_h=0, default_m=0):
    if not val or not str(val).strip():
        return default_h, default_m, ''
    m = re.match(r'^(\d{1,2}):(\d{2})', str(val).strip())
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        return h, mi, f'{h:02d}:{mi:02d}'
    return default_h, default_m, ''


def _parse_janela_datahora(data_inicio=None, data_fim=None, hora_inicio=None, hora_fim=None, data_legacy=None):
    """Monta janela [início, fim) com data/hora de início e data/hora de fim (America/Sao_Paulo)."""
    tz = _ravex_tz_baixados()
    di = (data_inicio or data_legacy or '').strip()[:10]
    df = (data_fim or '').strip()[:10]
    if not di and not df:
        dia = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
        return {
            'inicio': dia,
            'fim_excl': dia + timedelta(days=1),
            'data_inicio_iso': dia.strftime('%Y-%m-%d'),
            'data_fim_iso': dia.strftime('%Y-%m-%d'),
            'data_iso': dia.strftime('%Y-%m-%d'),
            'data': dia.strftime('%d/%m/%Y'),
            'hora_inicio': '',
            'hora_fim': '',
            'legivel': dia.strftime('%d/%m/%Y') + ' (dia inteiro)',
        }
    if di and not df:
        df = di
    if df and not di:
        di = df
    try:
        d_ini = datetime.strptime(di, '%Y-%m-%d').replace(tzinfo=tz)
    except ValueError:
        d_ini = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
        di = d_ini.strftime('%Y-%m-%d')
    try:
        d_fim = datetime.strptime(df, '%Y-%m-%d').replace(tzinfo=tz)
    except ValueError:
        d_fim = d_ini
        df = di
    tem_hi = bool(hora_inicio and str(hora_inicio).strip())
    tem_hf = bool(hora_fim and str(hora_fim).strip())
    hi, mi, hora_ini_out = _parse_hora_hhmm(hora_inicio, 0, 0)
    hf, mf, hora_fim_out = _parse_hora_hhmm(hora_fim, 23, 59)
    if not tem_hi:
        hora_ini_out = ''
    if not tem_hf:
        hora_fim_out = ''
        hf, mf = 23, 59
    inicio = d_ini.replace(hour=hi, minute=mi, second=0, microsecond=0)
    dia_fim = d_fim
    if di == df and tem_hi and tem_hf and (hf * 60 + mf) <= (hi * 60 + mi):
        dia_fim = d_fim + timedelta(days=1)
    fim_excl = dia_fim.replace(hour=hf, minute=mf, second=59, microsecond=999999) + timedelta(microseconds=1)
    legivel = d_ini.strftime('%d/%m/%Y') + ' · ' + (hora_ini_out or '00:00')
    legivel += ' → ' + dia_fim.strftime('%d/%m/%Y') + ' · ' + (hora_fim_out or '23:59')
    return {
        'inicio': inicio,
        'fim_excl': fim_excl,
        'data_inicio_iso': di,
        'data_fim_iso': df,
        'data_iso': di,
        'data': d_ini.strftime('%d/%m/%Y'),
        'hora_inicio': hora_ini_out,
        'hora_fim': hora_fim_out,
        'legivel': legivel,
    }


def _estoque_sp_parse_filtros_query():
    """Filtros comuns das abas Estoque SP (query string)."""
    return {
        'data_inicio': (request.args.get('data_inicio') or '').strip(),
        'data_fim': (request.args.get('data_fim') or '').strip(),
        'hora_inicio': (request.args.get('hora_inicio') or '').strip(),
        'hora_fim': (request.args.get('hora_fim') or '').strip(),
        'codigo_produto': (request.args.get('codigo_produto') or '').strip(),
        'codigo_barras': (request.args.get('codigo_barras') or '').strip(),
        'produto': (request.args.get('produto') or '').strip(),
    }


def _estoque_sp_bounds_datahora(filtros, conn):
    """Janela [início, fim) para filtro de data/hora em movimentações."""
    f = filtros or {}
    meta = _parse_janela_datahora(
        f.get('data_inicio'),
        f.get('data_fim'),
        f.get('hora_inicio'),
        f.get('hora_fim'),
    )
    if not (f.get('data_inicio') or f.get('data_fim') or f.get('hora_inicio') or f.get('hora_fim')):
        return None, None
    inicio = meta['inicio']
    fim_excl = meta['fim_excl']
    if getattr(conn, 'kind', None) == 'pg':
        return inicio, fim_excl
    tz = _ravex_tz_baixados()
    ini_s = inicio.astimezone(tz).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
    fim_s = fim_excl.astimezone(tz).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
    return ini_s, fim_s


def _estoque_sp_sql_filtro_data_hora(col, dt_ini, dt_fim, conn):
    """SQL extra + params para coluna de data/hora (fim exclusivo no PG)."""
    if not dt_ini and not dt_fim:
        return '', []
    is_pg = getattr(conn, 'kind', None) == 'pg'
    parts = []
    params = []
    if dt_ini is not None:
        parts.append(f' AND {col} >= ?')
        params.append(dt_ini)
    if dt_fim is not None:
        parts.append(f' AND {col} < ?' if is_pg else f' AND {col} <= ?')
        if is_pg:
            params.append(dt_fim)
        else:
            fim_txt = str(dt_fim).replace('T', ' ')[:19]
            params.append(fim_txt)
    return ''.join(parts), params


def _estoque_sp_filtrar_itens_lista(itens, filtros):
    """Filtro textual em listas já agregadas (código, barras, produto)."""
    cp = ((filtros or {}).get('codigo_produto') or '').strip().lower()
    cb = ((filtros or {}).get('codigo_barras') or '').strip().lower()
    pr = ((filtros or {}).get('produto') or '').strip().lower()
    if not cp and not cb and not pr:
        return itens
    out = []
    for it in itens or []:
        if cp and cp not in (it.get('codigo_produto') or '').lower():
            continue
        if cb and cb not in (it.get('codigo_barras') or '').lower():
            continue
        if pr and pr not in (it.get('produto') or '').lower():
            continue
        out.append(it)
    return out


def _estoque_sp_prepare_conn(conn):
    """Garante schema necessário; recupera transação PG após falha parcial."""
    for fn in (
        _ensure_devolucao_nf_schema,
        _ensure_pg_produtos_bipados_fluxo,
        _ensure_produtos_bipados_disposicao,
        _ensure_terceiros_schema,
    ):
        try:
            fn(conn)
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            fn(conn)


def _ensure_devolucao_nf_schema(conn):
    """Tabela de NF de devolução e vínculo em produtos_bipados."""
    is_pg = getattr(conn, 'kind', None) == 'pg'
    tbl_nf = 'public.devolucao_nota_fiscal' if is_pg else 'devolucao_nota_fiscal'
    tbl_pb = 'public.produtos_bipados' if is_pg else 'produtos_bipados'
    if is_pg:
        conn.execute(
            f'''CREATE TABLE IF NOT EXISTS {tbl_nf} (
                id BIGSERIAL PRIMARY KEY,
                id_viagem TEXT NOT NULL,
                numero_nf TEXT NOT NULL,
                motivo TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'em_andamento',
                doca TEXT,
                criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                concluida_em TIMESTAMPTZ,
                criado_por TEXT,
                concluida_por TEXT
            )'''
        )
        conn.execute(f'CREATE INDEX IF NOT EXISTS idx_devolucao_nf_viagem ON {tbl_nf} (id_viagem, status)')
        try:
            conn.execute(
                f'ALTER TABLE {tbl_pb} ADD COLUMN IF NOT EXISTS devolucao_nf_id BIGINT REFERENCES {tbl_nf}(id) ON DELETE SET NULL'
            )
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            conn.execute(f'ALTER TABLE {tbl_pb} ADD COLUMN IF NOT EXISTS devolucao_nf_id BIGINT')
        _ensure_devolucao_nf_status_cancelada_pg(conn)
    else:
        conn.execute(
            f'''CREATE TABLE IF NOT EXISTS {tbl_nf} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_viagem TEXT NOT NULL,
                numero_nf TEXT NOT NULL,
                motivo TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'em_andamento',
                doca TEXT,
                criado_em TEXT NOT NULL,
                concluida_em TEXT,
                criado_por TEXT,
                concluida_por TEXT
            )'''
        )
        try:
            conn.execute(f'ALTER TABLE {tbl_pb} ADD COLUMN devolucao_nf_id INTEGER')
        except Exception:
            pass
    try:
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


def _ensure_devolucao_nf_status_cancelada_pg(conn):
    """Permite status cancelada (parar NF / zerar conferência) no Postgres."""
    if getattr(conn, 'kind', None) != 'pg':
        return
    if 'dev_nf_status_cancelada' in _SCHEMA_ENSURE_DONE:
        return
    try:
        conn.execute(
            'ALTER TABLE public.devolucao_nota_fiscal DROP CONSTRAINT IF EXISTS devolucao_nota_fiscal_status_check'
        )
        conn.execute(
            """ALTER TABLE public.devolucao_nota_fiscal
               ADD CONSTRAINT devolucao_nota_fiscal_status_check
               CHECK (status IN ('em_andamento', 'concluida', 'cancelada'))"""
        )
        conn.commit()
        _SCHEMA_ENSURE_DONE.add('dev_nf_status_cancelada')
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


def _motivo_devolucao_valido(motivo):
    return (motivo or '').strip().lower() in ('parcial', 'total', 'reentrega')


def _motivo_devolucao_label(motivo):
    m = (motivo or '').strip().lower()
    return {
        'parcial': 'Devolução parcial',
        'total': 'Devolução total',
        'reentrega': 'Reentrega',
    }.get(m, motivo or '')


def _devolucao_status_label(status):
    s = (status or '').strip().lower()
    return {
        'em_andamento': 'Em andamento',
        'concluida': 'Concluída',
        'cancelada': 'Cancelada',
    }.get(s, status or '')


def _row_devolucao_nf_dict(row):
    if not row:
        return None
    g = row.get if hasattr(row, 'get') else None
    rid = g('id') if g else (row[0] if len(row) > 0 else None)
    return {
        'id': int(rid) if rid is not None else None,
        'id_viagem': (g('id_viagem') if g else row[1]) or '',
        'numero_nf': (g('numero_nf') if g else row[2]) or '',
        'motivo': (g('motivo') if g else row[3]) or '',
        'motivo_label': _motivo_devolucao_label((g('motivo') if g else row[3]) or ''),
        'status': (g('status') if g else row[4]) or 'em_andamento',
        'status_label': _devolucao_status_label((g('status') if g else row[4]) or 'em_andamento'),
        'doca': (g('doca') if g else (row[5] if len(row) > 5 else '')) or '',
        'criado_em': _fmt_datahora_br(g('criado_em') if g else (row[6] if len(row) > 6 else '')),
        'concluida_em': _fmt_datahora_br(g('concluida_em') if g else (row[7] if len(row) > 7 else '')),
        'criado_por': (g('criado_por') if g else (row[8] if len(row) > 8 else '')) or '',
        'concluida_por': (g('concluida_por') if g else (row[9] if len(row) > 9 else '')) or '',
    }


def _devolucao_conferencia_lista_nf(conn, id_viagem, devolucao_nf_id):
    """Itens bipados na NF de devolução ativa (somente retorno desta sessão)."""
    id_viagem = (id_viagem or '').strip()
    if not id_viagem or not devolucao_nf_id:
        return []
    try:
        nf_id = int(devolucao_nf_id)
    except (TypeError, ValueError):
        return []
    ids = _devolucao_ids_consulta(conn, id_viagem) or [id_viagem]
    ph = ','.join(['?'] * len(ids))
    is_pg = getattr(conn, 'kind', None) == 'pg'
    col_viagem = 'TRIM(COALESCE(id_viagem::text, \'\'))' if is_pg else 'TRIM(COALESCE(id_viagem, \'\'))'
    rows = conn.execute(
        f'''SELECT codigo_barras, codigo_interno, MAX(produto) AS produto,
                  SUM(quantidade) AS quantidade_bipada, MAX(unidade) AS unidade, MAX(peso) AS peso
           FROM produtos_bipados
           WHERE {col_viagem} IN ({ph})
             AND COALESCE(fluxo, 'carregamento') = 'devolucao' AND devolucao_nf_id = ?
           GROUP BY codigo_barras, codigo_interno
           ORDER BY MAX(data_hora) DESC''',
        (*ids, nf_id),
    ).fetchall()
    saida_rows = conn.execute(
        f'''SELECT codigo_barras, codigo_interno, SUM(quantidade) AS qtd_saida
           FROM produtos_bipados
           WHERE {col_viagem} IN ({ph})
             AND COALESCE(fluxo, 'carregamento') = 'carregamento'
           GROUP BY codigo_barras, codigo_interno''',
        tuple(ids),
    ).fetchall()
    saida_map = {}
    for sr in saida_rows or []:
        cb = ((sr.get('codigo_barras') if hasattr(sr, 'get') else sr[0]) or '').strip()
        ci = ((sr.get('codigo_interno') if hasattr(sr, 'get') else sr[1]) or '').strip()
        q = int((sr.get('qtd_saida') if hasattr(sr, 'get') else (sr[2] if len(sr) > 2 else 0)) or 0)
        if cb:
            saida_map[cb] = saida_map.get(cb, 0) + q
        if ci:
            saida_map['ci:' + ci] = saida_map.get('ci:' + ci, 0) + q
    resultado = []
    for r in rows or []:
        cb = ((r.get('codigo_barras') if hasattr(r, 'get') else r[0]) or '').strip()
        ci = ((r.get('codigo_interno') if hasattr(r, 'get') else r[1]) or '').strip()
        prod = (r.get('produto') if hasattr(r, 'get') else (r[2] if len(r) > 2 else '')) or ''
        q_bip = int((r.get('quantidade_bipada') if hasattr(r, 'get') else (r[3] if len(r) > 3 else 0)) or 0)
        un = (r.get('unidade') if hasattr(r, 'get') else (r[4] if len(r) > 4 else '')) or ''
        peso = (r.get('peso') if hasattr(r, 'get') else (r[5] if len(r) > 5 else '')) or ''
        q_saida = saida_map.get(cb, 0) or saida_map.get('ci:' + ci, 0) or 0
        resultado.append({
            'codigo_produto': ci,
            'codigo_barras': cb,
            'produto': prod,
            'quantidade_produto': q_saida,
            'quantidade_bipada': q_bip,
            'quantidade_falta': max(0, q_saida - q_bip) if q_saida > 0 else 0,
            'quantidade_sobra': max(0, q_bip - q_saida) if q_saida > 0 else 0,
            'quantidade_saida': q_saida,
            'unidade': un,
            'peso_bruto': peso,
            'status_bipado': 'COMPLETO' if q_bip > 0 and (q_saida <= 0 or q_bip >= q_saida) else ('PARCIAL' if q_bip > 0 else 'PENDENTE'),
            'aviso_sobra': ('Bipou %s a mais que a saída' % (q_bip - q_saida)) if q_saida > 0 and q_bip > q_saida else '',
            'id_viagem': id_viagem,
            'origem_romaneio': False,
            'modo_devolucao_nf': True,
        })
    return resultado


def _ensure_viagem_periodo_bipagem_table(conn):
    """Tabela auxiliar: início/fim real da bipagem (independente de data_hora dos itens gravados em lote)."""
    is_pg = getattr(conn, 'kind', None) == 'pg'
    tbl = 'public.viagem_periodo_bipagem' if is_pg else 'viagem_periodo_bipagem'
    if is_pg:
        conn.execute(
            f'''CREATE TABLE IF NOT EXISTS {tbl} (
                id_viagem TEXT NOT NULL,
                fluxo TEXT NOT NULL DEFAULT 'carregamento',
                inicio_em TIMESTAMPTZ NOT NULL,
                fim_em TIMESTAMPTZ NOT NULL,
                PRIMARY KEY (id_viagem, fluxo)
            )'''
        )
    else:
        conn.execute(
            f'''CREATE TABLE IF NOT EXISTS {tbl} (
                id_viagem TEXT NOT NULL,
                fluxo TEXT NOT NULL DEFAULT 'carregamento',
                inicio_em TEXT NOT NULL,
                fim_em TEXT NOT NULL,
                PRIMARY KEY (id_viagem, fluxo)
            )'''
        )
    try:
        if is_pg:
            conn.execute(
                f'ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS extrato_gerado_em TIMESTAMPTZ'
            )
        else:
            cols = [r[1] for r in conn.execute(f'PRAGMA table_info({tbl})').fetchall()]
            if 'extrato_gerado_em' not in cols:
                conn.execute(f'ALTER TABLE {tbl} ADD COLUMN extrato_gerado_em TEXT')
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


def _marcar_extrato_gerado(conn, id_viagem, fluxo='carregamento', momento=None, commit=False):
    """Marca que o extrato/comprovante da viagem foi gerado (carregamento concluído)."""
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm:
        return
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    _ensure_viagem_periodo_bipagem_table(conn)
    momento = momento or datetime.now(timezone.utc)
    arm = _datetime_para_armazenamento(conn, momento)
    tbl = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            f'''UPDATE {tbl} SET extrato_gerado_em = ?
                WHERE id_viagem = ? AND fluxo = ? AND extrato_gerado_em IS NULL''',
            (arm, id_norm, fluxo),
        )
        conn.execute(
            f'''INSERT INTO {tbl} (id_viagem, fluxo, inicio_em, fim_em, extrato_gerado_em)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT (id_viagem, fluxo) DO UPDATE SET
                  extrato_gerado_em = COALESCE({tbl}.extrato_gerado_em, excluded.extrato_gerado_em)''',
            (id_norm, fluxo, arm, arm, arm),
        )
    else:
        conn.execute(
            f'''UPDATE {tbl} SET extrato_gerado_em = ?
                WHERE id_viagem = ? AND fluxo = ? AND (extrato_gerado_em IS NULL OR extrato_gerado_em = '')''',
            (arm, id_norm, fluxo),
        )
        conn.execute(
            f'''INSERT INTO {tbl} (id_viagem, fluxo, inicio_em, fim_em, extrato_gerado_em)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(id_viagem, fluxo) DO UPDATE SET
                  extrato_gerado_em = COALESCE({tbl}.extrato_gerado_em, excluded.extrato_gerado_em)''',
            (id_norm, fluxo, arm, arm, arm),
        )
    if commit:
        conn.commit()


def _viagem_carregamento_concluido(extrato_gerado_em, bip_row):
    """Extrato gerado (comprovante) ou bipagem gravada em lote no comprovante."""
    if extrato_gerado_em:
        return True
    if not bip_row:
        return False
    total = int((bip_row.get('total_bipados') if hasattr(bip_row, 'get') else bip_row['total_bipados']) or 0)
    if total <= 0:
        return False
    inicio_b = bip_row.get('inicio') if hasattr(bip_row, 'get') else bip_row['inicio']
    fim_b = bip_row.get('fim') if hasattr(bip_row, 'get') else bip_row['fim']
    if inicio_b is None or fim_b is None:
        return False
    return str(inicio_b).strip() == str(fim_b).strip()


def _parse_datetime_iso(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return val.replace(tzinfo=timezone.utc)
        return val
    s = str(val).strip()
    if not s:
        return None
    s = s.replace('Z', '+00:00')
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def _datetime_para_armazenamento(conn, dt):
    if dt is None:
        return None
    if getattr(conn, 'kind', None) == 'pg':
        return dt
    return dt.astimezone(timezone.utc).isoformat()


def _formatar_data_hora_periodo(d):
    if d is None:
        return ''
    dt = None
    if hasattr(d, 'strftime'):
        dt = d
    else:
        s = str(d).strip()
        if not s:
            return ''
        dt = _parse_datetime_iso(s)
        if not dt:
            return s[:19] if len(s) > 19 else s
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    try:
        if ZoneInfo:
            dt = dt.astimezone(ZoneInfo('America/Sao_Paulo'))
    except Exception:
        pass
    return dt.strftime('%d/%m/%Y %H:%M:%S')


def _registrar_momento_bipagem(conn, id_viagem, fluxo='carregamento', momento=None, commit=False):
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm:
        return
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    _ensure_viagem_periodo_bipagem_table(conn)
    momento = momento or datetime.now(timezone.utc)
    arm = _datetime_para_armazenamento(conn, momento)
    tbl = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            f'''INSERT INTO {tbl} (id_viagem, fluxo, inicio_em, fim_em)
                VALUES (?, ?, ?, ?)
                ON CONFLICT (id_viagem, fluxo) DO UPDATE SET
                  inicio_em = LEAST({tbl}.inicio_em, excluded.inicio_em),
                  fim_em = GREATEST({tbl}.fim_em, excluded.fim_em)''',
            (id_norm, fluxo, arm, arm),
        )
    else:
        conn.execute(
            f'''INSERT INTO {tbl} (id_viagem, fluxo, inicio_em, fim_em)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(id_viagem, fluxo) DO UPDATE SET
                  inicio_em = MIN(inicio_em, excluded.inicio_em),
                  fim_em = MAX(fim_em, excluded.fim_em)''',
            (id_norm, fluxo, arm, arm),
        )
    if commit:
        conn.commit()


def _definir_periodo_bipagem(conn, id_viagem, fluxo, inicio, fim, commit=False):
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm or not inicio or not fim:
        return
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    if fim < inicio:
        fim = inicio
    _ensure_viagem_periodo_bipagem_table(conn)
    ai = _datetime_para_armazenamento(conn, inicio)
    af = _datetime_para_armazenamento(conn, fim)
    tbl = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
    conn.execute(
        f'''INSERT INTO {tbl} (id_viagem, fluxo, inicio_em, fim_em)
            VALUES (?, ?, ?, ?)
            ON CONFLICT (id_viagem, fluxo) DO UPDATE SET
              inicio_em = excluded.inicio_em,
              fim_em = excluded.fim_em''',
        (id_norm, fluxo, ai, af),
    )
    if commit:
        conn.commit()


def _limpar_periodo_bipagem(conn, id_viagem, fluxo='carregamento', commit=False):
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm:
        return
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    _ensure_viagem_periodo_bipagem_table(conn)
    tbl = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
    conn.execute(f'DELETE FROM {tbl} WHERE id_viagem = ? AND fluxo = ?', (id_norm, fluxo))
    if commit:
        conn.commit()


def _consultar_periodo_viagem_raw_conn(conn, id_viagem, fluxo='carregamento'):
    """Consulta início/fim da bipagem reutilizando conexão aberta (evita round-trip extra)."""
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm or not conn:
        return None, None
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    try:
        _ensure_viagem_periodo_bipagem_table(conn)
        tbl = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
        row_p = conn.execute(
            f'SELECT inicio_em, fim_em FROM {tbl} WHERE id_viagem = ? AND fluxo = ?',
            (id_norm, fluxo),
        ).fetchone()
        if row_p and row_p['inicio_em'] and row_p['fim_em']:
            return row_p['inicio_em'], row_p['fim_em']
        if getattr(conn, 'kind', None) == 'pg':
            row = conn.execute(
                "SELECT MIN(data_hora) as inicio, MAX(data_hora) as fim FROM produtos_bipados "
                "WHERE TRIM(COALESCE(id_viagem::text, '')) = ? AND COALESCE(fluxo, 'carregamento') = ?",
                (id_norm, fluxo),
            ).fetchone()
        else:
            row = conn.execute(
                "SELECT MIN(data_hora) as inicio, MAX(data_hora) as fim FROM produtos_bipados "
                "WHERE id_viagem = ? AND COALESCE(fluxo, 'carregamento') = ?",
                (id_norm, fluxo),
            ).fetchone()
        if row and row['inicio'] and row['fim']:
            return row['inicio'], row['fim']
    except Exception:
        pass
    return None, None


def _consultar_periodo_viagem_raw(id_viagem, fluxo='carregamento'):
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm:
        return None, None
    conn = get_db()
    try:
        return _consultar_periodo_viagem_raw_conn(conn, id_viagem, fluxo)
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _periodo_viagem_resposta(id_viagem, fluxo='carregamento'):
    inicio, fim = _consultar_periodo_viagem_raw(id_viagem, fluxo)
    inicio_str = _formatar_data_hora_periodo(inicio)
    fim_str = _formatar_data_hora_periodo(fim)
    return {
        'inicio_bipagem': inicio_str,
        'fim_bipagem': fim_str,
        'inicio_carregamento': inicio_str,
        'fim_carregamento': fim_str,
    }


def _tbl_terceiros_documentos(conn):
    return 'public.terceiros_documentos' if getattr(conn, 'kind', None) == 'pg' else 'terceiros_documentos'


def _tbl_terceiros_documento_itens(conn):
    return 'public.terceiros_documento_itens' if getattr(conn, 'kind', None) == 'pg' else 'terceiros_documento_itens'


def _tbl_terceiros_documento_eventos(conn):
    return 'public.terceiros_documento_eventos' if getattr(conn, 'kind', None) == 'pg' else 'terceiros_documento_eventos'


def _sql_cols_terceiros_documentos_listagem(alias):
    """Colunas para listagem de NFs — sem xml_conteudo (evita ler megabytes por linha e WORKER TIMEOUT no Gunicorn/Render)."""
    a = alias
    return (
        '%s.id, %s.area, %s.chave_nfe, %s.numero_nf, %s.serie_nf, %s.data_emissao, '
        '%s.remetente_nome, %s.remetente_cnpj, %s.destinatario_nome, %s.destinatario_cnpj, %s.destinatario_uf, '
        '%s.numero_pedido, %s.previsao_chegada, %s.arquivo_nome, '
        '%s.recebimento_concluido, %s.recebimento_concluido_em, %s.recebimento_concluido_por, '
        '%s.nota_lancada, %s.nota_lancada_em, %s.nota_lancada_por, '
        '%s.enviar_para_mg, %s.enviar_para_mg_em, %s.enviar_para_mg_por, '
        '%s.motorista_carreta, %s.motorista_carreta_em, %s.placa_carreta, '
        '%s.motorista_saida_mg, %s.motorista_saida_mg_em, %s.placa_saida_mg, '
        '%s.carga_recebida_mg, %s.carga_recebida_mg_em, %s.carga_recebida_mg_por, %s.recebedor_mg, '
        '%s.consumivel_sp, %s.recebedor_consumivel_sp, %s.consumivel_sp_historico, %s.consumivel_sp_historico_em, %s.consumivel_sp_historico_por, '
        '%s.motivo_nao_lancada, %s.motivo_nao_enviar_mg, %s.motivo_nao_recebida_mg, '
        '%s.criado_em, %s.criado_por, %s.atualizado_em, %s.atualizado_por'
    ) % ((a,) * 45)


def _terceiros_sql_join_resumo_itens(conn, tbl_i, alias_doc='d'):
    """Totais por NF: LATERAL no Postgres (usa índice por documento_id); subquery global no SQLite."""
    if getattr(conn, 'kind', None) == 'pg':
        return (
            ' LEFT JOIN LATERAL ('
            '   SELECT COUNT(id) AS total_itens,'
            '          COALESCE(SUM(quantidade_xml), 0) AS quantidade_total_xml,'
            '          COALESCE(SUM(quantidade_bipada), 0) AS quantidade_total_bipada,'
            '          SUM(CASE WHEN ABS(COALESCE(quantidade_xml, 0) - COALESCE(quantidade_bipada, 0)) > 0.000001'
            '               THEN 1 ELSE 0 END) AS itens_divergentes'
            '   FROM ' + tbl_i + ' i WHERE i.documento_id = ' + alias_doc + '.id'
            ') si ON TRUE'
        )
    return (
        ' LEFT JOIN ('
        '   SELECT documento_id,'
        '          COUNT(id) AS total_itens,'
        '          COALESCE(SUM(quantidade_xml), 0) AS quantidade_total_xml,'
        '          COALESCE(SUM(quantidade_bipada), 0) AS quantidade_total_bipada,'
        '          SUM(CASE WHEN ABS(COALESCE(quantidade_xml, 0) - COALESCE(quantidade_bipada, 0)) > 0.000001'
        '               THEN 1 ELSE 0 END) AS itens_divergentes'
        '   FROM ' + tbl_i + ' GROUP BY documento_id'
        ') si ON si.documento_id = ' + alias_doc + '.id'
    )


def _terceiros_cols_select_listagem(cols, com_resumo=True):
    if com_resumo:
        return (
            cols + ', COALESCE(si.total_itens, 0) AS total_itens,'
            ' COALESCE(si.quantidade_total_xml, 0) AS quantidade_total_xml,'
            ' COALESCE(si.quantidade_total_bipada, 0) AS quantidade_total_bipada,'
            ' COALESCE(si.itens_divergentes, 0) AS itens_divergentes'
        )
    return (
        cols + ', 0 AS total_itens, 0 AS quantidade_total_xml,'
        ' 0 AS quantidade_total_bipada, 0 AS itens_divergentes'
    )


_TERCEIROS_SCHEMA_PRONTO = False


def _ensure_terceiros_schema(conn, rodar_backfill=False):
    global _TERCEIROS_SCHEMA_PRONTO
    if _TERCEIROS_SCHEMA_PRONTO:
        return
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS public.terceiros_documentos (
                id BIGSERIAL PRIMARY KEY,
                area TEXT NOT NULL CHECK (area IN ('recebimento', 'expedicao', 'carreta')),
                chave_nfe TEXT,
                numero_nf TEXT,
                serie_nf TEXT,
                data_emissao TEXT,
                remetente_nome TEXT,
                remetente_cnpj TEXT,
                destinatario_nome TEXT,
                destinatario_cnpj TEXT,
                destinatario_uf TEXT,
                previsao_chegada TEXT,
                arquivo_nome TEXT,
                xml_conteudo TEXT,
                recebimento_concluido BOOLEAN NOT NULL DEFAULT FALSE,
                recebimento_concluido_em TIMESTAMPTZ,
                recebimento_concluido_por TEXT,
                nota_lancada TEXT,
                nota_lancada_em TIMESTAMPTZ,
                nota_lancada_por TEXT,
                enviar_para_mg TEXT,
                enviar_para_mg_em TIMESTAMPTZ,
                enviar_para_mg_por TEXT,
                motorista_carreta TEXT,
                motorista_carreta_em TIMESTAMPTZ,
                placa_carreta TEXT,
                motorista_saida_mg TEXT,
                motorista_saida_mg_em TIMESTAMPTZ,
                placa_saida_mg TEXT,
                carga_recebida_mg TEXT,
                carga_recebida_mg_em TIMESTAMPTZ,
                carga_recebida_mg_por TEXT,
                recebedor_mg TEXT,
                consumivel_sp TEXT,
                recebedor_consumivel_sp TEXT,
                consumivel_sp_historico TEXT,
                consumivel_sp_historico_em TIMESTAMPTZ,
                consumivel_sp_historico_por TEXT,
                criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                criado_por TEXT,
                atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                atualizado_por TEXT
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS public.terceiros_documento_itens (
                id BIGSERIAL PRIMARY KEY,
                documento_id BIGINT NOT NULL REFERENCES public.terceiros_documentos(id) ON DELETE CASCADE,
                n_item INTEGER,
                codigo_ean TEXT,
                codigo_produto_xml TEXT,
                descricao_xml TEXT,
                unidade_xml TEXT,
                quantidade_xml NUMERIC(14,3) NOT NULL DEFAULT 0,
                codigo_produto_base TEXT,
                codigo_barras_base TEXT,
                descricao_base TEXT,
                quantidade_bipada NUMERIC(14,3) NOT NULL DEFAULT 0,
                status_bipagem TEXT NOT NULL DEFAULT 'PENDENTE',
                ultimo_ean_bipado TEXT,
                atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                atualizado_por TEXT
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS public.terceiros_documento_eventos (
                id BIGSERIAL PRIMARY KEY,
                documento_id BIGINT NOT NULL REFERENCES public.terceiros_documentos(id) ON DELETE CASCADE,
                evento TEXT NOT NULL,
                valor_anterior TEXT,
                valor_novo TEXT,
                usuario TEXT,
                criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                detalhes TEXT
            )'''
        )
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_area ON public.terceiros_documentos(area, criado_em DESC)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_chave ON public.terceiros_documentos(chave_nfe)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_documento ON public.terceiros_documento_itens(documento_id)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_ean ON public.terceiros_documento_itens(codigo_ean)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_eventos_documento ON public.terceiros_documento_eventos(documento_id, criado_em DESC)')
    else:
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS terceiros_documentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                area TEXT NOT NULL,
                chave_nfe TEXT,
                numero_nf TEXT,
                serie_nf TEXT,
                data_emissao TEXT,
                remetente_nome TEXT,
                remetente_cnpj TEXT,
                destinatario_nome TEXT,
                destinatario_cnpj TEXT,
                destinatario_uf TEXT,
                previsao_chegada TEXT,
                arquivo_nome TEXT,
                xml_conteudo TEXT,
                recebimento_concluido INTEGER NOT NULL DEFAULT 0,
                recebimento_concluido_em TEXT,
                recebimento_concluido_por TEXT,
                nota_lancada TEXT,
                nota_lancada_em TEXT,
                nota_lancada_por TEXT,
                enviar_para_mg TEXT,
                enviar_para_mg_em TEXT,
                enviar_para_mg_por TEXT,
                motorista_carreta TEXT,
                motorista_carreta_em TEXT,
                placa_carreta TEXT,
                motorista_saida_mg TEXT,
                motorista_saida_mg_em TEXT,
                placa_saida_mg TEXT,
                carga_recebida_mg TEXT,
                carga_recebida_mg_em TEXT,
                carga_recebida_mg_por TEXT,
                recebedor_mg TEXT,
                consumivel_sp TEXT,
                recebedor_consumivel_sp TEXT,
                consumivel_sp_historico TEXT,
                consumivel_sp_historico_em TEXT,
                consumivel_sp_historico_por TEXT,
                motivo_nao_lancada TEXT,
                motivo_nao_enviar_mg TEXT,
                motivo_nao_recebida_mg TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                criado_por TEXT,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_por TEXT
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS terceiros_documento_itens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                documento_id INTEGER NOT NULL,
                n_item INTEGER,
                codigo_ean TEXT,
                codigo_produto_xml TEXT,
                descricao_xml TEXT,
                unidade_xml TEXT,
                quantidade_xml REAL NOT NULL DEFAULT 0,
                codigo_produto_base TEXT,
                codigo_barras_base TEXT,
                descricao_base TEXT,
                quantidade_bipada REAL NOT NULL DEFAULT 0,
                status_bipagem TEXT NOT NULL DEFAULT 'PENDENTE',
                ultimo_ean_bipado TEXT,
                atualizado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                atualizado_por TEXT,
                FOREIGN KEY (documento_id) REFERENCES terceiros_documentos(id) ON DELETE CASCADE
            )'''
        )
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS terceiros_documento_eventos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                documento_id INTEGER NOT NULL,
                evento TEXT NOT NULL,
                valor_anterior TEXT,
                valor_novo TEXT,
                usuario TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                detalhes TEXT,
                FOREIGN KEY (documento_id) REFERENCES terceiros_documentos(id) ON DELETE CASCADE
            )'''
        )
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_area ON terceiros_documentos(area, criado_em DESC)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documentos_chave ON terceiros_documentos(chave_nfe)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_documento ON terceiros_documento_itens(documento_id)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_itens_ean ON terceiros_documento_itens(codigo_ean)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_terceiros_documento_eventos_documento ON terceiros_documento_eventos(documento_id, criado_em DESC)')
    if getattr(conn, 'kind', None) == 'pg':
        try:
            conn.execute('ALTER TABLE public.terceiros_documentos DROP CONSTRAINT IF EXISTS terceiros_documentos_area_check')
            conn.execute(
                "ALTER TABLE public.terceiros_documentos ADD CONSTRAINT terceiros_documentos_area_check "
                "CHECK (area IN ('recebimento', 'expedicao', 'carreta'))"
            )
        except Exception:
            pass
    try:
        tbl_doc = _tbl_terceiros_documentos(conn)
        if getattr(conn, 'kind', None) == 'pg':
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS placa_carreta TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS motorista_saida_mg TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS motorista_saida_mg_em TIMESTAMPTZ')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS placa_saida_mg TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS recebedor_mg TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS destinatario_uf TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS numero_pedido TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS recebedor_consumivel_sp TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp_historico TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp_historico_em TIMESTAMPTZ')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp_historico_por TEXT')
        else:
            info = conn.execute('PRAGMA table_info(terceiros_documentos)').fetchall()
            nomes = [r[1] for r in (info or [])]
            if 'placa_carreta' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN placa_carreta TEXT')
            if 'motorista_saida_mg' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN motorista_saida_mg TEXT')
            if 'motorista_saida_mg_em' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN motorista_saida_mg_em TEXT')
            if 'placa_saida_mg' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN placa_saida_mg TEXT')
            if 'recebedor_mg' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN recebedor_mg TEXT')
            if 'destinatario_uf' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN destinatario_uf TEXT')
            if 'numero_pedido' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN numero_pedido TEXT')
            if 'consumivel_sp' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN consumivel_sp TEXT')
            if 'recebedor_consumivel_sp' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN recebedor_consumivel_sp TEXT')
            if 'consumivel_sp_historico' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN consumivel_sp_historico TEXT')
            if 'consumivel_sp_historico_em' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN consumivel_sp_historico_em TEXT')
            if 'consumivel_sp_historico_por' not in nomes:
                conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN consumivel_sp_historico_por TEXT')
            for col_m in ('motivo_nao_lancada', 'motivo_nao_enviar_mg', 'motivo_nao_recebida_mg'):
                if col_m not in nomes:
                    conn.execute('ALTER TABLE terceiros_documentos ADD COLUMN ' + col_m + ' TEXT')
    except Exception:
        pass
    if getattr(conn, 'kind', None) == 'pg':
        try:
            tbl_doc = _tbl_terceiros_documentos(conn)
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS motivo_nao_lancada TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS motivo_nao_enviar_mg TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS motivo_nao_recebida_mg TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS recebedor_mg TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS recebedor_consumivel_sp TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp_historico TEXT')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp_historico_em TIMESTAMPTZ')
            conn.execute('ALTER TABLE ' + tbl_doc + ' ADD COLUMN IF NOT EXISTS consumivel_sp_historico_por TEXT')
        except Exception:
            pass
    if rodar_backfill:
        try:
            _terceiros_backfill_campos_xml_pendentes(conn, limite=300)
        except Exception:
            pass
    try:
        tbl_it = _tbl_terceiros_documento_itens(conn)
        if getattr(conn, 'kind', None) == 'pg':
            conn.execute('ALTER TABLE ' + tbl_it + ' ADD COLUMN IF NOT EXISTS motivo TEXT')
        else:
            info_it = conn.execute('PRAGMA table_info(terceiros_documento_itens)').fetchall()
            nomes_it = [r[1] for r in (info_it or [])]
            if 'motivo' not in nomes_it:
                conn.execute('ALTER TABLE terceiros_documento_itens ADD COLUMN motivo TEXT')
    except Exception:
        pass
    conn.commit()
    _TERCEIROS_SCHEMA_PRONTO = True


def _usa_banco_para_dados():
    """True quando DATABASE_URL está definido: app usa apenas o banco (dados subidos por código), não a planilha."""
    return bool((os.environ.get('DATABASE_URL') or '').strip())


def _get_latest_dataset_id(conn):
    """Retorna o dataset_id ativo (base_codigo_barras) ou None."""
    if getattr(conn, 'kind', None) != 'pg':
        return None
    now = time.time()
    cached = getattr(_get_latest_dataset_id, '_cache', None)
    if cached and cached.get('exp', 0) > now:
        return cached.get('id')
    try:
        row = conn.execute(
            "SELECT dataset_id FROM excel_datasets WHERE ativo = ? LIMIT 1",
            (True,),
        ).fetchone()
        ds = row.get('dataset_id') if row else None
        _get_latest_dataset_id._cache = {'id': ds, 'exp': now + 60}
        return ds
    except Exception:
        return None


def _requer_login():
    """Retorna True se a rota atual requer login (e o usuário não está logado)."""
    if request.endpoint in (None, 'login', 'static', 'raiz', 'ravex_env_check'):
        return False
    if request.path.startswith('/api/login') or request.path.startswith('/api/cadastrar'):
        return False
    return 'usuario' not in session


_USUARIO_SESSAO_OK_TTL = 120  # segundos — evita ida ao banco a cada clique/navegação


def _usuario_ainda_existe(usuario):
    """Verifica se o usuário ainda existe no banco (não foi excluído do config)."""
    if not usuario:
        return False
    ok_user = session.get('_auth_ok_user')
    ok_ts = session.get('_auth_ok_ts')
    if ok_user == usuario and ok_ts is not None:
        try:
            if (time.time() - float(ok_ts)) < _USUARIO_SESSAO_OK_TTL:
                return True
        except (TypeError, ValueError):
            pass
    conn = get_db()
    try:
        row = conn.execute('SELECT 1 FROM usuarios WHERE usuario = ?', (usuario,)).fetchone()
        existe = row is not None
    finally:
        conn.close()
    if existe:
        session['_auth_ok_user'] = usuario
        session['_auth_ok_ts'] = time.time()
    else:
        session.pop('_auth_ok_user', None)
        session.pop('_auth_ok_ts', None)
    return existe


@app.errorhandler(Exception)
def _api_json_exception_handler(e):
    """APIs /api/* sempre retornam JSON (evita HTML 500 opaco no frontend)."""
    if isinstance(e, HTTPException):
        if request.path.startswith('/api/') and e.code and int(e.code) >= 400:
            msg = getattr(e, 'description', None) or str(e) or 'Erro HTTP'
            return jsonify({'ok': False, 'erro': msg}), int(e.code)
        return e
    if request.path.startswith('/api/'):
        try:
            app.logger.exception('Erro na API %s: %s', request.path, e)
        except Exception:
            pass
        return jsonify({'ok': False, 'erro': str(e) or 'Erro interno do servidor.'}), 500
    raise e


@app.before_request
def proteger_rotas():
    """Redireciona para /login se o usuário não estiver autenticado. Invalida sessão se o usuário foi excluído."""
    # Rotas que não exigem autenticação
    if request.endpoint in (None, 'login', 'static', 'raiz', 'ravex_env_check', 'api_health'):
        return None
    if request.path.startswith('/api/login') or request.path.startswith('/api/cadastrar'):
        return None
    # Se está logado, verificar se o usuário ainda existe no banco (não foi removido do config)
    if session.get('usuario'):
        try:
            usuario_ok = _usuario_ainda_existe(session.get('usuario'))
        except Exception as e:
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'erro': 'Falha ao validar sessão: %s' % (e,)}), 503
            return redirect(url_for('login'))
        if not usuario_ok:
            session.pop('usuario', None)
            session.pop('usuario_id', None)
            session.pop('_auth_ok_user', None)
            session.pop('_auth_ok_ts', None)
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'erro': 'Sessão encerrada. Usuário foi removido.'}), 401
            return redirect(url_for('login'))
    # Exige login
    if _requer_login():
        if request.path.startswith('/api/'):
            return jsonify({'ok': False, 'erro': 'Não autorizado'}), 401
        return redirect(url_for('login'))


@app.route('/login', methods=['GET'])
def login():
    """Página de login. Sempre exibe a tela de login; se já logado, mostra link para o painel."""
    return render_template('login.html', usuario=session.get('usuario') or '')


@app.route('/api/login', methods=['POST'])
def api_login():
    """Autentica usuário e inicia sessão."""
    data = request.get_json() or {}
    usuario = (data.get('usuario') or '').strip()
    senha = data.get('senha') or ''
    if not usuario or not senha:
        return jsonify({'ok': False, 'erro': 'Informe usuário e senha.'})
    conn = get_db()
    row = conn.execute('SELECT id, senha_hash FROM usuarios WHERE usuario = ?', (usuario,)).fetchone()
    conn.close()
    if not row or not check_password_hash(row['senha_hash'], senha):
        return jsonify({'ok': False, 'erro': 'Usuário ou senha incorretos.'})
    session['usuario'] = usuario
    session['usuario_id'] = row['id']
    session['_auth_ok_user'] = usuario
    session['_auth_ok_ts'] = time.time()
    return jsonify({'ok': True, 'redirect': url_for('entrada_modulos')})


@app.route('/api/cadastrar', methods=['POST'])
def api_cadastrar():
    """Cadastra novo usuário."""
    data = request.get_json() or {}
    usuario = (data.get('usuario') or '').strip()
    senha = data.get('senha') or ''
    confirmar = data.get('confirmar_senha') or ''
    if not usuario or not senha:
        return jsonify({'ok': False, 'erro': 'Informe usuário e senha.'})
    if len(usuario) < 2:
        return jsonify({'ok': False, 'erro': 'Usuário deve ter pelo menos 2 caracteres.'})
    if len(senha) < 4:
        return jsonify({'ok': False, 'erro': 'Senha deve ter pelo menos 4 caracteres.'})
    if senha != confirmar:
        return jsonify({'ok': False, 'erro': 'As senhas não coincidem.'})
    conn = get_db()
    try:
        existente = conn.execute('SELECT 1 FROM usuarios WHERE usuario = ?', (usuario,)).fetchone()
        if existente:
            return jsonify({'ok': False, 'erro': 'Este usuário já existe.'})
        conn.execute(
            'INSERT INTO usuarios (usuario, senha_hash, criado_em) VALUES (?, ?, ?)',
            (usuario, generate_password_hash(senha, method='pbkdf2:sha256'), datetime.now().isoformat())
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.rollback()
        return jsonify({'ok': False, 'erro': 'Este usuário já existe.'})
    except Exception as e:
        conn.rollback()
        msg = str(e or '').lower()
        if 'unique' in msg or 'duplicate key' in msg or 'usuarios_usuario_key' in msg:
            return jsonify({'ok': False, 'erro': 'Este usuário já existe.'})
        return jsonify({'ok': False, 'erro': f'Erro ao cadastrar usuário: {str(e)}'}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass
    adicionar_usuario_ao_config(usuario, senha)
    return jsonify({'ok': True, 'mensagem': 'Cadastro realizado. Faça login.'})


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """Encerra a sessão. Resposta mínima para sair rápido."""
    session.pop('usuario', None)
    session.pop('usuario_id', None)
    session.pop('_auth_ok_user', None)
    session.pop('_auth_ok_ts', None)
    return '', 204


@app.route('/api/ravex-env-check', methods=['GET'])
def ravex_env_check():
    """Diagnóstico: retorna se RAVEX_USER e RAVEX_PASSWORD estão definidos (sem mostrar valores). Acesso público."""
    u = bool((os.environ.get("RAVEX_USER") or os.environ.get("ravex_user") or "").strip())
    p = bool((os.environ.get("RAVEX_PASSWORD") or os.environ.get("ravex_password") or "").strip())
    # Listar chaves que contêm RAVEX (sem valores) para debug se o Render usar outro nome
    ravex_keys = [k for k in os.environ if "RAVEX" in k.upper()]
    return jsonify({"RAVEX_USER": u, "RAVEX_PASSWORD": p, "ok": u and p, "env_keys_com_ravex": ravex_keys})


@app.route('/api/usuarios', methods=['GET'])
def api_listar_usuarios():
    """Lista todos os usuários cadastrados (apenas nomes; senhas não são exibidas)."""
    conn = get_db()
    rows = conn.execute('SELECT usuario, criado_em FROM usuarios ORDER BY usuario').fetchall()
    conn.close()
    return jsonify([{'usuario': row['usuario'], 'criado_em': row['criado_em'] or ''} for row in rows])


@app.route('/')
def raiz():
    """Raiz do site: sempre abre na página de login."""
    return redirect(url_for('login'))


@app.route('/api/health')
def api_health():
    """Health check leve para proxy/hospedagem (Render)."""
    return jsonify({'ok': True, 'env': _app_env()}), 200


@app.route('/manifest.webmanifest')
def manifest_app():
    """Manifest PWA — ícone na tela inicial e instalação na barra de tarefas."""
    icons = [
        {
            'src': url_for('static', filename='icons/icon-192.png'),
            'sizes': '192x192',
            'type': 'image/png',
            'purpose': 'any',
        },
        {
            'src': url_for('static', filename='icons/icon-512.png'),
            'sizes': '512x512',
            'type': 'image/png',
            'purpose': 'any',
        },
        {
            'src': url_for('static', filename='icons/icon-512.png'),
            'sizes': '512x512',
            'type': 'image/png',
            'purpose': 'maskable',
        },
    ]
    return jsonify({
        'name': '%s WMS' % SYSTEM_NAME,
        'short_name': SYSTEM_NAME,
        'description': SYSTEM_TAGLINE,
        'start_url': url_for('entrada_modulos'),
        'scope': '/',
        'display': 'standalone',
        'background_color': '#f0f2f5',
        'theme_color': '#0369a1',
        'orientation': 'any',
        'lang': 'pt-BR',
        'icons': icons,
    }), 200, {'Content-Type': 'application/manifest+json; charset=utf-8'}


@app.route('/entrada')
def entrada_modulos():
    """Tela inicial pós-login: escolha do módulo (três botões)."""
    return render_template('entrada_modulos.html', usuario=session.get('usuario', ''))


@app.route('/painel')
def painel():
    """Página principal - Painel (requer login)."""
    return render_template('index.html', usuario=session.get('usuario', ''))


@app.route('/api/eventos-stream')
def eventos_stream():
    """Server-Sent Events: envia 'atualizar' quando alguém bipa. Conexão fecha em ~20s (Render free tier limita request ~30s); frontend reconecta."""
    def gerar():
        client_queue = queue.Queue()
        with _sse_lock:
            _sse_queues.append(client_queue)
        try:
            # Fechar em ~20s para ficar abaixo do limite de request do Render (free tier ~30s); cliente reconecta
            max_heartbeats = 2
            for _ in range(max_heartbeats):
                try:
                    msg = client_queue.get(timeout=10)
                    yield f"data: {msg}\n\n"
                except queue.Empty:
                    yield ": heartbeat\n\n"
        finally:
            with _sse_lock:
                if client_queue in _sse_queues:
                    _sse_queues.remove(client_queue)
    return Response(
        gerar(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no', 'Connection': 'keep-alive'}
    )


@app.route('/api/produtos', methods=['GET'])
def get_produtos():
    """Retorna todos os produtos bipados cadastrados no sistema"""
    conn = get_db()
    produtos = conn.execute('SELECT * FROM produtos_bipados ORDER BY data_hora DESC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in produtos])

def _coluna_base_eh_codigo_barras(header_str):
    """Coluna que contém código de barras (EAN, DUN ou Código de Barras)."""
    if not header_str:
        return False
    h = header_str.upper().strip()
    if ('CODIGO' in h or 'CÓDIGO' in h) and 'BARRAS' in h:
        return True
    if 'EAN' in h or 'DUN' in h:
        return True
    if h == 'CODIGO BARRAS' or h == 'CÓDIGO BARRAS':
        return True
    return False


def _coluna_base_eh_ean(header_str):
    """Coluna Cod. EAN-13."""
    if not header_str:
        return False
    h = header_str.upper().strip()
    return 'EAN' in h and '13' in h


def _coluna_base_eh_dun(header_str):
    """Coluna Cod. DUN-14."""
    if not header_str:
        return False
    h = header_str.upper().strip()
    return 'DUN' in h and '14' in h


def _coluna_base_eh_codigo_interno(header_str):
    """Coluna Codigo (código do produto, sem ser barras/EAN/DUN)."""
    if not header_str:
        return False
    h = header_str.upper().strip()
    if h == 'CODIGO' or h == 'CÓDIGO':
        return True
    if 'INTERNO' in h and ('CODIGO' in h or 'CÓDIGO' in h):
        return True
    if 'PRODUTO' in h and ('CODIGO' in h or 'CÓDIGO' in h) and 'BARRAS' not in h and 'EAN' not in h and 'DUN' not in h:
        return True
    return False


def _coluna_base_eh_descricao(header_str):
    if not header_str:
        return False
    h = header_str.upper().strip()
    if 'DESCRI' in h or 'DESCRIÇÃO' in h:
        return True
    if h == 'PRODUTO' or h == 'NOME' or h == 'NOME DO PRODUTO':
        return True
    return False


BASE_HEADERS_UI = ['Codigo', 'Descricao', 'Cod. EAN-13', 'Cod. DUN-14', 'Unidade', 'Peso Bruto']


def _parse_base_json_data(val):
    if isinstance(val, dict):
        return val
    if isinstance(val, str) and val.strip():
        try:
            return json.loads(val)
        except Exception:
            pass
    return {}


def _base_row_get(r, key, default=None):
    if r is None:
        return default
    if hasattr(r, 'get'):
        return r.get(key, default)
    try:
        return r[key]
    except (TypeError, KeyError, IndexError):
        return default


def _linha_base_frontend(r):
    """Monta linha da base para o painel: colunas indexadas + fallback no JSONB data."""
    data = _parse_base_json_data(_base_row_get(r, 'data'))

    def pick(col_name, *json_keys):
        val = _base_row_get(r, col_name)
        if val is not None and str(val).strip():
            return str(val).strip()
        for k in json_keys:
            v = data.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ''

    return {
        'Codigo': pick('codigo_interno', 'Codigo', 'codigo_interno'),
        'Descricao': pick('descricao', 'Descricao', 'descricao', 'Descrição'),
        'Cod. EAN-13': pick('ean', 'Cod. EAN-13', 'EAN-13', 'ean', 'EAN'),
        'Cod. DUN-14': pick('dun', 'Cod. DUN-14', 'DUN-14', 'dun', 'DUN'),
        'Unidade': pick('unidade', 'Unidade', 'unidade', 'und_ean'),
        'Peso Bruto': pick('peso', 'Peso Bruto', 'Peso', 'peso'),
        '_id': _base_row_get(r, 'id'),
    }


def _base_linha_passa_filtros(linha, filtro_codigo, filtro_codigo_interno, filtro_descricao, filtro_ean, filtro_dun, filtro_unidade):
    if filtro_codigo:
        fc = filtro_codigo.upper()
        cb = (linha.get('Cod. EAN-13') or '') + (linha.get('Cod. DUN-14') or '') + (linha.get('Codigo') or '')
        if fc not in cb.upper():
            return False
    if filtro_codigo_interno and filtro_codigo_interno.upper() not in (linha.get('Codigo') or '').upper():
        return False
    if filtro_descricao and filtro_descricao.upper() not in (linha.get('Descricao') or '').upper():
        return False
    if filtro_ean and filtro_ean.upper() not in (linha.get('Cod. EAN-13') or '').upper():
        return False
    if filtro_dun and filtro_dun.upper() not in (linha.get('Cod. DUN-14') or '').upper():
        return False
    if filtro_unidade and filtro_unidade.upper() not in (linha.get('Unidade') or '').upper():
        return False
    return True


@app.route('/api/base-planilha', methods=['GET'])
def get_base_planilha():
    """Retorna dados da aba BASE: do banco (base_codigo_barras) quando DATABASE_URL está definido; senão da planilha."""
    if _usa_banco_para_dados():
        conn = get_db()
        try:
            ds = _get_latest_dataset_id(conn)
            if not ds:
                return jsonify({'headers': BASE_HEADERS_UI, 'rows': []})
            filtro_codigo = request.args.get('codigo_barras', '').strip()
            filtro_codigo_interno = request.args.get('codigo_interno', '').strip()
            filtro_descricao = request.args.get('descricao', '').strip()
            filtro_ean = request.args.get('ean', '').strip()
            filtro_dun = request.args.get('dun', '').strip()
            filtro_unidade = request.args.get('unidade', '').strip()
            sql = """SELECT id, codigo_interno, ean, dun, descricao, unidade, peso, data
                     FROM base_codigo_barras WHERE dataset_id = ?"""
            params = [str(ds)]
            if getattr(conn, 'kind', None) == 'pg':
                if filtro_codigo:
                    sql += """ AND (COALESCE(ean, '') ILIKE ? OR COALESCE(dun, '') ILIKE ?
                               OR COALESCE(codigo_interno::text, '') ILIKE ?)"""
                    pct = '%' + filtro_codigo.replace('%', '\\%').replace('_', '\\_') + '%'
                    params.extend([pct, pct, pct])
                if filtro_codigo_interno:
                    sql += """ AND COALESCE(codigo_interno, '') ILIKE ?"""
                    params.append('%' + filtro_codigo_interno.replace('%', '\\%').replace('_', '\\_') + '%')
                if filtro_descricao:
                    sql += """ AND COALESCE(descricao, '') ILIKE ?"""
                    params.append('%' + filtro_descricao.replace('%', '\\%').replace('_', '\\_') + '%')
                if filtro_ean:
                    sql += """ AND COALESCE(ean, '') ILIKE ?"""
                    params.append('%' + filtro_ean.replace('%', '\\%').replace('_', '\\_') + '%')
                if filtro_dun:
                    sql += """ AND COALESCE(dun, '') ILIKE ?"""
                    params.append('%' + filtro_dun.replace('%', '\\%').replace('_', '\\_') + '%')
                if filtro_unidade:
                    sql += """ AND COALESCE(unidade, '') ILIKE ?"""
                    params.append('%' + filtro_unidade.replace('%', '\\%').replace('_', '\\_') + '%')
            sql += """ ORDER BY row_index LIMIT 2000"""
            try:
                rows_raw = conn.execute(sql, params).fetchall()
            except Exception:
                rows_raw = conn.execute(
                    """SELECT id, codigo_interno, ean, dun, descricao, unidade, peso, data
                       FROM base_codigo_barras WHERE dataset_id = ? ORDER BY row_index LIMIT 2000""",
                    (str(ds),),
                ).fetchall()
            rows = []
            for r in rows_raw:
                linha = _linha_base_frontend(r)
                if not any([linha.get('Codigo'), linha.get('Descricao'), linha.get('Cod. EAN-13'), linha.get('Cod. DUN-14')]):
                    continue
                if getattr(conn, 'kind', None) != 'pg' and not _base_linha_passa_filtros(
                    linha, filtro_codigo, filtro_codigo_interno, filtro_descricao, filtro_ean, filtro_dun, filtro_unidade
                ):
                    continue
                rows.append(linha)
            return jsonify({'headers': BASE_HEADERS_UI, 'rows': rows})
        except Exception as e:
            return jsonify({'erro': str(e)}), 500
        finally:
            try:
                conn.close()
            except Exception:
                pass

    # Sem planilha: dados vêm apenas do banco (DATABASE_URL)
    return jsonify({'headers': BASE_HEADERS_UI, 'rows': [], 'erro': 'Configure DATABASE_URL. Base de produtos vem apenas do banco.'})


def _base_item_payload_to_columns(payload):
    """Extrai do payload (dict do frontend) os valores para colunas e para o JSONB data."""
    if not payload or not isinstance(payload, dict):
        return None, None
    data = {str(k): (v.strftime('%Y-%m-%d %H:%M:%S') if isinstance(v, datetime) else v) for k, v in payload.items() if str(k) != '_id'}
    codigo_interno = str(payload.get('Codigo') or payload.get('codigo_interno') or payload.get('Codigo Interno') or '').strip() or None
    ean = str(payload.get('Cod. EAN-13') or payload.get('EAN-13') or payload.get('ean') or payload.get('EAN') or '').strip() or None
    dun = str(payload.get('Cod. DUN-14') or payload.get('DUN-14') or payload.get('dun') or payload.get('DUN') or '').strip() or None
    descricao = str(payload.get('Descricao') or payload.get('descricao') or payload.get('Descrição') or '').strip() or None
    unidade = str(payload.get('Unidade') or payload.get('unidade') or '').strip() or None
    peso = str(payload.get('Peso Bruto') or payload.get('Peso Líquido') or payload.get('Peso') or payload.get('peso') or '').strip() or None
    if not data and not any([codigo_interno, ean, dun, descricao]):
        return None, None
    if not data:
        data = {}
    if codigo_interno is not None:
        data['Codigo'] = codigo_interno
    if descricao is not None:
        data['Descricao'] = descricao
    if unidade is not None:
        data['Unidade'] = unidade
    if ean is not None:
        data['Cod. EAN-13'] = ean
    if dun is not None:
        data['Cod. DUN-14'] = dun
    if peso is not None:
        data['Peso Bruto'] = data.get('Peso Bruto') or peso
    return (codigo_interno, ean, dun, descricao, unidade, peso), data


@app.route('/api/base-item/vincular-codigo', methods=['POST'])
def api_base_item_vincular_codigo():
    """Vincula código bipado (EAN/DUN) a um código interno do romaneio e atualiza base_codigo_barras."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL.'}), 400
    payload = request.get_json() or {}
    codigo_interno = (payload.get('codigo_interno') or payload.get('codigo_produto') or '').strip()
    codigo_barras = _normalizar_codigo_barras(payload.get('codigo_barras') or payload.get('ean') or '')
    dun_in = re.sub(r'\D', '', str(payload.get('dun') or '').strip()) or None
    if dun_in == '':
        dun_in = None
    descricao = (payload.get('descricao') or payload.get('produto') or '').strip() or None
    unidade = (payload.get('unidade') or '').strip() or None
    peso = (payload.get('peso') or '').strip() or None
    tipo = (payload.get('tipo_codigo') or '').strip().upper()
    if not codigo_interno:
        return jsonify({'erro': 'Informe o código interno (item da lista).'}), 400
    if not codigo_barras and not dun_in:
        return jsonify({'erro': 'Informe o código de barras bipado (EAN ou DUN).'}), 400
    if not tipo:
        tipo = 'DUN' if len(codigo_barras or '') >= 14 else 'EAN'
    if tipo == 'EAN':
        ean_val = codigo_barras or None
        dun_val = dun_in
    else:
        ean_val = None
        dun_val = dun_in or codigo_barras or None
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) != 'pg':
            conn.close()
            return jsonify({'erro': 'Vincular código só disponível com Postgres (DATABASE_URL).'}), 400
        ds = _get_latest_dataset_id(conn)
        if not ds:
            conn.close()
            return jsonify({'erro': 'Nenhum dataset ativo. Importe a base primeiro.'}), 400
        rows = conn.execute(
            """SELECT id, ean, dun, descricao, unidade, peso, data
               FROM base_codigo_barras
               WHERE dataset_id = ? AND TRIM(COALESCE(codigo_interno, '')) = ?
               ORDER BY id""",
            (str(ds), codigo_interno),
        ).fetchall()
        item_id = None
        agora_iso = datetime.now(timezone.utc).isoformat()

        def _merge_data_vinculo(data_old):
            if isinstance(data_old, str):
                try:
                    data_old = json.loads(data_old)
                except Exception:
                    data_old = {}
            if not isinstance(data_old, dict):
                data_old = {}
            data_old['vinculado_de_conferencia'] = True
            data_old['vinculado_em'] = agora_iso
            data_old['Codigo'] = codigo_interno
            if descricao:
                data_old['Descricao'] = descricao
            if unidade:
                data_old['Unidade'] = unidade
            if ean_val:
                data_old['Cod. EAN-13'] = ean_val
            if dun_val:
                data_old['Cod. DUN-14'] = dun_val
            if peso:
                data_old['Peso Bruto'] = peso
            return data_old

        if rows:
            for alvo in rows:
                rid = alvo.get('id') if hasattr(alvo, 'get') else alvo[0]
                if item_id is None:
                    item_id = rid
                ean_at = str((alvo.get('ean') if hasattr(alvo, 'get') else (alvo[1] if len(alvo) > 1 else '')) or '').strip()
                dun_at = str((alvo.get('dun') if hasattr(alvo, 'get') else (alvo[2] if len(alvo) > 2 else '')) or '').strip()
                data_old = alvo.get('data') if hasattr(alvo, 'get') else (alvo[6] if len(alvo) > 6 else {})
                data_merged = _merge_data_vinculo(data_old)
                if tipo == 'EAN':
                    new_ean = ean_val or ean_at or None
                    new_dun = dun_val if dun_val else (dun_at or None)
                else:
                    new_dun = dun_val or dun_at or None
                    new_ean = ean_val if ean_val else (ean_at or None)
                conn.execute(
                    """UPDATE base_codigo_barras
                       SET ean = ?,
                           dun = ?,
                           descricao = COALESCE(?, descricao),
                           unidade = COALESCE(?, unidade),
                           peso = COALESCE(?, peso),
                           data = ?::jsonb
                       WHERE id = ?""",
                    (
                        new_ean,
                        new_dun,
                        descricao,
                        unidade,
                        peso,
                        json.dumps(data_merged, ensure_ascii=False, default=str),
                        rid,
                    ),
                )
        else:
            data_new = {
                'Codigo': codigo_interno,
                'Descricao': descricao or '',
                'Unidade': unidade or '',
                'Cod. EAN-13': ean_val or '',
                'Cod. DUN-14': dun_val or '',
                'vinculado_de_conferencia': True,
            }
            next_row = conn.execute(
                """SELECT COALESCE(MAX(row_index), 0) + 1 AS next_idx FROM base_codigo_barras WHERE dataset_id = ?""",
                (str(ds),),
            ).fetchone()
            row_index = next_row['next_idx'] if next_row else 1
            row_ins = conn.execute(
                """INSERT INTO base_codigo_barras
                   (dataset_id, row_index, codigo_interno, ean, dun, descricao, unidade, peso, data)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?::jsonb)
                   RETURNING id""",
                (
                    str(ds),
                    row_index,
                    codigo_interno,
                    ean_val,
                    dun_val,
                    descricao,
                    unidade,
                    peso,
                    json.dumps(data_new, ensure_ascii=False, default=str),
                ),
            ).fetchone()
            item_id = row_ins.get('id') if row_ins and hasattr(row_ins, 'get') else (row_ins[0] if row_ins else None)
        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'item_id': item_id,
            'mensagem': 'Código vinculado na base de produtos.',
            'tipo_codigo': tipo,
            'ean': ean_val,
            'dun': dun_val,
        })
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


@app.route('/api/base-item', methods=['POST'])
def api_base_item_create():
    """Cria um novo registro na base_codigo_barras (dataset ativo)."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL. Base de produtos vem apenas do banco.'}), 400
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) != 'pg':
            conn.close()
            return jsonify({'erro': 'Adicionar/editar base só disponível com Postgres (DATABASE_URL).'}), 400
        ds = _get_latest_dataset_id(conn)
        if not ds:
            conn.close()
            return jsonify({'erro': 'Nenhum dataset ativo. Importe uma planilha BASE primeiro.'}), 400
        payload = request.get_json() or {}
        cols, data = _base_item_payload_to_columns(payload)
        if cols is None:
            conn.close()
            return jsonify({'erro': 'Dados inválidos. Informe ao menos código ou descrição.'}), 400
        codigo_interno, ean, dun, descricao, unidade, peso = cols
        next_row = conn.execute(
            """SELECT COALESCE(MAX(row_index), 0) + 1 AS next_idx FROM base_codigo_barras WHERE dataset_id = ?""",
            (str(ds),),
        ).fetchone()
        row_index = next_row['next_idx'] if next_row else 1
        conn.execute(
            """INSERT INTO base_codigo_barras (dataset_id, row_index, codigo_interno, ean, dun, descricao, unidade, peso, data)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (str(ds), row_index, codigo_interno, ean, dun, descricao, unidade, peso, json.dumps(data, ensure_ascii=False, default=str)),
        )
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'mensagem': 'Produto cadastrado na base.'})
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


@app.route('/api/base-item/<int:item_id>', methods=['PUT'])
def api_base_item_update(item_id):
    """Atualiza um registro da base_codigo_barras."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL.'}), 400
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) != 'pg':
            conn.close()
            return jsonify({'erro': 'Editar base só disponível com Postgres.'}), 400
        payload = request.get_json() or {}
        cols, data = _base_item_payload_to_columns(payload)
        if cols is None:
            conn.close()
            return jsonify({'erro': 'Dados inválidos.'}), 400
        codigo_interno, ean, dun, descricao, unidade, peso = cols
        cur = conn.execute(
            """UPDATE base_codigo_barras SET codigo_interno = ?, ean = ?, dun = ?, descricao = ?, unidade = ?, peso = ?, data = ?
               WHERE id = ?""",
            (codigo_interno, ean, dun, descricao, unidade, peso, json.dumps(data, ensure_ascii=False, default=str), item_id),
        )
        n = getattr(cur, 'rowcount', 0) or 0
        conn.commit()
        conn.close()
        if n > 0:
            return jsonify({'ok': True, 'mensagem': 'Registro atualizado.'})
        return jsonify({'erro': 'Registro não encontrado.'}), 404
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


@app.route('/api/base-item/<int:item_id>', methods=['DELETE'])
def api_base_item_delete(item_id):
    """Exclui um registro da base_codigo_barras."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL.'}), 400
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) != 'pg':
            conn.close()
            return jsonify({'erro': 'Excluir da base só disponível com Postgres.'}), 400
        cur = conn.execute("""DELETE FROM base_codigo_barras WHERE id = ?""", (item_id,))
        n = getattr(cur, 'rowcount', 0) or 0
        conn.commit()
        conn.close()
        if n > 0:
            return jsonify({'ok': True, 'mensagem': 'Registro excluído.'})
        return jsonify({'erro': 'Registro não encontrado.'}), 404
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


def _normalizar_codigo_barras(codigo):
    """Se o leitor ler dois códigos grudados (um na frente do outro), usa só o primeiro."""
    if not codigo or len(codigo) < 4:
        return codigo
    s = str(codigo).strip()
    n = len(s)
    if n % 2 != 0:
        return s
    metade = n // 2
    if s[:metade] == s[metade:]:
        return s[:metade]
    return s


@app.route('/api/produtos', methods=['POST'])
def add_produto():
    """Adiciona um novo produto bipado. Se não estiver cadastrado na conferência, retorna produto_nao_cadastrado para o frontend perguntar se deseja adicionar."""
    data = request.json or {}
    forcar_adicionar = data.get('forcar_adicionar', False)
    fluxo = (data.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        _ensure_pg_produtos_bipados_fluxo(conn)
    _ensure_produtos_bipados_disposicao(conn)
    _ensure_devolucao_nf_schema(conn)
    codigo_barras = _normalizar_codigo_barras(data.get('codigo_barras', '') or '')
    id_viagem = (data.get('id_viagem') or '').strip()
    devolucao_nf_id = data.get('devolucao_nf_id')
    if fluxo == 'devolucao':
        try:
            devolucao_nf_id = int(devolucao_nf_id) if devolucao_nf_id is not None else None
        except (TypeError, ValueError):
            devolucao_nf_id = None
        if not devolucao_nf_id:
            conn.close()
            return jsonify({
                'success': False,
                'mensagem': 'Selecione a NF e o motivo da devolução antes de bipar o retorno.',
            }), 400
        nf_chk = conn.execute(
            "SELECT id, status FROM devolucao_nota_fiscal WHERE id = ?",
            (devolucao_nf_id,),
        ).fetchone()
        if not nf_chk:
            conn.close()
            return jsonify({'success': False, 'mensagem': 'NF de devolução não encontrada.'}), 400
        nf_st = (nf_chk.get('status') if hasattr(nf_chk, 'get') else nf_chk[1]) or ''
        if nf_st != 'em_andamento':
            conn.close()
            return jsonify({'success': False, 'mensagem': 'Esta NF já foi concluída. Inicie outra NF para bipar.'}), 400

    if id_viagem and not forcar_adicionar and fluxo != 'devolucao':
        g.conferencia_fluxo = fluxo
        try:
            ret = get_conferencia(id_viagem)
        finally:
            try:
                del g.conferencia_fluxo
            except Exception:
                pass
        resp = ret[0] if isinstance(ret, tuple) else ret
        body = resp.get_json() if hasattr(resp, 'get_json') else {}
        lista = body.get('lista') if isinstance(body, dict) else body
        if isinstance(lista, list):
            codigos_barras_conf = {
                str(item.get('codigo_barras') or '').strip()
                for item in lista
                if str(item.get('codigo_barras') or '').strip()
            }
            codigos_produto_conf = {
                str(item.get('codigo_produto') or '').strip()
                for item in lista
                if str(item.get('codigo_produto') or '').strip()
            }
            codigo_interno_chk = (data.get('codigo_interno') or '').strip()
            permitido = False
            if codigo_barras and codigo_barras in codigos_barras_conf:
                permitido = True
            elif codigo_interno_chk and codigo_interno_chk in codigos_produto_conf:
                permitido = True
            elif codigo_barras:
                info_conf = buscar_produto_na_planilha(codigo_barras)
                if info_conf and not info_conf.get('erro'):
                    cp_conf = str(info_conf.get('codigo_produto') or '').strip()
                    if cp_conf and cp_conf in codigos_produto_conf:
                        permitido = True
            if codigo_barras and not permitido:
                conn.close()
                return jsonify({
                    'success': False,
                    'produto_nao_cadastrado': True,
                    'mensagem': 'Código não reconhecido nesta viagem. Use «cadastrar» ao lado do campo ou vincule ao romaneio.',
                })

    doca = (data.get('doca') or '').strip()
    if fluxo == 'devolucao' and doca not in ('1', '2', '3', '4'):
        doca = '1'
    if doca not in ('1', '2', '3', '4'):
        return jsonify({'success': False, 'mensagem': 'Selecione a doca (1, 2, 3 ou 4) antes de bipar.'}), 400
    codigo_interno = (data.get('codigo_interno') or '').strip()
    codigo_dun = (data.get('codigo_dun') or '').strip()
    unidade = (data.get('unidade') or data.get('Unidade') or '').strip()
    peso = (data.get('peso') or '').strip()
    # EAN = Caixa; DUN = Pacote/Unidade conforme coluna Unidade do Romaneio por Item
    if not unidade and codigo_barras:
        info = buscar_produto_na_planilha(codigo_barras)
        if info and not info.get('erro'):
            if not codigo_interno and info.get('codigo_produto'):
                codigo_interno = str(info.get('codigo_produto', '')).strip()
            tipo = (info.get('tipo_codigo') or '').upper()
            if tipo == 'EAN':
                unidade = 'Caixa'
            elif tipo == 'DUN' and codigo_interno and id_viagem:
                unidade = get_unidade_romaneio(id_viagem, codigo_interno) or 'Pacote'
            if not unidade:
                unidade = get_unidade_romaneio(id_viagem, codigo_interno) if (codigo_interno and id_viagem) else 'Unidade'
    usuario_logado = session.get('usuario', '') or ''
    try:
        qtd_bipar = int(data.get('quantidade', 1))
    except (TypeError, ValueError):
        qtd_bipar = 1
    qtd_bipar = max(1, min(99999, qtd_bipar))
    disposicao_estoque = _normalizar_disposicao_estoque(
        data.get('disposicao_estoque'),
        fluxo=fluxo,
        devolucao_nf_id=devolucao_nf_id if fluxo == 'devolucao' else None,
        conn=conn,
    )
    conn.execute(
        '''INSERT INTO produtos_bipados 
           (codigo_barras, produto, quantidade, data_hora, veiculo, status, id_viagem, doca, codigo_interno, codigo_dun, unidade, peso, usuario_bipagem, fluxo, devolucao_nf_id, disposicao_estoque)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (
            codigo_barras,
            data.get('produto', ''),
            qtd_bipar,
            datetime.now(timezone.utc),
            data.get('veiculo', ''),
            data.get('status', 'PENDENTE'),
            id_viagem,
            doca,
            codigo_interno,
            codigo_dun,
            unidade,
            peso,
            usuario_logado,
            fluxo,
            devolucao_nf_id if fluxo == 'devolucao' else None,
            disposicao_estoque,
        )
    )
    if id_viagem:
        _registrar_momento_bipagem(conn, id_viagem, fluxo, datetime.now(timezone.utc))
    conn.commit()
    conn.close()
    _broadcast_atualizar()
    return jsonify({'success': True})

@app.route('/api/produtos/<int:produto_id>', methods=['PUT'])
def update_produto(produto_id):
    """Atualiza um produto"""
    data = request.json
    conn = get_db()
    
    conn.execute(
        '''UPDATE produtos_bipados 
           SET produto = ?, quantidade = ?, veiculo = ?, status = ?
           WHERE id = ?''',
        (
            data.get('produto', ''),
            data.get('quantidade', 1),
            data.get('veiculo', ''),
            data.get('status', 'PENDENTE'),
            produto_id
        )
    )
    
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/produtos/<int:produto_id>', methods=['DELETE'])
def delete_produto(produto_id):
    """Remove um produto"""
    conn = get_db()
    conn.execute('DELETE FROM produtos_bipados WHERE id = ?', (produto_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/conferencia/<id_viagem>/zerar', methods=['DELETE', 'POST'])
def zerar_bipagem_viagem(id_viagem):
    """Remove todos os registros de bipagem da viagem para permitir bipar novamente"""
    data = request.get_json(silent=True) or {}
    id_norm = _normalizar_id_viagem(
        id_viagem or data.get('id_viagem') or request.args.get('id_viagem')
    )
    if not id_norm:
        return jsonify({'erro': 'ID do roteiro não informado'}), 400
    fluxo = (request.args.get('fluxo') or data.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) == 'pg':
            cur = conn.execute(
                """DELETE FROM produtos_bipados
                   WHERE TRIM(COALESCE(id_viagem::text, '')) = ?
                     AND COALESCE(fluxo, 'carregamento') = ?""",
                (id_norm, fluxo),
            )
            removidos = getattr(cur, 'rowcount', None)
        else:
            cur = conn.execute(
                "DELETE FROM produtos_bipados WHERE id_viagem = ? AND COALESCE(fluxo, 'carregamento') = ?",
                (id_norm, fluxo),
            )
            removidos = cur.rowcount if cur else None
        _limpar_periodo_bipagem(conn, id_norm, fluxo)
        conn.commit()
    finally:
        conn.close()
    _broadcast_atualizar()
    return jsonify({
        'success': True,
        'mensagem': 'Bipagem zerada. Você pode bipar novamente.',
        'removidos': removidos,
    })


@app.route('/api/conferencia/remover', methods=['POST'])
def remover_itens_bipados():
    """Remove itens bipados (quando bipou errado). Body: id_viagem, codigo_barras, quantidade (ou 'tudo')"""
    data = request.json or {}
    id_viagem = (data.get('id_viagem') or '').strip()
    codigo_barras = (data.get('codigo_barras') or '').strip()
    quantidade = data.get('quantidade', 1)
    fluxo = (data.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    if not id_viagem or not codigo_barras:
        return jsonify({'erro': 'id_viagem e codigo_barras são obrigatórios'}), 400
    conn = get_db()

    def _row_id(r):
        return r.get('id') if hasattr(r, 'get') else (r[0] if len(r) > 0 else None)

    def _row_qtd(r):
        return r.get('quantidade') if hasattr(r, 'get') else (r[1] if len(r) > 1 else 0)

    rows = conn.execute(
        "SELECT id, quantidade FROM produtos_bipados WHERE id_viagem = ? AND codigo_barras = ? AND COALESCE(fluxo, 'carregamento') = ? ORDER BY id DESC",
        (id_viagem, codigo_barras, fluxo),
    ).fetchall()
    if not rows:
        conn.close()
        return jsonify({'success': True, 'mensagem': 'Nenhum item bipado para este produto nesta viagem.', 'removidos': 0})
    if quantidade == 'tudo' or quantidade == 'all':
        qtd_remover = sum(_row_qtd(r) for r in rows)
        conn.execute(
            "DELETE FROM produtos_bipados WHERE id_viagem = ? AND codigo_barras = ? AND COALESCE(fluxo, 'carregamento') = ?",
            (id_viagem, codigo_barras, fluxo),
        )
        conn.commit()
        conn.close()
        _broadcast_atualizar()
        return jsonify({'success': True, 'mensagem': f'{qtd_remover} unidade(s) removida(s).', 'removidos': qtd_remover})
    try:
        qtd_remover = int(quantidade)
    except (TypeError, ValueError):
        qtd_remover = 1
    if qtd_remover <= 0:
        conn.close()
        return jsonify({'success': True, 'removidos': 0})
    removidos = 0
    for row in rows:
        if removidos >= qtd_remover:
            break
        rid, qtd = _row_id(row), _row_qtd(row)
        falta_remover = qtd_remover - removidos
        if qtd <= falta_remover:
            conn.execute('DELETE FROM produtos_bipados WHERE id = ?', (rid,))
            removidos += qtd
        else:
            conn.execute('UPDATE produtos_bipados SET quantidade = quantidade - ? WHERE id = ?', (falta_remover, rid))
            removidos += falta_remover
    conn.commit()
    conn.close()
    _broadcast_atualizar()
    return jsonify({'success': True, 'mensagem': f'{removidos} unidade(s) removida(s).', 'removidos': removidos})


_PRODUTOS_BIPADOS_INSERT_SQL = '''INSERT INTO produtos_bipados
   (codigo_barras, produto, quantidade, data_hora, veiculo, status, id_viagem, doca,
    codigo_interno, codigo_dun, unidade, peso, usuario_bipagem, fluxo, disposicao_estoque)
   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''

_PRODUTOS_BIPADOS_INSERT_PG_TEMPLATE = (
    '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
)
_PRODUTOS_BIPADOS_INSERT_PG_SQL = """INSERT INTO produtos_bipados
   (codigo_barras, produto, quantidade, data_hora, veiculo, status, id_viagem, doca,
    codigo_interno, codigo_dun, unidade, peso, usuario_bipagem, fluxo, disposicao_estoque)
   VALUES %s"""


def _mapa_codigo_interno_por_barras(conn, codigos_barras):
    """Resolve codigo_interno por EAN/DUN em lote (evita N queries em gravar-bipagem)."""
    codigos = []
    vistos = set()
    for cb in codigos_barras or []:
        s = str(cb or '').strip()
        if s and s != '-' and s not in vistos:
            vistos.add(s)
            codigos.append(s)
    if not codigos or not _usa_banco_para_dados():
        return {}
    ds = _get_latest_dataset_id(conn)
    if not ds:
        return {}
    mapa = {}
    chunk_size = 150
    for i in range(0, len(codigos), chunk_size):
        chunk = codigos[i:i + chunk_size]
        ph = ','.join(['?'] * len(chunk))
        try:
            rows = conn.execute(
                """SELECT codigo_interno, ean, dun FROM base_codigo_barras
                   WHERE dataset_id = ?
                     AND (
                       TRIM(COALESCE(ean, '')) IN (""" + ph + """)
                       OR TRIM(COALESCE(dun, '')) IN (""" + ph + """)
                     )""",
                [str(ds)] + chunk + chunk,
            ).fetchall()
        except Exception:
            continue
        for row in rows or []:
            r = dict(row) if hasattr(row, 'keys') else row
            ci = str((r.get('codigo_interno') if hasattr(r, 'get') else (r[0] if len(r) > 0 else '')) or '').strip()
            if not ci:
                continue
            for campo in ('ean', 'dun'):
                val = str((r.get(campo) if hasattr(r, 'get') else '') or '').strip()
                if not val:
                    continue
                mapa[val] = ci
                dig = re.sub(r'\D', '', val)
                if dig:
                    mapa[dig] = ci
    for cb in codigos:
        if cb in mapa:
            continue
        dig = re.sub(r'\D', '', cb)
        if dig and dig in mapa:
            mapa[cb] = mapa[dig]
    return mapa


def _mapa_unidades_romaneio_viagem(conn, id_viagem):
    """Mapa codigo_produto -> unidade do romaneio (uma query; evita abrir planilha por item)."""
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm or not _usa_banco_para_dados():
        return {}
    mapa = {}
    try:
        ds = _get_latest_dataset_id(conn)
        if not ds:
            return {}
        if getattr(conn, 'kind', None) == 'pg':
            sql = """SELECT codigo_produto, unidade FROM romaneio_por_item
                     WHERE dataset_id = ? AND (
                       TRIM(COALESCE(id_viagem::text, '')) = ?
                       OR TRIM(COALESCE(id_roteiro::text, '')) = ?
                     )"""
        else:
            sql = """SELECT codigo_produto, unidade FROM romaneio_por_item
                     WHERE dataset_id = ? AND (
                       TRIM(COALESCE(id_viagem, '')) = ?
                       OR TRIM(COALESCE(id_roteiro, '')) = ?
                     )"""
        rows = conn.execute(sql, (str(ds), id_norm, id_norm)).fetchall()
        for row in rows:
            r = dict(row) if hasattr(row, 'keys') else row
            cp = str((r.get('codigo_produto') if hasattr(r, 'get') else r[0]) or '').strip()
            un = str((r.get('unidade') if hasattr(r, 'get') else (r[1] if len(r) > 1 else '')) or '').strip()
            if not cp or not un:
                continue
            if cp not in mapa:
                mapa[cp] = un
            cn = _normalizar_codigo_produto(cp)
            if cn and cn not in mapa:
                mapa[cn] = un
    except Exception:
        pass
    return mapa


def _bulk_insert_produtos_bipados(conn, params):
    """INSERT em lote (execute_values no Postgres)."""
    if not params:
        return
    if len(params) == 1:
        conn.execute(_PRODUTOS_BIPADOS_INSERT_SQL, params[0])
        return
    if getattr(conn, 'kind', None) == 'pg':
        try:
            from psycopg.extras import execute_values  # type: ignore
            raw = getattr(conn, '_conn', conn)
            with raw.cursor() as cur:
                execute_values(
                    cur,
                    _PRODUTOS_BIPADOS_INSERT_PG_SQL,
                    params,
                    template=_PRODUTOS_BIPADOS_INSERT_PG_TEMPLATE,
                    page_size=500,
                )
            return
        except Exception:
            pass
    conn.executemany(_PRODUTOS_BIPADOS_INSERT_SQL, params)


@app.route('/api/conferencia/<id_viagem>/gravar-bipagem', methods=['POST'])
def gravar_bipagem_viagem(id_viagem):
    """Grava de uma vez a bipagem exibida na tela (ao gerar comprovante, após rascunho local)."""
    id_norm = _normalizar_id_viagem(id_viagem)
    if not id_norm:
        return jsonify({'success': False, 'erro': 'ID do roteiro não informado'}), 400
    data = request.json or {}
    fluxo = (data.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    doca = (data.get('doca') or '').strip()
    if doca not in ('1', '2', '3', '4'):
        return jsonify({'success': False, 'erro': 'Selecione a doca (1, 2, 3 ou 4) antes de gerar o comprovante.'}), 400
    veiculo = (data.get('veiculo') or '').strip()
    status = (data.get('status') or 'PENDENTE').strip() or 'PENDENTE'
    usuario_logado = session.get('usuario', '') or ''
    itens = data.get('itens')
    if not isinstance(itens, list):
        itens = []
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        _ensure_pg_produtos_bipados_fluxo(conn)
    _ensure_produtos_bipados_disposicao(conn)
    gravados = 0
    pulados = 0
    agora = datetime.now(timezone.utc)
    mapa_unidades = _mapa_unidades_romaneio_viagem(conn, id_norm)
    precisa_lookup = []
    for it in itens:
        if not isinstance(it, dict):
            continue
        try:
            qtd = int(it.get('quantidade') or 0)
        except (TypeError, ValueError):
            qtd = 0
        if qtd <= 0:
            continue
        cb_pre = _normalizar_codigo_barras(str(it.get('codigo_barras') or '').strip())
        ci_pre = str(it.get('codigo_interno') or it.get('codigo_produto') or '').strip()
        if (not cb_pre or cb_pre == '-') and ci_pre:
            cb_pre = ci_pre
        if not ci_pre and cb_pre and cb_pre != '-':
            precisa_lookup.append(cb_pre)
    mapa_codigo_interno = _mapa_codigo_interno_por_barras(conn, precisa_lookup)
    params_insert = []
    try:
        _pg_auditoria_sync_desligar(conn)
        if getattr(conn, 'kind', None) == 'pg':
            conn.execute(
                """DELETE FROM produtos_bipados
                   WHERE TRIM(COALESCE(id_viagem::text, '')) = ?
                     AND COALESCE(fluxo, 'carregamento') = ?""",
                (id_norm, fluxo),
            )
        else:
            conn.execute(
                "DELETE FROM produtos_bipados WHERE id_viagem = ? AND COALESCE(fluxo, 'carregamento') = ?",
                (id_norm, fluxo),
            )
        for it in itens:
            if not isinstance(it, dict):
                pulados += 1
                continue
            try:
                qtd = int(it.get('quantidade') or 0)
            except (TypeError, ValueError):
                qtd = 0
            if qtd <= 0:
                continue
            qtd = max(1, min(99999, qtd))
            cb = _normalizar_codigo_barras(str(it.get('codigo_barras') or '').strip())
            codigo_interno = str(it.get('codigo_interno') or it.get('codigo_produto') or '').strip()
            if (not cb or cb == '-') and codigo_interno:
                cb = codigo_interno
            if not cb or cb == '-':
                pulados += 1
                continue
            produto = str(it.get('produto') or '').strip()
            codigo_dun = str(it.get('codigo_dun') or '').strip()
            unidade = str(it.get('unidade') or '').strip()
            if unidade in ('', '-'):
                unidade = ''
            peso = str(it.get('peso') or '').strip()
            if not codigo_interno and cb and cb != '-':
                codigo_interno = (
                    mapa_codigo_interno.get(cb)
                    or mapa_codigo_interno.get(re.sub(r'\D', '', cb))
                    or ''
                )
            if not unidade and codigo_interno:
                unidade = (
                    mapa_unidades.get(codigo_interno)
                    or mapa_unidades.get(_normalizar_codigo_produto(codigo_interno) or '')
                    or ''
                )
            disposicao_estoque = _normalizar_disposicao_estoque(it.get('disposicao_estoque'), fluxo=fluxo)
            params_insert.append(
                (
                    cb,
                    produto,
                    qtd,
                    agora,
                    veiculo,
                    status,
                    id_norm,
                    doca,
                    codigo_interno,
                    codigo_dun,
                    unidade,
                    peso,
                    usuario_logado,
                    fluxo,
                    disposicao_estoque,
                )
            )
            gravados += 1
        _bulk_insert_produtos_bipados(conn, params_insert)
        if gravados > 0:
            inicio_p = _parse_datetime_iso(data.get('inicio_carregamento'))
            fim_p = _parse_datetime_iso(data.get('fim_carregamento'))
            agora_fim = datetime.now(timezone.utc)
            if not fim_p:
                fim_p = agora_fim
            if inicio_p and fim_p and fim_p <= inicio_p:
                fim_p = inicio_p + timedelta(seconds=1)
            if inicio_p and fim_p:
                _definir_periodo_bipagem(conn, id_norm, fluxo, inicio_p, fim_p)
            else:
                _definir_periodo_bipagem(conn, id_norm, fluxo, agora_fim - timedelta(seconds=1), agora_fim)
            _marcar_extrato_gerado(conn, id_norm, fluxo, agora_fim)
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'erro': 'Erro ao gravar bipagem: %s' % str(e)}), 500
    finally:
        conn.close()
    _broadcast_atualizar()
    return jsonify({
        'success': True,
        'gravados': gravados,
        'pulados': pulados,
        'mensagem': 'Bipagem gravada (%s item(ns)).' % gravados if gravados else 'Nenhum item bipado para gravar.',
    })


@app.route('/api/viagem/<id_viagem>/periodo', methods=['GET'])
def get_periodo_viagem(id_viagem):
    """Retorna início e fim da bipagem/carregamento (data e hora) para a viagem."""
    if not id_viagem:
        return jsonify({'erro': 'ID do roteiro não informado'}), 400
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    return jsonify(_periodo_viagem_resposta(id_viagem, fluxo))


def _normalizar_id_viagem(val):
    """ID do roteiro (ou viagem) pode vir como número (555555 ou 555555.0) ou texto; retorna string unificada para comparação."""
    if val is None:
        return ''
    if isinstance(val, float):
        return str(int(val)) if val == int(val) else str(val).strip()
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _normalizar_codigo_produto(val):
    """Código do produto pode vir como número (3000, 3000.0) ou texto ('01.06.0001'); retorna string unificada para busca no mapa."""
    if val is None:
        return ''
    s = str(val).strip()
    if not s or s.startswith('='):
        return ''
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


def _variantes_codigo_produto(codigo):
    """Retorna lista de variantes do código para busca na BASE (ex.: '01.01.0001' -> ['01.01.0001', '1.1.1'])."""
    if not codigo or not isinstance(codigo, str):
        return []
    s = codigo.strip()
    if not s:
        return []
    variantes = [s]
    # Formato XX.XX.XXXX: também tentar sem zeros à esquerda (1.1.1)
    if '.' in s and s.replace('.', '').isdigit():
        partes = s.split('.')
        compact = '.'.join(str(int(p)) for p in partes if p.strip())
        if compact != s:
            variantes.append(compact)
    return variantes


def _formatar_data_celula(val):
    """Formata valor de célula como data DD/MM/YYYY quando for data."""
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%d/%m/%Y')
    s = str(val).strip()
    try:
        if isinstance(val, (int, float)):
            from datetime import datetime as dt, timedelta
            base = dt(1899, 12, 30)
            d = base + timedelta(days=int(float(val)))
            return d.strftime('%d/%m/%Y')
        if len(s) >= 8 and ('/' in s or '-' in s):
            return s
    except Exception:
        pass
    return s


def _get_viagem_info_planilha(id_viagem, conn=None):
    """Busca data expedição, placa, motorista, id_roteiro, identificador_rota: do banco. Sem planilha."""
    id_viagem = (id_viagem or '').strip()
    result = {'data_expedicao': '', 'placa': '', 'identificador_rota': '', 'motorista': '', 'id_roteiro': '', 'id_viagem': id_viagem}
    if not id_viagem:
        return result
    id_norm = _normalizar_id_viagem(id_viagem)
    if _usa_banco_para_dados():
        own_conn = conn is None
        if own_conn:
            conn = get_db()
        try:
            if getattr(conn, 'kind', None) == 'pg':
                rp = conn.execute(
                    "SELECT placa FROM viagem_placa WHERE TRIM(COALESCE(id_viagem::text, '')) = ?",
                    (id_norm,)
                ).fetchone()
                rm = conn.execute(
                    "SELECT motorista FROM viagem_motorista WHERE TRIM(COALESCE(id_viagem::text, '')) = ?",
                    (id_norm,)
                ).fetchone()
                if rp:
                    result['placa'] = (rp.get('placa') or rp[0] or '').strip()
                if rm:
                    result['motorista'] = (rm.get('motorista') or rm[0] or '').strip()
                ds = _get_latest_dataset_id(conn)
                if ds:
                    r = conn.execute(
                        """SELECT data_expedicao, id_roteiro, identificador_rota FROM romaneio_por_item
                           WHERE dataset_id = ? AND (TRIM(COALESCE(id_viagem::text, '')) = ? OR TRIM(COALESCE(id_roteiro::text, '')) = ?) LIMIT 1""",
                        (str(ds), id_norm, id_norm),
                    ).fetchone()
                    if r:
                        if r.get('data_expedicao') or (r[0] if hasattr(r, '__getitem__') else None):
                            result['data_expedicao'] = str(r.get('data_expedicao') or r[0] or '').strip()
                        if r.get('id_roteiro') or (r[1] if hasattr(r, '__getitem__') and len(r) > 1 else None):
                            result['id_roteiro'] = str(r.get('id_roteiro') or r[1] or '').strip()
                        if r.get('identificador_rota') is not None or (len(r) > 2 and r[2] is not None if hasattr(r, '__getitem__') else False):
                            result['identificador_rota'] = str(r.get('identificador_rota') or (r[2] if hasattr(r, '__getitem__') and len(r) > 2 else '') or '').strip()
                    if not result['identificador_rota'] or not result['id_roteiro']:
                        row_ir = conn.execute(
                            """SELECT id_roteiro, identificador_rota FROM id_roteiros WHERE dataset_id = ? AND TRIM(COALESCE(id_viagem::text, '')) = ? LIMIT 1""",
                            (str(ds), id_norm),
                        ).fetchone()
                        if row_ir:
                            if not result['id_roteiro']:
                                result['id_roteiro'] = str(row_ir.get('id_roteiro') if hasattr(row_ir, 'get') else (row_ir[0] if len(row_ir) > 0 else '') or '').strip()
                            if not result['identificador_rota']:
                                result['identificador_rota'] = str(row_ir.get('identificador_rota') if hasattr(row_ir, 'get') else (row_ir[1] if len(row_ir) > 1 else '') or '').strip()
            if own_conn:
                conn.close()
        except Exception:
            if own_conn:
                try:
                    conn.close()
                except Exception:
                    pass
        return result
    return result

def _get_viagem_info_planilha_LEGADO_NAO_USAR(id_viagem):
    """Legado: lê da planilha (desativado)."""
    caminho_planilha = encontrar_planilha()
    if not caminho_planilha:
        return None
    id_viagem = (id_viagem or '').strip()
    if not id_viagem:
        return None
    result = {'data_expedicao': '', 'placa': '', 'identificador_rota': '', 'motorista': ''}
    try:
        try:
            wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
        except Exception:
            wb = openpyxl.load_workbook(caminho_planilha, data_only=False)
        if 'ROMANEIO POR ITEM' not in wb.sheetnames:
            wb.close()
            return result
        ws = wb['ROMANEIO POR ITEM']
        max_col = ws.max_column or 50

        def get_cell_value(row_num, col_num):
            """Valor da célula; se estiver em mesclada, retorna o valor da primeira célula do merge."""
            cell = ws.cell(row=row_num, column=col_num)
            try:
                for merge_range in ws.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        return ws.cell(merge_range.min_row, merge_range.min_col).value
            except Exception:
                pass
            return cell.value

        def build_header_row(row_num):
            return [get_cell_value(row_num, c) for c in range(1, max_col + 1)]

        def norm(s):
            if s is None:
                return ''
            s = str(s).replace('\r', ' ').replace('\n', ' ').replace('\xa0', ' ')
            s = s.upper().strip()
            for old, new in [('Í', 'I'), ('É', 'E'), ('Á', 'A'), ('À', 'A'), ('Ã', 'A'), ('Â', 'A'), ('Ó', 'O'), ('Ô', 'O'), ('Õ', 'O'), ('Ú', 'U'), ('Ç', 'C')]:
                s = s.replace(old, new)
            return ' '.join(s.split())

        header_row = None
        header_row_num = 1
        for row_num in range(1, min(6, ws.max_row + 1)):
            row_vals = build_header_row(row_num)
            for val in row_vals:
                if val is not None and norm(val) == 'MOTORISTA':
                    header_row = row_vals
                    header_row_num = row_num
                    break
            if header_row is not None:
                break
        if header_row is None:
            header_row = build_header_row(1)
        coluna_id_viagem = None
        coluna_inicio_previsto = None
        coluna_inicio_previsto_exata = None
        coluna_placa = None
        coluna_identificador_rota = None
        coluna_motorista = None

        for idx, header in enumerate(header_row or []):
            if header is None:
                continue
            h_str = str(header).strip()
            if not h_str:
                continue
            hs = h_str.upper()
            hs_norm = hs.replace('Í', 'I').replace('É', 'E')
            hs_norm = ' '.join(hs_norm.split())
            h_norm = norm(header)
            if 'ID' in hs and 'VIAGEM' in hs and 'FATURADA' in hs:
                coluna_id_viagem = idx
            if hs_norm == 'INICIO PREVISTO':
                coluna_inicio_previsto_exata = idx
            elif ('INICIO' in hs_norm or 'INÍCIO' in hs) and 'PREVISTO' in hs_norm:
                coluna_inicio_previsto = idx
            # Coluna "Placa" (aba ROMANEIO POR ITEM)
            if coluna_placa is None:
                if (h_norm == 'PLACA' or hs_norm == 'PLACA' or
                    ('PLACA' in h_norm and 'EMPLACA' not in h_norm) or
                    'LICENSE' in h_norm and 'PLATE' in h_norm):
                    coluna_placa = idx
            # Coluna "Identificador de rota" (aba ROMANEIO POR ITEM)
            if coluna_identificador_rota is None:
                if (h_norm in ('IDENTIFICADOR DE ROTA', 'IDENTIFICADOR DA ROTA', 'IDENTIFICADOR ROTA') or
                    hs_norm in ('IDENTIFICADOR DE ROTA', 'IDENTIFICADOR DA ROTA', 'IDENTIFICADOR ROTA') or
                    ('IDENTIFICADOR' in h_norm and 'ROTA' in h_norm) or
                    ('ROTA' in h_norm and ('IDENTIFICADOR' in h_norm or 'CODIGO' in h_norm or 'ID ' in h_norm)) or
                    'ROUTE' in h_norm and ('ID' in h_norm or 'IDENTIFIER' in h_norm or 'CODE' in h_norm)):
                    coluna_identificador_rota = idx
            # Coluna "Motorista" (aba ROMANEIO POR ITEM)
            if coluna_motorista is None and (h_norm == 'MOTORISTA' or hs_norm == 'MOTORISTA'):
                coluna_motorista = idx
        # Mapeamento fixo da planilha ROMANEIO POR ITEM (analisada):
        # A=1 Data cadastro, B=2 Id roteiro, C=3 Identificador da rota, D=4 Id viagem faturada,
        # E=5 Entrega estimada, F=6 Placa, G=7 Pedido, ... U=21 Início previsto, ... Y=25 Motorista
        ncol = len(header_row or [])
        if ncol >= 25:
            coluna_id_viagem = 3
            coluna_identificador_rota = 2
            coluna_placa = 5
            coluna_inicio_previsto = 20
            coluna_motorista = 24
        else:
            if coluna_identificador_rota is None and ncol > 2:
                coluna_identificador_rota = 2
            if coluna_placa is None and ncol > 5:
                coluna_placa = 5
            if coluna_inicio_previsto is None and ncol > 20:
                coluna_inicio_previsto = 20
            if coluna_motorista is None and ncol > 24:
                coluna_motorista = 24
            if coluna_id_viagem is None and ncol > 3:
                coluna_id_viagem = 3
        coluna_inicio_previsto = coluna_inicio_previsto_exata if coluna_inicio_previsto_exata is not None else coluna_inicio_previsto
        if coluna_id_viagem is None:
            wb.close()
            return result
        id_viagem_norm = _normalizar_id_viagem(id_viagem)
        data_start_row = header_row_num + 1
        col_d = coluna_id_viagem + 1
        col_b = 2
        # Percorrer TODAS as linhas onde D (Id viagem faturada) OU B (Id roteiro) = ID e preencher cada campo
        for row_num in range(data_start_row, ws.max_row + 1):
            id_d = _normalizar_id_viagem(get_cell_value(row_num, col_d))
            id_b = _normalizar_id_viagem(get_cell_value(row_num, col_b))
            if id_d != id_viagem_norm and id_b != id_viagem_norm:
                continue
            if not result['data_expedicao'] and coluna_inicio_previsto is not None:
                v = get_cell_value(row_num, coluna_inicio_previsto + 1)
                if v is not None:
                    result['data_expedicao'] = _formatar_data_celula(v) or ''
            if not result['placa'] and coluna_placa is not None:
                v = get_cell_value(row_num, coluna_placa + 1)
                if v is not None and str(v).strip():
                    result['placa'] = str(v).strip()
            if not result['identificador_rota'] and coluna_identificador_rota is not None:
                v = get_cell_value(row_num, coluna_identificador_rota + 1)
                if v is not None and str(v).strip():
                    result['identificador_rota'] = str(v).strip()
            if not result['motorista'] and coluna_motorista is not None:
                v = get_cell_value(row_num, coluna_motorista + 1)
                if v is not None and str(v).strip():
                    result['motorista'] = str(v).strip()
        # Identificador da rota: só usa o valor da coluna correspondente; se estiver em branco na planilha, permanece em branco
        wb.close()
        return result
    except Exception:
        return result


def _get_data_expedicao_planilha(id_viagem):
    """Retorna apenas a data de expedição (compatibilidade)."""
    info = _get_viagem_info_planilha(id_viagem)
    return info.get('data_expedicao', '') if info else None


def _parse_data_expedicao(val):
    """Converte string de data (DD/MM/YYYY ou YYYY-MM-DD) para date. Retorna None se inválido."""
    if not val or not str(val).strip():
        return None
    s = str(val).strip()[:10]
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _get_id_viagens_por_data_expedicao(data_inicio, data_fim):
    """Retorna set de id_viagem cuja data de expedição está no intervalo [data_inicio, data_fim].
    data_inicio e data_fim em YYYY-MM-DD. Se ambos vazios, retorna None (sem filtro)."""
    if not (data_inicio or '').strip() and not (data_fim or '').strip():
        return None
    try:
        d0 = datetime.strptime((data_inicio or '').strip()[:10], '%Y-%m-%d').date() if (data_inicio or '').strip() else None
        d1 = datetime.strptime((data_fim or '').strip()[:10], '%Y-%m-%d').date() if (data_fim or '').strip() else None
    except ValueError:
        return None
    if d0 is None and d1 is None:
        return None
    if d0 is None:
        d0 = d1
    if d1 is None:
        d1 = d0
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT id_viagem FROM produtos_bipados WHERE id_viagem IS NOT NULL AND trim(id_viagem) != ''"
    ).fetchall()
    conn.close()
    ids_ok = set()
    for r in rows:
        vid = (r[0] or '').strip()
        if not vid:
            continue
        de = _get_data_expedicao_planilha(vid)
        dt = _parse_data_expedicao(de)
        if dt is None:
            continue
        if d0 <= dt <= d1:
            ids_ok.add(vid)
    return ids_ok


def _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem'):
    """Acrescenta filtro por id_viagem quando há intervalo de data de expedição."""
    if ids_filtro is None:
        return sql, params
    if len(ids_filtro) == 0:
        return sql + ' AND 1=0', params
    ph = ','.join('?' * len(ids_filtro))
    return sql + f' AND {col} IN ({ph})', params + list(ids_filtro)


COORDENADOR_PADRAO = 'ASTROGILDO RODRIGUES DOS SANTOS'


def _coordenador_resposta_api(val):
    """Não devolve o coordenador padrão legado como valor preenchido na tela."""
    s = (val or '').strip()
    if not s or s.upper() == COORDENADOR_PADRAO.upper():
        return ''
    return s


@app.route('/api/viagem/<id_viagem>/info', methods=['GET'])
def get_viagem_info(id_viagem):
    """Retorna data expedição, placa, identificador da rota, motorista e responsáveis (override ou planilha)."""
    if not id_viagem:
        return jsonify({
            'data_expedicao': '', 'placa': '', 'identificador_rota': '', 'motorista': '',
            'coordenador': '', 'conferente': '', 'ajudante1': '', 'ajudante2': '',
            'id_roteiro': '', 'id_viagem': ''
        })
    info = _get_viagem_info_planilha(id_viagem)
    if not info:
        info = {'data_expedicao': '', 'placa': '', 'identificador_rota': '', 'motorista': '', 'id_roteiro': '', 'id_viagem': id_viagem}
    info.setdefault('id_roteiro', '')
    info.setdefault('id_viagem', id_viagem)
    info.setdefault('coordenador', '')
    info.setdefault('conferente', '')
    info.setdefault('ajudante1', '')
    info.setdefault('ajudante2', '')
    id_norm = _normalizar_id_viagem(id_viagem)
    conn = get_db()
    row_p = conn.execute(
        'SELECT placa FROM viagem_placa WHERE id_viagem = ?',
        (id_norm,)
    ).fetchone()
    row_m = conn.execute(
        'SELECT motorista FROM viagem_motorista WHERE id_viagem = ?',
        (id_norm,)
    ).fetchone()
    row_r = conn.execute(
        'SELECT coordenador, conferente, ajudante1, ajudante2 FROM viagem_responsaveis WHERE id_viagem = ?',
        (id_norm,)
    ).fetchone()
    conn.close()
    if row_p:
        placa_val = (row_p.get('placa') if hasattr(row_p, 'get') else (row_p[0] if len(row_p) > 0 else None)) or ''
        if placa_val:
            info['placa'] = placa_val.strip() if isinstance(placa_val, str) else str(placa_val).strip()
    if row_m:
        motorista_val = (row_m.get('motorista') if hasattr(row_m, 'get') else (row_m[0] if len(row_m) > 0 else None)) or ''
        if motorista_val:
            info['motorista'] = motorista_val.strip() if isinstance(motorista_val, str) else str(motorista_val).strip()
    if row_r:
        raw_coord = (row_r.get('coordenador') if hasattr(row_r, 'get') else (row_r[0] if len(row_r) > 0 else None) or '').strip()
        info['coordenador'] = _coordenador_resposta_api(raw_coord)
        info['conferente'] = (row_r.get('conferente') if hasattr(row_r, 'get') else (row_r[1] if len(row_r) > 1 else None) or '').strip()
        info['ajudante1'] = (row_r.get('ajudante1') if hasattr(row_r, 'get') else (row_r[2] if len(row_r) > 2 else None) or '').strip()
        info['ajudante2'] = (row_r.get('ajudante2') if hasattr(row_r, 'get') else (row_r[3] if len(row_r) > 3 else None) or '').strip()
    else:
        info['coordenador'] = ''
    return jsonify(info)


def _ravex_romaneio_info_por_ids(conn, dataset_id, ids):
    """Conta itens no romaneio para uma lista de IDs (roteiro/viagem)."""
    ids = [str(x).strip() for x in (ids or []) if str(x or '').strip()]
    if not ids or not dataset_id:
        return None
    ph = ','.join(['?'] * len(ids))
    try:
        row = conn.execute(
            f"""SELECT COUNT(*)::int AS n, MAX(importado_em) AS ultimo,
                       MAX(TRIM(COALESCE(id_viagem::text, ''))) AS id_viagem,
                       MAX(TRIM(COALESCE(id_roteiro::text, ''))) AS id_roteiro
                FROM romaneio_por_item
                WHERE dataset_id = ?
                  AND (
                    TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
                    OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph})
                  )""",
            [str(dataset_id)] + ids + ids,
        ).fetchone()
    except Exception:
        try:
            row = conn.execute(
                f"""SELECT COUNT(*) AS n, MAX(importado_em) AS ultimo,
                           MAX(id_viagem) AS id_viagem,
                           MAX(id_roteiro) AS id_roteiro
                    FROM romaneio_por_item
                    WHERE dataset_id = ?
                      AND (
                        TRIM(COALESCE(id_viagem, '')) IN ({ph})
                        OR TRIM(COALESCE(id_roteiro, '')) IN ({ph})
                      )""",
                [str(dataset_id)] + ids + ids,
            ).fetchone()
        except Exception:
            return None
    if not row:
        return None
    n = int((row.get('n') if hasattr(row, 'get') else row[0]) or 0)
    if n <= 0:
        return None
    return {
        'itens_romaneio': n,
        'importado_em': row.get('ultimo') if hasattr(row, 'get') else (row[1] if len(row) > 1 else None),
        'fonte': 'romaneio',
        'id_viagem': str((row.get('id_viagem') if hasattr(row, 'get') else (row[2] if len(row) > 2 else '')) or '').strip(),
        'id_roteiro': str((row.get('id_roteiro') if hasattr(row, 'get') else (row[3] if len(row) > 3 else '')) or '').strip(),
    }


@app.route('/api/viagem/<id_viagem>/motorista', methods=['PUT', 'PATCH'])
def set_viagem_motorista(id_viagem):
    """Salva o motorista alterado para a viagem (override)."""
    if not id_viagem:
        return jsonify({'erro': 'ID do roteiro não informado'}), 400
    data = request.get_json(silent=True) or {}
    motorista = (data.get('motorista') or '').strip()
    id_norm = _normalizar_id_viagem(id_viagem)
    usuario = session.get('usuario', '')
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''INSERT INTO viagem_motorista (id_viagem, motorista, atualizado_por)
               VALUES (%s, %s, %s)
               ON CONFLICT (id_viagem) DO UPDATE SET motorista = EXCLUDED.motorista, atualizado_por = EXCLUDED.atualizado_por''',
            (id_norm, motorista, usuario)
        )
    else:
        conn.execute(
            '''INSERT INTO viagem_motorista (id_viagem, motorista) VALUES (?, ?)
               ON CONFLICT(id_viagem) DO UPDATE SET motorista = excluded.motorista''',
            (id_norm, motorista)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'motorista': motorista})


@app.route('/api/viagem/<id_viagem>/placa', methods=['PUT', 'PATCH'])
def set_viagem_placa(id_viagem):
    """Salva a placa alterada para a viagem (override)."""
    if not id_viagem:
        return jsonify({'erro': 'ID do roteiro não informado'}), 400
    data = request.get_json(silent=True) or {}
    placa = (data.get('placa') or '').strip()
    id_norm = _normalizar_id_viagem(id_viagem)
    usuario = session.get('usuario', '')
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''INSERT INTO viagem_placa (id_viagem, placa, atualizado_por)
               VALUES (%s, %s, %s)
               ON CONFLICT (id_viagem) DO UPDATE SET placa = EXCLUDED.placa, atualizado_por = EXCLUDED.atualizado_por''',
            (id_norm, placa, usuario)
        )
    else:
        conn.execute(
            '''INSERT INTO viagem_placa (id_viagem, placa) VALUES (?, ?)
               ON CONFLICT(id_viagem) DO UPDATE SET placa = excluded.placa''',
            (id_norm, placa)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'placa': placa})


@app.route('/api/viagem/<id_viagem>/responsaveis', methods=['PUT', 'PATCH'])
def set_viagem_responsaveis(id_viagem):
    """Salva os responsáveis da viagem: coordenador, conferente, ajudante1, ajudante2."""
    if not id_viagem:
        return jsonify({'erro': 'ID do roteiro não informado'}), 400
    data = request.get_json(silent=True) or {}
    coordenador = (data.get('coordenador') or '').strip()
    conferente = (data.get('conferente') or '').strip()
    ajudante1 = (data.get('ajudante1') or '').strip()
    ajudante2 = (data.get('ajudante2') or '').strip()
    id_norm = _normalizar_id_viagem(id_viagem)
    usuario = session.get('usuario', '')
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''INSERT INTO viagem_responsaveis (id_viagem, coordenador, conferente, ajudante1, ajudante2, atualizado_por)
               VALUES (%s, %s, %s, %s, %s, %s)
               ON CONFLICT (id_viagem) DO UPDATE SET
                 coordenador = EXCLUDED.coordenador,
                 conferente = EXCLUDED.conferente,
                 ajudante1 = EXCLUDED.ajudante1,
                 ajudante2 = EXCLUDED.ajudante2,
                 atualizado_por = EXCLUDED.atualizado_por''',
            (id_norm, coordenador, conferente, ajudante1, ajudante2, usuario)
        )
    else:
        conn.execute(
            '''INSERT INTO viagem_responsaveis (id_viagem, coordenador, conferente, ajudante1, ajudante2)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(id_viagem) DO UPDATE SET
                 coordenador = excluded.coordenador,
                 conferente = excluded.conferente,
                 ajudante1 = excluded.ajudante1,
                 ajudante2 = excluded.ajudante2''',
            (id_norm, coordenador, conferente, ajudante1, ajudante2)
        )
    conn.commit()
    conn.close()
    return jsonify({
        'ok': True,
        'coordenador': coordenador,
        'conferente': conferente,
        'ajudante1': ajudante1,
        'ajudante2': ajudante2
    })


def _lista_distinta_ordenada(valores):
    vistos = set()
    out = []
    for v in valores or []:
        s = str(v or '').strip()
        if not s:
            continue
        k = s.upper()
        if k in vistos:
            continue
        vistos.add(k)
        out.append(s)
    return sorted(out, key=lambda x: x.upper())


@app.route('/api/colaboradores-motoristas', methods=['GET'])
def get_colaboradores_motoristas():
    """Lista motoristas do cadastro + viagens/romaneio já gravados (sugestões na conferência)."""
    conn = get_db()
    nomes = []
    try:
        if getattr(conn, 'kind', None) == 'pg':
            rows = conn.execute(
                """SELECT DISTINCT TRIM(nome) AS nome FROM (
                       SELECT nome FROM motoristas WHERE nome IS NOT NULL AND TRIM(nome) <> ''
                       UNION
                       SELECT motorista AS nome FROM viagem_motorista
                         WHERE motorista IS NOT NULL AND TRIM(motorista) <> ''
                       UNION
                       SELECT motorista AS nome FROM romaneio_por_item
                         WHERE motorista IS NOT NULL AND TRIM(motorista) <> ''
                       UNION
                       SELECT nome FROM colaboradores
                         WHERE nome IS NOT NULL AND TRIM(nome) <> ''
                           AND (
                             UPPER(COALESCE(tipo, '')) LIKE 'MOTORISTA%%'
                             OR UPPER(COALESCE(centro_custo, '')) LIKE '%%TRANSPORTE%%'
                           )
                   ) AS t
                   WHERE TRIM(nome) <> ''
                   ORDER BY nome"""
            ).fetchall()
            nomes = [r.get('nome') or r[0] for r in rows or []]
        else:
            for sql in (
                '''SELECT nome FROM motoristas WHERE nome IS NOT NULL AND TRIM(nome) <> '' ORDER BY nome''',
                '''SELECT motorista FROM viagem_motorista WHERE motorista IS NOT NULL AND TRIM(motorista) <> '' ''',
                '''SELECT motorista FROM romaneio_por_item WHERE motorista IS NOT NULL AND TRIM(motorista) <> '' ''',
                '''SELECT nome FROM colaboradores WHERE ativo = 1 AND (
                       tipo = 'MOTORISTA' OR UPPER(centro_custo) LIKE '%TRANSPORTE GRU%'
                       OR UPPER(centro_custo) LIKE '%TRANSPORTE PPY%') ORDER BY nome''',
            ):
                try:
                    for row in conn.execute(sql).fetchall() or []:
                        nomes.append(row[0] if not hasattr(row, 'get') else (row.get('nome') or row.get('motorista') or row[0]))
                except Exception:
                    pass
        conn.close()
        return jsonify({'nomes': _lista_distinta_ordenada(nomes)})
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'nomes': []})


@app.route('/api/placas', methods=['GET'])
def get_placas():
    """Lista placas do cadastro + viagens/romaneio (sugestões na conferência)."""
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) == 'pg':
            rows = conn.execute(
                """SELECT DISTINCT TRIM(placa) AS placa FROM (
                       SELECT placa FROM placas WHERE placa IS NOT NULL AND TRIM(placa) <> ''
                       UNION
                       SELECT placa FROM viagem_placa
                         WHERE placa IS NOT NULL AND TRIM(placa) <> ''
                       UNION
                       SELECT placa FROM romaneio_por_item
                         WHERE placa IS NOT NULL AND TRIM(placa) <> ''
                   ) AS t
                   WHERE TRIM(placa) <> ''
                   ORDER BY placa"""
            ).fetchall()
        else:
            rows = []
            for sql in (
                """SELECT placa FROM placas WHERE placa IS NOT NULL AND TRIM(placa) <> '' ORDER BY placa""",
                """SELECT placa FROM viagem_placa WHERE placa IS NOT NULL AND TRIM(placa) <> '' """,
                """SELECT placa FROM romaneio_por_item WHERE placa IS NOT NULL AND TRIM(placa) <> '' """,
            ):
                try:
                    rows.extend(conn.execute(sql).fetchall() or [])
                except Exception:
                    pass
        lista = []
        for r in rows or []:
            p = (r.get('placa') if hasattr(r, 'get') else None) or (r[0] if r else None)
            if p:
                lista.append(p)
        conn.close()
        return jsonify({'placas': _lista_distinta_ordenada(lista)})
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'placas': []})


@app.route('/api/colaboradores', methods=['GET'])
def get_colaboradores():
    """Lista colaboradores ativos (ou todos se incluir_inativos=true)."""
    incluir_inativos = request.args.get('incluir_inativos', 'false').lower() == 'true'
    tipo = request.args.get('tipo', '').strip().upper()
    conn = get_db()
    sql = 'SELECT id, nome, funcao, centro_custo, tipo, cpf, telefone, email, ativo, observacoes FROM colaboradores'
    filtros = []
    params = []
    if not incluir_inativos:
        filtros.append('ativo = ?')
        params.append(1 if getattr(conn, 'kind', 'sqlite') == 'sqlite' else True)
    if tipo:
        filtros.append('UPPER(tipo) = ?')
        params.append(tipo)
    if filtros:
        sql += ' WHERE ' + ' AND '.join(filtros)
    sql += ' ORDER BY nome'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    colaboradores = [dict(row) for row in rows]
    return jsonify({'colaboradores': colaboradores})


@app.route('/api/colaboradores', methods=['POST'])
def add_colaborador():
    """Adiciona um novo colaborador."""
    data = request.get_json() or {}
    nome = (data.get('nome') or '').strip()
    if not nome:
        return jsonify({'success': False, 'erro': 'Nome é obrigatório'}), 400
    funcao = (data.get('funcao') or '').strip()
    centro_custo = (data.get('centro_custo') or '').strip()
    tipo = (data.get('tipo') or '').strip().upper()
    cpf = (data.get('cpf') or '').strip()
    telefone = (data.get('telefone') or '').strip()
    email = (data.get('email') or '').strip()
    observacoes = (data.get('observacoes') or '').strip()
    usuario = session.get('usuario', '')
    conn = get_db()
    try:
        if getattr(conn, 'kind', 'sqlite') == 'pg':
            conn.execute(
                '''INSERT INTO colaboradores (nome, funcao, centro_custo, tipo, cpf, telefone, email, observacoes, criado_por)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                (nome, funcao, centro_custo, tipo, cpf or None, telefone, email, observacoes, usuario)
            )
        else:
            conn.execute(
                '''INSERT INTO colaboradores (nome, funcao, centro_custo, tipo, cpf, telefone, email, observacoes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                (nome, funcao, centro_custo, tipo, cpf or None, telefone, email, observacoes)
            )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'mensagem': 'Colaborador adicionado com sucesso'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'erro': f'Erro ao adicionar: {str(e)}'}), 500


@app.route('/api/colaboradores/<int:colaborador_id>', methods=['PUT'])
def update_colaborador(colaborador_id):
    """Atualiza dados de um colaborador."""
    data = request.get_json() or {}
    nome = (data.get('nome') or '').strip()
    if not nome:
        return jsonify({'success': False, 'erro': 'Nome é obrigatório'}), 400
    funcao = (data.get('funcao') or '').strip()
    centro_custo = (data.get('centro_custo') or '').strip()
    tipo = (data.get('tipo') or '').strip().upper()
    cpf = (data.get('cpf') or '').strip()
    telefone = (data.get('telefone') or '').strip()
    email = (data.get('email') or '').strip()
    observacoes = (data.get('observacoes') or '').strip()
    usuario = session.get('usuario', '')
    conn = get_db()
    try:
        if getattr(conn, 'kind', 'sqlite') == 'pg':
            conn.execute(
                '''UPDATE colaboradores SET
                   nome = %s, funcao = %s, centro_custo = %s, tipo = %s,
                   cpf = %s, telefone = %s, email = %s, observacoes = %s, atualizado_por = %s
                   WHERE id = %s''',
                (nome, funcao, centro_custo, tipo, cpf or None, telefone, email, observacoes, usuario, colaborador_id)
            )
        else:
            conn.execute(
                '''UPDATE colaboradores SET
                   nome = ?, funcao = ?, centro_custo = ?, tipo = ?,
                   cpf = ?, telefone = ?, email = ?, observacoes = ?
                   WHERE id = ?''',
                (nome, funcao, centro_custo, tipo, cpf or None, telefone, email, observacoes, colaborador_id)
            )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'mensagem': 'Colaborador atualizado'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'erro': f'Erro ao atualizar: {str(e)}'}), 500


@app.route('/api/colaboradores/<int:colaborador_id>', methods=['DELETE'])
def delete_colaborador(colaborador_id):
    """Desativa um colaborador (soft delete: ativo = false)."""
    conn = get_db()
    try:
        if getattr(conn, 'kind', 'sqlite') == 'pg':
            conn.execute('UPDATE colaboradores SET ativo = FALSE WHERE id = %s', (colaborador_id,))
        else:
            conn.execute('UPDATE colaboradores SET ativo = 0 WHERE id = ?', (colaborador_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'mensagem': 'Colaborador desativado'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'erro': f'Erro ao desativar: {str(e)}'}), 500


@app.route('/api/colaboradores/importar-planilha', methods=['POST'])
def importar_colaboradores_planilha():
    """Importa colaboradores: planilha não é mais lida do disco; use cadastro pela tela ou scripts."""
    if _usa_banco_para_dados():
        return jsonify({'success': False, 'erro': 'Planilha não é mais usada. Cadastre pela tela de colaboradores ou use os scripts (gerar_sql_colaboradores.py) para subir dados.'}), 400
    caminho_planilha = encontrar_planilha()
    if not caminho_planilha:
        return jsonify({'success': False, 'erro': 'Planilha não encontrada'}), 404
    try:
        wb = openpyxl.load_workbook(os.path.abspath(caminho_planilha), data_only=True)
        nome_aba = next((s for s in wb.sheetnames if s.upper().strip() == 'COLABORADORES'), None)
        if not nome_aba:
            wb.close()
            return jsonify({'success': False, 'erro': 'Aba COLABORADORES não encontrada'}), 404
        ws = wb[nome_aba]
        # Colunas: A=Centro de Custo, B=Função, C=Colaborador (nome)
        col_centro_custo = 0
        col_funcao = 1
        col_nome = 2
        conn = get_db()
        importados = 0
        atualizados = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= col_nome:
                continue
            nome = str(row[col_nome] or '').strip()
            if not nome:
                continue
            funcao = str(row[col_funcao] or '').strip() if len(row) > col_funcao else ''
            centro_custo = str(row[col_centro_custo] or '').strip() if len(row) > col_centro_custo else ''
            # Identificar tipo baseado no centro de custo ou função
            tipo = ''
            cc_upper = centro_custo.upper()
            fn_upper = funcao.upper()
            if 'TRANSPORTE' in cc_upper or 'MOTORISTA' in fn_upper:
                tipo = 'MOTORISTA'
            elif 'CONFERENTE' in fn_upper:
                tipo = 'CONFERENTE'
            elif 'AJUDANTE' in fn_upper or 'AUXILIAR' in fn_upper:
                tipo = 'AJUDANTE'
            elif 'COORDENADOR' in fn_upper:
                tipo = 'COORDENADOR'
            # Verificar se já existe (por nome)
            existing = conn.execute('SELECT id FROM colaboradores WHERE nome = ?', (nome,)).fetchone()
            if existing:
                # Atualizar
                if getattr(conn, 'kind', None) == 'pg':
                    conn.execute(
                        'UPDATE colaboradores SET funcao = %s, centro_custo = %s, tipo = %s WHERE nome = %s',
                        (funcao, centro_custo, tipo, nome)
                    )
                else:
                    conn.execute(
                        'UPDATE colaboradores SET funcao = ?, centro_custo = ?, tipo = ? WHERE nome = ?',
                        (funcao, centro_custo, tipo, nome)
                    )
                atualizados += 1
            else:
                # Inserir
                if getattr(conn, 'kind', None) == 'pg':
                    conn.execute(
                        'INSERT INTO colaboradores (nome, funcao, centro_custo, tipo) VALUES (%s, %s, %s, %s)',
                        (nome, funcao, centro_custo, tipo)
                    )
                else:
                    conn.execute(
                        'INSERT INTO colaboradores (nome, funcao, centro_custo, tipo) VALUES (?, ?, ?, ?)',
                        (nome, funcao, centro_custo, tipo)
                    )
                importados += 1
        conn.commit()
        conn.close()
        wb.close()
        return jsonify({
            'success': True,
            'mensagem': f'Importação concluída: {importados} novos, {atualizados} atualizados',
            'importados': importados,
            'atualizados': atualizados
        })
    except Exception as e:
        return jsonify({'success': False, 'erro': f'Erro ao importar: {str(e)}'}), 500


@app.route('/api/viagem/<id_viagem>/data-expedicao', methods=['GET'])
def get_data_expedicao(id_viagem):
    """Retorna a data de expedição (Início Previsto) da aba ROMANEIO POR ITEM para a viagem."""
    if not id_viagem:
        return jsonify({'data_expedicao': ''})
    valor = _get_data_expedicao_planilha(id_viagem)
    return jsonify({'data_expedicao': valor if valor is not None else ''})


@app.route('/api/debug-romaneio-headers', methods=['GET'])
def debug_romaneio_headers():
    """Retorna os cabeçalhos do romaneio. Com DATABASE_URL, dados vêm do banco."""
    if _usa_banco_para_dados():
        return jsonify({'erro': 'Dados vêm do banco (romaneio_por_item).', 'headers': []})
    caminho_planilha = encontrar_planilha()
    if not caminho_planilha:
        return jsonify({'erro': 'Planilha não encontrada', 'headers': []})
    try:
        wb = openpyxl.load_workbook(caminho_planilha, data_only=False)
        if 'ROMANEIO POR ITEM' not in wb.sheetnames:
            wb.close()
            return jsonify({'erro': 'Aba ROMANEIO POR ITEM não encontrada', 'headers': []})
        ws = wb['ROMANEIO POR ITEM']
        max_col = min(ws.max_column or 30, 30)
        result = []
        for row_num in range(1, min(4, ws.max_row + 1)):
            row_headers = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col)
                val = cell.value
                try:
                    for merge_range in ws.merged_cells.ranges:
                        if cell.coordinate in merge_range:
                            val = ws.cell(merge_range.min_row, merge_range.min_col).value
                            break
                except Exception:
                    pass
                row_headers.append(str(val) if val is not None else '')
            result.append({'row': row_num, 'headers': row_headers})
        wb.close()
        return jsonify({'planilha': caminho_planilha, 'linhas': result})
    except Exception as e:
        return jsonify({'erro': str(e), 'headers': []})


def _codigos_produto_na_viagem(id_viagem):
    """Retorna set de códigos de produto (str) que estão no romaneio da viagem. Lê da tabela quando DATABASE_URL está definido."""
    if _usa_banco_para_dados():
        conn = get_db()
        try:
            if getattr(conn, 'kind', None) == 'pg':
                ds = _get_latest_dataset_id(conn)
                if ds:
                    id_norm = _normalizar_id_viagem(id_viagem)
                    rows = conn.execute(
                        """SELECT DISTINCT codigo_produto FROM romaneio_por_item
                           WHERE dataset_id = ? AND (id_viagem = ? OR id_roteiro = ?)""",
                        (str(ds), id_norm or id_viagem, id_norm or id_viagem),
                    ).fetchall()
                    codigos = set()
                    for r in rows or []:
                        cp = (r.get('codigo_produto') or r[0] or '').strip()
                        if cp:
                            codigos.add(cp)
                            codigos.add(_normalizar_codigo_produto(cp) or cp)
                    conn.close()
                    return codigos
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return set()
    return set()


@app.route('/api/conferencia/<id_viagem>/produto-na-lista', methods=['GET'])
def conferencia_produto_na_lista(id_viagem):
    """Verifica se o produto (codigo_produto ou codigo_barras) está na relação de itens da viagem."""
    if not id_viagem:
        return jsonify({'na_lista': False, 'erro': 'id_viagem obrigatório'}), 400
    codigo_produto = request.args.get('codigo_produto', '').strip()
    codigo_barras = request.args.get('codigo_barras', '').strip()
    if not codigo_produto and not codigo_barras:
        return jsonify({'na_lista': False})
    codigos_romaneio = _codigos_produto_na_viagem(id_viagem)
    if codigo_produto:
        cp_norm = _normalizar_codigo_produto(codigo_produto)
        na_lista = codigo_produto in codigos_romaneio or (cp_norm and cp_norm in codigos_romaneio)
        return jsonify({'na_lista': na_lista, 'codigo_produto': codigo_produto})
    codigo_produto_resolvido = None
    info = buscar_produto_na_planilha(codigo_barras)
    if info and not info.get('erro'):
        codigo_produto_resolvido = (info.get('codigo_produto') or '').strip()
    if not codigo_produto_resolvido:
        return jsonify({'na_lista': False, 'codigo_produto': None})
    cp_norm = _normalizar_codigo_produto(codigo_produto_resolvido)
    na_lista = codigo_produto_resolvido in codigos_romaneio or (cp_norm and cp_norm in codigos_romaneio)
    return jsonify({'na_lista': na_lista, 'codigo_produto': codigo_produto_resolvido})


def _conferencia_sets_codigos(resultado):
    """Conjuntos de código de barras e código produto já presentes na lista da conferência."""
    codigos_romaneio = set()
    codigos_prod = set()
    for it in resultado or []:
        cb = str(it.get('codigo_barras') or '').strip()
        if cb:
            codigos_romaneio.add(cb)
        cp = str(it.get('codigo_produto') or '').strip()
        if cp:
            codigos_prod.add(cp)
        for x in it.get('_todos_codigos_barras') or []:
            xc = str(x or '').strip()
            if xc:
                codigos_romaneio.add(xc)
    return codigos_romaneio, codigos_prod


def _conferencia_calc_status_item(quantidade_produto, quantidade_bipada):
    quantidade_falta = max(0, int(quantidade_produto or 0) - int(quantidade_bipada or 0))
    quantidade_sobra = max(0, int(quantidade_bipada or 0) - int(quantidade_produto or 0))
    if quantidade_sobra > 0:
        status_bipado = 'EXCEDENTE'
    elif quantidade_bipada >= quantidade_produto and quantidade_produto > 0:
        status_bipado = 'COMPLETO'
    elif quantidade_bipada > 0:
        status_bipado = 'PARCIAL'
    else:
        status_bipado = 'PENDENTE'
    aviso_sobra = ('Bipou %s a mais' % quantidade_sobra) if quantidade_sobra > 0 else ''
    return quantidade_falta, quantidade_sobra, status_bipado, aviso_sobra


def _conferencia_aviso_multiplos_codigos(todos_cb, aviso_atual=''):
    cods = sorted({str(x).strip() for x in (todos_cb or []) if str(x).strip()})
    if len(cods) <= 1:
        return aviso_atual or ''
    aviso_multi = '⚠️ Múltiplos códigos: ' + ', '.join(cods)
    if aviso_atual and 'Múltiplos códigos' not in aviso_atual:
        return (aviso_atual + ' — ' + aviso_multi) if aviso_atual else aviso_multi
    return aviso_multi


def _conferencia_incluir_extras_bipados(resultado, extras_rows, mapa_barras_to_codigo, id_viagem):
    """Itens bipados fora do romaneio (manual). Um registro por codigo_produto; meta não sobe ao bipar outro EAN/DUN."""
    codigos_romaneio, codigos_prod = _conferencia_sets_codigos(resultado)
    for row in extras_rows or []:
        cb = ((row.get('codigo_barras') or '') if hasattr(row, 'get') else (row[0] if len(row) > 0 else '')).strip()
        if not cb:
            continue
        cod_prod = (
            mapa_barras_to_codigo.get(cb)
            or ((row.get('codigo_interno') or '') if hasattr(row, 'get') else (row[1] if len(row) > 1 else '')).strip()
            or cb
        )
        qtd_bip = int((row.get('quantidade_bipada') or 0) if hasattr(row, 'get') else (row[3] if len(row) > 3 else 0))

        if cod_prod in codigos_prod:
            existente = None
            for it in resultado:
                if str(it.get('codigo_produto') or '').strip() == cod_prod:
                    existente = it
                    break
            if existente and existente.get('origem_romaneio'):
                continue
            if existente:
                existente['quantidade_bipada'] = max(int(existente.get('quantidade_bipada') or 0), qtd_bip)
                todos = list(existente.get('_todos_codigos_barras') or [])
                cb0 = str(existente.get('codigo_barras') or '').strip()
                if cb0 and cb0 not in todos:
                    todos.append(cb0)
                if cb not in todos:
                    todos.append(cb)
                existente['_todos_codigos_barras'] = todos
                codigos_romaneio.add(cb)
                qp = int(existente.get('quantidade_produto') or 0)
                qb = int(existente.get('quantidade_bipada') or 0)
                falta, sobra, st, aviso_bipou = _conferencia_calc_status_item(qp, qb)
                existente['quantidade_falta'] = falta
                existente['quantidade_sobra'] = sobra
                existente['status_bipado'] = st
                aviso_multi = _conferencia_aviso_multiplos_codigos(todos, '')
                if aviso_bipou:
                    existente['aviso_sobra'] = (aviso_multi + ' — ' + aviso_bipou) if aviso_multi else aviso_bipou
                else:
                    existente['aviso_sobra'] = aviso_multi or ''
            continue

        if cb in codigos_romaneio:
            continue

        descricao = (row.get('produto') or '').strip() if hasattr(row, 'get') else ''
        if not descricao and hasattr(row, '__getitem__'):
            try:
                descricao = (row['produto'] or '').strip()
            except Exception:
                descricao = ''
        unidade = (row.get('unidade') or '').strip() if hasattr(row, 'get') else '-'
        peso = (row.get('peso') or '').strip() if hasattr(row, 'get') else '-'
        falta, sobra, st, aviso = _conferencia_calc_status_item(qtd_bip, qtd_bip)
        resultado.append({
            'codigo_produto': cod_prod,
            'codigo_barras': cb,
            'produto': descricao or 'Item adicionado manualmente',
            'quantidade_produto': qtd_bip,
            'unidade': unidade or '-',
            'peso_bruto': peso or '-',
            'quantidade_bipada': qtd_bip,
            'quantidade_falta': falta,
            'quantidade_sobra': sobra,
            'aviso_sobra': aviso,
            'status_bipado': st,
            'id_viagem': id_viagem,
            'origem_romaneio': False,
            '_todos_codigos_barras': [cb],
        })
        codigos_romaneio.add(cb)
        codigos_prod.add(cod_prod)


def _codigo_barras_romaneio_linha(codigo_produto, unidade, mapa_ean, mapa_dun, mapa_codigo_barras):
    unidade_norm = (unidade or '').strip().upper()
    if unidade_norm in ('PT', 'UN'):
        return (mapa_ean.get(codigo_produto) or mapa_codigo_barras.get(codigo_produto, '') or '').strip()
    if unidade_norm == 'CX':
        return (mapa_dun.get(codigo_produto) or mapa_codigo_barras.get(codigo_produto, '') or '').strip()
    return (mapa_codigo_barras.get(codigo_produto, '') or '').strip()


def _agrupar_romaneio_rows_por_codigo_produto(romaneio_rows, mapa_ean, mapa_dun, mapa_codigo_barras):
    """Uma linha por codigo_produto no roteiro (soma quantidades das linhas repetidas)."""
    agreg = {}
    for r in romaneio_rows or []:
        codigo_produto = (r.get('codigo_produto') or '').strip()
        if not codigo_produto:
            continue
        descricao = (r.get('descricao') or '').strip()
        quantidade_produto = int(r.get('quantidade') or 0)
        unidade = (r.get('unidade') or '').strip() or '-'
        peso_bruto = (r.get('peso_bruto') or '').strip() or '-'
        cb = _codigo_barras_romaneio_linha(codigo_produto, unidade, mapa_ean, mapa_dun, mapa_codigo_barras)
        if codigo_produto not in agreg:
            agreg[codigo_produto] = {
                'codigo_produto': codigo_produto,
                'descricao': descricao,
                'quantidade_produto': 0,
                'unidades': set(),
                'peso_bruto': peso_bruto,
                'codigos_barras': set(),
            }
        g = agreg[codigo_produto]
        g['quantidade_produto'] += quantidade_produto
        if descricao and (not g['descricao'] or len(descricao) > len(g['descricao'])):
            g['descricao'] = descricao
        if unidade and unidade != '-':
            g['unidades'].add(unidade)
        if cb:
            g['codigos_barras'].add(cb)
    linhas = []
    for cod, g in agreg.items():
        unidades = sorted(g['unidades'], key=lambda x: str(x).upper())
        cods = sorted(g['codigos_barras'])
        linhas.append({
            'codigo_produto': cod,
            'descricao': g['descricao'],
            'quantidade_produto': g['quantidade_produto'],
            'unidade': ', '.join(unidades) if unidades else '-',
            'peso_bruto': g['peso_bruto'],
            'codigo_barras': cods[0] if cods else '',
            '_todos_codigos_barras': cods,
        })
    return linhas


_ROMANEIO_CONFERENCIA_COLS = """id_roteiro, id_viagem, identificador_rota, codigo_produto, descricao, quantidade, unidade, peso_bruto,
                  placa, motorista, data_expedicao"""


def _conferencia_ids_busca_variantes(id_input):
    """IDs equivalentes para buscar romaneio (roteiro ou viagem digitados)."""
    ids = []
    vistos = set()

    def _add(val):
        s = str(val or '').strip()
        if not s or s in vistos:
            return
        vistos.add(s)
        ids.append(s)

    _add(id_input)
    _add(_normalizar_id_viagem(id_input))
    return ids


def _conferencia_expandir_ids_romaneio(conn, ids, dataset_id=None):
    """Inclui o par viagem/roteiro gravado no romaneio (ex.: busca 741740 → também 19306305)."""
    ids = list(ids or [])
    vistos = set(ids)
    if not ids:
        return ids
    ph = ','.join(['?'] * len(ids))
    params = list(ids) + list(ids)
    sql = (
        f"""SELECT DISTINCT TRIM(COALESCE(id_viagem::text, '')) AS id_v,
                   TRIM(COALESCE(id_roteiro::text, '')) AS id_r
            FROM romaneio_por_item
            WHERE (
              TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
              OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph})
            )"""
    )
    if dataset_id:
        sql = (
            f"""SELECT DISTINCT TRIM(COALESCE(id_viagem::text, '')) AS id_v,
                       TRIM(COALESCE(id_roteiro::text, '')) AS id_r
                FROM romaneio_por_item
                WHERE dataset_id = ?
                  AND (
                    TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
                    OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph})
                  )"""
        )
        params = [str(dataset_id)] + params
    try:
        rows = conn.execute(sql, params).fetchall()
    except Exception:
        return ids
    for row in rows or []:
        for key in ('id_v', 'id_r'):
            val = str((row.get(key) if hasattr(row, 'get') else '') or '').strip()
            if val and val not in vistos:
                vistos.add(val)
                ids.append(val)
    return ids


def _conferencia_enriquecer_ids_de_importacao(conn, ids, dataset_id=None):
    """Completa IDs a partir do histórico ravex_importacoes (roteiro ↔ viagem)."""
    ids = list(ids or [])
    vistos = set(ids)
    if not ids:
        return ids
    ph = ','.join(['?'] * len(ids))
    params = list(ids) + list(ids) + list(ids)
    sql = f"""SELECT parametros FROM ravex_importacoes
              WHERE status IN ('OK', 'OK_COM_ERROS', 'DUPLICADO')
                AND (
                  TRIM(COALESCE(parametros->>'id_viagem', '')) IN ({ph})
                  OR TRIM(COALESCE(parametros->>'id_roteiro', '')) IN ({ph})
                  OR TRIM(COALESCE(parametros->>'id_informado', '')) IN ({ph})
                )
              ORDER BY criado_em DESC NULLS LAST
              LIMIT 12"""
    if dataset_id:
        sql = sql.replace(
            'WHERE status',
            'WHERE dataset_id = ?::uuid AND status',
            1,
        )
        params = [str(dataset_id)] + params
    try:
        rows = conn.execute(sql, params).fetchall()
    except Exception:
        return ids
    for row in rows or []:
        params_json = row.get('parametros') if hasattr(row, 'get') else (row[0] if row else None)
        if isinstance(params_json, str):
            try:
                params_json = json.loads(params_json)
            except Exception:
                params_json = {}
        if not isinstance(params_json, dict):
            continue
        for key in ('id_viagem', 'id_roteiro', 'id_informado'):
            _val = str(params_json.get(key) or '').strip()
            if _val and _val not in vistos:
                vistos.add(_val)
                ids.append(_val)
        for vid in params_json.get('viagens_importadas') or []:
            _val = str(vid or '').strip()
            if _val and _val not in vistos:
                vistos.add(_val)
                ids.append(_val)
    return ids


def _conferencia_resolver_dataset_id(conn, id_busca):
    """Dataset onde o romaneio deste roteiro/viagem está gravado (prioriza o ativo)."""
    ds_ativo = _get_latest_dataset_id(conn)
    ids = _conferencia_ids_busca_variantes(id_busca)
    ids = _conferencia_expandir_ids_romaneio(conn, ids)
    ids = _conferencia_enriquecer_ids_de_importacao(conn, ids)
    if not ids:
        return ds_ativo
    ph = ','.join(['?'] * len(ids))

    def _tem_no_dataset(ds):
        if not ds:
            return False
        try:
            row = conn.execute(
                f"""SELECT 1 FROM romaneio_por_item
                    WHERE dataset_id = ?
                      AND (
                        TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
                        OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph})
                      )
                    LIMIT 1""",
                [str(ds)] + ids + ids,
            ).fetchone()
        except Exception:
            return False
        return bool(row)

    if ds_ativo and _tem_no_dataset(ds_ativo):
        return ds_ativo
    try:
        row = conn.execute(
            f"""SELECT dataset_id FROM romaneio_por_item
                WHERE TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
                   OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph})
                ORDER BY importado_em DESC NULLS LAST
                LIMIT 1""",
            ids + ids,
        ).fetchone()
    except Exception:
        row = None
    if row:
        ds = row.get('dataset_id') if hasattr(row, 'get') else (row[0] if len(row) > 0 else None)
        if ds:
            return ds
    return ds_ativo


def _conferencia_buscar_romaneio_rows(conn, ds, id_busca, limit_romaneio):
    """Busca romaneio por roteiro ou viagem (TRIM, variantes e par roteiro↔viagem)."""
    id_busca = str(id_busca or '').strip()
    if not id_busca or not ds:
        return []
    ids = _conferencia_ids_busca_variantes(id_busca)
    ids = _conferencia_expandir_ids_romaneio(conn, ids, dataset_id=ds)
    ids = _conferencia_enriquecer_ids_de_importacao(conn, ids, dataset_id=ds)
    if not ids:
        return []
    ph = ','.join(['?'] * len(ids))
    try:
        rows = conn.execute(
            f"""SELECT {_ROMANEIO_CONFERENCIA_COLS}
                FROM romaneio_por_item
                WHERE dataset_id = ?
                  AND (
                    TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
                    OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph})
                  )
                ORDER BY row_index
                LIMIT ?""",
            [str(ds)] + ids + ids + [int(limit_romaneio)],
        ).fetchall()
    except Exception:
        rows = []
    return rows or []


def _conferencia_build_lista_pg(conn, id_viagem, fluxo_req='carregamento', limit_romaneio=400, somente_divergencias=False):
    """Monta lista de conferência no Postgres. Retorna (lista, meta, id_para_bipados, erro)."""
    id_viagem = (id_viagem or '').strip()
    if not id_viagem:
        return [], {}, '', 'ID do roteiro não informado'
    fluxo_req = (fluxo_req or 'carregamento').strip().lower()
    if fluxo_req not in ('carregamento', 'devolucao'):
        fluxo_req = 'carregamento'
    id_viagem_norm = _normalizar_id_viagem(id_viagem)
    _ensure_pg_produtos_bipados_fluxo(conn)
    id_busca = (id_viagem_norm or id_viagem or '').strip()
    ds = _conferencia_resolver_dataset_id(conn, id_busca)
    if not ds:
        return [], {}, '', 'Nenhum dataset ativo. Importe a base primeiro.'
    try:
        limit_romaneio = min(2000, max(200, int(limit_romaneio)))
    except (TypeError, ValueError):
        limit_romaneio = 400
    romaneio_rows = _conferencia_buscar_romaneio_rows(conn, ds, id_busca, limit_romaneio)
    id_para_bipados = id_busca
    if romaneio_rows:
        r0 = romaneio_rows[0]
        id_para_bipados = str(r0.get('id_viagem') or '').strip() or id_busca
    codigos_produto_romaneio = list({(str(r.get('codigo_produto') or '').strip()) for r in romaneio_rows if (r.get('codigo_produto') or '').strip()})
    if codigos_produto_romaneio:
        placeholders = ','.join(['?' for _ in codigos_produto_romaneio])
        base_rows = conn.execute(
            "SELECT codigo_interno, ean, dun FROM base_codigo_barras WHERE dataset_id = ? AND codigo_interno IN (" + placeholders + ")",
            [str(ds)] + codigos_produto_romaneio,
        ).fetchall()
    else:
        base_rows = []
    mapa_codigo_barras = {}
    mapa_ean = {}
    mapa_dun = {}
    mapa_barras_to_codigo = {}
    for r in base_rows or []:
        cod = ((r.get('codigo_interno') or '') if hasattr(r, 'get') else (r[0] if len(r) > 0 else '')).strip() if r else ''
        ean = ((r.get('ean') or '') if hasattr(r, 'get') else (r[1] if len(r) > 1 else '')).strip() if r else ''
        dun = ((r.get('dun') or '') if hasattr(r, 'get') else (r[2] if len(r) > 2 else '')).strip() if r else ''
        if cod:
            if ean:
                mapa_ean[cod] = ean
                mapa_barras_to_codigo[ean] = cod
            if dun:
                mapa_dun[cod] = dun
                mapa_barras_to_codigo[dun] = cod
            mapa_codigo_barras[cod] = ean or dun
    bipados_rows = conn.execute(
        """SELECT codigo_barras,
                  SUM(quantidade) as quantidade_bipada,
                  MAX(codigo_interno) as codigo_interno,
                  MAX(produto) as produto,
                  MAX(unidade) as unidade,
                  MAX(peso) as peso
           FROM produtos_bipados
           WHERE id_viagem = ? AND COALESCE(fluxo, 'carregamento') = ?
             AND codigo_barras IS NOT NULL AND trim(codigo_barras) != ''
           GROUP BY codigo_barras""",
        (id_para_bipados, fluxo_req),
    ).fetchall()
    bipados_dict = {}
    for row in bipados_rows or []:
        cb = ((row.get('codigo_barras') or '') if hasattr(row, 'get') else (row[0] if len(row) > 0 else '')).strip()
        q = int((row.get('quantidade_bipada') or 0) if hasattr(row, 'get') else (row[1] if len(row) > 1 else 0))
        bipados_dict[cb] = bipados_dict.get(cb, 0) + q
    bipados_por_codigo = {}
    for cb, qtd in bipados_dict.items():
        cod_prod = mapa_barras_to_codigo.get(cb) or cb
        bipados_por_codigo[cod_prod] = max(bipados_por_codigo.get(cod_prod, 0), qtd)
    resultado = []
    romaneio_agregado = _agrupar_romaneio_rows_por_codigo_produto(
        romaneio_rows, mapa_ean, mapa_dun, mapa_codigo_barras
    )
    for item_rom in romaneio_agregado:
        codigo_produto = item_rom['codigo_produto']
        descricao = item_rom['descricao']
        quantidade_produto = int(item_rom['quantidade_produto'] or 0)
        unidade = item_rom['unidade']
        peso_bruto = item_rom['peso_bruto']
        codigo_barras = item_rom['codigo_barras'] or ''
        quantidade_bipada = int(bipados_por_codigo.get(codigo_produto, 0) or 0)
        if not quantidade_bipada and codigo_barras:
            quantidade_bipada = int(bipados_dict.get(codigo_barras, 0) or 0)
        for cb_extra in item_rom.get('_todos_codigos_barras') or []:
            if cb_extra and cb_extra != codigo_barras:
                quantidade_bipada = max(
                    quantidade_bipada,
                    int(bipados_dict.get(cb_extra, 0) or 0),
                )
        quantidade_falta = max(0, quantidade_produto - quantidade_bipada)
        quantidade_sobra = max(0, quantidade_bipada - quantidade_produto)
        if somente_divergencias and quantidade_falta <= 0 and quantidade_sobra <= 0:
            continue
        if quantidade_sobra > 0:
            status_bipado = 'EXCEDENTE'
        elif quantidade_bipada >= quantidade_produto and quantidade_produto > 0:
            status_bipado = 'COMPLETO'
        elif quantidade_bipada > 0:
            status_bipado = 'PARCIAL'
        else:
            status_bipado = 'PENDENTE'
        todos_cb = item_rom.get('_todos_codigos_barras') or []
        aviso_sobra = ('Bipou %s a mais' % quantidade_sobra) if quantidade_sobra > 0 else ''
        if len(todos_cb) > 1:
            aviso_multi = '⚠️ Múltiplos códigos: ' + ', '.join(todos_cb)
            aviso_sobra = (aviso_sobra + ' — ' + aviso_multi) if aviso_sobra else aviso_multi
        resultado.append({
            'codigo_produto': codigo_produto,
            'codigo_barras': codigo_barras,
            'produto': descricao,
            'quantidade_produto': quantidade_produto,
            'unidade': unidade,
            'peso_bruto': peso_bruto,
            'quantidade_bipada': quantidade_bipada,
            'quantidade_falta': quantidade_falta,
            'quantidade_sobra': quantidade_sobra,
            'aviso_sobra': aviso_sobra,
            'status_bipado': status_bipado,
            'id_viagem': id_para_bipados,
            'origem_romaneio': True,
        })
    _conferencia_incluir_extras_bipados(resultado, bipados_rows, mapa_barras_to_codigo, id_para_bipados)
    if somente_divergencias:
        resultado = [
            it for it in resultado
            if (it.get('quantidade_falta') or 0) > 0 or (it.get('quantidade_sobra') or 0) > 0
        ]
    meta = {'id_roteiro': '', 'identificador_rota': '', 'placa': '', 'motorista': '', 'data_expedicao': ''}
    if romaneio_rows:
        r0 = romaneio_rows[0]
        meta['id_roteiro'] = str(r0.get('id_roteiro') or '').strip() or ''
        meta['identificador_rota'] = str(r0.get('identificador_rota') or '').strip() or ''
        meta['placa'] = str(r0.get('placa') or '').strip() or ''
        meta['motorista'] = str(r0.get('motorista') or '').strip() or ''
        meta['data_expedicao'] = str(r0.get('data_expedicao') or '').strip() or ''
    return resultado, meta, id_para_bipados, None


@app.route('/api/conferencia', methods=['GET'])
@app.route('/api/conferencia/<id_viagem>', methods=['GET'])
def get_conferencia(id_viagem=None):
    """Retorna itens do romaneio para uma viagem específica com status de bipado"""
    if not id_viagem:
        return jsonify({'erro': 'ID do roteiro não informado'}), 400
    
    id_viagem = (id_viagem or '').strip()
    id_viagem_norm = _normalizar_id_viagem(id_viagem)
    fluxo_req = getattr(g, 'conferencia_fluxo', None)
    if fluxo_req is None:
        fluxo_req = request.args.get('fluxo', 'carregamento') if request else 'carregamento'
    if fluxo_req not in ('carregamento', 'devolucao'):
        fluxo_req = 'carregamento'

    # Quando DATABASE_URL está definido: ler do banco (romaneio_por_item)
    if _usa_banco_para_dados():
        conn = get_db()
        try:
            if getattr(conn, 'kind', None) != 'pg':
                conn.close()
                return jsonify({'erro': 'Conferência disponível apenas com Postgres (DATABASE_URL).'}), 400
            id_busca = (id_viagem_norm or id_viagem or '').strip()
            try:
                limit_romaneio = min(2000, max(200, int(request.args.get('limit', 400))))
            except (TypeError, ValueError):
                limit_romaneio = 400
            resultado, meta, id_para_bipados, err_build = _conferencia_build_lista_pg(
                conn, id_viagem, fluxo_req, limit_romaneio=limit_romaneio, somente_divergencias=False,
            )
            if err_build:
                conn.close()
                return jsonify({'erro': err_build}), 400
            ds = _conferencia_resolver_dataset_id(conn, id_busca)
            id_para_lookup = id_para_bipados or id_busca
            meta = _conferencia_enriquecer_meta_pg(conn, ds, id_para_lookup, id_busca, meta)
            carregar_motivos = request.args.get('motivos', '1').strip().lower() not in ('0', 'false', 'no')
            if carregar_motivos:
                try:
                    _carregar_motivos_divergencia(resultado, conn=conn)
                except Exception:
                    pass
            id_viagem_resposta = str(id_para_bipados).strip() if id_para_bipados else ''
            id_norm_resp = _normalizar_id_viagem(id_para_lookup or id_busca)
            carregar_periodo = request.args.get('periodo_meta', '1').strip().lower() not in ('0', 'false', 'no')
            if carregar_periodo:
                inicio_p, fim_p = _consultar_periodo_viagem_raw_conn(conn, id_norm_resp or id_viagem_resposta, fluxo_req)
                inicio_str = _formatar_data_hora_periodo(inicio_p)
                fim_str = _formatar_data_hora_periodo(fim_p)
            else:
                inicio_str = ''
                fim_str = ''
            ja_baixado_ravex = True
            aviso_ravex = ''
            aviso_romaneio_vazio = ''
            info_b = None
            verificar_baixado = request.args.get('verificar_baixado', '1').strip().lower() not in ('0', 'false', 'no')
            if verificar_baixado:
                ja_b, info_b = _ravex_id_input_ja_baixado(conn, ds, id_busca)
                ja_baixado_ravex = bool(ja_b)
                if not ja_baixado_ravex:
                    aviso_ravex = _ravex_mensagem_nao_baixado(id_busca)
            total_romaneio = len(resultado)
            id_roteiro_resp = meta.get('id_roteiro') or ''
            if not id_roteiro_resp and info_b:
                id_roteiro_resp = str(info_b.get('id_roteiro') or '').strip()
            if not id_viagem_resposta and info_b:
                id_viagem_resposta = str(info_b.get('id_viagem') or '').strip()
            if total_romaneio == 0 and ja_baixado_ravex and info_b:
                n_rom = int(info_b.get('itens_romaneio') or 0)
                id_v_hist = str(info_b.get('id_viagem') or '').strip()
                id_r_hist = str(info_b.get('id_roteiro') or '').strip()
                if n_rom > 0 and id_v_hist and id_v_hist != id_busca:
                    aviso_romaneio_vazio = (
                        'Romaneio com %s itens vinculado à viagem %s (roteiro %s). '
                        'Recarregue a página e busque novamente; a bipagem usa o ID da viagem.'
                        % (n_rom, id_v_hist, id_r_hist or id_busca)
                    )
                elif info_b.get('fonte') == 'historico' and n_rom <= 0:
                    aviso_romaneio_vazio = (
                        'Este ID consta como baixado no Ravex, mas o romaneio está vazio no banco. '
                        'Reimporte pela aba Importar Ravex.'
                    )
            conn.close()
            return jsonify({
                'lista': resultado,
                'lista_ja_agregada': True,
                'id_roteiro': id_roteiro_resp,
                'id_viagem': id_viagem_resposta,
                'identificador_rota': meta.get('identificador_rota') or '',
                'placa': meta.get('placa') or '',
                'motorista': meta.get('motorista') or '',
                'data_expedicao': meta.get('data_expedicao') or '',
                'coordenador': _coordenador_resposta_api(meta.get('coordenador')),
                'conferente': meta.get('conferente') or '',
                'ajudante1': meta.get('ajudante1') or '',
                'ajudante2': meta.get('ajudante2') or '',
                'inicio_carregamento': inicio_str,
                'fim_carregamento': fim_str,
                'total_romaneio': total_romaneio,
                'limit_romaneio': limit_romaneio,
                'ja_baixado_ravex': ja_baixado_ravex,
                'aviso_ravex': aviso_ravex,
                'aviso_romaneio_vazio': aviso_romaneio_vazio,
            })
        except Exception as e:
            try:
                conn.close()
            except Exception:
                pass
            return jsonify({'erro': 'Erro ao carregar conferência: %s' % str(e)}), 500

    # Dados vêm apenas do banco; não usar mais planilha
    return jsonify({'erro': 'Configure DATABASE_URL. Os dados vêm apenas do banco de dados.'}), 400
    
    wb, from_cache = get_workbook_cached()
    if not wb:
        return jsonify({'erro': 'Planilha não encontrada'}), 404
    
    try:
        resultado = []
        
        # Buscar itens do romaneio (aba ROMANEIO POR ITEM) e código de barras na aba BASE
        nome_romaneio = next((s for s in wb.sheetnames if 'ROMANEIO' in s.upper() and 'ITEM' in s.upper()), None)
        if nome_romaneio:
            ws_romaneio = wb[nome_romaneio]
            
            # Ler cabeçalho para identificar todas as colunas
            header_row = list(ws_romaneio.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            coluna_id_viagem = None
            coluna_codigo_produto = None
            coluna_descricao_produto = None
            coluna_quantidade_produto = None
            coluna_unidade = None
            coluna_peso_bruto = None
            
            # Identificar colunas pelo cabeçalho
            for idx, header in enumerate(header_row):
                if header:
                    header_str = str(header).upper().strip()
                    
                    # ID ROTEIRO (coluna B) tem prioridade; fallback: ID VIAGEM FATURADA (coluna D)
                    if 'ID' in header_str and 'ROTEIRO' in header_str:
                        coluna_id_viagem = idx
                    elif coluna_id_viagem is None and 'ID' in header_str and 'VIAGEM' in header_str and 'FATURADA' in header_str:
                        coluna_id_viagem = idx
                    
                    # CÓDIGO DO PRODUTO
                    if ('CODIGO' in header_str or 'CÓDIGO' in header_str) and 'PRODUTO' in header_str:
                        if 'BARRAS' not in header_str:  # Não é código de barras
                            coluna_codigo_produto = idx
                    
                    # DESCRIÇÃO DO PRODUTO (prioridade)
                    if 'DESCRIÇÃO' in header_str or 'DESCRICAO' in header_str:
                        if 'PRODUTO' in header_str:
                            coluna_descricao_produto = idx
                    
                    # PRODUTO (fallback se não encontrar descrição)
                    if coluna_descricao_produto is None and 'PRODUTO' in header_str:
                        if 'CODIGO' not in header_str and 'CÓDIGO' not in header_str:  # Não é código do produto
                            coluna_descricao_produto = idx
                    
                    # QUANTIDADE DO PRODUTO
                    if 'QUANTIDADE' in header_str and 'PRODUTO' in header_str:
                        coluna_quantidade_produto = idx
                    # Fallback: QUANTIDADE ROMANEIO ou apenas QUANTIDADE
                    elif coluna_quantidade_produto is None and 'QUANTIDADE' in header_str:
                        if 'ROMANEIO' not in header_str:  # Evitar pegar quantidade romaneio
                            coluna_quantidade_produto = idx
                    
                    # UNIDADE: prioridade para "Unidade de medida do produto" (aba ROMANEIO POR ITEM)
                    if 'UNIDADE' in header_str and 'MEDIDA' in header_str:
                        coluna_unidade = idx  # Unidade de medida do produto
                    elif coluna_unidade is None and ('UNIDADE' in header_str or 'PACOTE' in header_str or header_str in ('CAIXA', 'UN')):
                        coluna_unidade = idx
                    
                    # PESO BRUTO PEDIDO (aba Romaneio por item)
                    if 'PESO' in header_str and 'BRUTO' in header_str and 'PEDIDO' in header_str:
                        coluna_peso_bruto = idx
                    elif coluna_peso_bruto is None and 'PESO' in header_str and 'BRUTO' in header_str:
                        coluna_peso_bruto = idx
            
            # ROMANEIO POR ITEM: coluna B = Id roteiro (1), O = Código do produto (14), P = Descrição (15), Q = Quantidade (16)
            num_cols = len(header_row) if header_row else 0
            if num_cols >= 17:
                coluna_id_viagem = 1  # Coluna B = Id roteiro (filtro por roteiro)
                coluna_codigo_produto = 14
                coluna_descricao_produto = 15
                coluna_quantidade_produto = 16
            else:
                if coluna_id_viagem is None and num_cols >= 2:
                    coluna_id_viagem = 1  # Coluna B = Id roteiro
                if coluna_codigo_produto is None:
                    coluna_codigo_produto = 0
                if coluna_descricao_produto is None:
                    coluna_descricao_produto = 1
                if coluna_quantidade_produto is None:
                    coluna_quantidade_produto = 2
            
            mapa_codigo_barras = {}
            mapa_barras_to_codigo_produto = {}
            bipados_por_codigo_produto = {}
            nome_aba_base = next((s for s in wb.sheetnames if s.upper().strip() == 'BASE'), None)
            if nome_aba_base:
                ws_base = wb[nome_aba_base]
                base_header = list(ws_base.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                col_codigo = next((i for i, h in enumerate(base_header or []) if _coluna_base_eh_codigo_interno(str(h or ''))), 0)
                col_ean = next((i for i, h in enumerate(base_header or []) if _coluna_base_eh_ean(str(h or ''))), None)
                col_dun = next((i for i, h in enumerate(base_header or []) if _coluna_base_eh_dun(str(h or ''))), None)
                for row in ws_base.iter_rows(min_row=2, values_only=True):
                    if col_codigo >= len(row):
                        continue
                    codigo_interno_base = _valor_celula(row, col_codigo)
                    if not codigo_interno_base:
                        continue
                    ean = _valor_celula(row, col_ean) if col_ean is not None else ''
                    dun = _valor_celula(row, col_dun) if col_dun is not None else ''
                    barcode_exibir = ean or dun
                    if not barcode_exibir:
                        continue
                    mapa_codigo_barras[codigo_interno_base] = barcode_exibir
                    norm = _normalizar_codigo_produto(codigo_interno_base)
                    if norm:
                        mapa_codigo_barras[norm] = barcode_exibir
                    for v in _variantes_codigo_produto(codigo_interno_base):
                        if v:
                            mapa_codigo_barras[v] = barcode_exibir
                    mapa_barras_to_codigo_produto[ean] = codigo_interno_base
                    mapa_barras_to_codigo_produto[dun] = codigo_interno_base
                    mapa_barras_to_codigo_produto[barcode_exibir] = codigo_interno_base

            conn = get_db()
            produtos_bipados = conn.execute(
                'SELECT codigo_barras, SUM(quantidade) as quantidade_bipada FROM produtos_bipados WHERE id_viagem = ? GROUP BY codigo_barras',
                (id_viagem,)
            ).fetchall()
            bipados_dict = {}
            for row in produtos_bipados:
                bipados_dict[row['codigo_barras']] = row['quantidade_bipada']
            bipados_por_codigo_produto = {}
            for cb, qtd in bipados_dict.items():
                codigo_prod = mapa_barras_to_codigo_produto.get(cb) or cb
                bipados_por_codigo_produto[codigo_prod] = max(bipados_por_codigo_produto.get(codigo_prod, 0), qtd)
            conn.close()
            
            # Ler itens do romaneio filtrando por ID do roteiro (coluna B)
            for idx, row in enumerate(ws_romaneio.iter_rows(min_row=2, values_only=True), start=2):
                # Verificar se o ID do roteiro corresponde (coluna B)
                if coluna_id_viagem is not None and coluna_id_viagem < len(row):
                    id_viagem_planilha = _normalizar_id_viagem(row[coluna_id_viagem])
                    id_viagem_busca = _normalizar_id_viagem(id_viagem)
                    if id_viagem_planilha and id_viagem_busca and id_viagem_planilha != id_viagem_busca:
                        continue
                
                # Ler código do produto usando a coluna identificada
                if coluna_codigo_produto < len(row) and row[coluna_codigo_produto]:
                    codigo_produto_valor = str(row[coluna_codigo_produto]).strip()
                    # Pular se for fórmula ou vazio
                    if not codigo_produto_valor or codigo_produto_valor.startswith('='):
                        continue
                else:
                    continue
                
                try:
                    codigo_produto = codigo_produto_valor
                    codigo_produto_norm = _normalizar_codigo_produto(row[coluna_codigo_produto])
                    if not codigo_produto_norm:
                        codigo_produto_norm = _normalizar_codigo_produto(codigo_produto_valor)
                    
                    # Consultar aba BASE: código do produto (Código Interno) -> código de barras (tentar valor, normalizado e variantes)
                    codigo_barras = mapa_codigo_barras.get(codigo_produto, '') or mapa_codigo_barras.get(codigo_produto_norm, '')
                    if not codigo_barras:
                        for v in _variantes_codigo_produto(codigo_produto) + _variantes_codigo_produto(codigo_produto_norm or ''):
                            if v and v in mapa_codigo_barras:
                                codigo_barras = mapa_codigo_barras[v]
                                break
                    
                    # Ler descrição do produto
                    produto = ''
                    if coluna_descricao_produto < len(row) and row[coluna_descricao_produto]:
                        produto = str(row[coluna_descricao_produto]).strip()
                    
                    # Quantidade do produto usando a coluna identificada
                    quantidade_produto = 0
                    if coluna_quantidade_produto < len(row) and row[coluna_quantidade_produto] is not None:
                        try:
                            valor = str(row[coluna_quantidade_produto]).strip()
                            if valor and not valor.startswith('='):
                                quantidade_produto = int(float(valor))
                        except:
                            quantidade_produto = 0
                    
                    # Unidade (pacote, caixa, etc.)
                    unidade = ''
                    if coluna_unidade is not None and coluna_unidade < len(row) and row[coluna_unidade]:
                        unidade = str(row[coluna_unidade]).strip()
                        if unidade.startswith('='):
                            unidade = ''
                    
                    # Peso Bruto (Peso Bruto Pedido do Romaneio por Item)
                    peso_bruto = ''
                    if coluna_peso_bruto is not None and coluna_peso_bruto < len(row) and row[coluna_peso_bruto] is not None:
                        v = row[coluna_peso_bruto]
                        if not str(v).strip().startswith('='):
                            peso_bruto = str(v).strip()
                    
                    # Quantidade já bipada: por código do produto (bip pode ser EAN ou DUN)
                    quantidade_bipada = bipados_por_codigo_produto.get(codigo_produto, 0) or bipados_por_codigo_produto.get(codigo_produto_norm or '', 0)
                    if quantidade_bipada == 0 and codigo_barras:
                        for v in _variantes_codigo_produto(codigo_produto) + _variantes_codigo_produto(codigo_produto_norm or ''):
                            if v and bipados_por_codigo_produto.get(v, 0) > 0:
                                quantidade_bipada = bipados_por_codigo_produto[v]
                                break
                    if quantidade_bipada == 0 and codigo_barras:
                        quantidade_bipada = bipados_dict.get(codigo_barras, 0)
                    
                    quantidade_falta = max(0, quantidade_produto - quantidade_bipada)
                    quantidade_sobra = max(0, quantidade_bipada - quantidade_produto)
                    aviso_sobra = ('Bipou %s a mais' % quantidade_sobra) if quantidade_sobra > 0 else ''
                    # Status: EXCEDENTE quando bipado a mais, COMPLETO quando certinho, PARCIAL quando falta, PENDENTE quando zero
                    if quantidade_sobra > 0:
                        status_bipado = 'EXCEDENTE'
                    elif quantidade_bipada >= quantidade_produto and quantidade_produto > 0:
                        status_bipado = 'COMPLETO'
                    elif quantidade_bipada > 0:
                        status_bipado = 'PARCIAL'
                    else:
                        status_bipado = 'PENDENTE'
                    
                    resultado.append({
                        'codigo_produto': codigo_produto,
                        'codigo_barras': codigo_barras,
                        'produto': produto,
                        'quantidade_produto': quantidade_produto,
                        'unidade': unidade or '-',
                        'peso_bruto': peso_bruto or '-',
                        'quantidade_bipada': quantidade_bipada,
                        'quantidade_falta': quantidade_falta,
                        'quantidade_sobra': quantidade_sobra,
                        'aviso_sobra': aviso_sobra,
                        'status_bipado': status_bipado,
                        'origem_romaneio': True,
                    })
                except Exception as e:
                    continue
        
        # Incluir itens bipados que NÃO estão no romaneio (adicionados via "adicionar na conferência")
        conn = get_db()
        conn.row_factory = sqlite3.Row
        extras = conn.execute('''
            SELECT codigo_barras, MAX(codigo_interno) as codigo_interno, MAX(produto) as produto, SUM(quantidade) as quantidade_bipada, MAX(unidade) as unidade, MAX(peso) as peso
            FROM produtos_bipados
            WHERE id_viagem = ? AND codigo_barras IS NOT NULL AND trim(codigo_barras) != ''
            GROUP BY codigo_barras
        ''', (id_viagem,)).fetchall()
        conn.close()
        for item in resultado:
            if item.get('codigo_produto') and not item.get('origem_romaneio'):
                item['origem_romaneio'] = True
        _conferencia_incluir_extras_bipados(resultado, extras, mapa_barras_to_codigo_produto, id_viagem)
        
        for item in resultado:
            item['id_viagem'] = id_viagem
        _carregar_motivos_divergencia(resultado)
        
        if not from_cache:
            wb.close()
        return jsonify(resultado)
        
    except Exception as e:
        if not from_cache and wb:
            try:
                wb.close()
            except Exception:
                pass
        return jsonify({'erro': f'Erro ao ler planilha: {str(e)}'}), 500


def _normalizar_cidade_nome(val):
    """Extrai o nome da cidade por extenso. Se val for objeto/dict (ex: {id, nome, uf, codigoIbge}), retorna nome. Se for string JSON, parse e retorna nome. Caso contrário retorna a string."""
    if val is None:
        return ''
    if isinstance(val, dict):
        return (val.get('nome') or val.get('nomeCidade') or '').strip() or ''
    if isinstance(val, str):
        s = val.strip()
        if s.startswith('{'):
            try:
                obj = json.loads(s)
                if isinstance(obj, dict):
                    return (obj.get('nome') or obj.get('nomeCidade') or '').strip() or ''
            except Exception:
                pass
        return s
    return str(val).strip()


def _romaneio_data_para_json(val):
    """Converte campo 'data' do romaneio para string JSON (psycopg não adapta dict)."""
    if val is None:
        return '{}'
    if isinstance(val, str):
        return val
    if isinstance(val, (dict, list)):
        return json.dumps(val)
    return json.dumps(val)  # qualquer outro tipo (evita cannot adapt type)


def _romaneio_linha_para_tuple_pg(ds, L):
    """Converte uma linha L para tupla de parâmetros seguros para psycopg (nenhum dict/list)."""
    def _str(v):
        if v is None:
            return None
        if isinstance(v, (dict, list)):
            return json.dumps(v)
        return str(v) if not isinstance(v, str) else (v or None)
    def _int(v):
        if v is None:
            return 0
        try:
            return int(v)
        except (TypeError, ValueError):
            return 0
    # cidade: gravar sempre como nome por extenso (não como objeto JSON)
    cidade_val = L.get('cidade')
    if cidade_val is not None and (isinstance(cidade_val, (dict, list)) or (isinstance(cidade_val, str) and cidade_val.strip().startswith('{'))):
        cidade_val = _normalizar_cidade_nome(cidade_val)
    # Garantir que o JSON "data" carregue o campo importado_em para aparecer como coluna na UI
    data_val = L.get('data')
    importado_em_val = L.get('importado_em')
    if importado_em_val:
        try:
            if isinstance(data_val, dict):
                if 'importado_em' not in data_val:
                    data_val = dict(data_val)
                    data_val['importado_em'] = importado_em_val
            elif isinstance(data_val, str) and data_val.strip().startswith('{'):
                dtmp = json.loads(data_val)
                if isinstance(dtmp, dict) and 'importado_em' not in dtmp:
                    dtmp['importado_em'] = importado_em_val
                    data_val = dtmp
        except Exception:
            pass
    data_json = _romaneio_data_para_json(data_val)
    return (
        str(ds),
        _int(L.get('row_index')),
        _str(L.get('id_roteiro')),
        _str(L.get('id_viagem')),
        _str(L.get('identificador_rota')),
        _str(L.get('codigo_produto')),
        _str(L.get('descricao')),
        _int(L.get('quantidade')),
        _str(L.get('unidade')),
        _str(L.get('peso_bruto')),
        _str(L.get('codigo_cliente')),
        _str(L.get('endereco')),
        _str(cidade_val if cidade_val is not None else L.get('cidade')),
        _str(L.get('placa')),
        _str(L.get('motorista')),
        _str(L.get('data_expedicao')),
        _str(L.get('importado_em')),
        data_json,
    )


def _ravex_parse_forcar_reimportar(data):
    """Permite rebaixar viagem já importada (body: forcar_reimportar, forcar ou reimportar)."""
    if not isinstance(data, dict):
        return False
    return bool(data.get('forcar_reimportar') or data.get('forcar') or data.get('reimportar'))


_RAVEX_IMPORT_WORKERS = max(1, min(12, int(os.environ.get('RAVEX_IMPORT_WORKERS', '6') or 6)))


def _ravex_identificador_rota_valido(identificador_rota):
    ident = (identificador_rota or '').strip()
    if not ident:
        return False
    if len(ident) >= 15 and ' - ' not in ident:
        return False
    return True


_ROMANEIO_INSERT_SQL = """INSERT INTO romaneio_por_item
   (dataset_id, row_index, id_roteiro, id_viagem, identificador_rota, codigo_produto, descricao, quantidade, unidade, peso_bruto,
    codigo_cliente, endereco, cidade, placa, motorista, data_expedicao, importado_em, data)
   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?::jsonb)"""

_ROMANEIO_INSERT_PG_TEMPLATE = (
    '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)'
)
_ROMANEIO_INSERT_PG_SQL = """INSERT INTO romaneio_por_item
   (dataset_id, row_index, id_roteiro, id_viagem, identificador_rota, codigo_produto, descricao, quantidade, unidade, peso_bruto,
    codigo_cliente, endereco, cidade, placa, motorista, data_expedicao, importado_em, data)
   VALUES %s"""


def _ravex_bulk_insert_romaneio_pg(conn, params):
    """INSERT em lote no Postgres (execute_values — mais rápido que executemany linha a linha)."""
    try:
        from psycopg.extras import execute_values  # type: ignore
    except ImportError:
        execute_values = None
    raw = getattr(conn, '_conn', conn)
    if execute_values:
        with raw.cursor() as cur:
            execute_values(
                cur,
                _ROMANEIO_INSERT_PG_SQL,
                params,
                template=_ROMANEIO_INSERT_PG_TEMPLATE,
                page_size=500,
            )
        return
    with raw.cursor() as cur:
        cur.executemany(_ROMANEIO_INSERT_SQL.replace('?', '%s'), params)


def _ravex_proximo_row_index_romaneio(conn, ds):
    """Próximo row_index livre no dataset (PK é dataset_id + row_index, global por dataset)."""
    row = conn.execute(
        """SELECT COALESCE(MAX(row_index), 0) AS mx FROM romaneio_por_item WHERE dataset_id = ?""",
        (str(ds),),
    ).fetchone()
    if row is None:
        return 0
    if hasattr(row, 'get'):
        return int(row.get('mx') or 0)
    return int(row[0] or 0)


def _ravex_insert_linhas_romaneio(conn, ds, id_viagem, linhas, id_roteiro=None):
    """Remove romaneio anterior da viagem e grava linhas com executemany."""
    if not linhas:
        return 0
    id_roteiro_del = str(id_roteiro or (linhas[0].get('id_roteiro') if linhas else '') or '').strip()
    id_viagem_del = str(id_viagem or '').strip()
    if id_roteiro_del and (not id_viagem_del or id_viagem_del == id_roteiro_del):
        conn.execute(
            """DELETE FROM romaneio_por_item
               WHERE dataset_id = ? AND TRIM(COALESCE(id_roteiro::text, '')) = ?""",
            (str(ds), id_roteiro_del),
        )
        id_viagem_grav = id_viagem_del or id_roteiro_del
    else:
        id_viagem_norm = str(_normalizar_id_viagem(id_viagem_del) or id_viagem_del).strip()
        conn.execute(
            """DELETE FROM romaneio_por_item
               WHERE dataset_id = ?
                 AND (
                   TRIM(COALESCE(id_viagem::text, '')) = ?
                   OR TRIM(COALESCE(id_viagem::text, '')) = ?
                 )""",
            (str(ds), id_viagem_del, id_viagem_norm),
        )
        id_viagem_grav = id_viagem_del
    next_row_index = _ravex_proximo_row_index_romaneio(conn, ds)
    importado_em = datetime.now(_ravex_tz_baixados()).isoformat()
    params = []
    for L in linhas:
        next_row_index += 1
        try:
            L['row_index'] = next_row_index
            L['importado_em'] = importado_em
        except Exception:
            pass
        params.append(_romaneio_linha_para_tuple_pg(ds, L))
    if len(params) == 1:
        conn.execute(_ROMANEIO_INSERT_SQL, params[0])
    elif getattr(conn, 'kind', None) == 'pg':
        _ravex_bulk_insert_romaneio_pg(conn, params)
    else:
        conn.executemany(_ROMANEIO_INSERT_SQL, params)
    _upsert_viagem_placa_motorista(conn, id_viagem_grav, linhas[0].get('placa'), linhas[0].get('motorista'))
    return len(params)


def _ravex_viagens_ja_baixadas_batch(conn, dataset_id, id_viagens):
    """Retorna set de id_viagem que já possuem romaneio importado no dataset."""
    ids_norm = []
    vistos = set()
    for id_v in id_viagens or []:
        id_norm = str(_normalizar_id_viagem(id_v) or id_v or '').strip()
        if id_norm and id_norm not in vistos:
            vistos.add(id_norm)
            ids_norm.append(id_norm)
    if not ids_norm or not dataset_id:
        return set()
    ph = ','.join(['?'] * len(ids_norm))
    try:
        rows = conn.execute(
            """SELECT DISTINCT TRIM(COALESCE(id_viagem::text, '')) AS id_v
               FROM romaneio_por_item
               WHERE dataset_id = ? AND TRIM(COALESCE(id_viagem::text, '')) IN (""" + ph + """)""",
            [str(dataset_id)] + ids_norm,
        ).fetchall()
    except Exception:
        return set()
    out = set()
    for r in rows or []:
        id_v = str((r.get('id_v') if hasattr(r, 'get') else (r[0] if len(r) > 0 else '')) or '').strip()
        if id_v:
            out.add(id_v)
    return out


def _ravex_id_input_ja_baixado(conn, dataset_id, id_input):
    """Verifica se o ID informado (roteiro ou viagem) já foi baixado (romaneio ou histórico)."""
    id_norm = str(id_input or '').strip()
    if not id_norm:
        return False, None
    ids = _conferencia_ids_busca_variantes(id_norm)
    ids = _conferencia_expandir_ids_romaneio(conn, ids)
    ids = _conferencia_enriquecer_ids_de_importacao(conn, ids)
    ds_lista = []
    if dataset_id:
        ds_lista.append(str(dataset_id))
    ds_resolvido = _conferencia_resolver_dataset_id(conn, id_norm)
    if ds_resolvido and str(ds_resolvido) not in ds_lista:
        ds_lista.append(str(ds_resolvido))
    for ds in ds_lista:
        info = _ravex_romaneio_info_por_ids(conn, ds, ids)
        if info:
            info['id_informado'] = id_norm
            return True, info
    ph = ','.join(['?'] * len(ids)) if ids else ''
    hist = None
    if ids:
        try:
            hist = conn.execute(
                f"""SELECT criado_em, total_itens, parametros
                   FROM ravex_importacoes
                   WHERE status IN ('OK', 'OK_COM_ERROS', 'DUPLICADO')
                     AND (
                       TRIM(COALESCE(parametros->>'id_viagem', '')) IN ({ph})
                       OR TRIM(COALESCE(parametros->>'id_roteiro', '')) IN ({ph})
                       OR TRIM(COALESCE(parametros->>'id_informado', '')) IN ({ph})
                     )
                   ORDER BY criado_em DESC NULLS LAST
                   LIMIT 1""",
                ids + ids + ids,
            ).fetchone()
        except Exception:
            hist = None
    if not hist and dataset_id and ids:
        try:
            hist = conn.execute(
                f"""SELECT criado_em, total_itens, parametros
                   FROM ravex_importacoes
                   WHERE dataset_id = ? AND status IN ('OK', 'OK_COM_ERROS', 'DUPLICADO')
                     AND (
                       TRIM(COALESCE(parametros->>'id_viagem', '')) IN ({ph})
                       OR TRIM(COALESCE(parametros->>'id_roteiro', '')) IN ({ph})
                       OR TRIM(COALESCE(parametros->>'id_informado', '')) IN ({ph})
                     )
                   ORDER BY criado_em DESC
                   LIMIT 1""",
                [str(dataset_id)] + ids + ids + ids,
            ).fetchone()
        except Exception:
            hist = None
    if hist:
        criado = hist.get('criado_em') if hasattr(hist, 'get') else (hist[0] if hist else None)
        total_h = int((hist.get('total_itens') if hasattr(hist, 'get') else (hist[1] if len(hist) > 1 else 0)) or 0)
        params = hist.get('parametros') if hasattr(hist, 'get') else (hist[2] if len(hist) > 2 else None)
        if isinstance(params, str):
            try:
                params = json.loads(params)
            except Exception:
                params = {}
        if not isinstance(params, dict):
            params = {}
        info = {
            'itens_romaneio': 0,
            'importado_em': criado,
            'total_itens_historico': total_h,
            'fonte': 'historico',
            'id_viagem': str(params.get('id_viagem') or '').strip(),
            'id_roteiro': str(params.get('id_roteiro') or '').strip(),
            'id_informado': id_norm,
        }
        return True, info
    id_viagem_norm = str(_normalizar_id_viagem(id_norm) or id_norm).strip()
    if id_viagem_norm and id_viagem_norm != id_norm:
        ja_v, info_v = _ravex_viagem_ja_baixada(conn, dataset_id, id_viagem_norm)
        if ja_v:
            if info_v and isinstance(info_v, dict):
                info_v = dict(info_v)
                info_v['id_informado'] = id_norm
            return True, info_v
    return False, None


def _ravex_mensagem_ja_baixado(id_input, info):
    """Texto de aviso quando ID de roteiro/viagem já foi importado."""
    id_inf = str(id_input or '').strip()
    info = info or {}
    id_v = str(info.get('id_viagem') or '').strip()
    id_r = str(info.get('id_roteiro') or '').strip()
    imp_em = info.get('importado_em')
    imp_txt = _formatar_criado_em_baixados_ravex(imp_em) if imp_em else ''
    if id_inf == id_v or not id_r:
        msg = 'Este ID de viagem (%s) já foi baixado do Ravex.' % (id_v or id_inf)
    elif id_inf == id_r:
        msg = 'Este ID de roteiro (%s) já foi baixado do Ravex.' % id_r
    else:
        msg = 'Este ID (%s) já foi baixado do Ravex.' % id_inf
    if id_v and id_r and id_inf not in (id_v, id_r):
        msg += ' (viagem %s, roteiro %s).' % (id_v, id_r)
    elif id_v and id_inf != id_v:
        msg += ' Viagem associada: %s.' % id_v
    if imp_txt:
        msg += ' Última importação: %s.' % imp_txt
    n_it = int(info.get('itens_romaneio') or 0)
    if n_it:
        msg += ' Itens no romaneio: %s.' % n_it
    msg += ' Para baixar de novo, marque «Forçar reimportação» abaixo.'
    return msg


def _ravex_mensagem_nao_baixado(id_input):
    """Texto quando o ID ainda não consta nos baixados/importações Ravex."""
    id_inf = str(id_input or '').strip()
    if not id_inf:
        return 'Roteiro ou viagem não informado.'
    return (
        'Este ID (%s) não foi baixado do Ravex. '
        'Importe o romaneio na aba «Importar Ravex» (ou confira em «Baixados Ravex») antes de conferir.'
    ) % id_inf


def _ravex_viagem_ja_baixada(conn, dataset_id, id_viagem):
    """Retorna (True, info) se a viagem já consta no romaneio ou em importação Ravex OK anterior."""
    id_norm = str(_normalizar_id_viagem(id_viagem) or id_viagem or '').strip()
    if not id_norm or not dataset_id:
        return False, None
    info = None
    ds = str(dataset_id)
    try:
        row = conn.execute(
            """SELECT COUNT(*)::int AS n, MAX(importado_em) AS ultimo
               FROM romaneio_por_item
               WHERE dataset_id = ? AND id_viagem::text = ?""",
            (ds, id_norm),
        ).fetchone()
    except Exception:
        row = conn.execute(
            """SELECT COUNT(*) AS n, MAX(importado_em) AS ultimo
               FROM romaneio_por_item
               WHERE dataset_id = ? AND id_viagem::text = ?""",
            (ds, id_norm),
        ).fetchone()
    n = 0
    ultimo = None
    if row:
        n = int((row.get('n') if hasattr(row, 'get') else row[0]) or 0)
        ultimo = row.get('ultimo') if hasattr(row, 'get') else (row[1] if len(row) > 1 else None)
    if n > 0:
        info = {'itens_romaneio': n, 'importado_em': ultimo, 'fonte': 'romaneio'}
        return True, info
    try:
        hist = conn.execute(
            """SELECT criado_em, total_itens
               FROM ravex_importacoes
               WHERE dataset_id = ? AND status IN ('OK', 'OK_COM_ERROS')
                 AND TRIM(COALESCE(parametros->>'id_viagem', '')) = ?
               ORDER BY criado_em DESC
               LIMIT 1""",
            (str(dataset_id), id_norm),
        ).fetchone()
    except Exception:
        hist = None
    if hist:
        criado = hist.get('criado_em') if hasattr(hist, 'get') else (hist[0] if hist else None)
        total_h = int((hist.get('total_itens') if hasattr(hist, 'get') else (hist[1] if len(hist) > 1 else 0)) or 0)
        info = {
            'itens_romaneio': 0,
            'importado_em': criado,
            'total_itens_historico': total_h,
            'fonte': 'historico',
        }
        return True, info
    return False, None


def _ravex_resposta_duplicado(conn, dataset_id, tipo, id_viagem, id_roteiro, info, id_input=None):
    """Registra tentativa duplicada e devolve JSON 409."""
    parametros = {'id_viagem': str(id_viagem), 'id_roteiro': str(id_roteiro or '')}
    if id_input:
        parametros['id_informado'] = str(id_input)
    if info:
        if info.get('importado_em'):
            parametros['importado_em'] = str(info.get('importado_em'))
        if info.get('itens_romaneio'):
            parametros['itens_romaneio'] = info.get('itens_romaneio')
    try:
        _registrar_importacao_ravex(
            conn,
            dataset_id=dataset_id,
            tipo=tipo,
            status='DUPLICADO',
            parametros=parametros,
            viagens_processadas=0,
            total_itens=0,
            erros=[{'mensagem': 'Viagem já importada anteriormente'}],
        )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    try:
        conn.close()
    except Exception:
        pass
    msg = _ravex_mensagem_ja_baixado(
        id_input or id_viagem,
        dict(info or {}, id_viagem=id_viagem, id_roteiro=id_roteiro),
    )
    return {
        'ok': False,
        'duplicado': True,
        'erro': msg,
        'id_viagem': id_viagem,
        'id_roteiro': id_roteiro,
        'detalhe': info or {},
    }, 409


_ravex_import_jobs = {}
_ravex_import_jobs_lock = threading.Lock()
_import_request_ctx = threading.local()


def _usuario_atual():
    """Usuário logado; em thread de importação usa valor copiado da requisição HTTP."""
    u = getattr(_import_request_ctx, 'usuario', None)
    if u is not None:
        return str(u)
    try:
        return session.get('usuario', '') or ''
    except RuntimeError:
        return ''


def _ravex_import_job_update(job_id, **fields):
    with _ravex_import_jobs_lock:
        job = dict(_ravex_import_jobs.get(job_id) or {})
        job.update(fields)
        job['updated_at'] = time.time()
        _ravex_import_jobs[job_id] = job


def _ravex_import_job_get(job_id):
    with _ravex_import_jobs_lock:
        return dict(_ravex_import_jobs.get(job_id) or {})


def _ravex_import_job_cleanup():
    """Remove jobs antigos da memória."""
    limite = time.time() - 7200
    with _ravex_import_jobs_lock:
        velhos = [k for k, v in _ravex_import_jobs.items() if (v.get('updated_at') or 0) < limite]
        for k in velhos:
            _ravex_import_jobs.pop(k, None)


def _registrar_importacao_ravex(conn, *, dataset_id, tipo, status, parametros=None, viagens_processadas=0, total_itens=0, erros=None):
    """Registra no banco a execução de importações Ravex (para aba 'Baixados'). Usa horário atual UTC para criado_em."""
    try:
        usuario = _usuario_atual() or None
        agora_utc = datetime.now(timezone.utc)
        conn.execute(
            """INSERT INTO public.ravex_importacoes (dataset_id, tipo, status, parametros, viagens_processadas, total_itens, usuario, erros, criado_em)
               VALUES (?, ?, ?, ?::jsonb, ?, ?, ?, ?::jsonb, ?)""",
            (
                str(dataset_id) if dataset_id else None,
                str(tipo or ''),
                str(status or 'OK'),
                json.dumps(parametros or {}),
                int(viagens_processadas or 0),
                int(total_itens or 0),
                usuario,
                json.dumps(erros or []),
                agora_utc,
            ),
        )
    except Exception as ex:
        try:
            app.logger.warning('Falha ao registrar ravex_importacoes: %s', ex)
        except Exception:
            pass


def _formatar_criado_em_baixados_ravex(criado):
    if criado is None:
        return ''
    if hasattr(criado, 'astimezone'):
        try:
            if getattr(criado, 'tzinfo', None) is None:
                criado = criado.replace(tzinfo=timezone.utc)
            if ZoneInfo:
                return criado.astimezone(ZoneInfo('America/Sao_Paulo')).strftime('%d/%m/%Y %H:%M:%S')
            return criado.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            return criado.strftime('%d/%m/%Y %H:%M:%S') if hasattr(criado, 'strftime') else str(criado)
    s = str(criado).strip()
    if len(s) >= 19 and s[4] == '-' and 'T' in s:
        try:
            dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
            if ZoneInfo:
                return dt.astimezone(ZoneInfo('America/Sao_Paulo')).strftime('%d/%m/%Y %H:%M:%S')
            return dt.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            pass
    return s


def _ravex_tz_baixados():
    if ZoneInfo:
        return ZoneInfo('America/Sao_Paulo')
    return timezone(timedelta(hours=-3))


def _ravex_filtro_periodo_datas(periodo, data_inicio, data_fim):
    """Retorna (inicio, fim_exclusivo) com timezone ou (None, None) se sem filtro de data."""
    tz = _ravex_tz_baixados()
    periodo = (periodo or '').strip().lower()
    di = (data_inicio or '').strip()[:10]
    df = (data_fim or '').strip()[:10]
    if di and df:
        try:
            inicio = datetime.strptime(di, '%Y-%m-%d').replace(tzinfo=tz)
            fim = datetime.strptime(df, '%Y-%m-%d').replace(tzinfo=tz) + timedelta(days=1)
            return inicio, fim
        except ValueError:
            pass
    elif di:
        try:
            inicio = datetime.strptime(di, '%Y-%m-%d').replace(tzinfo=tz)
            return inicio, inicio + timedelta(days=1)
        except ValueError:
            pass
    if periodo == 'hoje':
        now = datetime.now(tz)
        inicio = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return inicio, inicio + timedelta(days=1)
    if periodo == 'ontem':
        now = datetime.now(tz)
        fim = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return fim - timedelta(days=1), fim
    return None, None


def _ravex_aplicar_hora_janela(inicio, fim_excl, hora_inicio=None, hora_fim=None):
    """Ajusta janela de datas com hora HH:MM (America/Sao_Paulo). fim_excl é exclusivo."""
    if not inicio or not fim_excl:
        return inicio, fim_excl
    if not (hora_inicio and str(hora_inicio).strip()) and not (hora_fim and str(hora_fim).strip()):
        return inicio, fim_excl
    hi, mi = 0, 0
    hf, mf = 23, 59
    if hora_inicio and str(hora_inicio).strip():
        m = re.match(r'^(\d{1,2}):(\d{2})', str(hora_inicio).strip())
        if m:
            hi, mi = int(m.group(1)), int(m.group(2))
    if hora_fim and str(hora_fim).strip():
        m = re.match(r'^(\d{1,2}):(\d{2})', str(hora_fim).strip())
        if m:
            hf, mf = int(m.group(1)), int(m.group(2))
    span = (fim_excl - inicio).total_seconds()
    if span > 86400:
        new_inicio = inicio.replace(hour=hi, minute=mi, second=0, microsecond=0)
        last_day = fim_excl - timedelta(days=1)
        new_fim = last_day.replace(hour=hf, minute=mf, second=59, microsecond=999999) + timedelta(microseconds=1)
        return new_inicio, new_fim
    new_inicio = inicio.replace(hour=hi, minute=mi, second=0, microsecond=0)
    if hora_fim and str(hora_fim).strip():
        dia_fim = inicio
        if (hf * 60 + mf) <= (hi * 60 + mi):
            dia_fim = inicio + timedelta(days=1)
        new_fim = dia_fim.replace(hour=hf, minute=mf, second=59, microsecond=999999) + timedelta(microseconds=1)
    else:
        new_fim = fim_excl
    return new_inicio, new_fim


def _ravex_listar_usuarios_distintos(conn):
    try:
        rows = conn.execute(
            """SELECT DISTINCT TRIM(usuario) AS u
               FROM public.ravex_importacoes
               WHERE TRIM(COALESCE(usuario, '')) != ''
               ORDER BY u"""
        ).fetchall()
    except Exception:
        return []
    out = []
    for r in rows or []:
        u = (r.get('u') if hasattr(r, 'get') else (r[0] if r else None))
        if u:
            out.append(str(u).strip())
    return out


def _ravex_slim_parametros_listagem(params):
    """Reduz payload JSON na listagem (viagens_importadas enormes)."""
    if not isinstance(params, dict):
        return params or {}
    vi = params.get('viagens_importadas')
    if isinstance(vi, list) and len(vi) > 50:
        slim = dict(params)
        slim['viagens_importadas'] = vi[:50]
        slim['_viagens_total'] = len(vi)
        return slim
    return params


def _ravex_importacao_dict_from_row(r):
    def _get(k, idx):
        return (r.get(k) if hasattr(r, 'get') else (r[idx] if hasattr(r, '__getitem__') and len(r) > idx else None))

    params = _get('parametros', 3)
    errs = _get('erros', 7)
    erros_qtd = _get('erros_qtd', None)
    try:
        if isinstance(params, str):
            params = json.loads(params)
    except Exception:
        pass
    try:
        if isinstance(errs, str):
            errs = json.loads(errs)
    except Exception:
        pass
    if erros_qtd is None:
        erros_qtd = len(errs) if isinstance(errs, list) else 0
    else:
        try:
            erros_qtd = int(erros_qtd or 0)
        except (TypeError, ValueError):
            erros_qtd = 0
    return {
        'id': _get('id', 0),
        'tipo': _get('tipo', 1) or '',
        'status': _get('status', 2) or '',
        'parametros': _ravex_slim_parametros_listagem(params),
        'viagens_processadas': _get('viagens_processadas', 4) or 0,
        'total_itens': _get('total_itens', 5) or 0,
        'usuario': _get('usuario', 6) or '',
        'erros': [],
        'erros_count': erros_qtd,
        'criado_em': _formatar_criado_em_baixados_ravex(_get('criado_em', 8)),
    }


def _listar_baixados_ravex_de_romaneio(conn, limit=200):
    """Fallback: agrupa romaneio_por_item por lote de importado_em quando ravex_importacoes está vazio."""
    try:
        rows = conn.execute(
            """SELECT importado_em, dataset_id,
                      COUNT(DISTINCT id_viagem) AS viagens,
                      COUNT(*) AS total_itens
               FROM romaneio_por_item
               WHERE importado_em IS NOT NULL
               GROUP BY importado_em, dataset_id
               ORDER BY importado_em DESC
               LIMIT ?""",
            (limit,),
        ).fetchall()
    except Exception:
        return []
    out = []
    for r in rows or []:
        def _g(k, i):
            return (r.get(k) if hasattr(r, 'get') else (r[i] if hasattr(r, '__getitem__') and len(r) > i else None))

        imp = _g('importado_em', 0)
        ds = _g('dataset_id', 1)
        out.append({
            'id': None,
            'tipo': 'romaneio',
            'status': 'OK',
            'parametros': {'origem': 'romaneio_por_item', 'dataset_id': str(ds) if ds else None},
            'viagens_processadas': int(_g('viagens', 2) or 0),
            'total_itens': int(_g('total_itens', 3) or 0),
            'usuario': '',
            'erros': [],
            'criado_em': _formatar_criado_em_baixados_ravex(imp),
        })
    return out


def _ravex_normalizar_id_viagem_lista(ids):
    out = []
    vistos = set()
    for raw in ids or []:
        id_n = str(_normalizar_id_viagem(raw) or raw or '').strip()
        if id_n and id_n not in vistos:
            vistos.add(id_n)
            out.append(id_n)
    return out


def _ravex_viagens_por_janela_importacao(conn, dataset_id, criado_em, limite=None):
    """Fallback: viagens cujo romaneio foi gravado na mesma execução da importação."""
    if not dataset_id or not criado_em:
        return []
    ds = str(dataset_id)
    try:
        rows = conn.execute(
            """SELECT DISTINCT TRIM(COALESCE(id_viagem::text, '')) AS id_v
               FROM romaneio_por_item
               WHERE dataset_id = ?
                 AND importado_em IS NOT NULL
                 AND importado_em >= (?::timestamptz - INTERVAL '3 hours')
                 AND importado_em <= (?::timestamptz + INTERVAL '3 minutes')
                 AND TRIM(COALESCE(id_viagem::text, '')) != ''
               ORDER BY id_v""",
            (ds, criado_em, criado_em),
        ).fetchall()
    except Exception:
        return []
    ids = []
    for r in rows or []:
        id_v = str((r.get('id_v') if hasattr(r, 'get') else (r[0] if r else '')) or '').strip()
        if id_v:
            ids.append(id_v)
    if limite and len(ids) > int(limite):
        ids = ids[: int(limite)]
    return ids


def _ravex_resolver_viagens_da_importacao(conn, import_row):
    """Descobre quais id_viagem serão apagados ao excluir um registro de ravex_importacoes."""
    def _get(k, idx):
        return (import_row.get(k) if hasattr(import_row, 'get') else (import_row[idx] if hasattr(import_row, '__getitem__') and len(import_row) > idx else None))

    params = _get('parametros', 3) or {}
    if isinstance(params, str):
        try:
            params = json.loads(params)
        except Exception:
            params = {}
    if not isinstance(params, dict):
        params = {}

    dataset_id = _get('dataset_id', 1)
    tipo = str(_get('tipo', 2) or '').strip().lower()
    criado_em = _get('criado_em', 8)
    viagens_proc = int(_get('viagens_processadas', 4) or 0)

    viagens = set()
    for key in ('viagens_importadas', 'id_viagens'):
        viagens.update(_ravex_normalizar_id_viagem_lista(params.get(key)))

    id_v = str(params.get('id_viagem') or '').strip()
    if id_v:
        viagens.add(str(_normalizar_id_viagem(id_v) or id_v).strip())

    if not viagens:
        id_r = str(params.get('id_roteiro') or params.get('id_informado') or '').strip()
        if id_r and dataset_id:
            try:
                row_ir = conn.execute(
                    """SELECT TRIM(COALESCE(id_viagem::text, '')) AS id_v
                       FROM id_roteiros
                       WHERE dataset_id = ?
                         AND (TRIM(COALESCE(id_roteiro::text, '')) = ? OR TRIM(COALESCE(id_viagem::text, '')) = ?)
                       LIMIT 1""",
                    (str(dataset_id), id_r, id_r),
                ).fetchone()
                if row_ir:
                    id_v2 = str((row_ir.get('id_v') if hasattr(row_ir, 'get') else row_ir[0]) or '').strip()
                    if id_v2:
                        viagens.add(id_v2)
            except Exception:
                pass
            if not viagens:
                try:
                    row_rm = conn.execute(
                        """SELECT TRIM(COALESCE(id_viagem::text, '')) AS id_v
                           FROM romaneio_por_item
                           WHERE dataset_id = ?
                             AND (TRIM(COALESCE(id_roteiro::text, '')) = ? OR TRIM(COALESCE(id_viagem::text, '')) = ?)
                           LIMIT 1""",
                        (str(dataset_id), id_r, id_r),
                    ).fetchone()
                    if row_rm:
                        id_v3 = str((row_rm.get('id_v') if hasattr(row_rm, 'get') else row_rm[0]) or '').strip()
                        if id_v3:
                            viagens.add(id_v3)
                except Exception:
                    pass

    if not viagens and tipo in ('periodo', 'lista', 'romaneio') and dataset_id and criado_em:
        limite = viagens_proc if viagens_proc > 0 else None
        viagens.update(_ravex_viagens_por_janela_importacao(conn, dataset_id, criado_em, limite))

    return sorted(v for v in viagens if v)


def _excluir_dados_viagem_ravex(conn, dataset_id, id_viagem, id_roteiros_extra=None):
    """Remove romaneio, bipagem, metadados e roteiro ligados a uma viagem importada."""
    id_norm = str(_normalizar_id_viagem(id_viagem) or id_viagem or '').strip()
    if not id_norm:
        return {}
    ds = str(dataset_id) if dataset_id else None
    removidos = {}

    def _exec_del(sql, params, chave):
        try:
            cur = conn.execute(sql, params)
            n = getattr(cur, 'rowcount', None)
            removidos[chave] = int(n) if n is not None and n >= 0 else 0
        except Exception:
            removidos[chave] = 0

    if ds:
        _exec_del(
            """DELETE FROM romaneio_por_item
               WHERE dataset_id = ? AND TRIM(COALESCE(id_viagem::text, '')) = ?""",
            (ds, id_norm),
            'romaneio_por_item',
        )
    else:
        _exec_del(
            """DELETE FROM romaneio_por_item
               WHERE TRIM(COALESCE(id_viagem::text, '')) = ?""",
            (id_norm,),
            'romaneio_por_item',
        )

    _exec_del(
        """DELETE FROM produtos_bipados
           WHERE TRIM(COALESCE(id_viagem::text, '')) = ?""",
        (id_norm,),
        'produtos_bipados',
    )
    _exec_del('DELETE FROM viagem_placa WHERE id_viagem = ?', (id_norm,), 'viagem_placa')
    _exec_del('DELETE FROM viagem_motorista WHERE id_viagem = ?', (id_norm,), 'viagem_motorista')
    _exec_del('DELETE FROM viagem_responsaveis WHERE id_viagem = ?', (id_norm,), 'viagem_responsaveis')
    _exec_del('DELETE FROM divergencia_motivo WHERE id_viagem = ?', (id_norm,), 'divergencia_motivo')
    try:
        _ensure_viagem_periodo_bipagem_table(conn)
        tbl_per = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
        _exec_del(f'DELETE FROM {tbl_per} WHERE id_viagem = ?', (id_norm,), 'viagem_periodo_bipagem')
    except Exception:
        removidos['viagem_periodo_bipagem'] = 0

    extras = set(str(x).strip() for x in (id_roteiros_extra or []) if str(x).strip())
    if ds:
        if extras:
            ph = ','.join(['?'] * len(extras))
            _exec_del(
                f"""DELETE FROM id_roteiros
                    WHERE dataset_id = ?
                      AND (TRIM(COALESCE(id_viagem::text, '')) = ?
                           OR TRIM(COALESCE(id_roteiro::text, '')) IN ({ph}))""",
                tuple([ds, id_norm] + sorted(extras)),
                'id_roteiros',
            )
        else:
            _exec_del(
                """DELETE FROM id_roteiros
                   WHERE dataset_id = ? AND TRIM(COALESCE(id_viagem::text, '')) = ?""",
                (ds, id_norm),
                'id_roteiros',
            )

    return removidos


def _excluir_importacao_ravex_por_id(conn, import_id):
    row = conn.execute(
        """SELECT id, dataset_id, tipo, status, parametros, viagens_processadas, total_itens, usuario, erros, criado_em
           FROM public.ravex_importacoes
           WHERE id = ?""",
        (int(import_id),),
    ).fetchone()
    if not row:
        return None, {'erro': 'Importação não encontrada.'}

    params = row.get('parametros') if hasattr(row, 'get') else None
    if isinstance(params, str):
        try:
            params = json.loads(params)
        except Exception:
            params = {}
    if not isinstance(params, dict):
        params = {}

    dataset_id = row.get('dataset_id') if hasattr(row, 'get') else None
    viagens = _ravex_resolver_viagens_da_importacao(conn, row)
    id_roteiro_extra = str(params.get('id_roteiro') or params.get('id_informado') or '').strip() or None

    detalhes_viagens = []
    totais = {}
    for id_v in viagens:
        extras = [id_roteiro_extra] if id_roteiro_extra else []
        rem = _excluir_dados_viagem_ravex(conn, dataset_id, id_v, extras)
        detalhes_viagens.append({'id_viagem': id_v, 'removidos': rem})
        for k, v in rem.items():
            totais[k] = totais.get(k, 0) + int(v or 0)

    conn.execute('DELETE FROM public.ravex_importacoes WHERE id = ?', (int(import_id),))
    totais['ravex_importacoes'] = 1
    return {
        'ok': True,
        'import_id': int(import_id),
        'viagens_excluidas': viagens,
        'total_viagens': len(viagens),
        'removidos': totais,
        'detalhes': detalhes_viagens,
    }, None


def _normalizar_data_para_ravex(s):
    """Converte data para YYYY-MM-DD. Aceita YYYY-MM-DD, DD/MM/YYYY ou DD-MM-YYYY."""
    s = (s or '').strip()[:10]
    if not s:
        return ''
    if len(s) == 10 and s[4] == '-' and s[7] == '-':
        return s
    import re
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', s)
    if m:
        d, mes, a = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return '%s-%s-%s' % (a, mes, d)
    return s


def _ravex_roteiros_por_periodo_para_romaneio(token, data_inicio, data_fim):
    """
    Chama /api/roteiro/obter-roteiro-por-periodo e retorna lista de dicts com id_viagem, id_roteiro, identificador_rota
    (extraídos de viagemFaturada.id e identificadorRota de cada roteiro). Usado para preencher a tabela Romaneio por item.
    """
    if not obter_roteiro_por_periodo:
        return []
    p_ini = _normalizar_data_para_ravex(data_inicio)
    p_fim = _normalizar_data_para_ravex(data_fim)
    if not p_ini or not p_fim:
        return []
    try:
        roteiros = obter_roteiro_por_periodo(token, p_ini, p_fim)
    except Exception:
        return []
    resultado = []
    for rot in (roteiros or []):
        if not isinstance(rot, dict):
            continue
        rot_obj = rot.get('data') if rot.get('viagemFaturada') is None and rot.get('data') else rot
        vf = rot_obj.get('viagemFaturada') or rot_obj.get('viagem_faturada')
        id_viagem = None
        if isinstance(vf, dict):
            id_viagem = vf.get('id') or vf.get('Id')
        elif isinstance(vf, (int, float)):
            id_viagem = vf
        if id_viagem is None:
            id_viagem = rot_obj.get('viagemFaturadaId') or rot_obj.get('viagemFaturada_id') or rot_obj.get('viagem_faturada_id')
        if id_viagem is None:
            continue
        id_roteiro = str(rot_obj.get('id') or rot_obj.get('Id') or rot.get('id') or rot.get('Id') or '').strip()
        identificador_rota = (rot_obj.get('identificadorRota') or rot_obj.get('identificador_rota') or rot.get('identificadorRota') or rot.get('identificador_rota') or rot_obj.get('identificador') or rot_obj.get('nome') or '').strip()
        data_viagem = None
        if isinstance(vf, dict):
            criado = vf.get('criadoDatahora') or vf.get('criado_datahora') or vf.get('dataCriacao')
            if criado:
                data_viagem = criado
        resultado.append({
            'id_viagem': str(id_viagem).strip(),
            'id_roteiro': id_roteiro,
            'identificador_rota': identificador_rota,
            'data_viagem': data_viagem,
        })
    return resultado


def _ravex_extrair_identificador_rota(obj):
    if not isinstance(obj, dict):
        return ''
    return (
        obj.get('identificadorRota') or obj.get('identificador_rota')
        or obj.get('identificador') or obj.get('nome') or obj.get('descricao') or ''
    ).strip()


def _ravex_extrair_viagem_id_de_roteiro(roteiro):
    """Extrai id da viagem faturada do objeto roteiro da API Ravex."""
    if not isinstance(roteiro, dict):
        return ''
    vf = roteiro.get('viagemFaturada') or roteiro.get('viagem_faturada')
    if isinstance(vf, dict):
        return str(vf.get('id') or vf.get('Id') or '').strip()
    if vf not in (None, '', 0):
        return str(vf).strip()
    return str(roteiro.get('viagemFaturadaId') or roteiro.get('viagem_id') or '').strip()


def _ravex_resolver_id_db(conn, dataset_id, id_input):
    """Resolve id_viagem/id_roteiro pelo cadastro local id_roteiros (sem API)."""
    id_norm = str(id_input or '').strip()
    if not id_norm or not dataset_id:
        return None
    try:
        row = conn.execute(
            """SELECT id_viagem::text AS id_viagem, id_roteiro::text AS id_roteiro, identificador_rota
               FROM id_roteiros
               WHERE dataset_id = ? AND (id_viagem::text = ? OR id_roteiro::text = ?)
               LIMIT 1""",
            (str(dataset_id), id_norm, id_norm),
        ).fetchone()
    except Exception:
        return None
    if not row:
        return None
    id_viagem = str((row.get('id_viagem') if hasattr(row, 'get') else (row[0] if len(row) > 0 else '')) or '').strip()
    id_roteiro = str((row.get('id_roteiro') if hasattr(row, 'get') else (row[1] if len(row) > 1 else '')) or '').strip()
    identificador_rota = str((row.get('identificador_rota') if hasattr(row, 'get') else (row[2] if len(row) > 2 else '')) or '').strip()
    if not id_viagem and not id_roteiro:
        return None
    return {
        'id_viagem': id_viagem,
        'id_roteiro': id_roteiro or id_norm,
        'identificador_rota': identificador_rota,
        'viagem_full': None,
    }


def _ravex_resolver_id_unico_ravex(token, id_unico):
    """Resolve viagem/roteiro na API Ravex (consulta viagem e roteiro em paralelo)."""
    if not obter_viagem_por_id or not obter_roteiro_por_id:
        return None
    id_str = str(id_unico or '').strip()
    try:
        rid = int(id_str)
    except (TypeError, ValueError):
        return None
    viagem = None
    roteiro = None
    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_v = pool.submit(obter_viagem_por_id, token, rid)
        fut_r = pool.submit(obter_roteiro_por_id, token, rid)
        viagem = fut_v.result()
        roteiro = fut_r.result()
    if viagem:
        id_viagem = id_str
        id_roteiro = ''
        identificador_rota = ''
        roteiro_obj = viagem.get('roteiro') or viagem.get('roteiroFaturado')
        if isinstance(roteiro_obj, dict):
            id_roteiro = str(roteiro_obj.get('id') or roteiro_obj.get('Id') or '')
            identificador_rota = _ravex_extrair_identificador_rota(roteiro_obj)
        if not id_roteiro:
            id_roteiro = str(viagem.get('roteiroId') or viagem.get('idRoteiro') or '')
        if not id_roteiro and isinstance(roteiro, dict):
            id_roteiro = id_str
            identificador_rota = _ravex_extrair_identificador_rota(roteiro) or identificador_rota
        return {
            'id_viagem': id_viagem,
            'id_roteiro': id_roteiro or id_str,
            'identificador_rota': identificador_rota,
            'viagem_full': viagem,
        }
    if isinstance(roteiro, dict):
        id_viagem = _ravex_extrair_viagem_id_de_roteiro(roteiro)
        if id_viagem:
            return {
                'id_viagem': id_viagem,
                'id_roteiro': id_str,
                'identificador_rota': _ravex_extrair_identificador_rota(roteiro),
                'viagem_full': None,
            }
        return {
            'id_viagem': '',
            'id_roteiro': id_str,
            'identificador_rota': _ravex_extrair_identificador_rota(roteiro),
            'viagem_full': None,
            'roteiro_full': roteiro,
            'somente_roteiro': True,
        }
    return {'erro': 'ID não encontrado como viagem nem como roteiro na API Ravex.'}


def _ravex_resolver_importacao_unico(conn, token, dataset_id, id_unico):
    """Resolve IDs para importação única: banco local primeiro, depois API em paralelo."""
    local = _ravex_resolver_id_db(conn, dataset_id, id_unico)
    if local:
        if local.get('id_viagem'):
            return local
        if local.get('id_roteiro'):
            local['somente_roteiro'] = True
            return local
    api = _ravex_resolver_id_unico_ravex(token, id_unico)
    if not api:
        return {'erro': 'Não foi possível resolver o ID na API Ravex.'}
    return api


def _ravex_itens_produto_do_pedido(token, pedido_resumo, pedido_id, cache_pedido):
    """Monta linhas de produto a partir de compartimentos ou GET /api/pedido/{id}/itens."""
    pid = pedido_id
    if pid is None and isinstance(pedido_resumo, dict):
        pid = pedido_resumo.get('id') or pedido_resumo.get('Id')
    if pid is None:
        return [], None
    try:
        pid = int(pid)
    except (TypeError, ValueError):
        return [], None
    pk = str(pid)
    if pk in cache_pedido:
        return cache_pedido[pk]
    ref = None
    endereco = cidade = codigo_cliente = ''
    compartimentos = []
    produtos = []
    if isinstance(pedido_resumo, dict):
        compartimentos = pedido_resumo.get('compartimentos') or []
        ref = pedido_resumo.get('pontoReferencia') or pedido_resumo.get('ponto_referencia')
    if ref and isinstance(ref, dict):
        endereco = (ref.get('endereco') or ref.get('endereço') or ref.get('logradouro') or '').strip()
        cidade = _normalizar_cidade_nome(ref.get('cidade') or '')
        codigo_cliente = (ref.get('cnpj') or ref.get('Cnpj') or '').strip()
    if compartimentos:
        pass
    elif obter_itens_pedido:
        for item in obter_itens_pedido(token, pid) or []:
            if not isinstance(item, dict):
                continue
            prod = item.get('produto') or {}
            codigo_produto = str(item.get('codigo') or prod.get('codigo') or prod.get('Codigo') or '').strip()
            descricao = str(item.get('descricao') or item.get('descricaoItem') or prod.get('descricao') or prod.get('descrição') or '').strip()
            quantidade = int(item.get('quantidade', 0) or 0)
            if not codigo_produto and not descricao and quantidade <= 0:
                continue
            produtos.append({
                'codigo_produto': codigo_produto,
                'descricao': descricao,
                'quantidade': quantidade,
                'unidade': str(item.get('unidade') or prod.get('unidade') or '').strip(),
                'peso_bruto': str(item.get('pesoBruto') or item.get('pesoLiquido') or prod.get('pesoBruto') or '').strip(),
                'codigo_cliente': codigo_cliente,
                'endereco': endereco,
                'cidade': cidade,
            })
        if produtos:
            cache_pedido[pk] = (produtos, ref)
            return produtos, ref
    if not compartimentos and obter_pedido_por_id:
        pedido_full = obter_pedido_por_id(token, pid)
        if isinstance(pedido_full, dict):
            compartimentos = pedido_full.get('compartimentos') or []
            if not ref:
                ref = pedido_full.get('pontoReferencia') or pedido_full.get('ponto_referencia')
            if ref and isinstance(ref, dict):
                endereco = (ref.get('endereco') or ref.get('endereço') or ref.get('logradouro') or '').strip()
                cidade = _normalizar_cidade_nome(ref.get('cidade') or '')
                codigo_cliente = (ref.get('cnpj') or ref.get('Cnpj') or '').strip()
    for comp in compartimentos or []:
        if not isinstance(comp, dict):
            continue
        prod = comp.get('produto') or {}
        codigo_produto = str(prod.get('codigo') or prod.get('Codigo') or '').strip()
        descricao = str(prod.get('descricao') or prod.get('descrição') or '').strip()
        quantidade = int(comp.get('quantidade', 0) or 0)
        if not codigo_produto and not descricao and quantidade <= 0:
            continue
        produtos.append({
            'codigo_produto': codigo_produto,
            'descricao': descricao,
            'quantidade': quantidade,
            'unidade': '',
            'peso_bruto': '',
            'codigo_cliente': codigo_cliente,
            'endereco': endereco,
            'cidade': cidade,
        })
    cache_pedido[pk] = (produtos, ref)
    return produtos, ref


def _ravex_linhas_romaneio_roteiro(token, id_roteiro, roteiro_full=None, identificador_rota_pre=None):
    """Monta romaneio a partir do roteiro (itens/pedidos) quando ainda não há viagem faturada."""
    if not obter_roteiro_por_id or not obter_itens_roteiro:
        return (None, [])
    id_roteiro = str(id_roteiro or '').strip()
    if not id_roteiro:
        return (None, [])
    try:
        rid = int(id_roteiro)
    except (TypeError, ValueError):
        return (None, [])
    if not isinstance(roteiro_full, dict):
        roteiro_full = obter_roteiro_por_id(token, rid)
    if not isinstance(roteiro_full, dict):
        return (None, [])
    identificador_rota = (identificador_rota_pre or _ravex_extrair_identificador_rota(roteiro_full) or '').strip()
    if not _ravex_identificador_rota_valido(identificador_rota):
        identificador_rota = _ravex_extrair_identificador_rota(roteiro_full)
    placa = ''
    veiculo = roteiro_full.get('veiculo') or roteiro_full.get('veículo')
    if isinstance(veiculo, dict):
        placa = (veiculo.get('placa') or veiculo.get('Placa') or '').strip()
    motorista = ''
    mot = roteiro_full.get('motorista')
    if isinstance(mot, dict):
        motorista = (mot.get('nome') or mot.get('Nome') or '').strip()
    data_ini = (
        roteiro_full.get('dataInicioPrevisto') or roteiro_full.get('separacaoDatahora')
        or roteiro_full.get('dataFimPrevisto') or ''
    )
    data_expedicao = data_ini[:10] if isinstance(data_ini, str) and len(data_ini) >= 10 else str(data_ini or '')

    itens_roteiro = obter_itens_roteiro(token, rid) or []
    if not itens_roteiro:
        return (id_roteiro, [])

    cache_pedido = {}
    pedidos_fetch = []
    pedidos_vistos = set()
    for it in itens_roteiro:
        if not isinstance(it, dict):
            continue
        pedido = it.get('pedido')
        pid = (pedido.get('id') or pedido.get('Id')) if isinstance(pedido, dict) else None
        if pid is not None:
            pk = str(pid)
            if pk not in pedidos_vistos:
                pedidos_vistos.add(pk)
                pedidos_fetch.append((pid, pedido))

    def _fetch_pedido(pair):
        pid, pedido_resumo = pair
        try:
            prods, _ref = _ravex_itens_produto_do_pedido(token, pedido_resumo, pid, cache_pedido)
            return pid, prods
        except Exception:
            return pid, []

    produtos_por_pedido = {}
    if pedidos_fetch:
        workers = min(_RAVEX_IMPORT_WORKERS * 2, max(2, len(pedidos_fetch)), 16)
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for pid, prods in pool.map(_fetch_pedido, pedidos_fetch):
                if prods:
                    produtos_por_pedido[str(pid)] = prods

    linhas = []
    row_index = 0
    for it in itens_roteiro:
        if not isinstance(it, dict):
            continue
        pedido = it.get('pedido')
        pid = (pedido.get('id') or pedido.get('Id')) if isinstance(pedido, dict) else None
        prods = produtos_por_pedido.get(str(pid), []) if pid is not None else []
        for item in prods:
            row_index += 1
            codigo_produto = item.get('codigo_produto') or ''
            descricao = item.get('descricao') or ''
            quantidade = int(item.get('quantidade') or 0)
            data_row = {
                'id_roteiro': id_roteiro,
                'id_viagem': '',
                'codigo_produto': codigo_produto,
                'descricao': descricao,
                'quantidade': quantidade,
                'unidade': item.get('unidade') or '',
                'peso_bruto': item.get('peso_bruto') or '',
                'codigo_cliente': item.get('codigo_cliente') or '',
                'endereco': item.get('endereco') or '',
                'cidade': item.get('cidade') or '',
                'placa': placa,
                'motorista': motorista,
                'data_expedicao': data_expedicao,
            }
            linhas.append({
                'row_index': row_index,
                'id_roteiro': id_roteiro,
                'id_viagem': '',
                'identificador_rota': identificador_rota or None,
                'codigo_produto': codigo_produto or None,
                'descricao': descricao or None,
                'quantidade': quantidade,
                'unidade': item.get('unidade') or None,
                'peso_bruto': item.get('peso_bruto') or None,
                'codigo_cliente': item.get('codigo_cliente') or None,
                'endereco': item.get('endereco') or None,
                'cidade': item.get('cidade') or None,
                'placa': placa or None,
                'motorista': motorista or None,
                'data_expedicao': data_expedicao or None,
                'data': data_row,
            })
    return (id_roteiro, linhas)


def _ravex_linhas_romaneio_viagem(token, id_viagem, id_roteiro_pre=None, identificador_rota_pre=None, viagem_full_pre=None):
    """Dado token e id_viagem, busca na API Ravex e monta (id_roteiro, linhas) para romaneio_por_item. Retorna (None, []) em erro.
    Se id_roteiro_pre e identificador_rota_pre forem passados (ex.: vindos de obter-roteiro-por-periodo), usa-os e não chama roteiro por período."""
    if not obter_viagem_por_id or not obter_canhotos_viagem or not obter_notas_fiscais_viagem or not obter_itens_nota_fiscal or not obter_ponto_referencia:
        return (None, [])
    if viagem_full_pre and isinstance(viagem_full_pre, dict):
        viagem_full = viagem_full_pre
    else:
        viagem_full = obter_viagem_por_id(token, id_viagem)
    if not viagem_full:
        return (None, [])
    # Usar id_roteiro e identificador_rota já obtidos por obter-roteiro-por-periodo quando fornecidos
    id_roteiro = (id_roteiro_pre if id_roteiro_pre is not None else '') or ''
    identificador_rota = (identificador_rota_pre if identificador_rota_pre is not None else '') or ''
    if obter_roteiro_por_periodo and id_roteiro_pre is None:
        data_ini = (
            viagem_full.get('inicioDataHora') or viagem_full.get('dataInicioViagem') or viagem_full.get('dataHoraInicio')
            or viagem_full.get('dataInicio') or viagem_full.get('dataFim') or viagem_full.get('data') or ''
        )
        dt = None
        if isinstance(data_ini, str) and len(data_ini) >= 10:
            dt_str = data_ini.replace('Z', '').strip()
            try:
                if len(dt_str) >= 19:
                    dt = datetime.strptime(dt_str[:19], '%Y-%m-%dT%H:%M:%S')
                else:
                    dt = datetime.strptime(dt_str[:10], '%Y-%m-%d')
            except (ValueError, TypeError):
                try:
                    dt = datetime.strptime(dt_str[:10], '%Y-%m-%d')
                except (ValueError, TypeError):
                    pass
        if dt is None:
            dt = datetime.utcnow()
        try:
            # Janela ampla (3 dias cada lado); API aceita YYYY-MM-DD (ex: 2026-03-11, 2026-03-13)
            inicio = (dt - timedelta(days=3)).strftime('%Y-%m-%d')
            fim = (dt + timedelta(days=3)).strftime('%Y-%m-%d')
            roteiros_periodo = obter_roteiro_por_periodo(token, inicio, fim)
            if not roteiros_periodo:
                inicio_iso = (dt - timedelta(days=3)).strftime('%Y-%m-%dT00:00:00.000Z')
                fim_iso = (dt + timedelta(days=3)).strftime('%Y-%m-%dT23:59:59.000Z')
                roteiros_periodo = obter_roteiro_por_periodo(token, inicio_iso, fim_iso)
            id_viagem_int = None
            try:
                id_viagem_int = int(id_viagem)
            except (TypeError, ValueError):
                pass
            id_viagem_str = str(id_viagem).strip()
            for rot in (roteiros_periodo or []):
                if not isinstance(rot, dict):
                    continue
                rot_obj = rot.get('data') if rot.get('viagemFaturada') is None and rot.get('data') else rot
                vf = rot_obj.get('viagemFaturada') or rot_obj.get('viagem_faturada')
                vf_id = vf.get('id') or vf.get('Id') if isinstance(vf, dict) else None
                if vf_id is None and isinstance(vf, (int, float)):
                    vf_id = vf
                if vf_id is None:
                    vf_id = rot_obj.get('viagemFaturadaId') or rot_obj.get('viagemFaturada_id') or rot_obj.get('viagem_faturada_id')
                if vf_id is not None and (vf_id == id_viagem_int or str(vf_id) == id_viagem_str):
                    id_roteiro = str(rot_obj.get('id') or rot_obj.get('Id') or rot.get('id') or rot.get('Id') or '')
                    identificador_rota = (rot_obj.get('identificadorRota') or rot_obj.get('identificador_rota') or rot.get('identificadorRota') or rot.get('identificador_rota') or rot_obj.get('identificador') or rot_obj.get('nome') or '').strip()
                    break
        except Exception:
            pass
    roteiro_obj = viagem_full.get('roteiro') or viagem_full.get('roteiroFaturado')
    if not id_roteiro and isinstance(roteiro_obj, dict):
        id_roteiro = str(roteiro_obj.get('id') or roteiro_obj.get('Id') or roteiro_obj.get('roteiroId') or roteiro_obj.get('idRoteiro') or '')
    if not id_roteiro:
        id_roteiro = str(viagem_full.get('roteiroId') or viagem_full.get('idRoteiro') or viagem_full.get('roteiro_id') or viagem_full.get('id_roteiro') or '')
    # Via pedido da primeira NF quando a viagem não traz roteiro (igual BASE VIAGENS sync_to_excel)
    nfs_pre = None
    if not id_roteiro and obter_notas_fiscais_viagem and obter_pedido_por_id:
        nfs_pre = obter_notas_fiscais_viagem(token, id_viagem)
        for nf in (nfs_pre or [])[:5]:
            pedido = nf.get('pedido') if isinstance(nf, dict) else None
            id_pedido = (pedido.get('id') or pedido.get('Id')) if isinstance(pedido, dict) else None
            if not id_pedido:
                continue
            pedido_full = obter_pedido_por_id(token, id_pedido)
            if not isinstance(pedido_full, dict):
                continue
            id_roteiro = str(pedido_full.get('roteiroId') or pedido_full.get('idRoteiro') or pedido_full.get('roteiro_id') or '')
            if not id_roteiro:
                rot = pedido_full.get('roteiro') or pedido_full.get('roteiroFaturado')
                if isinstance(rot, dict):
                    id_roteiro = str(rot.get('id') or rot.get('Id') or '')
            if id_roteiro:
                break
    if not identificador_rota and isinstance(roteiro_obj, dict):
        identificador_rota = (roteiro_obj.get('identificadorRota') or roteiro_obj.get('identificador_rota') or roteiro_obj.get('identificador') or roteiro_obj.get('nome') or roteiro_obj.get('descricao') or '').strip()
    if not identificador_rota:
        identificador_rota = (viagem_full.get('identificadorRota') or viagem_full.get('identificador') or viagem_full.get('nome') or '').strip()
    # Não usar identificador da viagem quando for código técnico (ex: 260306090310SWB3F77); legível é tipo "38 - GRU - BARRA FUNDA"
    if identificador_rota and len(identificador_rota) >= 12 and ' - ' not in identificador_rota:
        identificador_rota = ''
    # GET /api/roteiro/{id} é a fonte oficial do identificador (ex: .../api/roteiro/660002); usar sempre que tivermos id_roteiro
    if id_roteiro and obter_roteiro_por_id and not _ravex_identificador_rota_valido(identificador_rota):
        try:
            rid = int(id_roteiro)
            roteiro_full = obter_roteiro_por_id(token, rid)
            if isinstance(roteiro_full, dict):
                ident = (roteiro_full.get('identificadorRota') or roteiro_full.get('identificador_rota') or roteiro_full.get('identificador') or roteiro_full.get('nome') or '').strip()
                if ident:
                    identificador_rota = ident
        except (TypeError, ValueError):
            pass
    # Identificador legível tem formato tipo "38 - GRU - BARRA FUNDA"; código longo alfanumérico sem espaços/traços é inválido
    identificador_invalido = (
        identificador_rota and len(identificador_rota) >= 15 and ' - ' not in identificador_rota
    )
    # Quando a viagem não vem com roteiro/identificadorRota: GET obter-roteiro-por-periodo e achar roteiro onde viagemFaturada.id == id_viagem (omitir se já veio por id_roteiro_pre)
    if (not id_roteiro or not identificador_rota or identificador_invalido) and obter_roteiro_por_periodo and id_roteiro_pre is None:
        if identificador_invalido:
            identificador_rota = ''
        data_ini = (
            viagem_full.get('inicioDataHora') or viagem_full.get('dataInicioViagem') or viagem_full.get('dataHoraInicio')
            or viagem_full.get('dataInicio') or viagem_full.get('dataFim') or viagem_full.get('data') or ''
        )
        dt = None
        if isinstance(data_ini, str) and len(data_ini) >= 10:
            dt_str = data_ini.replace('Z', '').strip()
            try:
                if len(dt_str) >= 26 and '.' in dt_str:
                    dt = datetime.strptime(dt_str[:26].ljust(26, '0')[:26], '%Y-%m-%dT%H:%M:%S.%f')
                elif len(dt_str) >= 19:
                    dt = datetime.strptime(dt_str[:19], '%Y-%m-%dT%H:%M:%S')
                else:
                    dt = datetime.strptime(dt_str[:10], '%Y-%m-%d')
            except (ValueError, TypeError):
                try:
                    dt = datetime.strptime(dt_str[:10], '%Y-%m-%d')
                except (ValueError, TypeError):
                    pass
        if dt is None:
            dt = datetime.utcnow()
        try:
            inicio = (dt - timedelta(days=3)).strftime('%Y-%m-%d')
            fim = (dt + timedelta(days=3)).strftime('%Y-%m-%d')
            roteiros_periodo = obter_roteiro_por_periodo(token, inicio, fim)
            if not roteiros_periodo:
                inicio_iso = (dt - timedelta(days=3)).strftime('%Y-%m-%dT00:00:00.000Z')
                fim_iso = (dt + timedelta(days=3)).strftime('%Y-%m-%dT23:59:59.000Z')
                roteiros_periodo = obter_roteiro_por_periodo(token, inicio_iso, fim_iso)
            id_viagem_int = None
            try:
                id_viagem_int = int(id_viagem)
            except (TypeError, ValueError):
                pass
            id_viagem_str = str(id_viagem).strip()
            for rot in (roteiros_periodo or []):
                if not isinstance(rot, dict):
                    continue
                rot_obj = rot.get('data') if rot.get('viagemFaturada') is None and rot.get('data') else rot
                vf = rot_obj.get('viagemFaturada') or rot_obj.get('viagem_faturada')
                vf_id = vf.get('id') or vf.get('Id') if isinstance(vf, dict) else None
                if vf_id is None and isinstance(vf, (int, float)):
                    vf_id = vf
                if vf_id is None:
                    vf_id = rot_obj.get('viagemFaturadaId') or rot_obj.get('viagemFaturada_id') or rot_obj.get('viagem_faturada_id')
                if vf_id is not None and (vf_id == id_viagem_int or str(vf_id) == id_viagem_str):
                    if not id_roteiro:
                        id_roteiro = str(rot_obj.get('id') or rot_obj.get('Id') or rot.get('id') or rot.get('Id') or '')
                    if not identificador_rota:
                        identificador_rota = (rot_obj.get('identificadorRota') or rot_obj.get('identificador_rota') or rot_obj.get('identificador') or rot_obj.get('nome') or rot.get('identificadorRota') or rot.get('identificador_rota') or '').strip()
                    break
        except Exception:
            pass
    # Canhotos + NFs em paralelo; itens de cada NF também em paralelo (romaneio import)
    cache_ref = {}

    def _ref_cached(ref_id):
        if not ref_id:
            return None
        k = str(ref_id)
        if k not in cache_ref:
            cache_ref[k] = obter_ponto_referencia(token, ref_id)
        return cache_ref[k]

    def _fetch_canhotos():
        return obter_canhotos_viagem(token, id_viagem, enriquecer=False)

    def _fetch_nfs():
        if nfs_pre is not None:
            return nfs_pre
        return obter_notas_fiscais_viagem(token, id_viagem)

    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_canhotos = pool.submit(_fetch_canhotos)
        fut_nfs = pool.submit(_fetch_nfs)
        entregas, meta_canhotos = fut_canhotos.result()
        nfs = fut_nfs.result() or []
    placa = (meta_canhotos.get('veiculo') or meta_canhotos.get('veículo') or '').strip()
    if not placa:
        veiculo_viagem = viagem_full.get('veiculo') or viagem_full.get('veículo')
        if isinstance(veiculo_viagem, dict):
            placa = (veiculo_viagem.get('placa') or veiculo_viagem.get('Placa') or '').strip()
        if not placa and isinstance(veiculo_viagem, dict) and (veiculo_viagem or {}).get('id') and obter_veiculo_por_id:
            v_api = obter_veiculo_por_id(token, veiculo_viagem.get('id'))
            if isinstance(v_api, dict):
                placa = (v_api.get('placa') or v_api.get('Placa') or '').strip()
    motorista = (meta_canhotos.get('motoristaNome') or meta_canhotos.get('motorista_nome') or '').strip()
    data_ini = viagem_full.get('inicioDataHora') or viagem_full.get('dataInicioViagem') or ''
    data_expedicao = data_ini[:10] if isinstance(data_ini, str) and len(data_ini) >= 10 else str(data_ini or '')

    nf_ids = []
    for nf in nfs:
        nf_id = nf.get('id') or nf.get('Id')
        if nf_id:
            nf_ids.append((nf, nf_id))

    itens_por_nf_id = {}
    if nf_ids:
        def _fetch_itens_nf(pair):
            nf_obj, nf_id = pair
            try:
                return nf_id, obter_itens_nota_fiscal(token, id_viagem, nf_id) or []
            except Exception:
                return nf_id, []

        workers = min(8, max(2, len(nf_ids)))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for nf_id, itens in pool.map(_fetch_itens_nf, nf_ids):
                itens_por_nf_id[nf_id] = itens

    linhas = []
    row_index = 0
    for nf in nfs or []:
        nf_id = nf.get('id') or nf.get('Id')
        numero_nf = str(nf.get('numero') or nf.get('numeroNF') or nf.get('numeroNf') or nf.get('nf') or '')
        pedido = nf.get('pedido') or {}
        info_ent = None
        for ent in (entregas or []):
            dados = ent.get('entrega') or ent.get('canhoto') or ent
            if not isinstance(dados, dict):
                dados = ent
            for nf_ent in (dados.get('notasFiscais') or dados.get('notas_fiscais') or []):
                if not isinstance(nf_ent, dict):
                    if nf_ent == numero_nf or str(nf_ent) == str(nf_id):
                        info_ent = dados
                        break
                    continue
                if (nf_ent.get('numero') or nf_ent.get('id')) in (numero_nf, nf_id, str(nf_id)):
                    info_ent = dados
                    break
            if info_ent:
                break
        dados_ent = info_ent if info_ent else ((entregas or [{}])[0] if entregas else {})
        if isinstance(dados_ent, dict) and not dados_ent:
            ent0 = (entregas or [{}])[0] if entregas else {}
            dados_ent = (ent0.get('entrega') or ent0.get('canhoto') or ent0) if isinstance(ent0, dict) else {}
        cliente = (dados_ent.get('cliente') or dados_ent.get('Cliente')) if isinstance(dados_ent, dict) else None
        ref_id = (dados_ent.get('referenciaId') or dados_ent.get('referencia_id')) if isinstance(dados_ent, dict) else None
        ref = _ref_cached(ref_id)
        cliente_nome = (cliente.get('nome', '') or cliente.get('razaoSocial', '')) if isinstance(cliente, dict) else ''
        cliente_razao = (cliente.get('razaoSocial', '') or cliente.get('nome', '')) if isinstance(cliente, dict) else ''
        endereco = (cliente.get('endereco') or cliente.get('endereço') or cliente.get('logradouro') or '') if isinstance(cliente, dict) else ''
        cidade = (cliente.get('cidade') or '') if isinstance(cliente, dict) else ''
        cidade = _normalizar_cidade_nome(cidade)
        if ref and isinstance(ref, dict):
            if not cliente_nome:
                cliente_nome = ref.get('nome') or ref.get('razaoSocial') or ''
            if not cliente_razao:
                cliente_razao = ref.get('razaoSocial') or ref.get('nome') or ''
            if not endereco:
                endereco = ref.get('endereco') or ref.get('endereço') or ref.get('logradouro') or ''
            if not cidade:
                cidade = _normalizar_cidade_nome(ref.get('cidade'))
        codigo_cliente = (dados_ent.get('cnpj') or dados_ent.get('Cnpj') or '') if isinstance(dados_ent, dict) else ''
        if not codigo_cliente and isinstance(pedido, dict):
            codigo_cliente = (pedido.get('cnpj') or pedido.get('Cnpj') or '').strip()
        if not codigo_cliente and ref and isinstance(ref, dict):
            codigo_cliente = (ref.get('cnpj') or ref.get('Cnpj') or '').strip()
        itens = itens_por_nf_id.get(nf_id, []) if nf_id else []
        for item in itens or []:
            prod = item.get('produto') or {}
            codigo_produto = str(item.get('codigo') or item.get('referenciaItem') or prod.get('codigo') or '').strip()
            descricao = str(item.get('descricaoItem') or item.get('descricao') or prod.get('descricao') or prod.get('descrição') or '').strip()
            quantidade = int(item.get('quantidade', 0) or 0)
            unidade = str(item.get('unidade') or prod.get('unidade') or '').strip()
            peso_bruto = str(item.get('pesoBruto') or item.get('pesoLiquido') or prod.get('pesoBruto') or '').strip()
            row_index += 1
            data_row = {
                'id_roteiro': id_roteiro,
                'id_viagem': id_viagem,
                'codigo_produto': codigo_produto,
                'descricao': descricao,
                'quantidade': quantidade,
                'unidade': unidade,
                'peso_bruto': peso_bruto,
                'codigo_cliente': codigo_cliente,
                'endereco': endereco,
                'cidade': cidade,
                'placa': placa,
                'motorista': motorista,
                'data_expedicao': data_expedicao,
            }
            linhas.append({
                'row_index': row_index,
                'id_roteiro': id_roteiro or None,
                'id_viagem': id_viagem,
                'identificador_rota': identificador_rota or None,
                'codigo_produto': codigo_produto or None,
                'descricao': descricao or None,
                'quantidade': quantidade,
                'unidade': unidade or None,
                'peso_bruto': peso_bruto or None,
                'codigo_cliente': codigo_cliente or None,
                'endereco': endereco or None,
                'cidade': cidade or None,
                'placa': placa or None,
                'motorista': motorista or None,
                'data_expedicao': data_expedicao or None,
                'data': data_row,
            })
    return (id_roteiro, linhas)


def _ravex_resolver_id_para_viagem(token, id_unico):
    """Dado um ID (roteiro ou viagem), resolve para (id_viagem, id_roteiro, somente_roteiro)."""
    api = _ravex_resolver_id_unico_ravex(token, id_unico)
    if not api or api.get('erro'):
        return (None, None, False)
    id_viagem = str(api.get('id_viagem') or '').strip()
    id_roteiro = str(api.get('id_roteiro') or '').strip()
    somente_roteiro = bool(api.get('somente_roteiro'))
    if somente_roteiro and id_roteiro:
        return (id_roteiro, id_roteiro, True)
    if not id_viagem:
        return (None, None, False)
    return (id_viagem, id_roteiro, False)


def _ravex_502_resposta(e):
    """Resposta 502 para falha de auth Ravex, com diagnóstico se as vars estão no servidor."""
    msg = getattr(e, 'args', [str(e)])[0] if getattr(e, 'args', None) else str(e)
    env_u = bool((os.environ.get("RAVEX_USER") or os.environ.get("ravex_user") or "").strip())
    env_p = bool((os.environ.get("RAVEX_PASSWORD") or os.environ.get("ravex_password") or "").strip())
    return jsonify({
        'erro': 'Falha ao autenticar na API Ravex: %s' % msg,
        'diagnostico': 'No servidor: RAVEX_USER=%s, RAVEX_PASSWORD=%s. Se os dois forem "não", no Render faça Save e depois Manual Deploy → Clear build cache & deploy. Teste: /api/ravex-env-check' % ('sim' if env_u else 'não', 'sim' if env_p else 'não')
    }), 502


def _upsert_viagem_placa_motorista(conn, id_viagem, placa, motorista):
    """Grava placa e motorista em viagem_placa e viagem_motorista para o id_viagem (upsert)."""
    if not id_viagem:
        return
    id_norm = _normalizar_id_viagem(id_viagem)
    usuario = _usuario_atual()
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''INSERT INTO viagem_placa (id_viagem, placa, atualizado_por) VALUES (?, ?, ?)
               ON CONFLICT (id_viagem) DO UPDATE SET placa = EXCLUDED.placa, atualizado_por = EXCLUDED.atualizado_por''',
            (id_norm, (placa or '').strip(), usuario)
        )
        conn.execute(
            '''INSERT INTO viagem_motorista (id_viagem, motorista, atualizado_por) VALUES (?, ?, ?)
               ON CONFLICT (id_viagem) DO UPDATE SET motorista = EXCLUDED.motorista, atualizado_por = EXCLUDED.atualizado_por''',
            (id_norm, (motorista or '').strip(), usuario)
        )
    else:
        conn.execute(
            '''INSERT INTO viagem_placa (id_viagem, placa) VALUES (?, ?) ON CONFLICT(id_viagem) DO UPDATE SET placa = excluded.placa''',
            (id_norm, (placa or '').strip())
        )
        conn.execute(
            '''INSERT INTO viagem_motorista (id_viagem, motorista) VALUES (?, ?) ON CONFLICT(id_viagem) DO UPDATE SET motorista = excluded.motorista''',
            (id_norm, (motorista or '').strip())
        )


@app.route('/api/ravex/verificar-baixado', methods=['GET'])
def api_ravex_verificar_baixado():
    """Verifica no banco se o ID (roteiro ou viagem) já foi importado, sem chamar a API Ravex."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL.', 'ja_baixado': False}), 400
    id_input = (
        request.args.get('id')
        or request.args.get('id_viagem')
        or request.args.get('id_roteiro')
        or ''
    ).strip()
    if not id_input:
        return jsonify({'erro': 'Informe o parâmetro id (roteiro ou viagem).', 'ja_baixado': False}), 400
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.close()
        return jsonify({'erro': 'Banco não é Postgres.', 'ja_baixado': False}), 400
    ds = _get_latest_dataset_id(conn)
    if not ds:
        conn.close()
        return jsonify({'erro': 'Nenhum dataset ativo.', 'ja_baixado': False}), 400
    ja, info = _ravex_id_input_ja_baixado(conn, ds, id_input)
    conn.close()
    msg = _ravex_mensagem_ja_baixado(id_input, info) if ja else ''
    return jsonify({
        'ok': True,
        'ja_baixado': ja,
        'duplicado': ja,
        'id_informado': id_input,
        'id_viagem': (info or {}).get('id_viagem') or '',
        'id_roteiro': (info or {}).get('id_roteiro') or '',
        'erro': msg if ja else None,
        'detalhe': info or {},
    })


@app.route('/api/ravex/verificar-baixado-lote', methods=['POST'])
def api_ravex_verificar_baixado_lote():
    """Verifica vários IDs de uma vez (evita N requisições HTTP na importação por lista)."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL.', 'ja_baixados': []}), 400
    data = request.get_json(silent=True) or {}
    ids_raw = data.get('ids') or data.get('id_list') or []
    if isinstance(ids_raw, str):
        ids_raw = ids_raw.replace(',', '\n').split('\n')
    ids = []
    vistos = set()
    for raw in ids_raw or []:
        s = str(raw or '').strip()
        if s and s not in vistos:
            vistos.add(s)
            ids.append(s)
    if not ids:
        return jsonify({'ok': True, 'ja_baixados': [], 'total': 0}), 200
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.close()
        return jsonify({'erro': 'Banco não é Postgres.', 'ja_baixados': []}), 400
    ds = _get_latest_dataset_id(conn)
    if not ds:
        conn.close()
        return jsonify({'erro': 'Nenhum dataset ativo.', 'ja_baixados': []}), 400
    ja_baixados = []
    try:
        for id_input in ids:
            ja, info = _ravex_id_input_ja_baixado(conn, ds, id_input)
            if ja:
                ja_baixados.append({
                    'id': id_input,
                    'id_viagem': (info or {}).get('id_viagem') or '',
                    'id_roteiro': (info or {}).get('id_roteiro') or '',
                    'erro': _ravex_mensagem_ja_baixado(id_input, info),
                })
    finally:
        conn.close()
    return jsonify({'ok': True, 'ja_baixados': ja_baixados, 'total': len(ja_baixados)})


def _ravex_importar_romaneio_executar(data, progress_cb=None):
    """Executa importação Ravex (romaneio). Retorna (body_dict, http_status)."""
    def _prog(pct, msg):
        if progress_cb:
            try:
                progress_cb(int(pct), msg or '')
            except Exception:
                pass

    if not _usa_banco_para_dados():
        return {'erro': 'Configure DATABASE_URL para usar importação Ravex.'}, 400
    if not ravex_get_token or not obter_viagem_por_id or not obter_roteiro_por_id:
        return {'erro': 'Módulo ravex_client não disponível. Instale: pip install requests'}, 500
    data = data or {}
    id_roteiro_in = (data.get('id_roteiro') or data.get('id_roteiro_str') or '').strip()
    id_viagem_in = (data.get('id_viagem') or data.get('id_viagem_str') or '').strip()
    id_unico = (data.get('id') or data.get('id_busca') or '').strip()
    if not id_roteiro_in and not id_viagem_in and not id_unico:
        return {'erro': 'Informe id_roteiro, id_viagem ou id (roteiro/viagem).'}, 400
    _prog(5, 'Autenticando na API Ravex...')
    try:
        token = ravex_get_token()
    except Exception as e:
        msg = getattr(e, 'args', [str(e)])[0] if getattr(e, 'args', None) else str(e)
        return {'erro': 'Falha ao autenticar na API Ravex: %s' % msg}, 502
    _prog(15, 'Conectando ao banco...')
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.close()
        return {'erro': 'Banco não é Postgres.'}, 400
    ds = _get_latest_dataset_id(conn)
    if not ds:
        conn.close()
        return {'erro': 'Nenhum dataset ativo. Importe a base (planilha) primeiro.'}, 400

    id_viagem = None
    id_roteiro = None
    identificador_rota = ''
    viagem_full = None
    roteiro_full = None
    somente_roteiro = False

    _prog(25, 'Resolvendo ID na Ravex...')
    if id_unico and not id_viagem_in and not id_roteiro_in:
        resolved = _ravex_resolver_importacao_unico(conn, token, ds, id_unico)
        if not resolved or resolved.get('erro'):
            conn.close()
            err = (resolved or {}).get('erro') or 'Não foi possível resolver o ID.'
            status = 404 if 'não encontrado' in err.lower() or 'nem como' in err.lower() else 400
            return {'erro': err}, status
        id_viagem = resolved.get('id_viagem')
        id_roteiro = resolved.get('id_roteiro') or id_unico
        identificador_rota = resolved.get('identificador_rota') or ''
        viagem_full = resolved.get('viagem_full')
        roteiro_full = resolved.get('roteiro_full')
        somente_roteiro = bool(resolved.get('somente_roteiro'))
    elif id_viagem_in:
        roteiro_full = None
        somente_roteiro = False
        try:
            vid = int(id_viagem_in)
        except (TypeError, ValueError):
            vid = id_viagem_in
        viagem_full = obter_viagem_por_id(token, vid)
        if viagem_full:
            id_viagem = str(vid)
            roteiro_obj = viagem_full.get('roteiro') or viagem_full.get('roteiroFaturado')
            if isinstance(roteiro_obj, dict):
                id_roteiro = str(roteiro_obj.get('id') or roteiro_obj.get('Id') or '')
                identificador_rota = _ravex_extrair_identificador_rota(roteiro_obj)
            if not id_roteiro:
                id_roteiro = str(viagem_full.get('roteiroId') or viagem_full.get('idRoteiro') or '')
    elif id_roteiro_in:
        try:
            rid = int(id_roteiro_in)
        except (TypeError, ValueError):
            rid = id_roteiro_in
        roteiro_full = obter_roteiro_por_id(token, rid)
        if not roteiro_full:
            conn.close()
            return {'erro': 'Roteiro não encontrado na API Ravex.'}, 404
        id_roteiro = str(rid)
        identificador_rota = _ravex_extrair_identificador_rota(roteiro_full)
        id_viagem = _ravex_extrair_viagem_id_de_roteiro(roteiro_full)
        if not id_viagem:
            somente_roteiro = True

    if not id_viagem and not somente_roteiro:
        conn.close()
        return {'erro': 'Não foi possível obter o ID da viagem ou do roteiro.'}, 400
    if somente_roteiro and not id_roteiro:
        conn.close()
        return {'erro': 'Não foi possível obter o ID do roteiro.'}, 400

    id_input_dup = id_unico or id_viagem_in or id_roteiro_in or id_roteiro
    forcar = _ravex_parse_forcar_reimportar(data)
    if not forcar:
        if somente_roteiro:
            ja_dup, info_dup = _ravex_id_input_ja_baixado(conn, ds, id_input_dup or id_roteiro)
        else:
            ja_dup, info_dup = _ravex_viagem_ja_baixada(conn, ds, id_viagem)
        if ja_dup:
            return _ravex_resposta_duplicado(
                conn, ds, 'id_unico', id_viagem or id_roteiro, id_roteiro or id_roteiro_in or id_unico, info_dup,
                id_input=id_input_dup,
            )
    _prog(40, 'Buscando itens na API Ravex...')
    if somente_roteiro:
        id_roteiro, linhas = _ravex_linhas_romaneio_roteiro(
            token, id_roteiro,
            roteiro_full=roteiro_full,
            identificador_rota_pre=identificador_rota,
        )
        if id_roteiro is None:
            conn.close()
            return {'erro': 'Roteiro não encontrado na API Ravex.'}, 404
        if not linhas:
            conn.close()
            return {'erro': 'Roteiro sem itens/pedidos na API Ravex.'}, 404
        chave_gravacao = id_roteiro
    else:
        id_roteiro, linhas = _ravex_linhas_romaneio_viagem(
            token, id_viagem,
            id_roteiro_pre=id_roteiro,
            identificador_rota_pre=identificador_rota,
            viagem_full_pre=viagem_full,
        )
        if id_roteiro is None:
            conn.close()
            return {'erro': 'Viagem não encontrada ou sem itens na API Ravex.'}, 404
        chave_gravacao = id_viagem
    _prog(75, 'Gravando %s itens no banco...' % len(linhas))
    try:
        n_itens = _ravex_insert_linhas_romaneio(
            conn, ds, '' if somente_roteiro else chave_gravacao, linhas, id_roteiro=id_roteiro,
        )
        params_ok = {
            'id_roteiro': id_roteiro or id_roteiro_in or id_unico,
            'somente_roteiro': somente_roteiro,
        }
        if somente_roteiro:
            params_ok['id_viagem'] = ''
            params_ok['viagens_importadas'] = []
        else:
            params_ok['id_viagem'] = id_viagem
            params_ok['viagens_importadas'] = [str(_normalizar_id_viagem(id_viagem) or id_viagem).strip()]
        _registrar_importacao_ravex(
            conn,
            dataset_id=ds,
            tipo='id_unico',
            status='OK',
            parametros=params_ok,
            viagens_processadas=1,
            total_itens=len(linhas),
            erros=[],
        )
        conn.commit()
        conn.close()
    except Exception as e:
        try:
            _registrar_importacao_ravex(
                conn,
                dataset_id=ds,
                tipo='id_unico',
                status='ERRO',
                parametros={
                    'id_viagem': '' if somente_roteiro else id_viagem,
                    'id_roteiro': id_roteiro or id_roteiro_in or id_unico,
                    'somente_roteiro': somente_roteiro,
                    'viagens_importadas': [] if somente_roteiro else [str(_normalizar_id_viagem(id_viagem) or id_viagem).strip()],
                },
                viagens_processadas=0,
                total_itens=0,
                erros=[{'erro': str(e)}],
            )
            conn.close()
        except Exception:
            pass
        return {'erro': 'Erro ao gravar romaneio: %s' % str(e)}, 500
    _prog(100, 'Importação concluída.')
    if somente_roteiro:
        msg_conf = 'Romaneio importado pelo roteiro %s (sem viagem faturada). Use este ID do roteiro na conferência.' % id_roteiro
    else:
        msg_conf = 'Romaneio importado. Use o ID da viagem (%s) para conferência.' % id_viagem
    return {
        'ok': True,
        'id_viagem': id_viagem or '',
        'id_roteiro': id_roteiro or id_viagem,
        'somente_roteiro': somente_roteiro,
        'total_itens': len(linhas),
        'mensagem': msg_conf,
    }, 200


def _ravex_job_worker(job_id, job_kind, data, app, usuario=''):
    """Thread em background: importação Ravex longa (evita timeout 502 do proxy)."""
    with app.app_context():
        _import_request_ctx.usuario = usuario or ''

        def progress_cb(pct, msg):
            _ravex_import_job_update(job_id, status='running', progress=pct, message=msg)

        try:
            if job_kind == 'periodo':
                body, code = _ravex_sincronizar_periodo_executar(data, progress_cb=progress_cb)
            elif job_kind == 'lista':
                body, code = _ravex_importar_lista_executar(data, progress_cb=progress_cb)
            else:
                body, code = _ravex_importar_romaneio_executar(data, progress_cb=progress_cb)
            _ravex_import_job_update(
                job_id,
                status='done',
                progress=100,
                message='Concluído',
                body=body,
                http_status=code,
            )
        except Exception as e:
            try:
                app.logger.exception('ravex job %s %s', job_kind, job_id)
            except Exception:
                pass
            _ravex_import_job_update(
                job_id,
                status='error',
                progress=100,
                message=str(e),
                body={'erro': str(e)},
                http_status=500,
            )


def _ravex_importar_romaneio_worker(job_id, data, app, usuario=''):
    """Compat: worker de importação unitária."""
    _ravex_job_worker(job_id, 'romaneio', data, app, usuario)


def _ravex_iniciar_job_background(job_kind, data):
    """Cria job assíncrono e retorna (job_id, 202)."""
    _ravex_import_job_cleanup()
    job_id = uuid.uuid4().hex[:16]
    _ravex_import_job_update(job_id, status='running', progress=0, message='Iniciando...', kind=job_kind)
    app = current_app._get_current_object()
    usuario_imp = session.get('usuario', '') or ''
    threading.Thread(
        target=_ravex_job_worker,
        args=(job_id, job_kind, dict(data), app, usuario_imp),
        daemon=True,
    ).start()
    return job_id


@app.route('/api/ravex/importar-romaneio', methods=['POST'])
def api_ravex_importar_romaneio():
    """Inicia importação Ravex em background; consulte status com GET .../status/<job_id>."""
    data = request.get_json() or {}
    id_roteiro_in = (data.get('id_roteiro') or data.get('id_roteiro_str') or '').strip()
    id_viagem_in = (data.get('id_viagem') or data.get('id_viagem_str') or '').strip()
    id_unico = (data.get('id') or data.get('id_busca') or '').strip()
    if not id_roteiro_in and not id_viagem_in and not id_unico:
        return jsonify({'erro': 'Informe id_roteiro, id_viagem ou id (roteiro/viagem).'}), 400
    if data.get('sync') or request.args.get('sync'):
        body, code = _ravex_importar_romaneio_executar(data)
        return jsonify(body), code
    job_id = _ravex_iniciar_job_background('romaneio', data)
    return jsonify({'ok': True, 'async': True, 'job_id': job_id}), 202


@app.route('/api/ravex/importar-romaneio/status/<job_id>', methods=['GET'])
def api_ravex_importar_romaneio_status(job_id):
    """Status da importação assíncrona (polling)."""
    job = _ravex_import_job_get((job_id or '').strip())
    if not job:
        return jsonify({'erro': 'Importação não encontrada ou expirada.'}), 404
    status = job.get('status') or 'running'
    if status == 'done':
        body = job.get('body') or {}
        return jsonify(body), int(job.get('http_status') or 200)
    if status == 'error':
        body = job.get('body') or {'erro': job.get('message') or 'Erro na importação.'}
        return jsonify(body), int(job.get('http_status') or 500)
    return jsonify({
        'ok': True,
        'async': True,
        'status': 'running',
        'progress': job.get('progress') or 0,
        'message': job.get('message') or 'Processando...',
    }), 202


def _ravex_sincronizar_periodo_executar(data, progress_cb=None):
    """Sincroniza romaneio do período Ravex. Retorna (body_dict, http_status)."""
    def _prog(pct, msg):
        if progress_cb:
            try:
                progress_cb(int(pct), msg or '')
            except Exception:
                pass

    if not _usa_banco_para_dados():
        return {'erro': 'Configure DATABASE_URL para usar sincronização Ravex.'}, 400
    if not ravex_get_token or not obter_roteiro_por_periodo:
        return {'erro': 'Módulo ravex_client não disponível ou obter_roteiro_por_periodo não implementado.'}, 500
    data = data or {}
    data_inicio = (data.get('data_inicio') or data.get('dataInicio') or '').strip()
    data_fim = (data.get('data_fim') or data.get('dataFim') or '').strip()
    if not data_inicio or not data_fim:
        return {'erro': 'Informe data_inicio e data_fim (ex: 2026-03-01 e 2026-03-05).'}, 400
    _prog(5, 'Autenticando na API Ravex...')
    try:
        token = ravex_get_token()
    except Exception as e:
        msg = getattr(e, 'args', [str(e)])[0] if getattr(e, 'args', None) else str(e)
        return {'erro': 'Falha ao autenticar na API Ravex: %s' % msg}, 502
    _prog(12, 'Buscando roteiros do período na Ravex...')
    roteiros_romaneio = _ravex_roteiros_por_periodo_para_romaneio(token, data_inicio, data_fim)
    # Evitar processar o mesmo id_viagem mais de uma vez (pode aparecer em mais de um roteiro no período)
    vistos = set()
    lista_viagens = []
    for r in roteiros_romaneio:
        id_v = (r.get('id_viagem') or '').strip()
        if id_v and id_v not in vistos:
            vistos.add(id_v)
            lista_viagens.append(r)
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.close()
        return {'erro': 'Banco não é Postgres.'}, 400
    ds = _get_latest_dataset_id(conn)
    if not ds:
        conn.close()
        return {'erro': 'Nenhum dataset ativo. Importe a base (planilha) primeiro.'}, 400
    viagens_processadas = 0
    total_itens = 0
    erros = []
    pulados_duplicados = []
    viagens_importadas = []
    forcar = _ravex_parse_forcar_reimportar(data)
    try:
        # Registrar todos os roteiros do período na tabela id_roteiros
        for r in roteiros_romaneio:
            id_roteiro = (r.get('id_roteiro') or '').strip()
            id_viagem_r = (r.get('id_viagem') or '').strip()
            if not id_roteiro or not id_viagem_r:
                continue
            identificador_rota = (r.get('identificador_rota') or '').strip() or None
            data_viagem = r.get('data_viagem')
            if data_viagem is not None and isinstance(data_viagem, str) and len(data_viagem) < 10:
                data_viagem = None
            conn.execute(
                """INSERT INTO id_roteiros (dataset_id, id_roteiro, id_viagem, identificador_rota, data_viagem)
                   VALUES (?, ?, ?, ?, ?::timestamptz)
                   ON CONFLICT (dataset_id, id_roteiro) DO UPDATE SET
                     id_viagem = EXCLUDED.id_viagem,
                     identificador_rota = EXCLUDED.identificador_rota,
                     data_viagem = EXCLUDED.data_viagem,
                     atualizado_em = NOW()""",
                (str(ds), id_roteiro, id_viagem_r, identificador_rota, data_viagem),
            )
        ja_baixadas = set()
        if not forcar:
            ja_baixadas = _ravex_viagens_ja_baixadas_batch(
                conn, ds, [v.get('id_viagem') for v in lista_viagens if v.get('id_viagem')]
            )
        viagens_a_processar = []
        for v in lista_viagens:
            id_viagem = v.get('id_viagem') or ''
            if not id_viagem:
                continue
            id_norm = str(_normalizar_id_viagem(id_viagem) or id_viagem).strip()
            if not forcar and id_norm in ja_baixadas:
                pulados_duplicados.append({
                    'id_viagem': id_viagem,
                    'id_roteiro': v.get('id_roteiro'),
                    'motivo': 'já importada',
                })
                continue
            viagens_a_processar.append(v)

        total_proc = len(viagens_a_processar)
        _prog(22, 'Puxando %d viagem(ns) da API Ravex...' % total_proc)

        def _fetch_linhas_viagem(v):
            id_viagem = v.get('id_viagem') or ''
            id_roteiro_pre = v.get('id_roteiro') or None
            identificador_rota_pre = v.get('identificador_rota') or None
            try:
                id_roteiro, linhas = _ravex_linhas_romaneio_viagem(
                    token, id_viagem,
                    id_roteiro_pre=id_roteiro_pre,
                    identificador_rota_pre=identificador_rota_pre,
                )
                return id_viagem, id_roteiro, linhas, None
            except Exception as e:
                return id_viagem, None, [], str(e)

        resultados_api = []
        workers = min(_RAVEX_IMPORT_WORKERS, max(1, len(viagens_a_processar)))
        if viagens_a_processar:
            done_api = 0
            with ThreadPoolExecutor(max_workers=workers) as pool:
                for res in pool.map(_fetch_linhas_viagem, viagens_a_processar):
                    resultados_api.append(res)
                    done_api += 1
                    if total_proc:
                        pct = 22 + int(68 * done_api / total_proc)
                        _prog(pct, 'API Ravex: %d/%d viagem(ns)...' % (done_api, total_proc))

        _prog(92, 'Gravando romaneio no banco...')
        for id_viagem, id_roteiro, linhas, err_api in resultados_api:
            if err_api:
                erros.append({'id_viagem': id_viagem, 'erro': err_api})
                continue
            if id_roteiro is None or not linhas:
                continue
            try:
                n_gravados = _ravex_insert_linhas_romaneio(conn, ds, id_viagem, linhas)
                viagens_processadas += 1
                total_itens += n_gravados
                id_norm_v = str(_normalizar_id_viagem(id_viagem) or id_viagem).strip()
                if id_norm_v:
                    viagens_importadas.append(id_norm_v)
            except Exception as e:
                erros.append({'id_viagem': id_viagem, 'erro': str(e)})
        _registrar_importacao_ravex(
            conn,
            dataset_id=ds,
            tipo='periodo',
            status='OK' if not erros else 'OK_COM_ERROS',
            parametros={
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'pulados_duplicados': len(pulados_duplicados),
                'viagens_importadas': viagens_importadas,
            },
            viagens_processadas=viagens_processadas,
            total_itens=total_itens,
            erros=erros,
        )
        conn.commit()
    except Exception as e:
        try:
            _registrar_importacao_ravex(
                conn,
                dataset_id=ds,
                tipo='periodo',
                status='ERRO',
                parametros={'data_inicio': data_inicio, 'data_fim': data_fim},
                viagens_processadas=viagens_processadas,
                total_itens=total_itens,
                erros=erros + [{'erro': str(e)}],
            )
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return {'erro': 'Erro ao sincronizar: %s' % str(e)}, 500
    conn.close()
    _prog(100, 'Sincronização concluída.')
    return {
        'ok': True,
        'viagens_processadas': viagens_processadas,
        'total_itens': total_itens,
        'roteiros_no_periodo': len(roteiros_romaneio),
        'roteiros_registrados': len([r for r in roteiros_romaneio if (r.get('id_roteiro') or '').strip() and (r.get('id_viagem') or '').strip()]),
        'viagens_listadas': len(lista_viagens),
        'pulados_duplicados': pulados_duplicados,
        'total_pulados_duplicados': len(pulados_duplicados),
        'erros': erros,
    }, 200


@app.route('/api/ravex/sincronizar-periodo', methods=['POST'])
def api_ravex_sincronizar_periodo():
    """Puxa roteiros do período (assíncrono por padrão; use ?sync=1 para resposta direta)."""
    data = request.get_json() or {}
    data_inicio = (data.get('data_inicio') or data.get('dataInicio') or '').strip()
    data_fim = (data.get('data_fim') or data.get('dataFim') or '').strip()
    if not data_inicio or not data_fim:
        return jsonify({'erro': 'Informe data_inicio e data_fim (ex: 2026-03-01 e 2026-03-05).'}), 400
    if data.get('sync') or request.args.get('sync'):
        body, code = _ravex_sincronizar_periodo_executar(data)
        return jsonify(body), code
    job_id = _ravex_iniciar_job_background('periodo', data)
    return jsonify({'ok': True, 'async': True, 'job_id': job_id}), 202


def _ravex_importar_lista_executar(data, progress_cb=None):
    """Importa lista de IDs Ravex. Retorna (body_dict, http_status)."""
    def _prog(pct, msg):
        if progress_cb:
            try:
                progress_cb(int(pct), msg or '')
            except Exception:
                pass

    if not _usa_banco_para_dados():
        return {'erro': 'Configure DATABASE_URL para usar importação Ravex.'}, 400
    if not ravex_get_token:
        return {'erro': 'Módulo ravex_client não disponível.'}, 500
    data = data or {}
    raw_ids = data.get('ids') or data.get('lista') or data.get('id_list') or []
    if isinstance(raw_ids, str):
        raw_ids = [s.strip() for s in raw_ids.replace(',', '\n').splitlines() if s.strip()]
    elif not isinstance(raw_ids, list):
        raw_ids = []
    ids = [str(x).strip() for x in raw_ids if str(x).strip()]
    if not ids:
        return {'erro': 'Informe uma lista de IDs (ids: ["18765009", "18765008", ...] ou um ID por linha).'}, 400
    _prog(5, 'Autenticando na API Ravex...')
    try:
        token = ravex_get_token()
    except Exception as e:
        msg = getattr(e, 'args', [str(e)])[0] if getattr(e, 'args', None) else str(e)
        return {'erro': 'Falha ao autenticar na API Ravex: %s' % msg}, 502
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.close()
        return {'erro': 'Banco não é Postgres.'}, 400
    ds = _get_latest_dataset_id(conn)
    if not ds:
        conn.close()
        return {'erro': 'Nenhum dataset ativo. Importe a base primeiro.'}, 400
    viagens_processadas = 0
    total_itens = 0
    erros = []
    pulados_duplicados = []
    viagens_importadas = []
    ids_vistos = set()
    forcar = _ravex_parse_forcar_reimportar(data)
    try:
        ids_unicos = []
        for id_unico in ids:
            if not id_unico or id_unico in ids_vistos:
                continue
            ids_vistos.add(id_unico)
            ids_unicos.append(id_unico)

        def _resolver_id(id_unico):
            try:
                id_viagem, id_roteiro, somente_roteiro = _ravex_resolver_id_para_viagem(token, id_unico)
                return id_unico, id_viagem, id_roteiro, somente_roteiro, None
            except Exception as e:
                return id_unico, None, None, False, str(e)

        resolvidos = []
        workers = min(_RAVEX_IMPORT_WORKERS, max(1, len(ids_unicos)))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for res in pool.map(_resolver_id, ids_unicos):
                resolvidos.append(res)

        ja_baixadas = set()
        if not forcar:
            ja_baixadas = _ravex_viagens_ja_baixadas_batch(
                conn, ds, [r[1] for r in resolvidos if r[1] and not r[3]]
            )

        tarefas = []
        for id_unico, id_viagem, id_roteiro, somente_roteiro, err_res in resolvidos:
            if err_res:
                erros.append({'id': id_unico, 'erro': err_res})
                continue
            if not id_viagem and not id_roteiro:
                erros.append({'id': id_unico, 'erro': 'ID não encontrado na API Ravex'})
                continue
            if somente_roteiro:
                if not forcar:
                    ja_r, _info = _ravex_id_input_ja_baixado(conn, ds, id_unico)
                    if ja_r:
                        pulados_duplicados.append({
                            'id': id_unico,
                            'id_viagem': '',
                            'id_roteiro': id_roteiro,
                            'motivo': 'já importada',
                        })
                        continue
            else:
                id_norm = str(_normalizar_id_viagem(id_viagem) or id_viagem).strip()
                if not forcar and id_norm in ja_baixadas:
                    pulados_duplicados.append({
                        'id': id_unico,
                        'id_viagem': id_viagem,
                        'id_roteiro': id_roteiro,
                        'motivo': 'já importada',
                    })
                    continue
            tarefas.append((id_unico, id_viagem, id_roteiro, somente_roteiro))

        total_tarefas = len(tarefas)
        _prog(18, 'Puxando %d ID(s) da API Ravex...' % total_tarefas)

        def _fetch_linhas_id(t):
            id_unico, id_viagem, id_roteiro, somente_roteiro = t
            try:
                if somente_roteiro:
                    rid, linhas = _ravex_linhas_romaneio_roteiro(token, id_roteiro or id_unico)
                else:
                    rid, linhas = _ravex_linhas_romaneio_viagem(
                        token, id_viagem, id_roteiro_pre=id_roteiro,
                    )
                return id_unico, id_viagem, rid, linhas, somente_roteiro, None
            except Exception as e:
                return id_unico, id_viagem, None, [], somente_roteiro, str(e)

        resultados_api = []
        workers = min(_RAVEX_IMPORT_WORKERS, max(1, len(tarefas)))
        if tarefas:
            done_api = 0
            with ThreadPoolExecutor(max_workers=workers) as pool:
                for res in pool.map(_fetch_linhas_id, tarefas):
                    resultados_api.append(res)
                    done_api += 1
                    if total_tarefas:
                        pct = 18 + int(72 * done_api / total_tarefas)
                        _prog(pct, 'API Ravex: %d/%d ID(s)...' % (done_api, total_tarefas))

        _prog(92, 'Gravando romaneio no banco...')
        for id_unico, id_viagem, id_roteiro, linhas, somente_roteiro, err_api in resultados_api:
            if err_api:
                erros.append({'id': id_unico, 'erro': err_api})
                continue
            if id_roteiro is None or not linhas:
                erros.append({'id': id_unico, 'erro': 'Sem itens na API'})
                continue
            try:
                n_gravados = _ravex_insert_linhas_romaneio(
                    conn, ds,
                    '' if somente_roteiro else id_viagem,
                    linhas,
                    id_roteiro=id_roteiro,
                )
                viagens_processadas += 1
                total_itens += n_gravados
                if somente_roteiro:
                    id_r_norm = str(id_roteiro or '').strip()
                    if id_r_norm:
                        viagens_importadas.append(id_r_norm)
                else:
                    id_norm_v = str(_normalizar_id_viagem(id_viagem) or id_viagem).strip()
                    if id_norm_v:
                        viagens_importadas.append(id_norm_v)
            except Exception as e:
                erros.append({'id': id_unico, 'erro': str(e)})
        _registrar_importacao_ravex(
            conn,
            dataset_id=ds,
            tipo='lista',
            status='OK' if not erros else 'OK_COM_ERROS',
            parametros={
                'ids_recebidos': len(ids),
                'pulados_duplicados': len(pulados_duplicados),
                'viagens_importadas': viagens_importadas,
            },
            viagens_processadas=viagens_processadas,
            total_itens=total_itens,
            erros=erros,
        )
        conn.commit()
    except Exception as e:
        try:
            _registrar_importacao_ravex(
                conn,
                dataset_id=ds,
                tipo='lista',
                status='ERRO',
                parametros={'ids_recebidos': len(ids)},
                viagens_processadas=viagens_processadas,
                total_itens=total_itens,
                erros=erros + [{'erro': str(e)}],
            )
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return {'erro': 'Erro ao importar lista: %s' % str(e)}, 500
    conn.close()
    _prog(100, 'Importação da lista concluída.')
    return {
        'ok': True,
        'viagens_processadas': viagens_processadas,
        'total_itens': total_itens,
        'ids_recebidos': len(ids),
        'pulados_duplicados': pulados_duplicados,
        'total_pulados_duplicados': len(pulados_duplicados),
        'erros': erros,
    }, 200


@app.route('/api/ravex/importar-lista', methods=['POST'])
def api_ravex_importar_lista():
    """Importa lista de IDs (assíncrono por padrão; use ?sync=1 para resposta direta)."""
    data = request.get_json() or {}
    raw_ids = data.get('ids') or data.get('lista') or data.get('id_list') or []
    if isinstance(raw_ids, str):
        raw_ids = [s.strip() for s in raw_ids.replace(',', '\n').splitlines() if s.strip()]
    elif not isinstance(raw_ids, list):
        raw_ids = []
    ids = [str(x).strip() for x in raw_ids if str(x).strip()]
    if not ids:
        return jsonify({'erro': 'Informe uma lista de IDs.'}), 400
    if data.get('sync') or request.args.get('sync'):
        body, code = _ravex_importar_lista_executar(data)
        return jsonify(body), code
    job_id = _ravex_iniciar_job_background('lista', data)
    return jsonify({'ok': True, 'async': True, 'job_id': job_id}), 202


@app.route('/api/extrato', methods=['GET'])
def get_extrato():
    """Retorna itens da carga agrupados (sem repetir): um por produto com quantidade total"""
    id_viagem = request.args.get('id_viagem')
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    conn = get_db()
    if id_viagem:
        produtos = conn.execute('''
            SELECT id_viagem, codigo_barras, produto,
                   SUM(quantidade) AS quantidade,
                   MAX(data_hora) AS data_hora,
                   MAX(veiculo) AS veiculo,
                   MAX(status) AS status
            FROM produtos_bipados
            WHERE id_viagem = ? AND COALESCE(fluxo, 'carregamento') = ?
            GROUP BY id_viagem, codigo_barras, produto
            ORDER BY produto
        ''', (id_viagem.strip(), fluxo)).fetchall()
    else:
        produtos = conn.execute('''
            SELECT id_viagem, codigo_barras, produto,
                   SUM(quantidade) AS quantidade,
                   MAX(data_hora) AS data_hora,
                   MAX(veiculo) AS veiculo,
                   MAX(status) AS status
            FROM produtos_bipados
            WHERE COALESCE(fluxo, 'carregamento') = ?
            GROUP BY id_viagem, codigo_barras, produto
            ORDER BY id_viagem, produto
        ''', (fluxo,)).fetchall()
    conn.close()
    return jsonify([dict(row) for row in produtos])


@app.route('/api/ravex/importacoes', methods=['GET'])
def api_ravex_listar_importacoes():
    """Lista histórico do que foi baixado/importado do Ravex (para exibir na nova aba)."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL para usar histórico de importações.', 'rows': []}), 400
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) != 'pg':
            conn.close()
            return jsonify({'erro': 'Banco não é Postgres.', 'rows': []}), 400
        limit = request.args.get('limit', '150')
        dataset_filtro = (request.args.get('dataset_id') or '').strip()
        periodo = (request.args.get('periodo') or '').strip().lower()
        data_inicio = (request.args.get('data_inicio') or '').strip()
        data_fim = (request.args.get('data_fim') or '').strip()
        hora_inicio = (request.args.get('hora_inicio') or '').strip()[:5]
        hora_fim = (request.args.get('hora_fim') or '').strip()[:5]
        usuario_filtro = (request.args.get('usuario') or '').strip()
        include_usuarios = request.args.get('usuarios', '1').strip().lower() not in ('0', 'false', 'no')
        try:
            limit_i = max(10, min(500, int(limit)))
        except Exception:
            limit_i = 150
        inicio_dt, fim_dt = None, None
        if data_inicio or data_fim or hora_inicio or hora_fim:
            meta = _parse_janela_datahora(data_inicio, data_fim, hora_inicio, hora_fim)
            inicio_dt, fim_dt = meta['inicio'], meta['fim_excl']
        else:
            inicio_dt, fim_dt = _ravex_filtro_periodo_datas(periodo, data_inicio, data_fim)
            if inicio_dt and fim_dt and (hora_inicio or hora_fim):
                inicio_dt, fim_dt = _ravex_aplicar_hora_janela(inicio_dt, fim_dt, hora_inicio, hora_fim)
        tem_filtros = bool(inicio_dt and fim_dt) or bool(usuario_filtro) or bool(hora_inicio or hora_fim)
        where = []
        params = []
        if dataset_filtro:
            where.append('dataset_id = ?::uuid')
            params.append(dataset_filtro)
        if inicio_dt and fim_dt:
            where.append('criado_em >= ? AND criado_em < ?')
            params.extend([inicio_dt, fim_dt])
        if usuario_filtro:
            where.append('LOWER(TRIM(COALESCE(usuario, \'\'))) = LOWER(?)')
            params.append(usuario_filtro)
        where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
        sql = f"""SELECT id, tipo, status, parametros, viagens_processadas, total_itens, usuario,
                         COALESCE(jsonb_array_length(erros), 0) AS erros_qtd, criado_em
                  FROM public.ravex_importacoes
                  {where_sql}
                  ORDER BY criado_em DESC NULLS LAST
                  LIMIT ?"""
        params.append(limit_i)
        rows = conn.execute(sql, tuple(params)).fetchall()
        out = [_ravex_importacao_dict_from_row(r) for r in (rows or [])]
        fonte = 'ravex_importacoes'
        if not out and not tem_filtros:
            tem_historico = conn.execute(
                'SELECT 1 FROM public.ravex_importacoes LIMIT 1'
            ).fetchone()
            if not tem_historico:
                out = _listar_baixados_ravex_de_romaneio(conn, limit_i)
                fonte = 'romaneio_por_item' if out else 'vazio'
            else:
                fonte = 'vazio'
        usuarios = _ravex_listar_usuarios_distintos(conn) if include_usuarios else []
        conn.close()
        return jsonify({'ok': True, 'rows': out, 'fonte': fonte, 'usuarios': usuarios})
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        err = str(e)
        if 'ravex_importacoes' in err.lower() and ('does not exist' in err.lower() or 'não existe' in err.lower()):
            return jsonify({
                'erro': 'Tabela ravex_importacoes não existe no Supabase. Execute supabase/migrate_ravex_importacoes.sql no SQL Editor.',
                'rows': [],
            }), 500
        return jsonify({'erro': err, 'rows': []}), 500


@app.route('/api/ravex/importacoes/<int:import_id>', methods=['DELETE'])
def api_ravex_excluir_importacao(import_id):
    """Exclui importação do histórico e apaga viagem/roteiro com romaneio, bipagem e metadados."""
    if not _usa_banco_para_dados():
        return jsonify({'erro': 'Configure DATABASE_URL para excluir importações.'}), 400
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) != 'pg':
            conn.close()
            return jsonify({'erro': 'Banco não é Postgres.'}), 400
        resultado, erro = _excluir_importacao_ravex_por_id(conn, import_id)
        if erro:
            conn.close()
            status = 404 if 'não encontrada' in str(erro.get('erro', '')).lower() else 400
            return jsonify(erro), status
        conn.commit()
        conn.close()
        return jsonify(resultado)
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


def _qual_coluna_filtro(header_str, tipo):
    """Retorna True se o cabeçalho corresponde ao tipo de filtro (data, pedido, nota_fiscal, cliente)."""
    if not header_str:
        return False
    h = header_str.upper().strip().replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ç', 'C')
    if tipo == 'data':
        return 'DATA' in h
    if tipo == 'pedido':
        return 'PEDIDO' in h and 'DATA' not in h
    if tipo == 'nota_fiscal':
        return ('NOTA' in h and 'FISCAL' in h) or h == 'NF' or (h.startswith('NOTA') and 'FISCAL' in h)
    if tipo == 'cliente':
        return 'CLIENTE' in h
    return False


def _qual_coluna_romaneio(header_str, tipo):
    """Retorna True se o cabeçalho corresponde ao filtro do romaneio (id_viagem, id_roteiro, codigo_cliente, codigo_produto, endereco, cidade)."""
    if not header_str:
        return False
    h = header_str.upper().strip().replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ã', 'A').replace('Ç', 'C').replace('Ó', 'O')
    if tipo == 'id_viagem':
        return ('ID' in h and 'VIAGEM' in h and 'FATURADA' in h) or (h == 'ID VIAGEM')
    if tipo == 'id_roteiro':
        return ('ID' in h and 'ROTEIRO' in h) or (h == 'ID ROTEIRO')
    if tipo == 'codigo_cliente':
        return 'CODIGO' in h and 'CLIENTE' in h
    if tipo == 'codigo_produto':
        return 'CODIGO' in h and 'PRODUTO' in h and 'BARRAS' not in h
    if tipo == 'endereco':
        return 'ENDERE' in h
    if tipo == 'cidade':
        return 'CIDADE' in h
    return False


@app.route('/api/romaneio', methods=['GET'])
def get_romaneio():
    """Retorna dados do romaneio: do banco (romaneio_por_item) quando DATABASE_URL está definido; senão planilha."""
    if _usa_banco_para_dados():
        conn = get_db()
        try:
            if getattr(conn, 'kind', None) != 'pg':
                conn.close()
                return jsonify({'headers': [], 'rows': [], 'erro': 'Configure DATABASE_URL.'})
            ds = _get_latest_dataset_id(conn)
            if not ds:
                conn.close()
                return jsonify({'headers': [], 'rows': []})
            filtro_id_viagem = request.args.get('id_viagem', '').strip()
            filtro_id_roteiro = request.args.get('id_roteiro', '').strip()
            filtro_codigo_cliente = request.args.get('codigo_cliente', '').strip()
            filtro_codigo_produto = request.args.get('codigo_produto', '').strip()
            filtro_endereco = request.args.get('endereco', '').strip()
            filtro_cidade = request.args.get('cidade', '').strip()
            sql = """SELECT data FROM romaneio_por_item WHERE dataset_id = ?"""
            params = [str(ds)]
            if filtro_id_viagem:
                sql += " AND id_viagem = ?"
                params.append(filtro_id_viagem)
            if filtro_id_roteiro:
                sql += " AND id_roteiro = ?"
                params.append(filtro_id_roteiro)
            if filtro_codigo_produto:
                sql += " AND codigo_produto ILIKE ?"
                params.append('%' + filtro_codigo_produto + '%')
            if filtro_codigo_cliente:
                sql += " AND COALESCE(codigo_cliente, '') ILIKE ?"
                params.append('%' + filtro_codigo_cliente + '%')
            if filtro_endereco:
                sql += " AND COALESCE(endereco, '') ILIKE ?"
                params.append('%' + filtro_endereco + '%')
            if filtro_cidade:
                sql += " AND COALESCE(cidade, '') ILIKE ?"
                params.append('%' + filtro_cidade + '%')
            sql += " ORDER BY row_index LIMIT 2000"
            rows_raw = conn.execute(sql, params).fetchall()
            conn.close()
            headers = []
            rows = []
            for r in rows_raw:
                data = r.get('data') if isinstance(r.get('data'), dict) else (json.loads(r['data']) if isinstance(r.get('data'), str) else {})
                if not data:
                    continue
                if not headers:
                    headers = [str(k) for k in data.keys()]
                row_dict = {str(k): (v.strftime('%Y-%m-%d %H:%M:%S') if isinstance(v, datetime) else v) for k, v in data.items()}
                if 'cidade' in row_dict:
                    row_dict['cidade'] = _normalizar_cidade_nome(row_dict.get('cidade'))
                rows.append(row_dict)
            return jsonify({'headers': headers or [], 'rows': rows})
        except Exception as e:
            try:
                conn.close()
            except Exception:
                pass
            return jsonify({'erro': str(e), 'headers': [], 'rows': []}), 500

    # Dados vêm apenas do banco; não usar mais planilha
    return jsonify({'headers': [], 'rows': [], 'erro': 'Configure DATABASE_URL. Os dados vêm apenas do banco de dados.'}), 400

    caminho_planilha = encontrar_planilha()
    if not caminho_planilha:
        return jsonify({'erro': 'Planilha não encontrada'}), 404
    
    filtro_id_viagem = request.args.get('id_viagem', '').strip()
    filtro_id_roteiro = request.args.get('id_roteiro', '').strip()
    filtro_codigo_cliente = request.args.get('codigo_cliente', '').strip()
    filtro_codigo_produto = request.args.get('codigo_produto', '').strip()
    filtro_endereco = request.args.get('endereco', '').strip()
    filtro_cidade = request.args.get('cidade', '').strip()
    
    try:
        try:
            wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
        except:
            wb = openpyxl.load_workbook(caminho_planilha, data_only=False)
        
        headers = []
        resultado = []
        
        if 'ROMANEIO POR ITEM' in wb.sheetnames:
            ws_romaneio = wb['ROMANEIO POR ITEM']
            header_row = list(ws_romaneio.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            headers = [str(h).strip() if h is not None else f'Col_{i}' for i, h in enumerate(header_row)]
            if not headers:
                headers = [f'Col_{i}' for i in range(100)]
            
            col_id_viagem = next((i for i, h in enumerate(headers) if _qual_coluna_romaneio(h, 'id_viagem')), None)
            col_id_roteiro = next((i for i, h in enumerate(headers) if _qual_coluna_romaneio(h, 'id_roteiro')), None)
            col_codigo_cliente = next((i for i, h in enumerate(headers) if _qual_coluna_romaneio(h, 'codigo_cliente')), None)
            col_codigo_produto = next((i for i, h in enumerate(headers) if _qual_coluna_romaneio(h, 'codigo_produto')), None)
            col_endereco = next((i for i, h in enumerate(headers) if _qual_coluna_romaneio(h, 'endereco')), None)
            col_cidade = next((i for i, h in enumerate(headers) if _qual_coluna_romaneio(h, 'cidade')), None)
            
            for row in ws_romaneio.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                try:
                    row_dict = {}
                    for i, val in enumerate(row):
                        key = headers[i] if i < len(headers) else f'Col_{i}'
                        if val is None:
                            row_dict[key] = ''
                        elif str(val).strip().startswith('='):
                            row_dict[key] = ''
                        else:
                            row_dict[key] = str(val).strip() if isinstance(val, (str, int, float)) else str(val)
                    
                    if not any(row_dict.get(h) for h in headers[:min(3, len(headers))]):
                        continue
                    
                    def valor_col(idx):
                        if idx is None or idx >= len(row):
                            return ''
                        return (str(row[idx]).strip() if row[idx] is not None else '').upper()
                    
                    if filtro_id_viagem and col_id_viagem is not None:
                        if filtro_id_viagem.upper() not in valor_col(col_id_viagem):
                            continue
                    if filtro_id_roteiro and col_id_roteiro is not None:
                        if filtro_id_roteiro.upper() not in valor_col(col_id_roteiro):
                            continue
                    if filtro_codigo_cliente and col_codigo_cliente is not None:
                        if filtro_codigo_cliente.upper() not in valor_col(col_codigo_cliente):
                            continue
                    if filtro_codigo_produto and col_codigo_produto is not None:
                        if filtro_codigo_produto.upper() not in valor_col(col_codigo_produto):
                            continue
                    if filtro_endereco and col_endereco is not None:
                        if filtro_endereco.upper() not in valor_col(col_endereco):
                            continue
                    if filtro_cidade and col_cidade is not None:
                        if filtro_cidade.upper() not in valor_col(col_cidade):
                            continue
                    
                    resultado.append(row_dict)
                except Exception:
                    continue
        
        wb.close()
        return jsonify({'headers': headers, 'rows': resultado})
        
    except Exception as e:
        return jsonify({'erro': f'Erro ao ler planilha: {str(e)}'}), 500

@app.route('/api/romaneio', methods=['POST'])
def update_romaneio():
    """Atualiza quantidade do romaneio"""
    data = request.json
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''INSERT INTO romaneio (codigo_barras, quantidade_romaneio)
               VALUES (%s, %s)
               ON CONFLICT (codigo_barras) DO UPDATE SET quantidade_romaneio = EXCLUDED.quantidade_romaneio''',
            (data['codigo_barras'], data['quantidade_romaneio'])
        )
    else:
        conn.execute(
            '''INSERT INTO romaneio (codigo_barras, quantidade_romaneio) VALUES (?, ?)
               ON CONFLICT(codigo_barras) DO UPDATE SET quantidade_romaneio = excluded.quantidade_romaneio''',
            (data['codigo_barras'], data['quantidade_romaneio'])
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

def _formatar_motivo_divergencia(raw):
    """Converte motivo salvo (texto livre ou JSON) para exibição legível."""
    if raw is None:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    if s.startswith('{'):
        try:
            o = json.loads(s)
            if isinstance(o, dict):
                parts = list(o.get('motivos') or [])
                obs = (o.get('observacao') or '').strip()
                if obs:
                    parts.append('Obs: ' + obs)
                return '; '.join(parts) if parts else ''
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    return s

def _carregar_motivos_divergencia(lista, conn=None):
    """Adiciona o campo 'motivo_divergencia' em cada item da lista (id_viagem + codigo_produto)."""
    if not lista:
        return lista
    pares = set()
    ids_viagem = set()
    for item in lista:
        vid = (item.get('id_viagem') or '').strip()
        cod = (item.get('codigo_produto') or '').strip()
        if vid and cod:
            id_norm = _normalizar_id_viagem(vid)
            pares.add((id_norm, cod))
            ids_viagem.add(id_norm)
    if not pares:
        for item in lista:
            item['motivo_divergencia'] = ''
        return lista
    own_conn = conn is None
    if own_conn:
        conn = get_db()
    motivos = {}
    try:
        if ids_viagem:
            ph = ','.join(['?'] * len(ids_viagem))
            params = list(ids_viagem)
            rows = conn.execute(
                f'SELECT id_viagem, codigo_produto, motivo FROM divergencia_motivo WHERE id_viagem IN ({ph})',
                tuple(params),
            ).fetchall()
            for r in rows or []:
                id_v = _normalizar_id_viagem(r.get('id_viagem') if hasattr(r, 'get') else r[0])
                cod = (r.get('codigo_produto') if hasattr(r, 'get') else r[1]) or ''
                cod = str(cod).strip()
                motivo = (r.get('motivo') if hasattr(r, 'get') else (r[2] if len(r) > 2 else '')) or ''
                motivos[(id_v, cod)] = str(motivo).strip()
    except Exception as ex:
        try:
            app.logger.warning('divergencia_motivo: %s', ex)
        except Exception:
            pass
    finally:
        if own_conn:
            try:
                conn.close()
            except Exception:
                pass
    for item in lista:
        vid = (item.get('id_viagem') or '').strip()
        cod = (item.get('codigo_produto') or '').strip()
        id_norm = _normalizar_id_viagem(vid) if vid else ''
        item['motivo_divergencia'] = motivos.get((id_norm, cod), '') if (vid and cod) else ''
    return lista


def _conferencia_enriquecer_meta_pg(conn, ds, id_para_lookup, id_busca, meta):
    """Preenche meta (roteiro, placa, motorista, responsáveis) em uma única consulta."""
    if not conn or not ds:
        return meta
    meta = dict(meta or {})
    id_t = str(id_para_lookup or id_busca or '').strip()
    if not id_t:
        return meta
    if meta.get('_extras_carregados'):
        return meta
    try:
        row = conn.execute(
            """SELECT ir.id_roteiro, ir.identificador_rota, vp.placa, vm.motorista,
                      vr.coordenador, vr.conferente, vr.ajudante1, vr.ajudante2
               FROM (SELECT ?::text AS id_v) q
               LEFT JOIN id_roteiros ir
                 ON ir.dataset_id = ? AND ir.id_viagem::text = q.id_v
               LEFT JOIN viagem_placa vp ON vp.id_viagem = q.id_v
               LEFT JOIN viagem_motorista vm ON vm.id_viagem = q.id_v
               LEFT JOIN viagem_responsaveis vr ON vr.id_viagem = q.id_v
               LIMIT 1""",
            (id_t, str(ds)),
        ).fetchone()
        if row:
            if not meta.get('id_roteiro'):
                meta['id_roteiro'] = str((row.get('id_roteiro') if hasattr(row, 'get') else (row[0] if len(row) > 0 else '')) or '').strip()
            if not meta.get('identificador_rota'):
                meta['identificador_rota'] = str((row.get('identificador_rota') if hasattr(row, 'get') else (row[1] if len(row) > 1 else '')) or '').strip()
            if not meta.get('placa'):
                meta['placa'] = str((row.get('placa') if hasattr(row, 'get') else (row[2] if len(row) > 2 else '')) or '').strip()
            if not meta.get('motorista'):
                meta['motorista'] = str((row.get('motorista') if hasattr(row, 'get') else (row[3] if len(row) > 3 else '')) or '').strip()
            coord = (row.get('coordenador') if hasattr(row, 'get') else (row[4] if len(row) > 4 else None) or '').strip()
            meta['coordenador'] = _coordenador_resposta_api(coord)
            meta['conferente'] = (row.get('conferente') if hasattr(row, 'get') else (row[5] if len(row) > 5 else None) or '').strip()
            meta['ajudante1'] = (row.get('ajudante1') if hasattr(row, 'get') else (row[6] if len(row) > 6 else None) or '').strip()
            meta['ajudante2'] = (row.get('ajudante2') if hasattr(row, 'get') else (row[7] if len(row) > 7 else None) or '').strip()
        else:
            meta.setdefault('coordenador', '')
            meta.setdefault('conferente', '')
            meta.setdefault('ajudante1', '')
            meta.setdefault('ajudante2', '')
    except Exception:
        meta.setdefault('coordenador', '')
        meta.setdefault('conferente', '')
        meta.setdefault('ajudante1', '')
        meta.setdefault('ajudante2', '')
    meta['_extras_carregados'] = True
    return meta


def _conferencia_lista_de_resposta(data):
    """Extrai itens da resposta JSON de get_conferencia (array legado ou objeto com lista)."""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        if data.get('erro'):
            return None
        lista = data.get('lista')
        if isinstance(lista, list):
            return lista
    return []


def _conferencia_lista_interna(id_viagem, fluxo='carregamento', conn=None, somente_divergencias=False):
    """Lista de itens da conferência sem passar pela resposta HTTP (uso interno)."""
    id_viagem = (id_viagem or '').strip()
    if not id_viagem:
        return []
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    fechar_conn = False
    if _usa_banco_para_dados():
        if conn is None:
            conn = get_db()
            fechar_conn = True
        try:
            if getattr(conn, 'kind', None) == 'pg':
                lista, _, _, err = _conferencia_build_lista_pg(
                    conn, id_viagem, fluxo, somente_divergencias=somente_divergencias,
                )
                if err:
                    return []
                return lista if isinstance(lista, list) else []
        finally:
            if fechar_conn:
                try:
                    conn.close()
                except Exception:
                    pass
    with app.test_request_context(query_string={'fluxo': fluxo}):
        result = get_conferencia(id_viagem)
    resp = result[0] if isinstance(result, tuple) else result
    status_code = result[1] if isinstance(result, tuple) and len(result) > 1 else 200
    if status_code and status_code >= 400:
        return []
    data = resp.get_json() if hasattr(resp, 'get_json') else None
    if isinstance(data, dict) and data.get('erro'):
        return []
    lista = _conferencia_lista_de_resposta(data)
    if somente_divergencias and isinstance(lista, list):
        lista = [
            it for it in lista
            if (it.get('quantidade_falta') or 0) > 0 or (it.get('quantidade_sobra') or 0) > 0
        ]
    return lista if isinstance(lista, list) else []


def _itens_divergentes_da_lista(lista, id_viagem):
    out = []
    for item in lista or []:
        if (item.get('quantidade_falta') or 0) > 0 or (item.get('quantidade_sobra') or 0) > 0:
            it = dict(item)
            it['id_viagem'] = id_viagem
            out.append(it)
    return out


def _ids_viagens_com_bipagem(conn, fluxo='carregamento', limit=80):
    """Viagens com bipagem, das mais recentes para as mais antigas."""
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    try:
        lim = max(1, min(200, int(limit)))
    except (TypeError, ValueError):
        lim = 80
    if getattr(conn, 'kind', None) == 'pg':
        sql = """
            SELECT TRIM(COALESCE(id_viagem::text, '')) AS id_viagem
            FROM produtos_bipados
            WHERE id_viagem IS NOT NULL AND TRIM(COALESCE(id_viagem::text, '')) != ''
              AND COALESCE(fluxo, 'carregamento') = ?
            GROUP BY TRIM(COALESCE(id_viagem::text, ''))
            ORDER BY MAX(data_hora) DESC
            LIMIT ?
        """
    else:
        sql = """
            SELECT id_viagem
            FROM produtos_bipados
            WHERE id_viagem IS NOT NULL AND trim(id_viagem) != ''
              AND COALESCE(fluxo, 'carregamento') = ?
            GROUP BY id_viagem
            ORDER BY MAX(data_hora) DESC
            LIMIT ?
        """
    rows = conn.execute(sql, (fluxo, lim)).fetchall()
    ids = []
    for row in rows or []:
        vid = (row.get('id_viagem') if hasattr(row, 'get') else row[0]) or ''
        vid = str(vid).strip()
        if vid:
            ids.append(vid)
    return ids


def _coletar_divergencias(fluxo='carregamento', id_viagem=None, limit_viagens=50):
    """Coleta itens divergentes (falta/sobra) de uma ou várias viagens."""
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    if id_viagem:
        id_v = (id_viagem or '').strip()
        lista = _conferencia_lista_interna(id_v, fluxo, somente_divergencias=True)
        todas = _itens_divergentes_da_lista(lista, id_v)
    else:
        conn = get_db()
        todas = []
        try:
            ids = _ids_viagens_com_bipagem(conn, fluxo, limit_viagens)
            for vid in ids:
                try:
                    lista = _conferencia_lista_interna(
                        vid, fluxo, conn=conn, somente_divergencias=True,
                    )
                    todas.extend(_itens_divergentes_da_lista(lista, vid))
                except Exception as ex:
                    try:
                        app.logger.warning('divergencia viagem %s: %s', vid, ex)
                    except Exception:
                        pass
        finally:
            try:
                conn.close()
            except Exception:
                pass
    todas.sort(key=lambda x: (str(x.get('id_viagem') or ''), str(x.get('produto') or '')))
    try:
        _carregar_motivos_divergencia(todas)
    except Exception as ex:
        try:
            app.logger.warning('motivos divergencia: %s', ex)
        except Exception:
            pass
    return todas


@app.route('/api/divergencias', methods=['GET'])
def get_divergencias():
    """Retorna itens com divergência na conferência: falta (faltar item) ou sobra (passar).
    Sem id_viagem: retorna divergências dos roteiros mais recentes (limite para evitar timeout).
    Com id_viagem: retorna apenas divergências daquele roteiro. Inclui motivo_divergencia."""
    try:
        id_viagem = request.args.get('id_viagem', '').strip()
        fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
        if fluxo not in ('carregamento', 'devolucao'):
            fluxo = 'carregamento'
        try:
            limit_v = int(request.args.get('limit_viagens', 50))
        except (TypeError, ValueError):
            limit_v = 50
        limit_v = max(1, min(120, limit_v))
        todas = _coletar_divergencias(
            fluxo=fluxo,
            id_viagem=id_viagem or None,
            limit_viagens=limit_v,
        )
        return jsonify(todas)
    except Exception as e:
        try:
            app.logger.exception('get_divergencias: %s', e)
        except Exception:
            pass
        return jsonify({'erro': 'Erro ao carregar divergências: %s' % str(e)}), 500


@app.route('/api/divergencias/motivo', methods=['PUT', 'PATCH', 'POST'])
def salvar_motivo_divergencia():
    """Salva o motivo da divergência para um item (id_viagem + codigo_produto). Body: id_viagem, codigo_produto, motivo."""
    data = request.get_json() or {}
    id_viagem = (data.get('id_viagem') or '').strip()
    codigo_produto = (data.get('codigo_produto') or '').strip()
    motivo = (data.get('motivo') or '').strip()
    if not id_viagem or not codigo_produto:
        return jsonify({'success': False, 'erro': 'id_viagem e codigo_produto são obrigatórios'}), 400
    id_norm = _normalizar_id_viagem(id_viagem)
    usuario = session.get('usuario', '')
    conn = get_db()
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''INSERT INTO divergencia_motivo (id_viagem, codigo_produto, motivo, registrado_por)
               VALUES (%s, %s, %s, %s)
               ON CONFLICT (id_viagem, codigo_produto) DO UPDATE SET motivo = EXCLUDED.motivo, registrado_por = EXCLUDED.registrado_por''',
            (id_norm, codigo_produto, motivo, usuario)
        )
    else:
        conn.execute(
            '''INSERT INTO divergencia_motivo (id_viagem, codigo_produto, motivo) VALUES (?, ?, ?)
               ON CONFLICT(id_viagem, codigo_produto) DO UPDATE SET motivo = excluded.motivo''',
            (id_norm, codigo_produto, motivo)
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'motivo': motivo})


def _get_lista_divergencias_todas(fluxo='carregamento'):
    """Retorna lista de itens divergentes dos roteiros recentes (exportação Excel). Inclui motivo_divergencia."""
    return _coletar_divergencias(fluxo=fluxo, limit_viagens=120)


def _get_viagem_info_dict(id_viagem):
    """Retorna dict com dados da viagem (para exportação)."""
    if not id_viagem:
        return {'data_expedicao': '', 'placa': '', 'identificador_rota': '', 'motorista': '',
                'coordenador': '', 'conferente': '', 'ajudante1': '', 'ajudante2': ''}
    info = _get_viagem_info_planilha(id_viagem) or {}
    info.setdefault('data_expedicao', '')
    info.setdefault('placa', '')
    info.setdefault('identificador_rota', '')
    info.setdefault('motorista', '')
    info.setdefault('coordenador', '')
    info.setdefault('conferente', '')
    info.setdefault('ajudante1', '')
    info.setdefault('ajudante2', '')
    id_norm = _normalizar_id_viagem(id_viagem)
    conn = get_db()
    row_p = conn.execute('SELECT placa FROM viagem_placa WHERE id_viagem = ?', (id_norm,)).fetchone()
    row_m = conn.execute('SELECT motorista FROM viagem_motorista WHERE id_viagem = ?', (id_norm,)).fetchone()
    row_r = conn.execute('SELECT coordenador, conferente, ajudante1, ajudante2 FROM viagem_responsaveis WHERE id_viagem = ?', (id_norm,)).fetchone()
    conn.close()
    if row_p and row_p[0]:
        info['placa'] = row_p[0].strip()
    if row_m and row_m[0]:
        info['motorista'] = row_m[0].strip()
    if row_r:
        info['coordenador'] = _coordenador_resposta_api((row_r[0] or '').strip())
        info['conferente'] = (row_r[1] or '').strip()
        info['ajudante1'] = (row_r[2] or '').strip()
        info['ajudante2'] = (row_r[3] or '').strip()
    return info


def _get_periodo_dict(id_viagem, fluxo='carregamento'):
    """Retorna dict com início e fim do carregamento."""
    resp = _periodo_viagem_resposta(id_viagem, fluxo)
    return {
        'inicio_carregamento': resp.get('inicio_carregamento') or '',
        'fim_carregamento': resp.get('fim_carregamento') or '',
    }


@app.route('/api/divergencias/excel', methods=['GET'])
def export_divergencias_excel():
    """Gera Excel: tipo=itens (só itens divergentes), tipo=roteiros (só dados dos roteiros), tipo=completo (página: roteiros + itens)."""
    from openpyxl import Workbook
    tipo = request.args.get('tipo', 'completo').strip().lower()
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    if tipo not in ('itens', 'roteiros', 'completo'):
        tipo = 'completo'
    data_exp_inicio = request.args.get('data_expedicao_inicio', '').strip()
    data_exp_fim = request.args.get('data_expedicao_fim', '').strip()
    divergencias = _get_lista_divergencias_todas(fluxo)
    ids_filtro = _get_id_viagens_por_data_expedicao(data_exp_inicio, data_exp_fim)
    if ids_filtro is not None:
        if not ids_filtro:
            divergencias = []
        else:
            divergencias = [d for d in divergencias if (d.get('id_viagem') or '').strip() in ids_filtro]
    ids_roteiros = sorted(set(d.get('id_viagem') or '' for d in divergencias if d.get('id_viagem')))
    wb = Workbook()
    header_font = Font(bold=True)
    if tipo == 'itens':
        ws = wb.active
        ws.title = 'Itens Divergentes'
        headers = ['ID Roteiro', 'Status', 'Código de Barras', 'Código do Produto', 'Produto', 'Qtd. Romaneio', 'Qtd. Bipada', 'Unidade', 'Peso Bruto', 'Qtd. Falta', 'Qtd. Sobra', 'Aviso', 'Motivo da divergência']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
        for row_idx, item in enumerate(divergencias, 2):
            status = item.get('status_bipado') or ''
            qtd_sobra = item.get('quantidade_sobra') or 0
            ws.cell(row=row_idx, column=1, value=item.get('id_viagem') or '')
            ws.cell(row=row_idx, column=2, value=status)
            ws.cell(row=row_idx, column=3, value=item.get('codigo_barras') or '')
            ws.cell(row=row_idx, column=4, value=item.get('codigo_produto') or '')
            ws.cell(row=row_idx, column=5, value=item.get('produto') or '')
            ws.cell(row=row_idx, column=6, value=item.get('quantidade_produto'))
            ws.cell(row=row_idx, column=7, value=item.get('quantidade_bipada'))
            ws.cell(row=row_idx, column=8, value=item.get('unidade') or '')
            ws.cell(row=row_idx, column=9, value=item.get('peso_bruto') or '')
            ws.cell(row=row_idx, column=10, value=item.get('quantidade_falta'))
            ws.cell(row=row_idx, column=11, value=qtd_sobra)
            ws.cell(row=row_idx, column=12, value=item.get('aviso_sobra') or '')
            ws.cell(row=row_idx, column=13, value=_formatar_motivo_divergencia(item.get('motivo_divergencia')))
        nome_arquivo = 'divergencias_itens.xlsx'
    elif tipo == 'roteiros':
        ws = wb.active
        ws.title = 'Dados dos Roteiros Divergentes'
        headers = ['ID Roteiro', 'Identificador da Rota', 'Data Expedição', 'Placa', 'Motorista', 'Início Carregamento', 'Fim Carregamento', 'Coordenador', 'Conferente', 'Auxiliar 1', 'Auxiliar 2']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
        for row_idx, vid in enumerate(ids_roteiros, 2):
            info = _get_viagem_info_dict(vid)
            periodo = _get_periodo_dict(vid)
            ws.cell(row=row_idx, column=1, value=vid)
            ws.cell(row=row_idx, column=2, value=info.get('identificador_rota') or '')
            ws.cell(row=row_idx, column=3, value=info.get('data_expedicao') or '')
            ws.cell(row=row_idx, column=4, value=info.get('placa') or '')
            ws.cell(row=row_idx, column=5, value=info.get('motorista') or '')
            ws.cell(row=row_idx, column=6, value=periodo.get('inicio_carregamento') or '')
            ws.cell(row=row_idx, column=7, value=periodo.get('fim_carregamento') or '')
            ws.cell(row=row_idx, column=8, value=info.get('coordenador') or '')
            ws.cell(row=row_idx, column=9, value=info.get('conferente') or '')
            ws.cell(row=row_idx, column=10, value=info.get('ajudante1') or '')
            ws.cell(row=row_idx, column=11, value=info.get('ajudante2') or '')
        nome_arquivo = 'divergencias_roteiros.xlsx'
    else:
        # completo: uma única aba com, para cada roteiro, DADOS DO ROTEIRO + ITENS DIVERGENTES (como na tela)
        ws = wb.active
        ws.title = 'Divergências - Página Completa'
        headers_rot = ['ID Roteiro', 'Identificador da Rota', 'Data Expedição', 'Placa', 'Motorista', 'Início', 'Fim', 'Coordenador', 'Conferente', 'Auxiliar 1', 'Auxiliar 2']
        headers_itens = ['Status', 'Código de Barras', 'Código do Produto', 'Produto', 'Qtd. Romaneio', 'Qtd. Bipada', 'Unidade', 'Peso Bruto', 'Aviso', 'Qtd. Falta', 'Qtd. Sobra', 'Motivo da divergência']
        linha = 1
        for vid in ids_roteiros:
            itens_roteiro = [d for d in divergencias if (d.get('id_viagem') or '') == vid]
            # Título do roteiro
            ws.cell(row=linha, column=1, value='Roteiro: ' + str(vid))
            ws.cell(row=linha, column=1).font = Font(bold=True, size=12)
            linha += 1
            # DADOS DO ROTEIRO
            ws.cell(row=linha, column=1, value='DADOS DO ROTEIRO')
            ws.cell(row=linha, column=1).font = header_font
            linha += 1
            for col, h in enumerate(headers_rot, 1):
                c = ws.cell(row=linha, column=col, value=h)
                c.font = header_font
            linha += 1
            info = _get_viagem_info_dict(vid)
            periodo = _get_periodo_dict(vid)
            ws.cell(row=linha, column=1, value=vid)
            ws.cell(row=linha, column=2, value=info.get('identificador_rota') or '')
            ws.cell(row=linha, column=3, value=info.get('data_expedicao') or '')
            ws.cell(row=linha, column=4, value=info.get('placa') or '')
            ws.cell(row=linha, column=5, value=info.get('motorista') or '')
            ws.cell(row=linha, column=6, value=periodo.get('inicio_carregamento') or '')
            ws.cell(row=linha, column=7, value=periodo.get('fim_carregamento') or '')
            ws.cell(row=linha, column=8, value=info.get('coordenador') or '')
            ws.cell(row=linha, column=9, value=info.get('conferente') or '')
            ws.cell(row=linha, column=10, value=info.get('ajudante1') or '')
            ws.cell(row=linha, column=11, value=info.get('ajudante2') or '')
            linha += 2  # linha em branco antes dos itens
            # ITENS DIVERGENTES
            ws.cell(row=linha, column=1, value='ITENS DIVERGENTES')
            ws.cell(row=linha, column=1).font = header_font
            linha += 1
            for col, h in enumerate(headers_itens, 1):
                c = ws.cell(row=linha, column=col, value=h)
                c.font = header_font
            linha += 1
            for item in itens_roteiro:
                ws.cell(row=linha, column=1, value=item.get('status_bipado') or '')
                ws.cell(row=linha, column=2, value=item.get('codigo_barras') or '')
                ws.cell(row=linha, column=3, value=item.get('codigo_produto') or '')
                ws.cell(row=linha, column=4, value=item.get('produto') or '')
                ws.cell(row=linha, column=5, value=item.get('quantidade_produto'))
                ws.cell(row=linha, column=6, value=item.get('quantidade_bipada'))
                ws.cell(row=linha, column=7, value=item.get('unidade') or '')
                ws.cell(row=linha, column=8, value=item.get('peso_bruto') or '')
                ws.cell(row=linha, column=9, value=item.get('aviso_sobra') or '')
                ws.cell(row=linha, column=10, value=item.get('quantidade_falta'))
                ws.cell(row=linha, column=11, value=item.get('quantidade_sobra') or 0)
                ws.cell(row=linha, column=12, value=_formatar_motivo_divergencia(item.get('motivo_divergencia')))
                linha += 1
            linha += 2  # linhas em branco entre roteiros
        nome_arquivo = 'divergencias_pagina_completa.xlsx'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=nome_arquivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _get_mapa_motivos_divergencia():
    """Retorna dict (id_viagem_norm, codigo_produto) -> motivo para uso nos relatórios."""
    conn = get_db()
    rows = conn.execute('SELECT id_viagem, codigo_produto, motivo FROM divergencia_motivo').fetchall()
    conn.close()
    mapa = {}
    for r in rows:
        vid = _normalizar_id_viagem(r['id_viagem'] or '')
        cod = (r['codigo_produto'] or '').strip()
        if vid and cod:
            mapa[(vid, cod)] = (r['motivo'] or '').strip()
    return mapa


@app.route('/api/relatorios/excel/bipados', methods=['GET'])
def export_relatorio_bipados():
    """Gera Excel com todos os registros bipados (tudo que foi bipado). Conferente = do roteiro. Inclui Motivo da divergência."""
    from openpyxl import Workbook
    data_inicio = (request.args.get('data_expedicao_inicio') or request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_expedicao_fim') or request.args.get('data_fim') or '').strip()
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    ids_filtro = _get_id_viagens_por_data_expedicao(data_inicio, data_fim)
    conn = get_db()
    sql = '''
        SELECT p.*, r.conferente as conferente_roteiro
        FROM produtos_bipados p
        LEFT JOIN viagem_responsaveis r ON p.id_viagem = r.id_viagem
    '''
    params = []
    conditions = ["COALESCE(p.fluxo, 'carregamento') = ?"]
    params = [fluxo]
    sql += ' WHERE ' + ' AND '.join(conditions)
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='p.id_viagem')
    sql += ' ORDER BY p.data_hora DESC, p.id DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    rows = [dict(r) for r in rows]
    mapa_motivos = _get_mapa_motivos_divergencia()
    for row in rows:
        vid = _normalizar_id_viagem(row.get('id_viagem') or '')
        cod = (row.get('codigo_interno') or row.get('codigo_barras') or '').strip()
        row['motivo_divergencia'] = mapa_motivos.get((vid, cod), '') or mapa_motivos.get((vid, (row.get('codigo_barras') or '').strip()), '')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tudo que foi bipado'
    header_font = Font(bold=True)
    col_map = [
        ('id', 'ID'),
        ('id_viagem', 'ID Roteiro'),
        ('codigo_barras', 'Código de Barras'),
        ('codigo_interno', 'Código Interno'),
        ('produto', 'Produto'),
        ('quantidade', 'Quantidade'),
        ('data_hora', 'Data/Hora'),
        ('veiculo', 'Veículo'),
        ('conferente_roteiro', 'Conferente (roteiro)'),
        ('motivo_divergencia', 'Motivo da divergência'),
        ('status', 'Status'),
        ('doca', 'Doca'),
        ('codigo_dun', 'Código DUN'),
        ('unidade', 'Unidade'),
        ('peso', 'Peso'),
    ]
    keys_disponiveis = list(rows[0].keys()) if rows else [c[0] for c in col_map]
    cols_in_order = [(k, label) for k, label in col_map if k in keys_disponiveis]
    if not cols_in_order:
        cols_in_order = [(k, label) for k, label in col_map]
    for col_idx, (_, label) in enumerate(cols_in_order, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(cols_in_order, 1):
            val = row.get(key)
            if key == 'motivo_divergencia':
                val = _formatar_motivo_divergencia(val)
            ws.cell(row=row_idx, column=col_idx, value=val)
    nome_arquivo = 'relatorio_tudo_que_foi_bipado.xlsx'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=nome_arquivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _relatorio_bipados_periodo(data_inicio, data_fim, data_expedicao_inicio=None, data_expedicao_fim=None):
    """Gera workbook com bipados filtrados por período (data_hora) e opcionalmente por data de expedição."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = '''
        SELECT p.*, r.conferente as conferente_roteiro
        FROM produtos_bipados p
        LEFT JOIN viagem_responsaveis r ON p.id_viagem = r.id_viagem
    '''
    conditions, params = [], []
    if data_inicio and data_fim:
        d0 = (data_inicio.strip() + ' 00:00:00') if len(data_inicio.strip()) <= 10 else data_inicio.strip()
        d1 = (data_fim.strip() + ' 23:59:59') if len(data_fim.strip()) <= 10 else data_fim.strip()
        conditions.append('p.data_hora >= ? AND p.data_hora <= ?')
        params.extend([d0, d1])
    if conditions:
        sql += ' WHERE ' + ' AND '.join(conditions)
    else:
        sql += ' WHERE 1=1'
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='p.id_viagem')
    sql += ' ORDER BY p.data_hora DESC, p.id DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    rows = [dict(r) for r in rows]
    mapa_motivos = _get_mapa_motivos_divergencia()
    for row in rows:
        vid = _normalizar_id_viagem(row.get('id_viagem') or '')
        cod = (row.get('codigo_interno') or row.get('codigo_barras') or '').strip()
        row['motivo_divergencia'] = mapa_motivos.get((vid, cod), '') or mapa_motivos.get((vid, (row.get('codigo_barras') or '').strip()), '')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bipados no período'
    header_font = Font(bold=True)
    col_map = [
        ('id', 'ID'), ('id_viagem', 'ID Roteiro'), ('codigo_barras', 'Código de Barras'), ('codigo_interno', 'Código Interno'),
        ('produto', 'Produto'), ('quantidade', 'Quantidade'), ('data_hora', 'Data/Hora'), ('veiculo', 'Veículo'),
        ('conferente_roteiro', 'Conferente (roteiro)'), ('motivo_divergencia', 'Motivo da divergência'), ('status', 'Status'), ('doca', 'Doca'), ('codigo_dun', 'Código DUN'), ('unidade', 'Unidade'), ('peso', 'Peso'),
    ]
    keys = list(rows[0].keys()) if rows else [c[0] for c in col_map]
    cols = [(k, lb) for k, lb in col_map if k in keys]
    if not cols:
        cols = [(k, k) for k, lb in col_map]
    for col_idx, (_, lb) in enumerate(cols, 1):
        ws.cell(row=1, column=col_idx, value=lb).font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(cols, 1):
            val = row.get(key)
            if key == 'motivo_divergencia':
                val = _formatar_motivo_divergencia(val)
            ws.cell(row=row_idx, column=col_idx, value=val)
    return wb


def _relatorio_resumo_roteiro(data_expedicao_inicio=None, data_expedicao_fim=None, fluxo='carregamento'):
    """Uma linha por viagem: ID Roteiro, Placa, Motorista, Data expedição, Início/Fim, Duração (min), Total itens, Faltas, Responsáveis."""
    from openpyxl import Workbook
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = '''
        SELECT id_viagem, SUM(quantidade) as total_bipados, MIN(data_hora) as inicio, MAX(data_hora) as fim
        FROM produtos_bipados WHERE id_viagem IS NOT NULL AND trim(id_viagem) != '' AND COALESCE(fluxo, 'carregamento') = ?
    '''
    params = [fluxo]
    sql += ' GROUP BY id_viagem ORDER BY MAX(data_hora) DESC'
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    viagens = []
    for r in rows:
        inicio = r['inicio'] or ''
        fim = r['fim'] or ''
        d_min = None
        if inicio and fim:
            t0, t1 = _parse_datetime(inicio), _parse_datetime(fim)
            if t0 and t1:
                d_min = max(0, int((t1 - t0).total_seconds() / 60))
        viagens.append({
            'id_viagem': r['id_viagem'], 'total_bipados': r['total_bipados'] or 0,
            'inicio': inicio, 'fim': fim, 'duracao_minutos': d_min
        })
    total_faltas_map = {}
    for v in viagens[:300]:
        try:
            with app.test_request_context(query_string={'fluxo': fluxo, 'limit': '2000'}):
                ret = get_conferencia(v['id_viagem'])
                resp = ret[0] if isinstance(ret, tuple) else ret
                data = resp.get_json() if hasattr(resp, 'get_json') else None
                lista_conf = _conferencia_lista_de_resposta(data)
                total_faltas_map[v['id_viagem']] = sum((item.get('quantidade_falta') or 0) for item in (lista_conf or []))
        except Exception:
            total_faltas_map[v['id_viagem']] = 0
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumo por Roteiro'
    header_font = Font(bold=True)
    headers = ['ID Roteiro', 'Placa', 'Motorista', 'Data Expedição', 'Início Carregamento', 'Fim Carregamento', 'Duração (min)', 'Total Itens Bipados', 'Faltas', 'Coordenador', 'Conferente', 'Auxiliar 1', 'Auxiliar 2']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, v in enumerate(viagens, 2):
        info = _get_viagem_info_dict(v.get('id_viagem') or '')
        periodo = _get_periodo_dict(v.get('id_viagem') or '', fluxo)
        faltas = total_faltas_map.get(v['id_viagem'], 0)
        ws.cell(row=row_idx, column=1, value=v.get('id_viagem') or '')
        ws.cell(row=row_idx, column=2, value=info.get('placa') or '')
        ws.cell(row=row_idx, column=3, value=info.get('motorista') or '')
        ws.cell(row=row_idx, column=4, value=info.get('data_expedicao') or '')
        ws.cell(row=row_idx, column=5, value=periodo.get('inicio_carregamento') or '')
        ws.cell(row=row_idx, column=6, value=periodo.get('fim_carregamento') or '')
        ws.cell(row=row_idx, column=7, value=v.get('duracao_minutos'))
        ws.cell(row=row_idx, column=8, value=v.get('total_bipados', 0))
        ws.cell(row=row_idx, column=9, value=faltas)
        ws.cell(row=row_idx, column=10, value=info.get('coordenador') or '')
        ws.cell(row=row_idx, column=11, value=info.get('conferente') or '')
        ws.cell(row=row_idx, column=12, value=info.get('ajudante1') or '')
        ws.cell(row=row_idx, column=13, value=info.get('ajudante2') or '')
    return wb


def _relatorio_tempo_placa(data_expedicao_inicio=None, data_expedicao_fim=None):
    """Tempo de carregamento por placa (total minutos e quantidade de viagens)."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = '''
        SELECT id_viagem, SUM(quantidade) as total_bipados, MIN(data_hora) as inicio, MAX(data_hora) as fim
        FROM produtos_bipados WHERE id_viagem IS NOT NULL AND trim(id_viagem) != ''
    '''
    params = []
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    sql += ' GROUP BY id_viagem'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    placa_minutos = {}
    placa_viagens = {}
    for r in rows:
        inicio, fim = r['inicio'] or '', r['fim'] or ''
        d_min = None
        if inicio and fim:
            t0, t1 = _parse_datetime(inicio), _parse_datetime(fim)
            if t0 and t1:
                d_min = max(0, int((t1 - t0).total_seconds() / 60))
        info = _get_viagem_info_dict(r['id_viagem'] or '')
        placa = (info.get('placa') or '').strip() or 'Sem placa'
        placa_minutos[placa] = placa_minutos.get(placa, 0) + (d_min or 0)
        placa_viagens[placa] = placa_viagens.get(placa, 0) + 1
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tempo por Placa'
    header_font = Font(bold=True)
    for col, h in enumerate(['Placa', 'Total Minutos', 'Qtd. Viagens'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, (placa, minutos) in enumerate(sorted(placa_minutos.items(), key=lambda x: -x[1]), 2):
        ws.cell(row=row_idx, column=1, value=placa)
        ws.cell(row=row_idx, column=2, value=minutos)
        ws.cell(row=row_idx, column=3, value=placa_viagens.get(placa, 0))
    return wb


def _relatorio_itens_por_roteiro(data_expedicao_inicio=None, data_expedicao_fim=None):
    """Resumo: por roteiro e produto, quantidade bipada (uma linha por id_viagem + codigo_barras). Inclui conferente e motivo da divergência."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = '''
        SELECT id_viagem, codigo_barras, MAX(codigo_interno) as codigo_interno, produto, SUM(quantidade) as total
        FROM produtos_bipados WHERE id_viagem IS NOT NULL AND trim(id_viagem) != ''
    '''
    params = []
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    sql += ' GROUP BY id_viagem, codigo_barras ORDER BY id_viagem, total DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    vids = sorted(set((r['id_viagem'] or '').strip() for r in rows if (r['id_viagem'] or '').strip()))
    conferente_por_viagem = {vid: (_get_viagem_info_dict(vid).get('conferente') or '').strip() or '' for vid in vids}
    mapa_motivos = _get_mapa_motivos_divergencia()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Itens por Roteiro'
    header_font = Font(bold=True)
    for col, h in enumerate(['ID Roteiro', 'Conferente', 'Motivo da divergência', 'Código de Barras', 'Produto', 'Quantidade Bipada'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, r in enumerate(rows, 2):
        vid = (r['id_viagem'] or '').strip()
        vid_norm = _normalizar_id_viagem(vid)
        cod_prod = (r.get('codigo_interno') or r.get('codigo_barras') or '').strip()
        motivo = mapa_motivos.get((vid_norm, cod_prod), '') or mapa_motivos.get((vid_norm, (r.get('codigo_barras') or '').strip()), '')
        ws.cell(row=row_idx, column=1, value=vid)
        ws.cell(row=row_idx, column=2, value=conferente_por_viagem.get(vid, ''))
        ws.cell(row=row_idx, column=3, value=motivo)
        ws.cell(row=row_idx, column=4, value=r['codigo_barras'] or '')
        ws.cell(row=row_idx, column=5, value=r['produto'] or '')
        ws.cell(row=row_idx, column=6, value=r['total'] or 0)
    return wb


def _relatorio_itens_mais_bipados(data_expedicao_inicio=None, data_expedicao_fim=None):
    """Ranking de produtos por quantidade total bipada."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = 'SELECT codigo_barras, produto, SUM(quantidade) as total, COUNT(DISTINCT id_viagem) as num_viagens FROM produtos_bipados WHERE 1=1'
    params = []
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    sql += ' GROUP BY codigo_barras ORDER BY total DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Itens mais bipados'
    header_font = Font(bold=True)
    for col, h in enumerate(['Código de Barras', 'Produto', 'Total Bipado', 'Qtd. Viagens'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r['codigo_barras'] or '')
        ws.cell(row=row_idx, column=2, value=(r['produto'] or '')[:80])
        ws.cell(row=row_idx, column=3, value=r['total'] or 0)
        ws.cell(row=row_idx, column=4, value=r['num_viagens'] or 0)
    return wb


def _relatorio_resumo_produto(data_expedicao_inicio=None, data_expedicao_fim=None, fluxo='carregamento'):
    """Por produto: total bipado, em quantas viagens, primeira e última data."""
    from openpyxl import Workbook
    fluxo = (fluxo or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = '''SELECT codigo_barras, produto, SUM(quantidade) as total,
               COUNT(DISTINCT id_viagem) as num_viagens, MIN(data_hora) as primeira, MAX(data_hora) as ultima
        FROM produtos_bipados WHERE COALESCE(fluxo, 'carregamento') = ?'''
    params = [fluxo]
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    sql += ' GROUP BY codigo_barras ORDER BY total DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumo por Produto'
    header_font = Font(bold=True)
    for col, h in enumerate(['Código de Barras', 'Produto', 'Total Bipado', 'Qtd. Viagens', 'Primeira Data', 'Última Data'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r['codigo_barras'] or '')
        ws.cell(row=row_idx, column=2, value=(r['produto'] or '')[:80])
        ws.cell(row=row_idx, column=3, value=r['total'] or 0)
        ws.cell(row=row_idx, column=4, value=r['num_viagens'] or 0)
        ws.cell(row=row_idx, column=5, value=r['primeira'] or '')
        ws.cell(row=row_idx, column=6, value=r['ultima'] or '')
    return wb


def _relatorio_roteiros_divergencia(data_expedicao_inicio=None, data_expedicao_fim=None, fluxo='carregamento'):
    """Roteiros que têm divergência, com totais de faltas e sobras. Inclui conferente do roteiro."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    divergencias = _get_lista_divergencias_todas(fluxo)
    if ids_filtro is not None:
        if len(ids_filtro) == 0:
            divergencias = []
        else:
            divergencias = [d for d in divergencias if (d.get('id_viagem') or '').strip() in ids_filtro]
    by_roteiro = {}
    for d in divergencias:
        vid = d.get('id_viagem') or ''
        if vid not in by_roteiro:
            by_roteiro[vid] = {'faltas': 0, 'sobras': 0}
        by_roteiro[vid]['faltas'] += (d.get('quantidade_falta') or 0)
        by_roteiro[vid]['sobras'] += (d.get('quantidade_sobra') or 0)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Roteiros com Divergência'
    header_font = Font(bold=True)
    for col, h in enumerate(['ID Roteiro', 'Conferente', 'Total Faltas', 'Total Sobras'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, (vid, tot) in enumerate(sorted(by_roteiro.items()), 2):
        info = _get_viagem_info_dict(vid)
        conferente = (info.get('conferente') or '').strip() or ''
        ws.cell(row=row_idx, column=1, value=vid)
        ws.cell(row=row_idx, column=2, value=conferente)
        ws.cell(row=row_idx, column=3, value=tot['faltas'])
        ws.cell(row=row_idx, column=4, value=tot['sobras'])
    return wb


def _relatorio_carregamento_veiculo(data_expedicao_inicio=None, data_expedicao_fim=None):
    """Por veículo (placa/veículo no bipado): total itens e quantidade de viagens."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = '''
        SELECT veiculo, SUM(quantidade) as total, COUNT(DISTINCT id_viagem) as num_viagens
        FROM produtos_bipados WHERE veiculo IS NOT NULL AND trim(veiculo) != ''
    '''
    params = []
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    sql += ' GROUP BY veiculo ORDER BY total DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Carregamento por Veículo'
    header_font = Font(bold=True)
    for col, h in enumerate(['Veículo', 'Total Itens', 'Qtd. Viagens'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r['veiculo'] or '')
        ws.cell(row=row_idx, column=2, value=r['total'] or 0)
        ws.cell(row=row_idx, column=3, value=r['num_viagens'] or 0)
    return wb


def _relatorio_peso_viagem_placa(data_expedicao_inicio=None, data_expedicao_fim=None):
    """Peso carregado por viagem e por placa. Peso unitário: do banco (base_codigo_barras) ou planilha."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = 'SELECT id_viagem, codigo_barras, quantidade FROM produtos_bipados WHERE 1=1'
    params = []
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    rows = conn.execute(sql, params).fetchall()
    mapa_peso = {}
    mapa_barras_codigo = {}
    if _usa_banco_para_dados() and getattr(conn, 'kind', None) == 'pg':
        try:
            ds = _get_latest_dataset_id(conn)
            if ds:
                base_rows = conn.execute("SELECT codigo_interno, ean, dun, peso FROM base_codigo_barras WHERE dataset_id = ?", (str(ds),)).fetchall()
                for r in base_rows:
                    ci = (r.get('codigo_interno') or r[0] or '').strip()
                    ean = (r.get('ean') or r[1] or '').strip()
                    dun = (r.get('dun') or r[2] or '').strip()
                    try:
                        p = float(r.get('peso') or r[3] or 0)
                    except (TypeError, ValueError):
                        p = 0
                    if ci:
                        mapa_peso[ci] = p
                    if ean:
                        mapa_barras_codigo[ean] = ci
                    if dun:
                        mapa_barras_codigo[dun] = ci
        except Exception:
            pass
    conn.close()
    if not mapa_peso and not mapa_barras_codigo:
        caminho_planilha = encontrar_planilha()
    if caminho_planilha:
        try:
            wb_plan = openpyxl.load_workbook(caminho_planilha, data_only=True)
            mapa_peso = _build_mapa_peso_romaneio(wb_plan)
            mapa_barras_codigo = _build_mapa_barras_to_codigo_produto(wb_plan)
            wb_plan.close()
        except Exception:
            pass
    peso_por_viagem = {}
    peso_por_placa = {}
    for r in rows:
        vid = (r['id_viagem'] or '').strip()
        codigo_barras = (r['codigo_barras'] or '').strip()
        qtd = int(r['quantidade'] or 0)
        codigo_produto = mapa_barras_codigo.get(codigo_barras) or codigo_barras
        peso_unit = mapa_peso.get(codigo_produto, 0) or mapa_peso.get(codigo_barras, 0)
        p = qtd * peso_unit
        if vid:
            peso_por_viagem[vid] = peso_por_viagem.get(vid, 0.0) + p
        info = _get_viagem_info_dict(vid)
        placa = (info.get('placa') or '').strip() or 'Sem placa'
        peso_por_placa[placa] = peso_por_placa.get(placa, 0.0) + p
    wb = Workbook()
    header_font = Font(bold=True)
    ws1 = wb.active
    ws1.title = 'Peso por Viagem'
    for col, h in enumerate(['ID Roteiro', 'Conferente', 'Peso Total (kg)'], 1):
        ws1.cell(row=1, column=col, value=h).font = header_font
    for row_idx, (vid, p) in enumerate(sorted(peso_por_viagem.items(), key=lambda x: -x[1]), 2):
        info = _get_viagem_info_dict(vid)
        conferente = (info.get('conferente') or '').strip() or ''
        ws1.cell(row=row_idx, column=1, value=vid)
        ws1.cell(row=row_idx, column=2, value=conferente)
        ws1.cell(row=row_idx, column=3, value=round(p, 2))
    ws2 = wb.create_sheet('Peso por Placa')
    for col, h in enumerate(['Placa', 'Peso Total (kg)'], 1):
        ws2.cell(row=1, column=col, value=h).font = header_font
    for row_idx, (placa, p) in enumerate(sorted(peso_por_placa.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=row_idx, column=1, value=placa)
        ws2.cell(row=row_idx, column=2, value=round(p, 2))
    return wb


def _relatorio_responsaveis_viagem(data_expedicao_inicio=None, data_expedicao_fim=None):
    """Por roteiro: coordenador, conferente, auxiliares 1 e 2."""
    from openpyxl import Workbook
    ids_filtro = _get_id_viagens_por_data_expedicao(data_expedicao_inicio or '', data_expedicao_fim or '')
    conn = get_db()
    sql = "SELECT DISTINCT id_viagem FROM produtos_bipados WHERE id_viagem IS NOT NULL AND trim(id_viagem) != ''"
    params = []
    sql, params = _aplicar_filtro_expedicao_sql(sql, params, ids_filtro, col='id_viagem')
    sql += ' ORDER BY id_viagem'
    ids = conn.execute(sql, params).fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Responsáveis por Viagem'
    header_font = Font(bold=True)
    for col, h in enumerate(['ID Roteiro', 'Coordenador', 'Conferente', 'Auxiliar 1', 'Auxiliar 2'], 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, (row,) in enumerate(ids, 2):
        info = _get_viagem_info_dict(row or '')
        ws.cell(row=row_idx, column=1, value=row or '')
        ws.cell(row=row_idx, column=2, value=info.get('coordenador') or '')
        ws.cell(row=row_idx, column=3, value=info.get('conferente') or '')
        ws.cell(row=row_idx, column=4, value=info.get('ajudante1') or '')
        ws.cell(row=row_idx, column=5, value=info.get('ajudante2') or '')
    return wb


@app.route('/api/relatorios/excel/bipados_periodo', methods=['GET'])
def export_relatorio_bipados_periodo():
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    data_exp_inicio = request.args.get('data_expedicao_inicio', '').strip()
    data_exp_fim = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_bipados_periodo(data_inicio, data_fim, data_exp_inicio, data_exp_fim)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_bipados_periodo.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/resumo_roteiro', methods=['GET'])
def export_relatorio_resumo_roteiro():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    wb = _relatorio_resumo_roteiro(de, ate, fluxo)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_resumo_por_roteiro.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/tempo_placa', methods=['GET'])
def export_relatorio_tempo_placa():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_tempo_placa(de, ate)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_tempo_por_placa.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/itens_por_roteiro', methods=['GET'])
def export_relatorio_itens_por_roteiro():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_itens_por_roteiro(de, ate)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_itens_por_roteiro.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/itens_mais_bipados', methods=['GET'])
def export_relatorio_itens_mais_bipados():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_itens_mais_bipados(de, ate)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_itens_mais_bipados.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/resumo_produto', methods=['GET'])
def export_relatorio_resumo_produto():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    wb = _relatorio_resumo_produto(de, ate, fluxo)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_resumo_por_produto.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/roteiros_divergencia', methods=['GET'])
def export_relatorio_roteiros_divergencia():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    wb = _relatorio_roteiros_divergencia(de, ate, fluxo)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_roteiros_com_divergencia.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/carregamento_veiculo', methods=['GET'])
def export_relatorio_carregamento_veiculo():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_carregamento_veiculo(de, ate)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_carregamento_por_veiculo.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/peso_viagem_placa', methods=['GET'])
def export_relatorio_peso_viagem_placa():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_peso_viagem_placa(de, ate)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_peso_por_viagem_placa.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/responsaveis_viagem', methods=['GET'])
def export_relatorio_responsaveis_viagem():
    de = request.args.get('data_expedicao_inicio', '').strip()
    ate = request.args.get('data_expedicao_fim', '').strip()
    wb = _relatorio_responsaveis_viagem(de, ate)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_responsaveis_por_viagem.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _relatorio_romaneio_guarulhos():
    """Gera workbook com os dados do Romaneio CD Guarulhos: Resumo, Quantidade por item, Peso por carro."""
    from openpyxl import Workbook
    wb_plan, _ = get_workbook_cached()
    if not wb_plan:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Resumo'
        ws.cell(row=1, column=1, value='Planilha não encontrada. Carregue a planilha e tente novamente.')
        return wb
    try:
        stats = _estatisticas_romaneio_por_item(wb_plan)
    except Exception:
        stats = {}
    wb = Workbook()
    header_font = Font(bold=True)
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'
    ws_resumo.cell(row=1, column=1, value='Romaneio CD Guarulhos Ultrapão (Distribuidora) - Resumo').font = Font(bold=True, size=12)
    ws_resumo.cell(row=2, column=1, value='Indicador')
    ws_resumo.cell(row=2, column=2, value='Valor')
    ws_resumo.cell(row=2, column=1).font = header_font
    ws_resumo.cell(row=2, column=2).font = header_font
    ws_resumo.cell(row=3, column=1, value='Quantidade de roteiros')
    ws_resumo.cell(row=3, column=2, value=stats.get('qtd_roteiros', 0))
    ws_resumo.cell(row=4, column=1, value='Quantidade de veículos')
    ws_resumo.cell(row=4, column=2, value=stats.get('qtd_veiculos', 0))
    ws_resumo.cell(row=5, column=1, value='Quantidade total de itens (soma)')
    ws_resumo.cell(row=5, column=2, value=stats.get('quantidade_total_itens', 0))
    ws_resumo.cell(row=6, column=1, value='Peso total (kg)')
    ws_resumo.cell(row=6, column=2, value=stats.get('peso_total_geral', 0))
    ws_itens = wb.create_sheet('Quantidade por item')
    ws_itens.cell(row=1, column=1, value='Código do produto').font = header_font
    ws_itens.cell(row=1, column=2, value='Descrição do produto').font = header_font
    ws_itens.cell(row=1, column=3, value='Quantidade total').font = header_font
    itens = stats.get('itens_total_por_codigo') or {}
    descricoes = stats.get('itens_descricao_por_codigo') or {}
    for row_idx, (cod, qtd) in enumerate(sorted(itens.items(), key=lambda x: -x[1]), 2):
        ws_itens.cell(row=row_idx, column=1, value=cod)
        ws_itens.cell(row=row_idx, column=2, value=descricoes.get(cod, '') or '')
        ws_itens.cell(row=row_idx, column=3, value=qtd)
    ws_peso = wb.create_sheet('Peso por carro')
    ws_peso.cell(row=1, column=1, value='Placa / Veículo').font = header_font
    ws_peso.cell(row=1, column=2, value='Peso total (kg)').font = header_font
    peso_carro = stats.get('peso_por_carro') or {}
    for row_idx, (placa, peso) in enumerate(peso_carro.items(), 2):
        ws_peso.cell(row=row_idx, column=1, value=placa)
        ws_peso.cell(row=row_idx, column=2, value=peso)
    return wb


@app.route('/api/relatorios/excel/romaneio_guarulhos', methods=['GET'])
def export_relatorio_romaneio_guarulhos():
    wb = _relatorio_romaneio_guarulhos()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='relatorio_romaneio_cd_guarulhos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/excel/extrato', methods=['GET'])
def export_relatorio_extrato_excel():
    """Gera Excel com o extrato (comprovante) da carga de um roteiro: mesmas colunas da aba Extrato."""
    from openpyxl import Workbook
    id_viagem = (request.args.get('id_viagem') or '').strip()
    fluxo = (request.args.get('fluxo') or 'carregamento').strip().lower()
    if fluxo not in ('carregamento', 'devolucao'):
        fluxo = 'carregamento'
    if not id_viagem:
        return jsonify({'erro': 'Informe o ID do roteiro (id_viagem)'}), 400
    try:
        with app.test_request_context(query_string={'fluxo': fluxo, 'limit': '2000'}):
            result = get_conferencia(id_viagem)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500
    resp = result[0] if isinstance(result, tuple) else result
    status = result[1] if isinstance(result, tuple) and len(result) > 1 else 200
    data = resp.get_json() if hasattr(resp, 'get_json') else None
    if isinstance(data, dict) and data.get('erro'):
        return jsonify(data), status if status != 200 else 400
    itens = _conferencia_lista_de_resposta(data)
    if itens is None:
        err = (data or {}).get('erro') if isinstance(data, dict) else 'Erro ao carregar conferência'
        return jsonify({'erro': err or 'Erro ao carregar conferência'}), 400
    if not itens:
        return jsonify({'erro': 'Nenhum item encontrado para este roteiro'}), 404
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extrato'
    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value='Extrato - ID Roteiro: ' + id_viagem).font = Font(bold=True, size=12)
    headers = ['Status', 'Motivo da divergência', 'Código de Barras', 'Código do Produto', 'Produto', 'Qtd. Produto', 'Qtd. Bipada', 'Unidade', 'Peso Bruto', 'Aviso', 'Qtd. Falta']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
    for row_idx, item in enumerate(itens, 4):
        ws.cell(row=row_idx, column=1, value=item.get('status_bipado') or '')
        ws.cell(row=row_idx, column=2, value=_formatar_motivo_divergencia(item.get('motivo_divergencia')))
        ws.cell(row=row_idx, column=3, value=item.get('codigo_barras') or '')
        ws.cell(row=row_idx, column=4, value=item.get('codigo_produto') or '')
        ws.cell(row=row_idx, column=5, value=item.get('produto') or '')
        ws.cell(row=row_idx, column=6, value=item.get('quantidade_produto'))
        ws.cell(row=row_idx, column=7, value=item.get('quantidade_bipada'))
        ws.cell(row=row_idx, column=8, value=item.get('unidade') or '')
        ws.cell(row=row_idx, column=9, value=item.get('peso_bruto') or '')
        ws.cell(row=row_idx, column=10, value=item.get('aviso_sobra') or '')
        ws.cell(row=row_idx, column=11, value=item.get('quantidade_falta'))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = 'extrato_roteiro_{}.xlsx'.format(id_viagem.replace('/', '_').replace('\\', '_')[:50])
    return send_file(buf, as_attachment=True, download_name=nome_arquivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def allowed_file(filename):
    """Verifica se o arquivo tem extensão permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_SPREADSHEET_EXTENSIONS


def allowed_xml_file(filename):
    """Verifica se o arquivo XML tem extensão permitida."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_XML_EXTENSIONS

def encontrar_planilha():
    """Encontra a planilha Excel no diretório do app, na pasta do exe (quando instalado) ou no diretório atual"""
    arquivos_possiveis = [
        'CONTROLE DE CARREGAMENTO ULTRAPAO.xlsx',
        'CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx'
    ]
    app_dir = os.path.dirname(os.path.abspath(__file__))
    diretorios = [app_dir, _EXE_DIR, _BASE_DIR, '.']
    for diretorio in diretorios:
        try:
            for arquivo in os.listdir(diretorio):
                if arquivo.endswith('.xlsx') and 'ULTRAPAO' in arquivo.upper():
                    path = os.path.join(diretorio, arquivo)
                    if os.path.isfile(path):
                        return path
            for arquivo in arquivos_possiveis:
                path = os.path.join(diretorio, arquivo)
                if os.path.isfile(path):
                    return path
        except OSError:
            continue
    return None

# Cache da planilha (até 6 segundos) para agilizar várias requisições seguidas (ex.: painel + conferência)
_workbook_cache = {}

def get_workbook_cached():
    """Retorna (wb, from_cache). Se from_cache=True, não feche o wb. TTL 6 segundos."""
    path = encontrar_planilha()
    if not path:
        return (None, False)
    path = os.path.abspath(path)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return (None, False)
    now = time.time()
    c = _workbook_cache
    if c.get('path') == path and c.get('mtime') == mtime and (now - c.get('ts', 0)) < 6:
        return (c['wb'], True)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(path, data_only=False)
    _workbook_cache.clear()
    _workbook_cache['path'] = path
    _workbook_cache['mtime'] = mtime
    _workbook_cache['wb'] = wb
    _workbook_cache['ts'] = now
    return (wb, False)

def _valor_celula(row, col):
    if col is None or col >= len(row) or row[col] is None:
        return ''
    v = row[col]
    if str(v).strip().startswith('='):
        return ''
    return str(int(v)) if isinstance(v, float) and v == int(v) else str(v).strip()


def get_unidade_romaneio(id_viagem, codigo_produto):
    """Retorna a unidade (Pacote, Caixa, Unidade, etc.) do produto no Romaneio por Item para o id_viagem/roteiro."""
    caminho_planilha = encontrar_planilha()
    if not caminho_planilha or not (id_viagem and codigo_produto):
        return ''
    try:
        try:
            wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
        except Exception:
            wb = openpyxl.load_workbook(caminho_planilha, data_only=False)
        nome_romaneio = next((s for s in wb.sheetnames if 'ROMANEIO' in s.upper() and 'ITEM' in s.upper()), None)
        if not nome_romaneio:
            wb.close()
            return ''
        ws = wb[nome_romaneio]
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        coluna_id_viagem = None
        coluna_codigo_produto = None
        coluna_unidade = None
        for idx, header in enumerate(header_row or []):
            if not header:
                continue
            hs = str(header).upper().strip()
            if 'ID' in hs and 'ROTEIRO' in hs:
                coluna_id_viagem = idx
            elif coluna_id_viagem is None and 'ID' in hs and 'VIAGEM' in hs and 'FATURADA' in hs:
                coluna_id_viagem = idx
            if ('CODIGO' in hs or 'CÓDIGO' in hs) and 'PRODUTO' in hs and 'BARRAS' not in hs:
                coluna_codigo_produto = idx
            if 'UNIDADE' in hs and 'MEDIDA' in hs:
                coluna_unidade = idx
            elif coluna_unidade is None and ('UNIDADE' in hs or 'PACOTE' in hs or hs in ('CAIXA', 'UN')):
                coluna_unidade = idx
        if coluna_codigo_produto is None:
            coluna_codigo_produto = 14
        if coluna_id_viagem is None:
            coluna_id_viagem = 1
        id_busca = _normalizar_id_viagem(id_viagem)
        cod_busca = str(codigo_produto).strip()
        cod_norm = _normalizar_codigo_produto(codigo_produto)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if coluna_id_viagem >= len(row) or coluna_codigo_produto >= len(row):
                continue
            id_plan = _normalizar_id_viagem(row[coluna_id_viagem])
            if id_plan and id_busca and id_plan != id_busca:
                continue
            cp = _valor_celula(row, coluna_codigo_produto)
            if not cp or (cp != cod_busca and (not cod_norm or _normalizar_codigo_produto(cp) != cod_norm)):
                continue
            if coluna_unidade is not None and coluna_unidade < len(row):
                un = _valor_celula(row, coluna_unidade)
                if un:
                    wb.close()
                    return un
            wb.close()
            return 'Unidade'
        wb.close()
        return 'Unidade'
    except Exception:
        return 'Unidade'


def buscar_produto_na_planilha(codigo_barras):
    """Busca produto por código de barras: no banco (base_codigo_barras) quando DATABASE_URL está definido; senão na planilha."""
    if _usa_banco_para_dados():
        conn = get_db()
        try:
            ds = _get_latest_dataset_id(conn)
            if not ds:
                conn.close()
                return None
            codigo_busca = str(codigo_barras or '').strip()
            if not codigo_busca:
                conn.close()
                return None
            codigo_digits = re.sub(r'\D', '', codigo_busca)
            row = conn.execute(
                """SELECT codigo_interno, ean, dun, descricao, unidade, peso, data FROM base_codigo_barras
                   WHERE dataset_id = ?
                     AND (
                       TRIM(COALESCE(ean, '')) = ?
                       OR TRIM(COALESCE(dun, '')) = ?
                       OR (? != '' AND regexp_replace(TRIM(COALESCE(dun, '')), '[^0-9]', '', 'g') = ?)
                     )
                   LIMIT 1""",
                (str(ds), codigo_busca, codigo_busca, codigo_digits, codigo_digits),
            ).fetchone()
            conn.close()
            if not row:
                return None
            r = dict(row) if hasattr(row, 'keys') else row
            tipo_codigo = 'EAN' if (r.get('ean') or '') == codigo_busca else 'DUN'
            peso_bruto = r.get('peso') or ''
            return {
                'codigo_barras': codigo_busca,
                'codigo_produto': r.get('codigo_interno') or '',
                'produto': r.get('descricao') or '',
                'quantidade': 1,
                'veiculo': '',
                'status': 'PENDENTE',
                'peso_bruto': peso_bruto,
                'tipo_codigo': tipo_codigo,
            }
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
            return None

    # Sem planilha: dados vêm apenas do banco (DATABASE_URL)
    return None


def _buscar_produto_por_codigo_interno(codigo_interno):
    """Busca produto por código interno: no banco (base_codigo_barras) quando DATABASE_URL está definido; senão na planilha."""
    if _usa_banco_para_dados():
        conn = get_db()
        try:
            ds = _get_latest_dataset_id(conn)
            if not ds:
                conn.close()
                return None
            ci = str(codigo_interno or '').strip()
            if not ci:
                conn.close()
                return None
            row = conn.execute(
                """SELECT codigo_interno, ean, dun, descricao, unidade, peso FROM base_codigo_barras
                   WHERE dataset_id = ? AND codigo_interno = ? LIMIT 1""",
                (str(ds), ci),
            ).fetchone()
            conn.close()
            if not row:
                return None
            r = dict(row) if hasattr(row, 'keys') else row
            return {
                'codigo_barras': r.get('ean') or r.get('dun') or ci,
                'codigo_produto': r.get('codigo_interno') or ci,
                'produto': r.get('descricao') or '',
                'quantidade': 1,
                'veiculo': '',
                'status': 'PENDENTE',
                'peso_bruto': r.get('peso') or '',
                'tipo_codigo': 'EAN' if r.get('ean') else 'DUN',
            }
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
            return None

    # Sem planilha: dados vêm apenas do banco
    return None


@app.route('/api/buscar-produto/<codigo_barras>', methods=['GET'])
def buscar_produto(codigo_barras):
    """Endpoint para buscar produto na planilha por código de barras"""
    resultado = buscar_produto_na_planilha(codigo_barras)
    
    if resultado is None:
        return jsonify({'encontrado': False})
    
    if 'erro' in resultado:
        return jsonify(resultado), 500
    
    return jsonify({
        'encontrado': True,
        'produto': resultado
    })


@app.route('/api/buscar-produto-por-codigo-interno/<codigo_interno>', methods=['GET'])
def api_buscar_produto_por_codigo_interno(codigo_interno):
    """Endpoint para buscar produto na planilha por código do produto (código interno)"""
    resultado = _buscar_produto_por_codigo_interno(codigo_interno)
    if resultado is None:
        return jsonify({'encontrado': False})
    return jsonify({
        'encontrado': True,
        'produto': resultado
    })

def importar_planilha_excel(caminho_arquivo):
    """Importa dados da planilha Excel"""
    try:
        # Tentar carregar com data_only=True primeiro (valores calculados)
        # Se não funcionar, tentar sem data_only para pegar fórmulas
        try:
            wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        except:
            wb = openpyxl.load_workbook(caminho_arquivo, data_only=False)
        
        conn = get_db()
        resultados = {
            'base': 0,
            'romaneio': 0,
            'erros': []
        }
        
        # Importar dados da aba BASE
        if 'BASE' in wb.sheetnames:
            ws_base = wb['BASE']
            linhas_importadas = 0
            
            # Pular cabeçalho (linha 1) e ler a partir da linha 2
            for idx, row in enumerate(ws_base.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and str(row[0]).strip() and not str(row[0]).startswith('='):  # Se tem código de barras e não é fórmula
                    try:
                        codigo_barras = str(row[0]).strip()
                        produto = str(row[1]).strip() if row[1] else ''
                        
                        # Tentar converter quantidade
                        quantidade = 1
                        if row[2] is not None:
                            try:
                                quantidade = int(float(str(row[2])))
                            except:
                                quantidade = 1
                        
                        data_hora = str(row[3]).strip() if row[3] else datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                        veiculo = str(row[4]).strip() if row[4] else ''
                        status = str(row[5]).strip() if row[5] else 'PENDENTE'
                        
                        # Verificar se já existe (evitar duplicatas)
                        existing = conn.execute(
                            'SELECT * FROM produtos_bipados WHERE codigo_barras = ? AND data_hora = ? AND quantidade = ?',
                            (codigo_barras, data_hora, quantidade)
                        ).fetchone()
                        
                        if not existing:
                            conn.execute(
                                '''INSERT INTO produtos_bipados 
                                   (codigo_barras, produto, quantidade, data_hora, veiculo, status)
                                   VALUES (?, ?, ?, ?, ?, ?)''',
                                (codigo_barras, produto, quantidade, data_hora, veiculo, status)
                            )
                            linhas_importadas += 1
                    except Exception as e:
                        resultados['erros'].append(f"Erro na linha BASE {idx}: {str(e)}")
            
            resultados['base'] = linhas_importadas
            conn.commit()
        
        # Importar dados da aba ROMANEIO POR ITEM
        if 'ROMANEIO POR ITEM' in wb.sheetnames:
            ws_romaneio = wb['ROMANEIO POR ITEM']
            linhas_importadas = 0
            
            # Pular cabeçalho (linha 1) e ler a partir da linha 2
            for idx, row in enumerate(ws_romaneio.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] and str(row[0]).strip() and not str(row[0]).startswith('='):  # Se tem código de barras e não é fórmula
                    try:
                        codigo_barras = str(row[0]).strip()
                        
                        # Coluna D (índice 3) é a quantidade do romaneio
                        quantidade_romaneio = 0
                        if row[3] is not None:
                            try:
                                # Tentar converter para número
                                valor = str(row[3]).strip()
                                if valor and not valor.startswith('='):
                                    quantidade_romaneio = int(float(valor))
                            except:
                                quantidade_romaneio = 0
                        
                        # Importar mesmo se quantidade for 0, para manter registro do código
                        if codigo_barras:
                            # Inserir ou atualizar romaneio
                            if getattr(conn, 'kind', 'sqlite') == 'pg':
                                conn.execute(
                                    '''INSERT INTO romaneio (codigo_barras, quantidade_romaneio)
                                       VALUES (%s, %s)
                                       ON CONFLICT (codigo_barras) DO UPDATE SET quantidade_romaneio = EXCLUDED.quantidade_romaneio''',
                                    (codigo_barras, quantidade_romaneio)
                                )
                            else:
                                conn.execute(
                                    '''INSERT INTO romaneio (codigo_barras, quantidade_romaneio) VALUES (?, ?)
                                       ON CONFLICT(codigo_barras) DO UPDATE SET quantidade_romaneio = excluded.quantidade_romaneio''',
                                    (codigo_barras, quantidade_romaneio)
                                )
                            linhas_importadas += 1
                    except Exception as e:
                        resultados['erros'].append(f"Erro na linha ROMANEIO {idx}: {str(e)}")
            
            resultados['romaneio'] = linhas_importadas
            conn.commit()
        
        # NOTA: A aba DIVERGÊNCIAS não é importada porque é sempre calculada dinamicamente
        # a partir da comparação entre produtos bipados e romaneio
        
        conn.close()
        return resultados
        
    except Exception as e:
        import traceback
        return {'erro': f'Erro ao importar planilha: {str(e)}', 'traceback': traceback.format_exc()}

@app.route('/api/importar-planilha', methods=['POST'])
def importar_planilha():
    """Endpoint para importar planilha Excel"""
    if 'file' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        resultados = importar_planilha_excel(filepath)
        
        # Remover arquivo após importação
        try:
            os.remove(filepath)
        except:
            pass
        
        return jsonify(resultados)
    
    return jsonify({'erro': 'Formato de arquivo não permitido'}), 400

@app.route('/api/importar-planilha-local', methods=['POST'])
def importar_planilha_local():
    """Importa planilha do diretório local"""
    data = request.json
    caminho_arquivo = data.get('caminho', '')
    
    # Tentar encontrar a planilha no diretório atual
    arquivos_possiveis = [
        'CONTROLE DE CARREGAMENTO ULTRAPAO.xlsx',
        'CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx',
        caminho_arquivo
    ]
    
    arquivo_encontrado = None
    for arquivo in arquivos_possiveis:
        if os.path.exists(arquivo):
            arquivo_encontrado = arquivo
            break
    
    if not arquivo_encontrado:
        # Procurar qualquer arquivo .xlsx no diretório atual
        for arquivo in os.listdir('.'):
            if arquivo.endswith('.xlsx') and 'ULTRAPAO' in arquivo.upper():
                arquivo_encontrado = arquivo
                break
    
    if arquivo_encontrado:
        resultados = importar_planilha_excel(arquivo_encontrado)
        return jsonify({
            'sucesso': True,
            'arquivo': arquivo_encontrado,
            **resultados
        })
    else:
        return jsonify({
            'erro': 'Planilha não encontrada. Certifique-se de que o arquivo está no mesmo diretório do servidor.',
            'arquivos_procurados': arquivos_possiveis
        }), 404

def _parse_datetime(s):
    """Tenta parsear data/hora em texto para calcular duração."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y'):
        try:
            return datetime.strptime(s[:19].replace('T', ' '), fmt)
        except (ValueError, TypeError):
            continue
    return None


@app.route('/api/painel-graficos', methods=['GET'])
def get_painel_graficos():
    """Retorna dados por viagem para gráficos: tempo de carregamento, itens bipados, faltas."""
    conn = get_db()
    rows = conn.execute('''
        SELECT id_viagem,
               SUM(quantidade) as total_bipados,
               MIN(data_hora) as inicio,
               MAX(data_hora) as fim
        FROM produtos_bipados
        WHERE id_viagem IS NOT NULL AND id_viagem != ''
        GROUP BY id_viagem
        ORDER BY MAX(data_hora) DESC
        LIMIT 25
    ''').fetchall()
    conn.close()
    
    viagens = []
    for r in rows:
        inicio = r['inicio'] or ''
        fim = r['fim'] or ''
        d_min = None
        if inicio and fim:
            t0 = _parse_datetime(inicio)
            t1 = _parse_datetime(fim)
            if t0 and t1:
                delta = t1 - t0
                d_min = max(0, int(delta.total_seconds() / 60))
        viagens.append({
            'id_viagem': r['id_viagem'],
            'total_bipados': r['total_bipados'] or 0,
            'inicio': inicio,
            'fim': fim,
            'duracao_minutos': d_min
        })
    
    # Calcular total_faltas só para as 2 primeiras viagens (conferência é pesada; cache da planilha agiliza)
    try:
        for v in viagens[:2]:
            ret = get_conferencia(v['id_viagem'])
            resp = ret[0] if isinstance(ret, tuple) else ret
            data = resp.get_json() if hasattr(resp, 'get_json') else None
            lista_conf = _conferencia_lista_de_resposta(data)
            if lista_conf is not None:
                v['total_faltas'] = sum((item.get('quantidade_falta') or 0) for item in lista_conf)
            else:
                v['total_faltas'] = 0
        for v in viagens[2:]:
            v['total_faltas'] = 0
    except Exception:
        for v in viagens:
            if 'total_faltas' not in v:
                v['total_faltas'] = 0
    
    # Tempo de carregamento por placa (agrupa duração por placa de cada viagem)
    placa_to_minutos = {}
    for v in viagens:
        info = _get_viagem_info_dict(v.get('id_viagem') or '')
        placa = (info.get('placa') or '').strip() or 'Sem placa'
        if placa not in placa_to_minutos:
            placa_to_minutos[placa] = 0
        if v.get('duracao_minutos') is not None:
            placa_to_minutos[placa] += v['duracao_minutos']
    tempo_por_placa = [
        {'placa': p, 'total_minutos': m}
        for p, m in sorted(placa_to_minutos.items(), key=lambda x: -x[1])
    ]
    
    return jsonify({'viagens': viagens, 'tempo_por_placa': tempo_por_placa})


# Unidade/CD a considerar nos dados do painel (romaneio). None = todas; ou texto que deve constar na coluna da unidade.
UNIDADE_CD_FILTRO = 'Unidade CD Guarulhos Ultrapão (Distribuidora)'


def _estatisticas_romaneio_por_item_banco(conn, janela_inicio=None, janela_fim_excl=None):
    """
    Lê da tabela romaneio_por_item (dataset ativo) e retorna o mesmo formato de _estatisticas_romaneio_por_item:
    qtd_roteiros, qtd_veiculos, itens_total_por_codigo, itens_descricao_por_codigo, peso_por_carro, peso_total_geral, quantidade_total_itens, id_viagem_to_placa.
    Usado quando DATABASE_URL está definido para o painel usar a tabela em vez da planilha.
    """
    out = {
        'qtd_roteiros': 0,
        'qtd_veiculos': 0,
        'itens_total_por_codigo': {},
        'itens_descricao_por_codigo': {},
        'peso_por_carro': {},
        'peso_total_geral': 0.0,
        'quantidade_total_itens': 0,
        'id_viagem_to_placa': {}
    }
    if getattr(conn, 'kind', None) != 'pg':
        return out
    ds = _get_latest_dataset_id(conn)
    if not ds:
        return out
    ds_s = str(ds)
    filtro_janela, params_janela = _painel_romaneio_janela_sql(conn, janela_inicio, janela_fim_excl)
    try:
        row_totais = conn.execute(
            f"""SELECT
                 COUNT(DISTINCT COALESCE(NULLIF(TRIM(id_viagem::text), ''), NULLIF(TRIM(id_roteiro::text), '')))::int AS qtd_roteiros,
                 COUNT(DISTINCT NULLIF(TRIM(placa), ''))::int AS qtd_veiculos,
                 COALESCE(SUM(quantidade), 0)::bigint AS quantidade_total
               FROM romaneio_por_item WHERE dataset_id = ?{filtro_janela}""",
            (ds_s,) + params_janela,
        ).fetchone()
        rows_cod = conn.execute(
            f"""SELECT codigo_produto, MAX(descricao) AS descricao, SUM(quantidade)::bigint AS qtd
               FROM romaneio_por_item
               WHERE dataset_id = ? AND TRIM(COALESCE(codigo_produto::text, '')) != ''{filtro_janela}
               GROUP BY codigo_produto""",
            (ds_s,) + params_janela,
        ).fetchall()
        rows_placa = conn.execute(
            f"""SELECT DISTINCT ON (COALESCE(NULLIF(TRIM(id_viagem::text), ''), TRIM(id_roteiro::text)))
                      COALESCE(NULLIF(TRIM(id_viagem::text), ''), TRIM(id_roteiro::text)) AS id_v,
                      placa
               FROM romaneio_por_item
               WHERE dataset_id = ?
                 AND TRIM(COALESCE(placa, '')) != ''
                 AND COALESCE(NULLIF(TRIM(id_viagem::text), ''), TRIM(id_roteiro::text)) IS NOT NULL{filtro_janela}""",
            (ds_s,) + params_janela,
        ).fetchall()
        rows_peso = conn.execute(
            f"""SELECT COALESCE(NULLIF(TRIM(placa), ''), 'Sem placa') AS placa_eff,
                      SUM(
                        CASE WHEN peso_bruto IS NOT NULL AND TRIM(COALESCE(peso_bruto::text, '')) != ''
                        THEN (REPLACE(TRIM(peso_bruto::text), ',', '.')::numeric)
                             * GREATEST(1, COALESCE(quantidade, 1))
                        ELSE 0 END
                      ) AS peso_total
               FROM romaneio_por_item WHERE dataset_id = ?{filtro_janela}
               GROUP BY COALESCE(NULLIF(TRIM(placa), ''), 'Sem placa')""",
            (ds_s,) + params_janela,
        ).fetchall()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        return out
    qtd_roteiros = int((row_totais.get('qtd_roteiros') if row_totais else 0) or 0)
    qtd_veiculos = int((row_totais.get('qtd_veiculos') if row_totais else 0) or 0)
    quantidade_total = int((row_totais.get('quantidade_total') if row_totais else 0) or 0)
    itens_por_codigo = {}
    itens_descricao = {}
    for r in rows_cod or []:
        cod = (r.get('codigo_produto') or '').strip()
        if not cod:
            continue
        itens_por_codigo[cod] = int(r.get('qtd') or 0)
        desc = (r.get('descricao') or '').strip()
        if desc:
            itens_descricao[cod] = desc
    id_viagem_to_placa = {}
    for r in rows_placa or []:
        id_v = (r.get('id_v') or '').strip()
        placa = (r.get('placa') or '').strip()
        if id_v and placa:
            id_viagem_to_placa[id_v] = placa
    peso_por_placa = {}
    peso_total = 0.0
    for r in rows_peso or []:
        placa_eff = (r.get('placa_eff') or 'Sem placa').strip() or 'Sem placa'
        try:
            p = float(r.get('peso_total') or 0)
        except (TypeError, ValueError):
            p = 0.0
        peso_por_placa[placa_eff] = peso_por_placa.get(placa_eff, 0.0) + p
        peso_total += p
    out['qtd_roteiros'] = qtd_roteiros
    out['qtd_veiculos'] = qtd_veiculos
    out['itens_total_por_codigo'] = itens_por_codigo
    out['itens_descricao_por_codigo'] = itens_descricao
    out['peso_por_carro'] = {k: round(v, 2) for k, v in sorted(peso_por_placa.items(), key=lambda x: -x[1])}
    out['peso_total_geral'] = round(peso_total, 2)
    out['quantidade_total_itens'] = quantidade_total
    out['id_viagem_to_placa'] = id_viagem_to_placa
    return out


def _estatisticas_romaneio_por_item(wb):
    """
    Lê a aba ROMANEIO POR ITEM e retorna: qtd roteiros, qtd veículos,
    quantidade total por item (código), peso por carro (placa), peso total geral.
    Considera apenas linhas da unidade UNIDADE_CD_FILTRO quando a planilha tiver coluna de unidade/CD.
    """
    out = {
        'qtd_roteiros': 0,
        'qtd_veiculos': 0,
        'itens_total_por_codigo': {},
        'itens_descricao_por_codigo': {},
        'peso_por_carro': {},
        'peso_total_geral': 0.0,
        'quantidade_total_itens': 0
    }
    if not wb or 'ROMANEIO POR ITEM' not in wb.sheetnames:
        return out
    ws = wb['ROMANEIO POR ITEM']
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] or []
    col_id = None
    col_placa = None
    col_codigo = None
    col_descricao = None
    col_qtd = None
    col_peso = None
    peso_bruto_eh_total_linha = False  # True = coluna já é peso total da linha (ex: Peso Bruto Pedido)
    col_unidade_cd = None
    for idx, h in enumerate(header_row):
        if not h:
            continue
        hs = str(h).upper().strip()
        if ('ID' in hs and 'ROTEIRO' in hs) or (col_id is None and 'ID' in hs and 'VIAGEM' in hs and 'FATURADA' in hs):
            col_id = idx
        if col_placa is None and ('PLACA' in hs and 'EMPLACA' not in hs):
            col_placa = idx
        if ('CODIGO' in hs or 'CÓDIGO' in hs) and 'PRODUTO' in hs and 'BARRAS' not in hs:
            col_codigo = idx
        if col_descricao is None and ('DESCRI' in hs and 'PRODUTO' in hs or hs == 'DESCRICAO' or hs == 'DESCRIÇÃO'):
            col_descricao = idx
        if 'QUANTIDADE' in hs and 'PRODUTO' in hs:
            col_qtd = idx
        elif col_qtd is None and 'QUANTIDADE' in hs and 'ROMANEIO' not in hs:
            col_qtd = idx
        if 'PESO' in hs and 'BRUTO' in hs:
            col_peso = idx
            if 'PEDIDO' in hs or 'TOTAL' in hs or 'LINHA' in hs:
                peso_bruto_eh_total_linha = True
        if col_unidade_cd is None and (hs == 'UNIDADE' or 'DISTRIBUIDORA' in hs or (('CD' in hs or 'UNIDADE' in hs) and 'MEDIDA' not in hs and 'PRODUTO' not in hs)):
            col_unidade_cd = idx
    ncol = len(header_row)
    if ncol >= 17:
        col_id = 1 if col_id is None else col_id
        col_placa = 5 if col_placa is None else col_placa
        col_codigo = 14 if col_codigo is None else col_codigo
        col_qtd = 16 if col_qtd is None else col_qtd
    if col_id is None:
        col_id = 1
    if col_placa is None and ncol > 5:
        col_placa = 5
    if col_codigo is None:
        col_codigo = 14 if ncol >= 15 else 0
    if col_descricao is None and ncol >= 16:
        col_descricao = 15  # coluna P = Descrição típica na planilha
    if col_qtd is None:
        col_qtd = 16 if ncol >= 17 else 2
    roteiros = set()
    veiculos = set()
    itens_por_codigo = {}
    itens_descricao_por_codigo = {}
    peso_por_placa = {}
    id_viagem_to_placa = {}
    peso_total = 0.0
    quantidade_total = 0
    filtro_unidade = (UNIDADE_CD_FILTRO or '').strip()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if filtro_unidade and col_unidade_cd is not None and col_unidade_cd < len(row):
            cel_unidade = (str(row[col_unidade_cd]).strip() if row[col_unidade_cd] else '') or ''
            if filtro_unidade not in cel_unidade and 'GUARULHOS' not in cel_unidade.upper():
                continue
        id_v = _normalizar_id_viagem(row[col_id] if col_id < len(row) else '')
        placa = (str(row[col_placa]).strip() if col_placa is not None and col_placa < len(row) and row[col_placa] else '') or ''
        if id_v and placa:
            id_viagem_to_placa[id_v] = placa
        cod = str(row[col_codigo]).strip() if col_codigo < len(row) and row[col_codigo] else ''
        if cod and cod.startswith('='):
            cod = ''
        try:
            qtd = int(float(str(row[col_qtd]).replace(',', '.').strip() or 0)) if col_qtd < len(row) else 0
        except (ValueError, TypeError):
            qtd = 0
        try:
            peso = float(str(row[col_peso]).replace(',', '.').strip() or 0) if col_peso is not None and col_peso < len(row) and row[col_peso] else 0
        except (ValueError, TypeError):
            peso = 0
        if peso_bruto_eh_total_linha:
            p = peso
        else:
            p = qtd * peso
        if id_v:
            roteiros.add(id_v)
        if placa:
            veiculos.add(placa)
        if cod:
            itens_por_codigo[cod] = itens_por_codigo.get(cod, 0) + qtd
            if cod not in itens_descricao_por_codigo and col_descricao is not None and col_descricao < len(row):
                desc = (str(row[col_descricao]).strip() if row[col_descricao] else '') or ''
                itens_descricao_por_codigo[cod] = desc
        quantidade_total += qtd
        placa_eff = placa or id_viagem_to_placa.get(id_v, 'Sem placa')
        if placa_eff:
            peso_por_placa[placa_eff] = peso_por_placa.get(placa_eff, 0.0) + p
        peso_total += p
    out['qtd_roteiros'] = len(roteiros)
    out['qtd_veiculos'] = len(veiculos)
    out['itens_total_por_codigo'] = itens_por_codigo
    out['itens_descricao_por_codigo'] = itens_descricao_por_codigo
    out['peso_por_carro'] = {k: round(v, 2) for k, v in sorted(peso_por_placa.items(), key=lambda x: -x[1])}
    out['peso_total_geral'] = round(peso_total, 2)
    out['quantidade_total_itens'] = quantidade_total
    out['id_viagem_to_placa'] = id_viagem_to_placa
    return out


def _build_mapa_peso_romaneio(wb):
    """Retorna dict: codigo_produto -> peso unitário (float). Usa aba ROMANEIO POR ITEM."""
    mapa_codigo_to_peso = {}
    if 'ROMANEIO POR ITEM' not in wb.sheetnames:
        return mapa_codigo_to_peso
    ws = wb['ROMANEIO POR ITEM']
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] or []
    col_codigo = None
    col_peso = None
    col_qtd = None
    peso_bruto_eh_total = False
    for idx, h in enumerate(header_row):
        if not h:
            continue
        hs = str(h).upper().strip()
        if ('CODIGO' in hs or 'CÓDIGO' in hs) and 'PRODUTO' in hs and 'BARRAS' not in hs:
            col_codigo = idx
        if 'PESO' in hs and 'BRUTO' in hs:
            col_peso = idx
            if 'PEDIDO' in hs or 'TOTAL' in hs or 'LINHA' in hs:
                peso_bruto_eh_total = True
        if 'QUANTIDADE' in hs and 'PRODUTO' in hs:
            col_qtd = idx
        elif col_qtd is None and 'QUANTIDADE' in hs and 'ROMANEIO' not in hs:
            col_qtd = idx
    ncol = len(header_row)
    if col_codigo is None:
        col_codigo = 14 if ncol >= 15 else 0
    if col_peso is None:
        return mapa_codigo_to_peso
    if col_qtd is None:
        col_qtd = 16 if ncol >= 17 else 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_codigo < len(row) and col_peso < len(row):
            cod = str(row[col_codigo]).strip() if row[col_codigo] else ''
            if not cod or cod.startswith('='):
                continue
            try:
                p = float(str(row[col_peso]).replace(',', '.').strip() or 0)
            except (ValueError, TypeError):
                p = 0
            if peso_bruto_eh_total and col_qtd < len(row):
                try:
                    qtd = int(float(str(row[col_qtd]).replace(',', '.').strip() or 0))
                except (ValueError, TypeError):
                    qtd = 1
                if qtd > 0:
                    p = p / qtd
            mapa_codigo_to_peso[cod] = p
    return mapa_codigo_to_peso


def _build_mapa_barras_to_codigo_produto(wb):
    """Retorna dict: codigo_barras (EAN-13 ou DUN-14) -> codigo_interno (Codigo). Da aba BASE."""
    mapa = {}
    ws = next((wb[s] for s in wb.sheetnames if s.upper().strip() == 'BASE'), None)
    if ws is None:
        return mapa
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_codigo = next((i for i, h in enumerate(header_row or []) if _coluna_base_eh_codigo_interno(str(h or ''))), 0)
    col_ean = next((i for i, h in enumerate(header_row or []) if _coluna_base_eh_ean(str(h or ''))), None)
    col_dun = next((i for i, h in enumerate(header_row or []) if _coluna_base_eh_dun(str(h or ''))), None)
    colunas_barras = [i for i, h in enumerate(header_row or []) if _coluna_base_eh_codigo_barras(str(h or ''))]
    if col_ean is not None and col_ean not in colunas_barras:
        colunas_barras.append(col_ean)
    if col_dun is not None and col_dun not in colunas_barras:
        colunas_barras.append(col_dun)
    if not colunas_barras:
        colunas_barras = [0]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_codigo >= len(row):
            continue
        ci = _valor_celula(row, col_codigo)
        if not ci:
            continue
        for col in colunas_barras:
            if col >= len(row):
                continue
            cb = _valor_celula(row, col)
            if cb:
                mapa[cb] = ci
    return mapa


@app.route('/api/painel-graficos-extras', methods=['GET'])
def get_painel_graficos_extras():
    """Retorna dados para gráficos: itens mais bipados, carros com mais itens, carros com mais peso."""
    conn = get_db()
    conn.row_factory = sqlite3.Row
    # Item mais bipado (agrupa por codigo_barras, soma quantidade; label = produto ou codigo_barras)
    rows_itens = conn.execute('''
        SELECT produto, codigo_barras, SUM(quantidade) as total
        FROM produtos_bipados
        GROUP BY codigo_barras
        ORDER BY total DESC
        LIMIT 15
    ''').fetchall()
    top_itens = []
    for r in rows_itens:
        label = (r['produto'] or '').strip() or (r['codigo_barras'] or '')
        if not label:
            label = r['codigo_barras'] or '-'
        top_itens.append({'label': label[:50], 'total': r['total'] or 0})
    # Carros com mais itens (soma de quantidade por veículo)
    rows_carros = conn.execute('''
        SELECT veiculo, SUM(quantidade) as total
        FROM produtos_bipados
        WHERE veiculo IS NOT NULL AND trim(veiculo) != ''
        GROUP BY veiculo
        ORDER BY total DESC
        LIMIT 15
    ''').fetchall()
    carros_itens = [{'veiculo': r['veiculo'] or '', 'total': r['total'] or 0} for r in rows_carros]
    # Carros com mais peso: precisa da planilha (codigo_barras -> codigo produto -> peso)
    carros_peso = []
    wb_extras, from_cache_extras = get_workbook_cached()
    if wb_extras:
        try:
            mapa_peso = _build_mapa_peso_romaneio(wb_extras)
            mapa_barras_codigo = _build_mapa_barras_to_codigo_produto(wb_extras)
            if not from_cache_extras:
                wb_extras.close()
            rows_bipados = conn.execute(
                "SELECT veiculo, codigo_barras, quantidade FROM produtos_bipados WHERE veiculo IS NOT NULL AND trim(veiculo) != ''"
            ).fetchall()
            peso_por_veiculo = {}
            for r in rows_bipados:
                veic = (r['veiculo'] or '').strip()
                if not veic:
                    continue
                codigo_barras = (r['codigo_barras'] or '').strip()
                qtd = int(r['quantidade'] or 0)
                codigo_produto = mapa_barras_codigo.get(codigo_barras) or codigo_barras
                peso_unit = mapa_peso.get(codigo_produto, 0) or mapa_peso.get(codigo_barras, 0)
                if veic not in peso_por_veiculo:
                    peso_por_veiculo[veic] = 0.0
                peso_por_veiculo[veic] += qtd * peso_unit
            carros_peso = [{'veiculo': v, 'peso_total': round(p, 2)} for v, p in sorted(peso_por_veiculo.items(), key=lambda x: -x[1])[:15]]
        except Exception:
            if not from_cache_extras and wb_extras:
                try:
                    wb_extras.close()
                except Exception:
                    pass
    conn.close()
    return jsonify({
        'top_itens_bipados': top_itens,
        'carros_mais_itens': carros_itens,
        'carros_mais_peso': carros_peso
    })


def _calcular_duracao_minutos(inicio, fim):
    """Retorna duração em minutos entre dois timestamps (ou None)."""
    if not inicio or not fim:
        return None
    t0, t1 = _parse_datetime_iso(inicio), _parse_datetime_iso(fim)
    if not t0 and hasattr(inicio, 'strftime'):
        t0 = inicio
    if not t1 and hasattr(fim, 'strftime'):
        t1 = fim
    if not t0:
        t0 = _parse_datetime(inicio)
    if not t1:
        t1 = _parse_datetime(fim)
    if t0 and t1:
        return max(0, int((t1 - t0).total_seconds() / 60))
    return None


def _formatar_duracao_legivel(minutos):
    if minutos is None:
        return ''
    m = int(minutos)
    if m < 60:
        return '%s min' % m
    h, r = divmod(m, 60)
    return '%sh %smin' % (h, r) if r else '%sh' % h


def _build_mapas_peso_para_bipagem(conn):
    """Mapas código→peso e barras→código para cálculo de peso bipado."""
    mapa_peso = {}
    mapa_barras_codigo = {}
    if _usa_banco_para_dados() and getattr(conn, 'kind', None) == 'pg':
        try:
            ds = _get_latest_dataset_id(conn)
            if ds:
                base_rows = conn.execute(
                    "SELECT codigo_interno, ean, dun, peso FROM base_codigo_barras WHERE dataset_id = ?",
                    (str(ds),),
                ).fetchall()
                for r in base_rows:
                    ci = (r.get('codigo_interno') or r[0] or '').strip()
                    ean = (r.get('ean') or r[1] or '').strip()
                    dun = (r.get('dun') or r[2] or '').strip()
                    try:
                        p = float(r.get('peso') or r[3] or 0)
                    except (TypeError, ValueError):
                        p = 0
                    if ci:
                        mapa_peso[ci] = p
                    if ean:
                        mapa_barras_codigo[ean] = ci
                    if dun:
                        mapa_barras_codigo[dun] = ci
        except Exception:
            pass
    if not mapa_peso:
        caminho_planilha = encontrar_planilha()
        if caminho_planilha:
            try:
                wb_plan = openpyxl.load_workbook(caminho_planilha, data_only=True)
                mapa_peso = _build_mapa_peso_romaneio(wb_plan)
                mapa_barras_codigo = _build_mapa_barras_to_codigo_produto(wb_plan)
                wb_plan.close()
            except Exception:
                pass
    return mapa_peso, mapa_barras_codigo


def _parse_painel_janela(data_inicio=None, data_fim=None, hora_inicio=None, hora_fim=None, data_ref=None):
    """Janela de filtro do painel (data/hora início e fim) em America/Sao_Paulo."""
    return _parse_janela_datahora(data_inicio, data_fim, hora_inicio, hora_fim, data_legacy=data_ref)


def _painel_sql_janela_params(inicio, fim_excl, conn):
    """Parâmetros de data/hora para comparação no SQL (PG aware, SQLite como texto)."""
    if getattr(conn, 'kind', None) == 'pg':
        return inicio, fim_excl
    tz = _ravex_tz_baixados()
    if inicio.tzinfo:
        inicio = inicio.astimezone(tz).replace(tzinfo=None)
    if fim_excl.tzinfo:
        fim_excl = fim_excl.astimezone(tz).replace(tzinfo=None)
    return inicio.strftime('%Y-%m-%d %H:%M:%S'), fim_excl.strftime('%Y-%m-%d %H:%M:%S')


def _painel_romaneio_janela_sql(conn, janela_inicio, janela_fim_excl):
    """Filtro do painel em romaneio_por_item: importado_em ou data_expedicao na janela."""
    if janela_inicio is None or janela_fim_excl is None:
        return '', ()
    ini_p, fim_p = _painel_sql_janela_params(janela_inicio, janela_fim_excl, conn)
    if getattr(conn, 'kind', None) != 'pg':
        return ' AND importado_em >= ? AND importado_em < ?', (ini_p, fim_p)
    tz = _ravex_tz_baixados()
    di = janela_inicio.astimezone(tz) if janela_inicio.tzinfo else janela_inicio.replace(tzinfo=tz)
    df = janela_fim_excl.astimezone(tz) if janela_fim_excl.tzinfo else janela_fim_excl.replace(tzinfo=tz)
    d_ini = di.date().isoformat()
    d_fim_excl = df.date().isoformat()
    sql = """ AND (
            (importado_em >= ? AND importado_em < ?)
            OR (
                CASE
                    WHEN TRIM(COALESCE(data_expedicao::text, '')) ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
                        THEN LEFT(TRIM(data_expedicao::text), 10)::date
                    WHEN TRIM(COALESCE(data_expedicao::text, '')) ~ '^[0-9]{2}/[0-9]{2}/[0-9]{4}'
                        THEN to_date(LEFT(TRIM(data_expedicao::text), 10), 'DD/MM/YYYY')
                    ELSE NULL
                END >= ?::date
                AND CASE
                    WHEN TRIM(COALESCE(data_expedicao::text, '')) ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
                        THEN LEFT(TRIM(data_expedicao::text), 10)::date
                    WHEN TRIM(COALESCE(data_expedicao::text, '')) ~ '^[0-9]{2}/[0-9]{2}/[0-9]{4}'
                        THEN to_date(LEFT(TRIM(data_expedicao::text), 10), 'DD/MM/YYYY')
                    ELSE NULL
                END < ?::date
            )
        )"""
    return sql, (ini_p, fim_p, d_ini, d_fim_excl)


def _painel_request_janela_meta(conn):
    """Lê data/hora início e fim do request e devolve janela + parâmetros SQL + resumo legível."""
    data_inicio = (request.args.get('data_inicio') or request.args.get('data') or '').strip()[:10] or None
    data_fim = (request.args.get('data_fim') or request.args.get('data') or '').strip()[:10] or None
    hora_inicio = (request.args.get('hora_inicio') or '').strip()[:5] or None
    hora_fim = (request.args.get('hora_fim') or '').strip()[:5] or None
    meta = _parse_painel_janela(data_inicio, data_fim, hora_inicio, hora_fim)
    ini_p, fim_p = _painel_sql_janela_params(meta['inicio'], meta['fim_excl'], conn)
    filtros_resp = {
        'data': meta['data'],
        'data_iso': meta['data_iso'],
        'data_inicio': meta['data_inicio_iso'],
        'data_fim': meta['data_fim_iso'],
        'hora_inicio': meta['hora_inicio'],
        'hora_fim': meta['hora_fim'],
        'legivel': meta['legivel'],
    }
    return meta, ini_p, fim_p, filtros_resp


def _terceiros_sql_periodo_janela(alias, conn, data_inicio=None, data_fim=None, hora_inicio=None, hora_fim=None):
    """Cláusulas SQL de período (data/hora) em criado_em, com suporte a turno noturno."""
    data_inicio = (data_inicio or '').strip()[:10] or None
    data_fim = (data_fim or '').strip()[:10] or None
    hora_inicio = (hora_inicio or '').strip()[:5] or None
    hora_fim = (hora_fim or '').strip()[:5] or None
    meta = _parse_janela_datahora(data_inicio, data_fim, hora_inicio, hora_fim)
    ini_p, fim_p = _painel_sql_janela_params(meta['inicio'], meta['fim_excl'], conn)
    return [alias + '.criado_em >= ?', alias + '.criado_em < ?'], [ini_p, fim_p], meta


def _coletar_placas_baixadas_dia(conn, data_ref=None, janela_inicio=None, janela_fim_excl=None):
    """
    Lista viagens/placas baixadas do Ravex no dia (America/Sao_Paulo) com status de carregamento,
    período, duração e peso.
    """
    tz = _ravex_tz_baixados()
    if janela_inicio is not None and janela_fim_excl is not None:
        dia = janela_inicio.astimezone(tz) if janela_inicio.tzinfo else janela_inicio.replace(tzinfo=tz)
        dia = dia.replace(hour=0, minute=0, second=0, microsecond=0)
        fim_dia = janela_fim_excl
        inicio_q, fim_q = _painel_sql_janela_params(janela_inicio, janela_fim_excl, conn)
    elif data_ref:
        try:
            dia = datetime.strptime(str(data_ref).strip()[:10], '%Y-%m-%d').replace(tzinfo=tz)
        except ValueError:
            dia = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
        fim_dia = dia + timedelta(days=1)
        inicio_q, fim_q = _painel_sql_janela_params(dia, fim_dia, conn)
    else:
        dia = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
        fim_dia = dia + timedelta(days=1)
        inicio_q, fim_q = _painel_sql_janela_params(dia, fim_dia, conn)
    data_label = dia.strftime('%d/%m/%Y')
    rom_janela_ini = janela_inicio if janela_inicio is not None else dia
    rom_janela_fim = janela_fim_excl if janela_fim_excl is not None else fim_dia

    viagens_base = []
    if getattr(conn, 'kind', None) == 'pg' and _usa_banco_para_dados():
        ds = _get_latest_dataset_id(conn)
        if ds:
            try:
                filtro_rom, params_rom = _painel_romaneio_janela_sql(conn, rom_janela_ini, rom_janela_fim)
                rows = conn.execute(
                    """SELECT TRIM(COALESCE(id_viagem::text, '')) AS id_viagem,
                              MAX(TRIM(COALESCE(id_roteiro::text, ''))) AS id_roteiro,
                              MAX(TRIM(COALESCE(placa::text, ''))) AS placa,
                              MAX(importado_em) AS baixado_em,
                              COALESCE(SUM(
                                  CASE WHEN peso_bruto IS NOT NULL AND TRIM(COALESCE(peso_bruto::text, '')) != ''
                                  THEN (REPLACE(TRIM(peso_bruto::text), ',', '.')::numeric)
                                       * GREATEST(1, COALESCE(quantidade, 1))
                                  ELSE 0 END
                              ), 0) AS peso_romaneio
                       FROM romaneio_por_item
                       WHERE dataset_id = ?
                         AND TRIM(COALESCE(id_viagem::text, '')) != ''"""
                    + filtro_rom +
                    """ GROUP BY TRIM(COALESCE(id_viagem::text, ''))
                       ORDER BY MAX(importado_em) DESC""",
                    (str(ds),) + params_rom,
                ).fetchall()
                for r in rows or []:
                    vid = (r.get('id_viagem') if hasattr(r, 'get') else r[0]) or ''
                    vid = str(vid).strip()
                    if not vid:
                        continue
                    try:
                        pr = float(r.get('peso_romaneio') if hasattr(r, 'get') else (r[4] if len(r) > 4 else 0) or 0)
                    except (TypeError, ValueError):
                        pr = 0.0
                    viagens_base.append({
                        'id_viagem': vid,
                        'id_roteiro': str((r.get('id_roteiro') if hasattr(r, 'get') else (r[1] if len(r) > 1 else '')) or '').strip(),
                        'placa': str((r.get('placa') if hasattr(r, 'get') else (r[2] if len(r) > 2 else '')) or '').strip(),
                        'baixado_em': r.get('baixado_em') if hasattr(r, 'get') else (r[3] if len(r) > 3 else None),
                        'peso_romaneio_kg': round(pr, 2),
                    })
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass

    ids = [v['id_viagem'] for v in viagens_base]
    if not ids and getattr(conn, 'kind', None) != 'pg':
        try:
            rows = conn.execute(
                """SELECT id_viagem, MAX(veiculo) AS placa,
                          MIN(data_hora) AS inicio, MAX(data_hora) AS fim,
                          SUM(quantidade) AS total_bipados
                   FROM produtos_bipados
                   WHERE id_viagem IS NOT NULL AND id_viagem != ''
                     AND COALESCE(fluxo, 'carregamento') = 'carregamento'
                   GROUP BY id_viagem
                   ORDER BY MAX(data_hora) DESC
                   LIMIT 100"""
            ).fetchall()
            for r in rows or []:
                vid = (r['id_viagem'] or '').strip()
                if not vid:
                    continue
                viagens_base.append({
                    'id_viagem': vid,
                    'id_roteiro': '',
                    'placa': (r['placa'] or r.get('veiculo') or '').strip() if hasattr(r, 'get') else '',
                    'baixado_em': None,
                    'peso_romaneio_kg': 0.0,
                })
                ids.append(vid)
        except Exception:
            pass

    if not viagens_base:
        return {
            'data': data_label,
            'data_iso': dia.strftime('%Y-%m-%d'),
            'rows': [],
            'resumo': {'total': 0, 'carregados': 0, 'nao_carregados': 0, 'em_andamento': 0, 'peso_total_kg': 0},
        }

    ids_norm = [_normalizar_id_viagem(i) for i in ids]
    periodo_map = {}
    extrato_map = {}
    bip_map = {}
    placa_override = {}

    _ensure_viagem_periodo_bipagem_table(conn)
    tbl_per = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
    ph = ','.join('?' * len(ids_norm))
    try:
        rows_per = conn.execute(
            f"SELECT id_viagem, inicio_em, fim_em, extrato_gerado_em FROM {tbl_per} WHERE fluxo = 'carregamento' AND id_viagem IN ({ph})",
            tuple(ids_norm),
        ).fetchall()
        for r in rows_per or []:
            vid = _normalizar_id_viagem(r.get('id_viagem') if hasattr(r, 'get') else r[0])
            periodo_map[vid] = (
                r.get('inicio_em') if hasattr(r, 'get') else r[1],
                r.get('fim_em') if hasattr(r, 'get') else r[2],
            )
            extrato_map[vid] = r.get('extrato_gerado_em') if hasattr(r, 'get') else (r[3] if len(r) > 3 else None)
    except Exception:
        pass

    try:
        if getattr(conn, 'kind', None) == 'pg':
            sql_bip = f"""SELECT TRIM(COALESCE(id_viagem::text, '')) AS id_viagem,
                                 SUM(quantidade) AS total_bipados,
                                 MIN(data_hora) AS inicio, MAX(data_hora) AS fim
                          FROM produtos_bipados
                          WHERE COALESCE(fluxo, 'carregamento') = 'carregamento'
                            AND TRIM(COALESCE(id_viagem::text, '')) IN ({ph})
                          GROUP BY TRIM(COALESCE(id_viagem::text, ''))"""
        else:
            sql_bip = f"""SELECT id_viagem, SUM(quantidade) AS total_bipados,
                                 MIN(data_hora) AS inicio, MAX(data_hora) AS fim
                          FROM produtos_bipados
                          WHERE COALESCE(fluxo, 'carregamento') = 'carregamento'
                            AND id_viagem IN ({ph})
                          GROUP BY id_viagem"""
        for r in conn.execute(sql_bip, tuple(ids_norm)).fetchall() or []:
            vid = _normalizar_id_viagem(r.get('id_viagem') if hasattr(r, 'get') else r['id_viagem'])
            bip_map[vid] = r
    except Exception:
        pass

    try:
        rows_placa = conn.execute(
            f"SELECT id_viagem, placa FROM viagem_placa WHERE id_viagem IN ({ph})",
            tuple(ids_norm),
        ).fetchall()
        for r in rows_placa or []:
            vid = _normalizar_id_viagem(r.get('id_viagem') if hasattr(r, 'get') else r[0])
            pl = (r.get('placa') if hasattr(r, 'get') else r[1]) or ''
            if str(pl).strip():
                placa_override[vid] = str(pl).strip()
    except Exception:
        pass

    mapa_peso, mapa_barras = _build_mapas_peso_para_bipagem(conn)
    peso_bipado_map = {}
    try:
        if getattr(conn, 'kind', None) == 'pg':
            sql_p = f"""SELECT TRIM(COALESCE(id_viagem::text, '')) AS id_viagem, codigo_barras, quantidade
                        FROM produtos_bipados
                        WHERE COALESCE(fluxo, 'carregamento') = 'carregamento'
                          AND TRIM(COALESCE(id_viagem::text, '')) IN ({ph})"""
        else:
            sql_p = f"""SELECT id_viagem, codigo_barras, quantidade FROM produtos_bipados
                        WHERE COALESCE(fluxo, 'carregamento') = 'carregamento' AND id_viagem IN ({ph})"""
        for r in conn.execute(sql_p, tuple(ids_norm)).fetchall() or []:
            vid = _normalizar_id_viagem(r.get('id_viagem') if hasattr(r, 'get') else r['id_viagem'])
            cb = (r.get('codigo_barras') if hasattr(r, 'get') else r[1]) or ''
            qtd = int((r.get('quantidade') if hasattr(r, 'get') else r[2]) or 0)
            cp = mapa_barras.get(str(cb).strip()) or str(cb).strip()
            pu = mapa_peso.get(cp, 0) or mapa_peso.get(str(cb).strip(), 0)
            peso_bipado_map[vid] = peso_bipado_map.get(vid, 0.0) + qtd * float(pu or 0)
    except Exception:
        pass

    rows_out = []
    resumo = {'total': 0, 'carregados': 0, 'nao_carregados': 0, 'em_andamento': 0, 'peso_total_kg': 0.0}
    for vb in viagens_base:
        vid = _normalizar_id_viagem(vb['id_viagem'])
        placa = placa_override.get(vid) or vb.get('placa') or ''
        id_roteiro = vb.get('id_roteiro') or ''
        if not placa or not id_roteiro:
            info = _get_viagem_info_planilha(vid, conn=conn)
            if not placa:
                placa = (info.get('placa') or '').strip()
            if not id_roteiro:
                id_roteiro = (info.get('id_roteiro') or '').strip()

        inicio_raw, fim_raw = periodo_map.get(vid, (None, None))
        bip = bip_map.get(vid)
        if not inicio_raw and bip:
            inicio_raw = bip.get('inicio') if hasattr(bip, 'get') else bip['inicio']
        if not fim_raw and bip:
            fim_raw = bip.get('fim') if hasattr(bip, 'get') else bip['fim']
        total_bip = 0
        if bip:
            total_bip = int((bip.get('total_bipados') if hasattr(bip, 'get') else bip['total_bipados']) or 0)

        dur_min = _calcular_duracao_minutos(inicio_raw, fim_raw)
        inicio_fmt = _formatar_data_hora_periodo(inicio_raw)
        fim_fmt = _formatar_data_hora_periodo(fim_raw)

        if total_bip <= 0:
            status = 'Não carregado'
            resumo['nao_carregados'] += 1
        elif _viagem_carregamento_concluido(extrato_map.get(vid), bip):
            status = 'Concluído'
            resumo['carregados'] += 1
        elif total_bip > 0:
            status = 'Em andamento'
            resumo['em_andamento'] += 1
        else:
            status = 'Não carregado'
            resumo['nao_carregados'] += 1

        peso_bip = round(peso_bipado_map.get(vid, 0.0), 2)
        peso_rom = vb.get('peso_romaneio_kg') or 0.0
        peso_exibir = peso_bip if peso_bip > 0 else peso_rom

        rows_out.append({
            'placa': placa or '—',
            'id_roteiro': id_roteiro or '—',
            'id_viagem': vid,
            'status': status,
            'inicio_carregamento': inicio_fmt or '—',
            'fim_carregamento': fim_fmt or '—',
            'duracao_minutos': dur_min,
            'duracao_legivel': _formatar_duracao_legivel(dur_min) or '—',
            'peso_kg': peso_exibir,
            'peso_bipado_kg': peso_bip,
            'peso_romaneio_kg': peso_rom,
            'total_bipados': total_bip,
            'baixado_em': _formatar_criado_em_baixados_ravex(vb.get('baixado_em')) if vb.get('baixado_em') else '',
        })
        resumo['total'] += 1
        resumo['peso_total_kg'] = round(resumo['peso_total_kg'] + float(peso_exibir or 0), 2)

    return {
        'data': data_label,
        'data_iso': dia.strftime('%Y-%m-%d'),
        'rows': rows_out,
        'resumo': resumo,
    }


def _painel_metricas_eficiencia_viagem(lista):
    """Unidades do romaneio, bipadas, erros e % de eficiência da conferência."""
    un_rom = un_bip = un_falta = un_sobra = 0
    for it in lista or []:
        qp = int(it.get('quantidade_produto') or 0)
        qb = int(it.get('quantidade_bipada') or 0)
        un_rom += qp
        un_bip += qb
        un_falta += int(it.get('quantidade_falta') or 0)
        un_sobra += int(it.get('quantidade_sobra') or 0)
    un_erro = un_falta + un_sobra
    if un_rom > 0:
        efic = round(max(0.0, min(100.0, 100.0 * (un_rom - un_erro) / un_rom)), 1)
    elif un_bip > 0:
        efic = 100.0 if un_erro == 0 else 0.0
    else:
        efic = 0.0
    return {
        'unidades_romaneio': un_rom,
        'unidades_bipadas': un_bip,
        'unidades_erro': un_erro,
        'total_faltas': un_falta,
        'total_sobras': un_sobra,
        'eficiencia_pct': efic,
    }


def _painel_coletar_erros_carregamento(conn, ids_viagem, id_viagem_placa=None):
    """Divergências (faltas/sobras) das viagens do período para o painel."""
    id_viagem_placa = id_viagem_placa or {}
    rows_out = []
    por_viagem = {}
    correlacao = []
    faltas_map = {}
    resumo = {
        'total_itens_erro': 0,
        'total_faltas': 0,
        'total_sobras': 0,
        'viagens_com_erro': 0,
        'unidades_romaneio': 0,
        'unidades_bipadas': 0,
        'unidades_erro': 0,
        'eficiencia_geral_pct': 100.0,
    }
    ids_norm = []
    for vid in ids_viagem or []:
        v = _normalizar_id_viagem(vid)
        if v and v not in ids_norm:
            ids_norm.append(v)
    for vid in ids_norm[:25]:
        placa = (id_viagem_placa.get(vid) or id_viagem_placa.get(str(vid)) or '').strip()
        lista_full = []
        try:
            lista_full = _conferencia_lista_interna(vid, 'carregamento', conn=conn, somente_divergencias=False)
        except Exception:
            lista_full = []
        metricas = _painel_metricas_eficiencia_viagem(lista_full)
        faltas_v = int(metricas['total_faltas'])
        sobras_v = int(metricas['total_sobras'])
        faltas_map[vid] = faltas_v
        if metricas['unidades_bipadas'] > 0 or metricas['unidades_romaneio'] > 0:
            correlacao.append({
                'id_viagem': vid,
                'placa': placa or '—',
                'unidades_romaneio': metricas['unidades_romaneio'],
                'unidades_bipadas': metricas['unidades_bipadas'],
                'unidades_erro': metricas['unidades_erro'],
                'total_faltas': faltas_v,
                'total_sobras': sobras_v,
                'eficiencia_pct': metricas['eficiencia_pct'],
            })
        divergentes = _itens_divergentes_da_lista(lista_full, vid)
        try:
            _carregar_motivos_divergencia(divergentes, conn=conn)
        except Exception:
            pass
        if not divergentes:
            continue
        resumo['viagens_com_erro'] += 1
        por_viagem[vid] = {
            'id_viagem': vid,
            'placa': placa or '—',
            'total_faltas': faltas_v,
            'total_sobras': sobras_v,
            'qtd_itens': len(divergentes),
            'eficiencia_pct': metricas['eficiencia_pct'],
            'unidades_bipadas': metricas['unidades_bipadas'],
            'unidades_erro': metricas['unidades_erro'],
        }
        for it in divergentes:
            falta = int(it.get('quantidade_falta') or 0)
            sobra = int(it.get('quantidade_sobra') or 0)
            if falta <= 0 and sobra <= 0:
                continue
            tipo = 'Falta' if falta > 0 and sobra <= 0 else ('Sobra' if sobra > 0 and falta <= 0 else 'Falta e sobra')
            rows_out.append({
                'id_viagem': vid,
                'placa': placa or '—',
                'codigo_produto': (it.get('codigo_produto') or it.get('codigo_interno') or '').strip(),
                'produto': ((it.get('produto') or '').strip() or '—')[:100],
                'quantidade_falta': falta,
                'quantidade_sobra': sobra,
                'tipo_erro': tipo,
                'motivo': _formatar_motivo_divergencia(it.get('motivo_divergencia') or ''),
            })
            resumo['total_itens_erro'] += 1
            resumo['total_faltas'] += falta
            resumo['total_sobras'] += sobra
    tot_rom = sum(int(c.get('unidades_romaneio') or 0) for c in correlacao)
    tot_bip = sum(int(c.get('unidades_bipadas') or 0) for c in correlacao)
    tot_err = sum(int(c.get('unidades_erro') or 0) for c in correlacao)
    resumo['unidades_romaneio'] = tot_rom
    resumo['unidades_bipadas'] = tot_bip
    resumo['unidades_erro'] = tot_err
    if tot_rom > 0:
        resumo['eficiencia_geral_pct'] = round(max(0.0, min(100.0, 100.0 * (tot_rom - tot_err) / tot_rom)), 1)
    elif tot_bip > 0:
        resumo['eficiencia_geral_pct'] = 100.0 if tot_err == 0 else 0.0
    else:
        resumo['eficiencia_geral_pct'] = 100.0
    correlacao.sort(key=lambda x: (-(x.get('eficiencia_pct') or 0), -(x.get('unidades_bipadas') or 0)))
    rows_out.sort(
        key=lambda r: (-(r.get('quantidade_falta') or 0) - (r.get('quantidade_sobra') or 0), r.get('id_viagem') or ''),
    )
    return {
        'resumo': resumo,
        'rows': rows_out[:120],
        'por_viagem': sorted(por_viagem.values(), key=lambda x: -(x.get('total_faltas') or 0) - (x.get('total_sobras') or 0)),
        'correlacao': correlacao[:25],
        'faltas_por_viagem': faltas_map,
    }


@app.route('/api/painel-completo', methods=['GET'])
def get_painel_completo():
    """Um único request: estatísticas + viagens + gráficos. Estatísticas em 1 query para carregar mais rápido."""
    data_inicio = (request.args.get('data_inicio') or request.args.get('data') or '').strip()[:10] or None
    data_fim = (request.args.get('data_fim') or request.args.get('data') or '').strip()[:10] or None
    hora_inicio = (request.args.get('hora_inicio') or '').strip()[:5] or None
    hora_fim = (request.args.get('hora_fim') or '').strip()[:5] or None
    janela_meta = _parse_painel_janela(data_inicio, data_fim, hora_inicio, hora_fim)
    inicio_j = janela_meta['inicio']
    fim_excl = janela_meta['fim_excl']
    filtros_resp = {
        'data': janela_meta['data'],
        'data_iso': janela_meta['data_iso'],
        'data_inicio': janela_meta['data_inicio_iso'],
        'data_fim': janela_meta['data_fim_iso'],
        'hora_inicio': janela_meta['hora_inicio'],
        'hora_fim': janela_meta['hora_fim'],
        'legivel': janela_meta['legivel'],
    }
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.row_factory = sqlite3.Row
    try:
        ini_p, fim_p = _painel_sql_janela_params(inicio_j, fim_excl, conn)
        filtro_dt = " AND COALESCE(fluxo, 'carregamento') = 'carregamento' AND data_hora >= ? AND data_hora < ?"
        params_dt = (ini_p, fim_p)
        # Estatísticas em uma única query (menos ida e volta ao DB)
        row_stats = conn.execute(f'''
        SELECT
            (SELECT COUNT(*) FROM produtos_bipados WHERE 1=1{filtro_dt}) AS total_bipados,
            (SELECT COUNT(*) FROM produtos_bipados WHERE status = 'CARREGADO'{filtro_dt}) AS total_carregados,
            (SELECT COUNT(DISTINCT codigo_barras) FROM produtos_bipados WHERE 1=1{filtro_dt}) AS total_unicos,
            (SELECT COUNT(DISTINCT id_viagem) FROM produtos_bipados WHERE id_viagem IS NOT NULL AND trim(COALESCE(id_viagem,'')) != ''{filtro_dt}) AS total_viagens,
            (SELECT COALESCE(SUM(quantidade), 0) FROM produtos_bipados WHERE 1=1{filtro_dt}) AS soma_quantidades
    ''', params_dt * 5).fetchone()
        veiculos_rows = conn.execute(
            f"SELECT veiculo, COUNT(*) as total FROM produtos_bipados WHERE status = ? AND trim(COALESCE(veiculo,'')) != ''{filtro_dt} GROUP BY veiculo",
            ('CARREGADO',) + params_dt,
        ).fetchall()
        def _val(r, k):
            if r is None:
                return 0
            try:
                v = r[k]
                return v if v is not None else 0
            except (TypeError, KeyError, IndexError):
                return 0
        estatisticas = {
            'total_bipados': _val(row_stats, 'total_bipados'),
            'total_carregados': _val(row_stats, 'total_carregados'),
            'total_unicos': _val(row_stats, 'total_unicos'),
            'total_divergencias': 0,
            'total_viagens': _val(row_stats, 'total_viagens'),
            'soma_quantidades': _val(row_stats, 'soma_quantidades'),
            'veiculos': [dict(r) for r in veiculos_rows]
        }
        # Viagens
        rows = conn.execute(f'''
        SELECT id_viagem, SUM(quantidade) as total_bipados, MIN(data_hora) as inicio, MAX(data_hora) as fim
        FROM produtos_bipados
        WHERE id_viagem IS NOT NULL AND id_viagem != ''{filtro_dt}
        GROUP BY id_viagem
        ORDER BY MAX(data_hora) DESC
        LIMIT 25
        ''', params_dt).fetchall()
        viagens = []
        ids_viagem = []
        for r in rows:
            vid = _normalizar_id_viagem(r['id_viagem'])
            if vid:
                ids_viagem.append(vid)
        periodo_map = {}
        if ids_viagem:
            _ensure_viagem_periodo_bipagem_table(conn)
            tbl_per = 'public.viagem_periodo_bipagem' if getattr(conn, 'kind', None) == 'pg' else 'viagem_periodo_bipagem'
            ph = ','.join('?' * len(ids_viagem))
            try:
                rows_per = conn.execute(
                    f"SELECT id_viagem, inicio_em, fim_em FROM {tbl_per} WHERE fluxo = 'carregamento' AND id_viagem IN ({ph})",
                    tuple(ids_viagem),
                ).fetchall()
                for rp in rows_per or []:
                    vid_p = _normalizar_id_viagem(rp.get('id_viagem') if hasattr(rp, 'get') else rp[0])
                    periodo_map[vid_p] = (
                        rp.get('inicio_em') if hasattr(rp, 'get') else rp[1],
                        rp.get('fim_em') if hasattr(rp, 'get') else rp[2],
                    )
            except Exception:
                pass
        for r in rows:
            vid = _normalizar_id_viagem(r['id_viagem'])
            inicio_raw, fim_raw = periodo_map.get(vid, (None, None))
            if not inicio_raw:
                inicio_raw = r['inicio'] or ''
            if not fim_raw:
                fim_raw = r['fim'] or ''
            d_min = _calcular_duracao_minutos(inicio_raw, fim_raw)
            viagens.append({
                'id_viagem': r['id_viagem'],
                'total_bipados': r['total_bipados'] or 0,
                'inicio': _formatar_data_hora_periodo(inicio_raw) or '',
                'fim': _formatar_data_hora_periodo(fim_raw) or '',
                'duracao_minutos': d_min
            })
        wb = None
        from_cache = False
        id_viagem_placa = {}
        romaneio_stats = {}
        if _usa_banco_para_dados() and getattr(conn, 'kind', None) == 'pg':
            try:
                romaneio_stats = _estatisticas_romaneio_por_item_banco(conn, inicio_j, fim_excl)
                id_viagem_placa = romaneio_stats.get('id_viagem_to_placa') or {}
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
                pass
            finally:
                # Em PostgreSQL, se _estatisticas_romaneio_por_item_banco falhar internamente e retornar sem raise,
                # a conexão fica em "transaction aborted". Garantir estado limpo para as queries seguintes.
                if getattr(conn, 'kind', None) == 'pg':
                    try:
                        conn.rollback()
                    except Exception:
                        pass
        if wb and not romaneio_stats:
            try:
                romaneio_stats = _estatisticas_romaneio_por_item(wb)
                id_viagem_placa = romaneio_stats.get('id_viagem_to_placa') or {}
            except Exception:
                pass
        placa_to_minutos = {}
        for v in viagens:
            vid = (v.get('id_viagem') or '').strip()
            placa = (id_viagem_placa.get(vid) or '').strip() or 'Sem placa'
            placa_to_minutos[placa] = placa_to_minutos.get(placa, 0) + (v.get('duracao_minutos') or 0)
        tempo_por_placa = [{'placa': p, 'total_minutos': m} for p, m in sorted(placa_to_minutos.items(), key=lambda x: -x[1])]
        # Top itens: GROUP BY codigo_barras, produto (PostgreSQL exige todas as colunas não agregadas no GROUP BY)
        rows_itens = conn.execute(
            f'SELECT produto, codigo_barras, SUM(quantidade) as total FROM produtos_bipados WHERE 1=1{filtro_dt} GROUP BY codigo_barras, produto ORDER BY total DESC LIMIT 15',
            params_dt,
        ).fetchall()
        top_itens = []
        for r in rows_itens:
            label = (r['produto'] or '').strip() or (r['codigo_barras'] or '') or '-'
            top_itens.append({'label': label[:50], 'total': r['total'] or 0})
        rows_carros = conn.execute(
            f"SELECT veiculo, SUM(quantidade) as total FROM produtos_bipados WHERE veiculo IS NOT NULL AND trim(veiculo) != ''{filtro_dt} GROUP BY veiculo ORDER BY total DESC LIMIT 15",
            params_dt,
        ).fetchall()
        carros_itens = [{'veiculo': r['veiculo'] or '', 'total': r['total'] or 0} for r in rows_carros]
        carros_peso = []
        if wb:
            try:
                mapa_peso = _build_mapa_peso_romaneio(wb)
                mapa_barras_codigo = _build_mapa_barras_to_codigo_produto(wb)
                rows_bipados = conn.execute(
                    f"SELECT veiculo, codigo_barras, quantidade FROM produtos_bipados WHERE veiculo IS NOT NULL AND trim(veiculo) != ''{filtro_dt}",
                    params_dt,
                ).fetchall()
                peso_por_veiculo = {}
                for r in rows_bipados:
                    veic = (r['veiculo'] or '').strip()
                    if not veic:
                        continue
                    cb = (r['codigo_barras'] or '').strip()
                    qtd = int(r['quantidade'] or 0)
                    cp = mapa_barras_codigo.get(cb) or cb
                    peso_unit = mapa_peso.get(cp, 0) or mapa_peso.get(cb, 0)
                    peso_por_veiculo[veic] = peso_por_veiculo.get(veic, 0.0) + qtd * peso_unit
                carros_peso = [{'veiculo': v, 'peso_total': round(p, 2)} for v, p in sorted(peso_por_veiculo.items(), key=lambda x: -x[1])[:15]]
            except Exception:
                pass
        elif romaneio_stats and romaneio_stats.get('peso_por_carro'):
            carros_peso = [{'veiculo': k, 'peso_total': v} for k, v in sorted(romaneio_stats['peso_por_carro'].items(), key=lambda x: -x[1])[:15]]
        erros_carregamento = {'resumo': {'total_itens_erro': 0, 'total_faltas': 0, 'total_sobras': 0, 'viagens_com_erro': 0, 'unidades_romaneio': 0, 'unidades_bipadas': 0, 'unidades_erro': 0, 'eficiencia_geral_pct': 100.0}, 'rows': [], 'por_viagem': [], 'correlacao': []}
        try:
            erros_carregamento = _painel_coletar_erros_carregamento(conn, ids_viagem, id_viagem_placa)
            faltas_map = erros_carregamento.get('faltas_por_viagem') or {}
            for v in viagens:
                vid_n = _normalizar_id_viagem(v.get('id_viagem'))
                v['total_faltas'] = int(faltas_map.get(vid_n, 0) or 0)
            estatisticas['total_divergencias'] = int((erros_carregamento.get('resumo') or {}).get('total_itens_erro') or 0)
        except Exception as ex_err:
            try:
                app.logger.warning('painel erros_carregamento: %s', ex_err)
            except Exception:
                pass
            for v in viagens:
                v['total_faltas'] = 0
        if wb and not from_cache:
            try:
                wb.close()
            except Exception:
                pass
        placas_dia = {'data': '', 'data_iso': '', 'rows': [], 'resumo': {'total': 0, 'carregados': 0, 'nao_carregados': 0, 'em_andamento': 0, 'peso_total_kg': 0}}
        try:
            placas_dia = _coletar_placas_baixadas_dia(conn, None, inicio_j, fim_excl)
            placas_dia['filtros'] = filtros_resp
        except Exception as ex_placas:
            try:
                app.logger.warning('painel placas_baixadas_dia: %s', ex_placas)
            except Exception:
                pass
        conn.close()
        romaneio_resp = {k: v for k, v in romaneio_stats.items() if k != 'id_viagem_to_placa'}
        return jsonify({
            'estatisticas': estatisticas,
            'viagens': viagens,
            'tempo_por_placa': tempo_por_placa,
            'top_itens_bipados': top_itens,
            'carros_mais_itens': carros_itens,
            'carros_mais_peso': carros_peso,
            'romaneio': romaneio_resp,
            'placas_baixadas_dia': placas_dia,
            'erros_carregamento': {
                'resumo': erros_carregamento.get('resumo') or {},
                'rows': erros_carregamento.get('rows') or [],
                'por_viagem': erros_carregamento.get('por_viagem') or [],
                'correlacao': erros_carregamento.get('correlacao') or [],
            },
            'filtros': filtros_resp,
        })
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        # Retorna 200 com dados vazios para o painel carregar; frontend pode exibir erro
        empty = {'total_bipados': 0, 'total_carregados': 0, 'total_unicos': 0, 'total_divergencias': 0, 'total_viagens': 0, 'soma_quantidades': 0, 'veiculos': []}
        return jsonify({
            'estatisticas': empty,
            'viagens': [],
            'tempo_por_placa': [],
            'top_itens_bipados': [],
            'carros_mais_itens': [],
            'carros_mais_peso': [],
            'romaneio': {},
            'placas_baixadas_dia': {'data': '', 'data_iso': '', 'rows': [], 'resumo': {'total': 0, 'carregados': 0, 'nao_carregados': 0, 'em_andamento': 0, 'peso_total_kg': 0}},
            'erros_carregamento': {'resumo': {'total_itens_erro': 0, 'total_faltas': 0, 'total_sobras': 0, 'viagens_com_erro': 0, 'unidades_romaneio': 0, 'unidades_bipadas': 0, 'unidades_erro': 0, 'eficiencia_geral_pct': 100.0}, 'rows': [], 'por_viagem': [], 'correlacao': []},
            'erro': str(e)
        }), 200

def _devolucao_nf_numero_de_obj(nf):
    if not isinstance(nf, dict):
        return ''
    return str(
        nf.get('numero') or nf.get('numeroNF') or nf.get('numeroNf') or nf.get('nf') or ''
    ).strip()


def _devolucao_item_codigo_produto(item):
    if not isinstance(item, dict):
        return ''
    prod = item.get('produto') or {}
    if not isinstance(prod, dict):
        prod = {}
    ref = item.get('referenciaItem')
    if isinstance(ref, dict):
        ref = ref.get('codigo') or ref.get('referencia') or ''
    raw = (
        item.get('codigo') or item.get('codigoProduto') or ref
        or prod.get('codigo') or prod.get('codigoInterno') or ''
    )
    return _normalizar_codigo_produto(raw) or str(raw or '').strip()


def _devolucao_codigos_equivalentes(cod_a, cod_b):
    a = _normalizar_codigo_produto(cod_a)
    b = _normalizar_codigo_produto(cod_b)
    if not a or not b:
        return False
    if a == b:
        return True
    va = set(_variantes_codigo_produto(a))
    vb = set(_variantes_codigo_produto(b))
    return bool(va & vb)


def _devolucao_buscar_em_mapa_codigo(cod, cod_map):
    cod = (cod or '').strip()
    if not cod or not cod_map:
        return None
    if cod in cod_map:
        return cod_map[cod]
    nc = _normalizar_codigo_produto(cod)
    if nc and nc in cod_map:
        return cod_map[nc]
    for k, v in cod_map.items():
        if _devolucao_codigos_equivalentes(k, cod):
            return v
    return None


def _devolucao_qtd_bipada_mapa(cod, bipado_map):
    cod = (cod or '').strip()
    if not cod:
        return 0
    q = int(bipado_map.get(cod, 0) or 0)
    if q > 0:
        return q
    nc = _normalizar_codigo_produto(cod)
    if nc:
        q = int(bipado_map.get(nc, 0) or 0)
        if q > 0:
            return q
    for k, v in (bipado_map or {}).items():
        if _devolucao_codigos_equivalentes(k, cod):
            q = max(q, int(v or 0))
    return q


def _devolucao_inferir_motivo_nf(codigos_nf, diverg_map, bipado_map):
    codigos_nf = [c for c in (codigos_nf or []) if c]
    if not codigos_nf:
        return 'parcial'

    def _item_bipado(cod):
        return _devolucao_qtd_bipada_mapa(cod, bipado_map) > 0

    for cod in codigos_nf:
        div = _devolucao_buscar_em_mapa_codigo(cod, diverg_map) or {}
        mot_raw = (div.get('motivo_divergencia') or '').lower()
        if 'reentrega' in mot_raw:
            return 'reentrega'
    if not any(_item_bipado(c) for c in codigos_nf):
        return 'total'
    return 'parcial'


def _devolucao_ids_consulta(conn, id_viagem):
    """IDs equivalentes (roteiro/viagem) para consultas de bipagem e Ravex."""
    ids = []
    seen = set()

    def _add(val):
        v = str(val or '').strip()
        if v and v not in seen:
            seen.add(v)
            ids.append(v)

    raw = (id_viagem or '').strip()
    _add(raw)
    _add(_normalizar_id_viagem(raw))
    try:
        ds = _get_latest_dataset_id(conn)
    except Exception:
        ds = None
    if ds and raw:
        try:
            rows = conn.execute(
                """SELECT DISTINCT TRIM(COALESCE(id_viagem::text, '')) AS iv,
                                  TRIM(COALESCE(id_roteiro::text, '')) AS ir
                   FROM romaneio_por_item
                   WHERE dataset_id = ? AND (
                       TRIM(COALESCE(id_viagem::text, '')) = ? OR TRIM(COALESCE(id_roteiro::text, '')) = ?
                   )""",
                (str(ds), raw, raw),
            ).fetchall()
        except Exception:
            rows = conn.execute(
                """SELECT DISTINCT TRIM(COALESCE(id_viagem, '')) AS iv,
                                  TRIM(COALESCE(id_roteiro, '')) AS ir
                   FROM romaneio_por_item
                   WHERE dataset_id = ? AND (
                       TRIM(COALESCE(id_viagem, '')) = ? OR TRIM(COALESCE(id_roteiro, '')) = ?
                   )""",
                (str(ds), raw, raw),
            ).fetchall()
        for r in rows or []:
            _add(r.get('iv') if hasattr(r, 'get') else (r[0] if len(r) > 0 else ''))
            _add(r.get('ir') if hasattr(r, 'get') else (r[1] if len(r) > 1 else ''))
        local = _ravex_resolver_id_db(conn, ds, raw)
        if local:
            _add(local.get('id_viagem'))
            _add(local.get('id_roteiro'))
    return ids


def _devolucao_diverg_por_codigo(cod, diverg_map, divergent_codes):
    cod = (cod or '').strip()
    if not cod:
        return None, False
    div = _devolucao_buscar_em_mapa_codigo(cod, diverg_map)
    if div:
        return div, True
    nc = _normalizar_codigo_produto(cod)
    for dc in (divergent_codes or set()):
        if _devolucao_codigos_equivalentes(dc, cod):
            return diverg_map.get(dc) or _devolucao_buscar_em_mapa_codigo(dc, diverg_map), True
    return None, False


def _devolucao_pseudo_div_nf_item(item, cod, bipado_map):
    prod = item.get('produto') or {} if isinstance(item, dict) else {}
    desc = str(
        item.get('descricaoItem') or item.get('descricao')
        or prod.get('descricao') or prod.get('descrição') or ''
    ).strip()
    qtd_nf = int(item.get('quantidade') or item.get('qtd') or 0)
    q_bip = _devolucao_qtd_bipada_mapa(cod, bipado_map)
    return {
        'codigo_produto': cod,
        'produto': desc,
        'quantidade_produto': qtd_nf,
        'quantidade_bipada': q_bip,
        'quantidade_falta': max(0, qtd_nf - q_bip),
        'codigo_barras': '-',
        'unidade': str(item.get('unidade') or prod.get('unidade') or '-'),
        'peso_bruto': '',
        'motivo_divergencia': '',
    }


def _devolucao_resolver_vid_ravex(conn, token, id_viagem):
    """Resolve id de viagem na Ravex e retorna (vid_api, nfs, ids_tentados)."""
    ids = _devolucao_ids_consulta(conn, id_viagem)
    if not ids:
        ids = [(id_viagem or '').strip()]
    try:
        id_v, id_r, _somente = _ravex_resolver_id_para_viagem(token, id_viagem)
        if id_v and id_v not in ids:
            ids.insert(0, id_v)
        if id_r and id_r not in ids:
            ids.append(id_r)
    except Exception:
        pass
    if not obter_notas_fiscais_viagem:
        return None, [], ids
    for vid in ids:
        try:
            vid_api = int(vid)
        except (TypeError, ValueError):
            vid_api = vid
        try:
            nfs = obter_notas_fiscais_viagem(token, vid_api) or []
            if nfs:
                return vid_api, nfs, ids
        except Exception:
            continue
    return None, [], ids


def _devolucao_merge_esperados_bipados(esperados, bipados):
    by_cod = {}

    def _key(cod):
        return _normalizar_codigo_produto(cod) or (cod or '').strip()

    for it in esperados or []:
        k = _key(it.get('codigo_produto'))
        if k:
            by_cod[k] = dict(it)
    for it in bipados or []:
        k = _key(it.get('codigo_produto'))
        if not k:
            continue
        q_bip = int(it.get('quantidade_bipada') or 0)
        if k in by_cod:
            exp = by_cod[k]
            q_ret = int(exp.get('quantidade_produto') or 0)
            exp['quantidade_bipada'] = q_bip
            exp['quantidade_falta'] = max(0, q_ret - q_bip)
            if it.get('codigo_barras'):
                exp['codigo_barras'] = it.get('codigo_barras')
            if it.get('produto') and not exp.get('produto'):
                exp['produto'] = it.get('produto')
            if q_bip > 0:
                exp['status_bipado'] = 'COMPLETO' if q_bip >= q_ret else 'PARCIAL'
            else:
                exp['status_bipado'] = 'PENDENTE'
        else:
            by_cod[k] = dict(it)
    return list(by_cod.values())


def _devolucao_mapa_bipado_carregamento(conn, id_viagem, id_norm=None):
    out = {}
    ids = _devolucao_ids_consulta(conn, id_viagem) or []
    if id_norm and id_norm not in ids:
        ids.append(id_norm)
    if not ids:
        ids = [(id_viagem or '').strip(), (id_norm or '').strip()]
    ids = [x for x in ids if x]
    if not ids:
        return out
    ph = ','.join(['?'] * len(ids))
    is_pg = getattr(conn, 'kind', None) == 'pg'
    col_viagem = 'TRIM(COALESCE(id_viagem::text, \'\'))' if is_pg else 'TRIM(COALESCE(id_viagem, \'\'))'
    rows = conn.execute(
        f'''SELECT codigo_interno, codigo_barras, SUM(quantidade) AS qtd
           FROM produtos_bipados
           WHERE {col_viagem} IN ({ph})
             AND COALESCE(fluxo, 'carregamento') = 'carregamento'
           GROUP BY codigo_interno, codigo_barras''',
        tuple(ids),
    ).fetchall()
    for r in rows or []:
        ci = ((r.get('codigo_interno') if hasattr(r, 'get') else r[0]) or '').strip()
        cb = ((r.get('codigo_barras') if hasattr(r, 'get') else (r[1] if len(r) > 1 else '')) or '').strip()
        q = int((r.get('qtd') if hasattr(r, 'get') else (r[2] if len(r) > 2 else 0)) or 0)
        for key in (ci, cb):
            if not key:
                continue
            out[key] = out.get(key, 0) + q
            nk = _normalizar_codigo_produto(key)
            if nk:
                out[nk] = out.get(nk, 0) + q
    return out


def _devolucao_lista_divergencias_carregamento(conn, id_viagem):
    id_viagem = (id_viagem or '').strip()
    id_norm = _normalizar_id_viagem(id_viagem) or id_viagem
    lista = _conferencia_lista_interna(
        id_viagem, 'carregamento', conn=conn, somente_divergencias=True,
    )
    if not lista and id_norm != id_viagem:
        lista = _conferencia_lista_interna(
            id_norm, 'carregamento', conn=conn, somente_divergencias=True,
        )
    _carregar_motivos_divergencia(lista, conn)
    diverg_map = {}
    divergent_codes = set()
    for it in lista or []:
        cod = (it.get('codigo_produto') or '').strip()
        if cod:
            divergent_codes.add(cod)
            diverg_map[cod] = dict(it)
    return id_norm, divergent_codes, diverg_map


def _devolucao_item_esperado_linha(cod, div, qtd_nf, id_viagem):
    q_rom = int(div.get('quantidade_produto') or 0) if div else int(qtd_nf or 0)
    q_bip = int(div.get('quantidade_bipada') or 0) if div else 0
    q_falta = int(div.get('quantidade_falta') or 0) if div else max(0, q_rom - q_bip)
    q_sobra = int(div.get('quantidade_sobra') or 0) if div else max(0, q_bip - q_rom)
    q_retorno = q_sobra if q_sobra > 0 else (q_falta if q_falta > 0 else int(qtd_nf or 0))
    if q_retorno <= 0:
        q_retorno = int(qtd_nf or 0) or q_rom
    return {
        'codigo_produto': cod,
        'codigo_barras': (div.get('codigo_barras') or '-') if div else '-',
        'produto': (div.get('produto') or '') if div else '',
        'quantidade_produto': q_retorno,
        'quantidade_bipada': 0,
        'quantidade_falta': q_retorno,
        'quantidade_sobra': 0,
        'quantidade_saida': q_bip,
        'unidade': (div.get('unidade') or '-') if div else '-',
        'peso_bruto': (div.get('peso_bruto') or '') if div else '',
        'status_bipado': 'PENDENTE',
        'aviso_sobra': '',
        'id_viagem': id_viagem,
        'origem_romaneio': True,
        'modo_devolucao_nf': True,
        'motivo_divergencia': (div.get('motivo_divergencia') or '') if div else '',
    }


def _devolucao_montar_itens_nf(conn, id_viagem, numero_nf, motivo_nf='', itens_ravex=None):
    """Itens esperados para bipagem de retorno de uma NF (Ravex ou fallback romaneio)."""
    id_viagem = (id_viagem or '').strip()
    numero_nf = str(numero_nf or '').strip()
    if not id_viagem or not numero_nf:
        return []
    motivo_nf = (motivo_nf or '').strip().lower()
    id_norm, divergent_codes, diverg_map = _devolucao_lista_divergencias_carregamento(conn, id_viagem)
    bipado_map = _devolucao_mapa_bipado_carregamento(conn, id_viagem, id_norm)
    id_grav = id_norm or id_viagem

    def _montar_de_itens(itens, motivo_efetivo):
        if not itens:
            return []
        if not motivo_efetivo:
            codigos = [_devolucao_item_codigo_produto(x) for x in itens]
            codigos = [c for c in codigos if c]
            motivo_efetivo = _devolucao_inferir_motivo_nf(codigos, diverg_map, bipado_map)
        incluir_todos = motivo_efetivo == 'total'
        saida = []
        for item in itens:
            cod = _devolucao_item_codigo_produto(item)
            if not cod:
                continue
            qtd_nf = int(item.get('quantidade') or item.get('qtd') or 0)
            div, is_div = _devolucao_diverg_por_codigo(cod, diverg_map, divergent_codes)
            if not incluir_todos and not is_div:
                continue
            if not div:
                div = _devolucao_pseudo_div_nf_item(item, cod, bipado_map)
            linha = _devolucao_item_esperado_linha(cod, div, qtd_nf, id_grav)
            if not linha.get('produto') and div.get('produto'):
                linha['produto'] = div['produto']
            saida.append(linha)
        return saida

    if itens_ravex is not None:
        return _montar_de_itens(itens_ravex, motivo_nf)

    if obter_notas_fiscais_viagem and obter_itens_nota_fiscal and ravex_get_token:
        try:
            token = ravex_get_token()
            vid_api, nfs, _ids = _devolucao_resolver_vid_ravex(conn, token, id_viagem)
            if vid_api and nfs:
                nf_id = None
                itens_nf = None
                for nf in nfs:
                    num = _devolucao_nf_numero_de_obj(nf)
                    if num == numero_nf or num.lstrip('0') == numero_nf.lstrip('0'):
                        nf_id = nf.get('id') or nf.get('Id')
                        if nf_id:
                            itens_nf = obter_itens_nota_fiscal(token, vid_api, nf_id) or []
                        break
                if itens_nf is not None:
                    montados = _montar_de_itens(itens_nf, motivo_nf)
                    if montados:
                        return montados
        except Exception:
            pass

    incluir_todos = motivo_nf == 'total'
    lista = _conferencia_lista_interna(
        id_viagem, 'carregamento', conn=conn, somente_divergencias=not incluir_todos,
    )
    if not lista and id_norm != id_viagem:
        lista = _conferencia_lista_interna(
            id_norm, 'carregamento', conn=conn, somente_divergencias=not incluir_todos,
        )
    _carregar_motivos_divergencia(lista, conn)
    saida = []
    for it in lista or []:
        cod = (it.get('codigo_produto') or '').strip()
        if not cod:
            continue
        if not incluir_todos:
            _div, is_div = _devolucao_diverg_por_codigo(cod, diverg_map, divergent_codes)
            if not is_div:
                continue
        q_rom = int(it.get('quantidade_produto') or 0)
        saida.append(_devolucao_item_esperado_linha(cod, it, q_rom, id_grav))
    return saida


def _devolucao_notas_divergentes_ravex(id_viagem, conn=None, numero_nf_filtro=None):
    """NFs da viagem (Ravex) que possuem ao menos um item divergente no carregamento."""
    id_viagem = (id_viagem or '').strip()
    if not id_viagem:
        return {'notas': [], 'aviso': 'Informe o roteiro/viagem.', 'total_divergencias': 0, 'id_viagem': ''}
    fechar = False
    if conn is None:
        conn = get_db()
        fechar = True
    try:
        id_norm, divergent_codes, diverg_map = _devolucao_lista_divergencias_carregamento(conn, id_viagem)
        bipado_map = _devolucao_mapa_bipado_carregamento(conn, id_viagem, id_norm)
        if not divergent_codes:
            return {
                'notas': [],
                'aviso': 'Nenhuma divergência encontrada no carregamento deste roteiro.',
                'total_divergencias': 0,
                'id_viagem': id_norm,
            }
        if not obter_notas_fiscais_viagem or not obter_itens_nota_fiscal or not ravex_get_token:
            return {
                'notas': [],
                'aviso': 'API Ravex indisponível para listar notas fiscais.',
                'total_divergencias': len(divergent_codes),
                'id_viagem': id_norm,
            }
        try:
            token = ravex_get_token()
        except Exception as e:
            return {
                'notas': [],
                'aviso': 'Falha ao autenticar na Ravex: %s' % str(e),
                'total_divergencias': len(divergent_codes),
                'id_viagem': id_norm,
            }
        vid_api, nfs, _ids = _devolucao_resolver_vid_ravex(conn, token, id_viagem)
        if not vid_api or not nfs:
            return {
                'notas': [],
                'aviso': 'Nenhuma NF encontrada na Ravex para este roteiro/viagem.',
                'total_divergencias': len(divergent_codes),
                'id_viagem': id_norm,
            }
        nf_pairs = []
        for nf in nfs:
            nf_id = nf.get('id') or nf.get('Id')
            numero = _devolucao_nf_numero_de_obj(nf)
            if nf_id and numero:
                if numero_nf_filtro and numero != str(numero_nf_filtro).strip():
                    if numero.lstrip('0') != str(numero_nf_filtro).strip().lstrip('0'):
                        continue
                nf_pairs.append((nf, nf_id, numero))
        itens_por_nf = {}
        if nf_pairs:
            def _fetch(pair):
                _nf, nf_id, _num = pair
                try:
                    return nf_id, obter_itens_nota_fiscal(token, vid_api, nf_id) or []
                except Exception:
                    return nf_id, []

            from concurrent.futures import ThreadPoolExecutor
            workers = min(8, max(2, len(nf_pairs)))
            with ThreadPoolExecutor(max_workers=workers) as pool:
                for nf_id, itens in pool.map(_fetch, nf_pairs):
                    itens_por_nf[nf_id] = itens
        notas_saida = []
        for nf, nf_id, numero in nf_pairs:
            itens = itens_por_nf.get(nf_id, [])
            codigos_nf = []
            tem_div = False
            for item in itens or []:
                cod = _devolucao_item_codigo_produto(item)
                if not cod:
                    continue
                codigos_nf.append(cod)
                _div, is_div = _devolucao_diverg_por_codigo(cod, diverg_map, divergent_codes)
                if is_div:
                    tem_div = True
            if not codigos_nf:
                continue
            motivo_sug = _devolucao_inferir_motivo_nf(codigos_nf, diverg_map, bipado_map)
            if not tem_div and motivo_sug != 'total':
                continue
            itens_div = _devolucao_montar_itens_nf(
                conn, id_viagem, numero, motivo_sug, itens_ravex=itens,
            )
            if not itens_div:
                continue
            notas_saida.append({
                'numero_nf': numero,
                'nota_fiscal_id': nf_id,
                'motivo_sugerido': motivo_sug,
                'motivo_label': _motivo_devolucao_label(motivo_sug),
                'qtd_itens_divergentes': len(itens_div),
                'qtd_itens_nf': len(itens or []),
                'itens': itens_div,
            })
        notas_saida.sort(key=lambda x: str(x.get('numero_nf') or ''))
        aviso = ''
        if not notas_saida and divergent_codes:
            aviso = 'Há divergência no carregamento, mas nenhuma NF da Ravex contém esses itens.'
        return {
            'notas': notas_saida,
            'aviso': aviso,
            'total_divergencias': len(divergent_codes),
            'id_viagem': id_norm,
        }
    finally:
        if fechar:
            try:
                conn.close()
            except Exception:
                pass


def _devolucao_itens_esperados_nf(conn, id_viagem, numero_nf, motivo_nf=''):
    return _devolucao_montar_itens_nf(conn, id_viagem, numero_nf, motivo_nf)


@app.route('/api/devolucoes/itens-nf', methods=['GET'])
def api_devolucoes_itens_nf():
    """Itens divergentes esperados para bipagem de retorno de uma NF (pré-visualização ou conferência)."""
    id_viagem = (request.args.get('id_viagem') or '').strip()
    numero_nf = (request.args.get('numero_nf') or '').strip()
    motivo = (request.args.get('motivo') or '').strip()
    if not id_viagem or not numero_nf:
        return jsonify({'erro': 'Informe id_viagem e numero_nf'}), 400
    conn = get_db()
    try:
        itens = _devolucao_itens_esperados_nf(conn, id_viagem, numero_nf, motivo)
        motivo_ef = motivo
        if not motivo_ef:
            id_norm, diverg_map, _ = _devolucao_lista_divergencias_carregamento(conn, id_viagem)
            bipado_map = _devolucao_mapa_bipado_carregamento(conn, id_viagem, id_norm)
            codigos = [(it.get('codigo_produto') or '').strip() for it in itens if (it.get('codigo_produto') or '').strip()]
            motivo_ef = _devolucao_inferir_motivo_nf(codigos, diverg_map, bipado_map)
        conn.close()
        return jsonify({
            'ok': True,
            'itens': itens,
            'numero_nf': numero_nf,
            'motivo_sugerido': motivo_ef,
            'motivo_label': _motivo_devolucao_label(motivo_ef),
            'qtd_itens': len(itens or []),
        })
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e), 'itens': []}), 500


@app.route('/api/devolucoes/notas-divergentes', methods=['GET'])
def api_devolucoes_notas_divergentes():
    """NFs do roteiro na Ravex que contêm itens divergentes no carregamento."""
    id_viagem = (request.args.get('id_viagem') or '').strip()
    if not id_viagem:
        return jsonify({'erro': 'Informe id_viagem'}), 400
    try:
        return jsonify(_devolucao_notas_divergentes_ravex(id_viagem))
    except Exception as e:
        try:
            app.logger.exception('notas-divergentes: %s', e)
        except Exception:
            pass
        return jsonify({'erro': str(e), 'notas': []}), 500


@app.route('/api/devolucoes/notas', methods=['GET'])
def api_devolucoes_notas_listar():
    """Lista NFs de devolução de um roteiro/viagem."""
    id_viagem = (request.args.get('id_viagem') or '').strip()
    if not id_viagem:
        return jsonify({'erro': 'Informe id_viagem'}), 400
    conn = get_db()
    try:
        _ensure_devolucao_nf_schema(conn)
        _ensure_pg_produtos_bipados_fluxo(conn)
        id_norm = _normalizar_id_viagem(id_viagem) or id_viagem
        ids = _devolucao_ids_consulta(conn, id_viagem) or [id_viagem, id_norm]
        ids = list(dict.fromkeys([x for x in ids if x]))
        if not ids:
            ids = [id_viagem]
        ph = ','.join(['?'] * len(ids))
        is_pg = getattr(conn, 'kind', None) == 'pg'
        col_viagem = 'TRIM(COALESCE(id_viagem::text, \'\'))' if is_pg else 'TRIM(COALESCE(id_viagem, \'\'))'
        rows = conn.execute(
            f'''SELECT id, id_viagem, numero_nf, motivo, status, doca, criado_em, concluida_em, criado_por, concluida_por
               FROM devolucao_nota_fiscal
               WHERE {col_viagem} IN ({ph})
               ORDER BY CASE WHEN status = 'em_andamento' THEN 0 ELSE 1 END, id DESC''',
            tuple(ids),
        ).fetchall()
        notas = []
        for row in rows or []:
            d = dict(row) if hasattr(row, 'keys') else {}
            if not d and row:
                d = {
                    'id': row[0], 'id_viagem': row[1], 'numero_nf': row[2], 'motivo': row[3],
                    'status': row[4], 'doca': row[5], 'criado_em': row[6], 'concluida_em': row[7],
                    'criado_por': row[8], 'concluida_por': row[9],
                }
            d['motivo_label'] = _motivo_devolucao_label(d.get('motivo'))
            d['status_label'] = _devolucao_status_label(d.get('status'))
            d['criado_em'] = _fmt_datahora_br(d.get('criado_em'))
            d['concluida_em'] = _fmt_datahora_br(d.get('concluida_em'))
            notas.append(d)
        return jsonify({'notas': notas, 'id_viagem': id_viagem})
    except Exception as e:
        try:
            app.logger.exception('devolucoes/notas listar: %s', e)
        except Exception:
            pass
        return jsonify({'erro': str(e), 'notas': []}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/devolucoes/notas', methods=['POST'])
def api_devolucoes_notas_criar():
    """Inicia sessão de bipagem para uma NF de devolução."""
    data = request.get_json() or {}
    id_viagem = (data.get('id_viagem') or '').strip()
    numero_nf = (data.get('numero_nf') or '').strip()
    motivo = (data.get('motivo') or '').strip().lower()
    doca = (data.get('doca') or '').strip()
    if not id_viagem or not numero_nf:
        return jsonify({'success': False, 'erro': 'Informe o roteiro/viagem e o número da NF.'}), 400
    if not _motivo_devolucao_valido(motivo):
        return jsonify({'success': False, 'erro': 'Selecione o motivo: parcial, total ou reentrega.'}), 400
    if doca not in ('1', '2', '3', '4'):
        doca = '1'
    conn = get_db()
    _ensure_devolucao_nf_schema(conn)
    id_norm = _normalizar_id_viagem(id_viagem) or id_viagem
    aberta = conn.execute(
        '''SELECT id FROM devolucao_nota_fiscal
           WHERE TRIM(COALESCE(id_viagem, '')) IN (?, ?) AND status = 'em_andamento' LIMIT 1''',
        (id_viagem, id_norm),
    ).fetchone()
    if aberta:
        conn.close()
        return jsonify({
            'success': False,
            'erro': 'Já existe uma NF em andamento neste roteiro. Conclua-a antes de iniciar outra.',
            'nota_aberta_id': aberta.get('id') if hasattr(aberta, 'get') else aberta[0],
        }), 400
    usuario = session.get('usuario', '') or ''
    agora = datetime.now(timezone.utc)
    if getattr(conn, 'kind', None) == 'pg':
        row = conn.execute(
            '''INSERT INTO devolucao_nota_fiscal (id_viagem, numero_nf, motivo, status, doca, criado_em, criado_por)
               VALUES (%s, %s, %s, 'em_andamento', %s, %s, %s) RETURNING id''',
            (id_norm, numero_nf, motivo, doca, agora, usuario),
        ).fetchone()
        nf_id = row.get('id') if hasattr(row, 'get') else row[0]
    else:
        conn.execute(
            '''INSERT INTO devolucao_nota_fiscal (id_viagem, numero_nf, motivo, status, doca, criado_em, criado_por)
               VALUES (?, ?, ?, 'em_andamento', ?, ?, ?)''',
            (id_norm, numero_nf, motivo, doca, _agora_iso(), usuario),
        )
        nf_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.commit()
    conn.close()
    return jsonify({
        'success': True,
        'nota': {
            'id': int(nf_id),
            'id_viagem': id_norm,
            'numero_nf': numero_nf,
            'motivo': motivo,
            'motivo_label': _motivo_devolucao_label(motivo),
            'status': 'em_andamento',
            'doca': doca,
        },
    })


@app.route('/api/devolucoes/notas/<int:nf_id>/concluir', methods=['POST'])
def api_devolucoes_notas_concluir(nf_id):
    """Marca NF de devolução como concluída."""
    conn = get_db()
    _ensure_devolucao_nf_schema(conn)
    row = conn.execute(
        'SELECT id, id_viagem, numero_nf, motivo, status FROM devolucao_nota_fiscal WHERE id = ?',
        (nf_id,),
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'erro': 'NF não encontrada.'}), 404
    st = (row.get('status') if hasattr(row, 'get') else row[4]) or ''
    if st == 'concluida':
        conn.close()
        return jsonify({'success': True, 'mensagem': 'NF já estava concluída.'})
    usuario = session.get('usuario', '') or ''
    agora = datetime.now(timezone.utc)
    if getattr(conn, 'kind', None) == 'pg':
        conn.execute(
            '''UPDATE devolucao_nota_fiscal SET status = 'concluida', concluida_em = %s, concluida_por = %s WHERE id = %s''',
            (agora, usuario, nf_id),
        )
    else:
        conn.execute(
            '''UPDATE devolucao_nota_fiscal SET status = 'concluida', concluida_em = ?, concluida_por = ? WHERE id = ?''',
            (_agora_iso(), usuario, nf_id),
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': nf_id})


@app.route('/api/devolucoes/notas/<int:nf_id>/cancelar', methods=['POST'])
def api_devolucoes_notas_cancelar(nf_id):
    """Cancela NF em andamento: apaga bipagens desta NF e libera início de outra."""
    conn = get_db()
    _ensure_devolucao_nf_schema(conn)
    _ensure_pg_produtos_bipados_fluxo(conn)
    try:
        row = conn.execute(
            'SELECT id, id_viagem, numero_nf, status FROM devolucao_nota_fiscal WHERE id = ?',
            (nf_id,),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'erro': 'NF não encontrada.'}), 404
        st = (row.get('status') if hasattr(row, 'get') else row[3]) or ''
        if st != 'em_andamento':
            conn.close()
            return jsonify({'success': False, 'erro': 'Só é possível cancelar NF em andamento.'}), 400
        conn.execute(
            """DELETE FROM produtos_bipados
               WHERE devolucao_nf_id = ? AND COALESCE(fluxo, 'carregamento') = 'devolucao'""",
            (nf_id,),
        )
        conn.execute(
            """UPDATE devolucao_nota_fiscal SET status = 'cancelada' WHERE id = ?""",
            (nf_id,),
        )
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            app.logger.exception('devolucoes/notas cancelar: %s', e)
        except Exception:
            pass
        return jsonify({'success': False, 'erro': str(e)}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass
    _broadcast_atualizar()
    return jsonify({
        'success': True,
        'id': nf_id,
        'mensagem': 'NF cancelada. Você pode iniciar outra nota fiscal.',
    })


@app.route('/api/devolucoes/conferencia/<id_viagem>/zerar', methods=['POST'])
def api_devolucoes_zerar_conferencia(id_viagem):
    """Zera toda a conferência de devoluções do roteiro/viagem e cancela NFs em andamento."""
    id_viagem = (id_viagem or '').strip()
    if not id_viagem:
        return jsonify({'erro': 'Informe o roteiro/viagem.'}), 400
    conn = get_db()
    _ensure_devolucao_nf_schema(conn)
    _ensure_pg_produtos_bipados_fluxo(conn)
    removidos = 0
    try:
        ids = _devolucao_ids_consulta(conn, id_viagem) or [id_viagem]
        id_norm = _normalizar_id_viagem(id_viagem) or id_viagem
        if id_norm not in ids:
            ids.append(id_norm)
        ids = [x for x in ids if x]
        if not ids:
            ids = [id_viagem]
        ph = ','.join(['?'] * len(ids))
        is_pg = getattr(conn, 'kind', None) == 'pg'
        col_viagem = 'TRIM(COALESCE(id_viagem::text, \'\'))' if is_pg else 'TRIM(COALESCE(id_viagem, \'\'))'
        cur = conn.execute(
            f"""DELETE FROM produtos_bipados
               WHERE {col_viagem} IN ({ph})
                 AND COALESCE(fluxo, 'carregamento') = 'devolucao'""",
            tuple(ids),
        )
        removidos = getattr(cur, 'rowcount', None) if cur else 0
        conn.execute(
            f"""UPDATE devolucao_nota_fiscal SET status = 'cancelada'
               WHERE {col_viagem} IN ({ph}) AND status = 'em_andamento'""",
            tuple(ids),
        )
        for vid in ids:
            _limpar_periodo_bipagem(conn, vid, 'devolucao')
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            app.logger.exception('devolucoes/conferencia zerar: %s', e)
        except Exception:
            pass
        return jsonify({'success': False, 'erro': str(e)}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass
    _broadcast_atualizar()
    return jsonify({
        'success': True,
        'mensagem': 'Conferência de devoluções zerada. NFs em andamento foram canceladas.',
        'removidos': removidos,
    })


@app.route('/api/devolucoes/conferencia/<id_viagem>', methods=['GET'])
def api_devolucoes_conferencia_nf(id_viagem):
    """Conferência do retorno apenas da NF ativa (sem carregar romaneio inteiro)."""
    devolucao_nf_id = request.args.get('devolucao_nf_id', '').strip()
    if not devolucao_nf_id:
        return jsonify({'erro': 'Informe devolucao_nf_id'}), 400
    id_viagem = (id_viagem or '').strip()
    conn = get_db()
    _ensure_devolucao_nf_schema(conn)
    _ensure_pg_produtos_bipados_fluxo(conn)
    try:
        nf_row = conn.execute(
            'SELECT id, id_viagem, numero_nf, motivo, status, doca FROM devolucao_nota_fiscal WHERE id = ?',
            (int(devolucao_nf_id),),
        ).fetchone()
        if not nf_row:
            conn.close()
            return jsonify({'erro': 'NF não encontrada.'}), 404
        nf_d = dict(nf_row) if hasattr(nf_row, 'keys') else {
            'id': nf_row[0], 'id_viagem': nf_row[1], 'numero_nf': nf_row[2],
            'motivo': nf_row[3], 'status': nf_row[4], 'doca': nf_row[5],
        }
        motivo_nf = nf_d.get('motivo') or ''
        ids_viagem = _devolucao_ids_consulta(conn, id_viagem) or [id_viagem]
        nf_id_v = (nf_d.get('id_viagem') or '').strip()
        if nf_id_v and nf_id_v not in ids_viagem:
            ids_viagem.insert(0, nf_id_v)
        lista_bip = []
        for vid in ids_viagem:
            lb = _devolucao_conferencia_lista_nf(conn, vid, devolucao_nf_id)
            if lb:
                lista_bip = lb
                break
        esperados = _devolucao_montar_itens_nf(
            conn, id_viagem, nf_d.get('numero_nf'), motivo_nf,
        )
        if not esperados and nf_id_v:
            esperados = _devolucao_montar_itens_nf(
                conn, nf_id_v, nf_d.get('numero_nf'), motivo_nf,
            )
        if esperados:
            lista = _devolucao_merge_esperados_bipados(esperados, lista_bip)
        else:
            lista = lista_bip
        id_norm = _normalizar_id_viagem(id_viagem) or id_viagem
        info = _get_viagem_info_dict(id_norm)
        meta = {
            'placa': info.get('placa', ''),
            'motorista': info.get('motorista', ''),
            'identificador_rota': info.get('identificador_rota', ''),
            'data_expedicao': info.get('data_expedicao', ''),
            'id_roteiro': info.get('id_roteiro', '') if info.get('id_roteiro') else '',
        }
        conn.close()
        return jsonify({
            'lista': lista,
            'id_viagem': id_viagem,
            'devolucao_nf': {
                'id': nf_d.get('id'),
                'numero_nf': nf_d.get('numero_nf'),
                'motivo': nf_d.get('motivo'),
                'motivo_label': _motivo_devolucao_label(nf_d.get('motivo')),
                'status': nf_d.get('status'),
                'doca': nf_d.get('doca'),
            },
            'modo_devolucao_nf': True,
            'placa': meta.get('placa', ''),
            'motorista': meta.get('motorista', ''),
            'identificador_rota': meta.get('identificador_rota', ''),
            'data_expedicao': meta.get('data_expedicao', ''),
            'id_roteiro': meta.get('id_roteiro', ''),
        })
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


def _estoque_sp_rows_to_list(rows):
    out = []
    for r in rows or []:
        d = dict(r) if hasattr(r, 'keys') else {}
        if not d and r:
            d = {
                'codigo_produto': r[0],
                'produto': r[1] if len(r) > 1 else '',
                'codigo_barras': r[2] if len(r) > 2 else '',
                'quantidade': r[3] if len(r) > 3 else 0,
                'registros': r[4] if len(r) > 4 else 0,
            }
        out.append({
            'codigo_produto': (d.get('codigo_produto') or '').strip(),
            'produto': (d.get('produto') or '').strip(),
            'codigo_barras': (d.get('codigo_barras') or '').strip(),
            'quantidade': float(d.get('quantidade') or 0),
            'registros': int(d.get('registros') or d.get('nfs') or 0),
        })
    return out


def _estoque_sp_chave_item(item):
    cod = (item.get('codigo_produto') or '').strip()
    if cod:
        return 'p:' + cod
    cb = (item.get('codigo_barras') or '').strip()
    if cb:
        return 'b:' + cb
    return ''


def _estoque_sp_calcular_atual(saida_l, dev_l, ter_l):
    """Saldo = entradas (devolução + terceiros) − saída (expedição), por produto."""
    mapa = {}

    def _slot(chave):
        if chave not in mapa:
            mapa[chave] = {
                'codigo_produto': '',
                'produto': '',
                'codigo_barras': '',
                'qtd_saida': 0.0,
                'qtd_entrada_devolucao': 0.0,
                'qtd_entrada_terceiros': 0.0,
                'saldo': 0.0,
            }
        return mapa[chave]

    for it in saida_l or []:
        ch = _estoque_sp_chave_item(it)
        if not ch:
            continue
        s = _slot(ch)
        s['codigo_produto'] = s['codigo_produto'] or it.get('codigo_produto') or ''
        s['produto'] = s['produto'] or it.get('produto') or ''
        s['codigo_barras'] = s['codigo_barras'] or it.get('codigo_barras') or ''
        s['qtd_saida'] += float(it.get('quantidade') or 0)

    for it in dev_l or []:
        ch = _estoque_sp_chave_item(it)
        if not ch:
            continue
        s = _slot(ch)
        s['codigo_produto'] = s['codigo_produto'] or it.get('codigo_produto') or ''
        s['produto'] = s['produto'] or it.get('produto') or ''
        s['codigo_barras'] = s['codigo_barras'] or it.get('codigo_barras') or ''
        s['qtd_entrada_devolucao'] += float(it.get('quantidade') or 0)

    for it in ter_l or []:
        ch = _estoque_sp_chave_item(it)
        if not ch:
            continue
        s = _slot(ch)
        s['codigo_produto'] = s['codigo_produto'] or it.get('codigo_produto') or ''
        s['produto'] = s['produto'] or it.get('produto') or ''
        s['codigo_barras'] = s['codigo_barras'] or it.get('codigo_barras') or ''
        s['qtd_entrada_terceiros'] += float(it.get('quantidade') or 0)

    itens = []
    for s in mapa.values():
        s['saldo'] = s['qtd_entrada_devolucao'] + s['qtd_entrada_terceiros'] - s['qtd_saida']
        if abs(s['saldo']) < 1e-9 and s['qtd_saida'] <= 0 and s['qtd_entrada_devolucao'] <= 0 and s['qtd_entrada_terceiros'] <= 0:
            continue
        itens.append(s)
    itens.sort(key=lambda x: (-abs(x['saldo']), x.get('produto') or ''))
    return {
        'titulo': 'Estoque atual (entradas − saídas)',
        'itens': itens,
        'total_linhas': len(itens),
        'total_saldo': sum(x['saldo'] for x in itens),
        'total_saida': sum(x['qtd_saida'] for x in itens),
        'total_entrada_devolucao': sum(x['qtd_entrada_devolucao'] for x in itens),
        'total_entrada_terceiros': sum(x['qtd_entrada_terceiros'] for x in itens),
    }


def _estoque_sp_coletar_movimentos(conn, filtros=None):
    _ensure_produtos_bipados_disposicao(conn)
    filtro_disp = _estoque_sp_sql_apenas_disposicao_normal()
    dt_ini, dt_fim = _estoque_sp_bounds_datahora(filtros or {}, conn)
    filtro_data_saida, params_saida = _estoque_sp_sql_filtro_data_hora('data_hora', dt_ini, dt_fim, conn)
    filtro_data_dev = filtro_data_saida
    params_dev = list(params_saida)

    saida = conn.execute(
        f'''SELECT COALESCE(codigo_interno, '') AS codigo_produto, COALESCE(MAX(produto), '') AS produto,
                   COALESCE(codigo_barras, '') AS codigo_barras, COALESCE(SUM(quantidade), 0) AS quantidade,
                   COUNT(*) AS registros
            FROM produtos_bipados
            WHERE COALESCE(fluxo, 'carregamento') = 'carregamento' {filtro_disp} {filtro_data_saida}
            GROUP BY codigo_interno, codigo_barras
            ORDER BY quantidade DESC
            LIMIT 2000''',
        tuple(params_saida),
    ).fetchall()

    entrada_dev = conn.execute(
        f'''SELECT COALESCE(codigo_interno, '') AS codigo_produto, COALESCE(MAX(produto), '') AS produto,
                   COALESCE(codigo_barras, '') AS codigo_barras, COALESCE(SUM(quantidade), 0) AS quantidade,
                   COUNT(*) AS registros
            FROM produtos_bipados
            WHERE COALESCE(fluxo, 'carregamento') = 'devolucao' {filtro_disp} {filtro_data_dev}
            GROUP BY codigo_interno, codigo_barras
            ORDER BY quantidade DESC
            LIMIT 2000''',
        tuple(params_dev),
    ).fetchall()

    tbl_d = _tbl_terceiros_documentos(conn)
    tbl_i = _tbl_terceiros_documento_itens(conn)
    cod_prod_ter = "COALESCE(NULLIF(TRIM(i.codigo_produto_base), ''), NULLIF(TRIM(i.codigo_produto_xml), ''), '')"
    desc_prod_ter = "COALESCE(NULLIF(TRIM(i.descricao_base), ''), NULLIF(TRIM(i.descricao_xml), ''), '')"
    filtro_data_ter, params_ter = _estoque_sp_sql_filtro_data_hora('d.recebimento_concluido_em', dt_ini, dt_fim, conn)
    entrada_ter = conn.execute(
        f'''SELECT COALESCE(i.codigo_ean, '') AS codigo_barras,
                   {cod_prod_ter} AS codigo_produto,
                   COALESCE(MAX({desc_prod_ter}), '') AS produto,
                   COALESCE(SUM(i.quantidade_bipada), 0) AS quantidade,
                   COUNT(DISTINCT d.id) AS nfs
            FROM {tbl_i} i
            INNER JOIN {tbl_d} d ON d.id = i.documento_id
            WHERE (
                    d.recebimento_concluido IS TRUE
                    OR LOWER(TRIM(COALESCE(CAST(d.recebimento_concluido AS TEXT), ''))) IN ('1', 'true', 'sim', 't', 'yes')
                  )
              AND LOWER(TRIM(COALESCE(d.consumivel_sp, ''))) NOT IN ('sim', 's', 'true', '1', 'yes')
              AND LOWER(TRIM(COALESCE(d.enviar_para_mg, ''))) NOT IN ('sim', 's', 'true', '1', 'yes')
              AND COALESCE(i.quantidade_bipada, 0) > 0
              {filtro_data_ter}
            GROUP BY i.codigo_ean, {cod_prod_ter}
            ORDER BY quantidade DESC
            LIMIT 2000''',
        tuple(params_ter),
    ).fetchall()

    saida_l = _estoque_sp_filtrar_itens_lista(_estoque_sp_rows_to_list(saida), filtros)
    dev_l = _estoque_sp_filtrar_itens_lista(_estoque_sp_rows_to_list(entrada_dev), filtros)
    ter_l = _estoque_sp_filtrar_itens_lista(_estoque_sp_rows_to_list(entrada_ter), filtros)
    return saida_l, dev_l, ter_l


def _estoque_sp_coletar_por_disposicao(conn, disposicao, filtros=None):
    """Itens classificados (reentrega, avaria, descarte_perdas, palete_bloqueado)."""
    _ensure_produtos_bipados_disposicao(conn)
    _ensure_devolucao_nf_schema(conn)
    is_pg = getattr(conn, 'kind', None) == 'pg'
    tbl_nf = 'public.devolucao_nota_fiscal' if is_pg else 'devolucao_nota_fiscal'
    dt_ini, dt_fim = _estoque_sp_bounds_datahora(filtros or {}, conn)
    filtro_data, params = _estoque_sp_sql_filtro_data_hora('pb.data_hora', dt_ini, dt_fim, conn)
    if disposicao == 'reentrega':
        where = f"""(
            COALESCE(pb.disposicao_estoque, '') = 'reentrega'
            OR EXISTS (
                SELECT 1 FROM {tbl_nf} nf
                WHERE nf.id = pb.devolucao_nf_id
                  AND LOWER(COALESCE(nf.motivo, '')) = 'reentrega'
            )
        )"""
    else:
        where = "COALESCE(pb.disposicao_estoque, '') = ?"
        params.insert(0, disposicao)

    rows = conn.execute(
        f'''SELECT COALESCE(pb.codigo_interno, '') AS codigo_produto,
                   COALESCE(MAX(pb.produto), '') AS produto,
                   COALESCE(pb.codigo_barras, '') AS codigo_barras,
                   COALESCE(SUM(pb.quantidade), 0) AS quantidade,
                   COUNT(*) AS registros,
                   COALESCE(SUM(CASE WHEN COALESCE(pb.fluxo, 'carregamento') = 'devolucao'
                        THEN pb.quantidade ELSE 0 END), 0) AS qtd_entrada,
                   COALESCE(SUM(CASE WHEN COALESCE(pb.fluxo, 'carregamento') = 'carregamento'
                        THEN pb.quantidade ELSE 0 END), 0) AS qtd_saida
            FROM produtos_bipados pb
            WHERE {where} {filtro_data}
            GROUP BY pb.codigo_interno, pb.codigo_barras
            ORDER BY quantidade DESC
            LIMIT 2000''',
        tuple(params),
    ).fetchall()
    out = []
    for r in rows or []:
        d = _row_dict(r) if hasattr(r, 'keys') else None
        if not d and r:
            d = {
                'codigo_produto': r[0],
                'produto': r[1] if len(r) > 1 else '',
                'codigo_barras': r[2] if len(r) > 2 else '',
                'quantidade': r[3] if len(r) > 3 else 0,
                'registros': r[4] if len(r) > 4 else 0,
                'qtd_entrada': r[5] if len(r) > 5 else 0,
                'qtd_saida': r[6] if len(r) > 6 else 0,
            }
        out.append({
            'codigo_produto': (d.get('codigo_produto') or '').strip(),
            'produto': (d.get('produto') or '').strip(),
            'codigo_barras': (d.get('codigo_barras') or '').strip(),
            'quantidade': float(d.get('quantidade') or 0),
            'registros': int(d.get('registros') or 0),
            'qtd_entrada': float(d.get('qtd_entrada') or 0),
            'qtd_saida': float(d.get('qtd_saida') or 0),
        })
    return _estoque_sp_filtrar_itens_lista(out, filtros)


def _estoque_sp_secao_disposicao(titulo, itens):
    return {
        'titulo': titulo,
        'itens': itens,
        'total_quantidade': sum(x['quantidade'] for x in itens),
        'total_linhas': len(itens),
    }


@app.route('/api/estoque-sp/tempo-real', methods=['GET'])
def api_estoque_sp_tempo_real():
    """Estoque atual em tempo real: saldo por produto (todas as movimentações)."""
    conn = get_db()
    filtros = _estoque_sp_parse_filtros_query()
    try:
        _estoque_sp_prepare_conn(conn)
        saida_l, dev_l, ter_l = _estoque_sp_coletar_movimentos(conn, filtros)
        atual = _estoque_sp_calcular_atual(saida_l, dev_l, ter_l)
        conn.close()
        return jsonify({
            'atualizado_em': _fmt_datahora_br(datetime.now(timezone.utc)),
            'estoque_atual': atual,
            'filtros': filtros,
        })
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


@app.route('/api/estoque-sp/resumo', methods=['GET'])
def api_estoque_sp_resumo():
    """Estoque SP: saída (carregamento), entrada devoluções, entrada terceiros (sem MG, sem consumível)."""
    conn = get_db()
    filtros = _estoque_sp_parse_filtros_query()
    try:
        _estoque_sp_prepare_conn(conn)
        saida_l, dev_l, ter_l = _estoque_sp_coletar_movimentos(conn, filtros)
        reent_l = _estoque_sp_coletar_por_disposicao(conn, 'reentrega', filtros)
        avar_l = _estoque_sp_coletar_por_disposicao(conn, 'avaria', filtros)
        desc_l = _estoque_sp_coletar_por_disposicao(conn, 'descarte_perdas', filtros)
        pal_l = _estoque_sp_coletar_por_disposicao(conn, 'palete_bloqueado', filtros)
        conn.close()
        return jsonify({
            'filtros': filtros,
            'saida': {
                'titulo': 'Saída (expedição bipada)',
                'itens': saida_l,
                'total_quantidade': sum(x['quantidade'] for x in saida_l),
                'total_linhas': len(saida_l),
            },
            'entrada_devolucao': {
                'titulo': 'Entrada (devoluções bipadas)',
                'itens': dev_l,
                'total_quantidade': sum(x['quantidade'] for x in dev_l),
                'total_linhas': len(dev_l),
            },
            'entrada_terceiros': {
                'titulo': 'Entrada recebimento terceiros (sem MG, sem consumível SP)',
                'itens': ter_l,
                'total_quantidade': sum(x['quantidade'] for x in ter_l),
                'total_linhas': len(ter_l),
            },
            'reentregas': _estoque_sp_secao_disposicao('Reentregas', reent_l),
            'avaria': _estoque_sp_secao_disposicao('Avaria (descarte avariado)', avar_l),
            'descarte_perdas': _estoque_sp_secao_disposicao('Descarte / perdas', desc_l),
            'palete_bloqueado': _estoque_sp_secao_disposicao('Palete bloqueado', pal_l),
        })
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'erro': str(e)}), 500


@app.route('/api/devolucoes/painel', methods=['GET'])
def get_painel_devolucoes():
    """Resumo do painel de devoluções, separado por fluxo=devolucao."""
    conn = get_db()
    if getattr(conn, 'kind', None) != 'pg':
        conn.row_factory = sqlite3.Row
    try:
        _, ini_p, fim_p, filtros_resp = _painel_request_janela_meta(conn)
        filtro_dt = ' AND data_hora >= ? AND data_hora < ?'
        params_dt = (ini_p, fim_p)
        fluxo = 'devolucao'
        row_stats = conn.execute(f'''
        SELECT
            COUNT(*) AS total_bipados,
            COALESCE(SUM(quantidade), 0) AS soma_quantidades,
            COUNT(DISTINCT codigo_barras) AS total_unicos,
            COUNT(DISTINCT CASE WHEN id_viagem IS NOT NULL AND trim(COALESCE(id_viagem,'')) != '' THEN id_viagem END) AS total_viagens,
            COUNT(DISTINCT CASE WHEN doca IS NOT NULL AND trim(COALESCE(doca,'')) != '' THEN doca END) AS total_docas,
            COUNT(DISTINCT CASE WHEN usuario_bipagem IS NOT NULL AND trim(COALESCE(usuario_bipagem,'')) != '' THEN usuario_bipagem END) AS total_usuarios
        FROM produtos_bipados
        WHERE COALESCE(fluxo, 'carregamento') = ?{filtro_dt}
        ''', (fluxo,) + params_dt).fetchone()

        viagens_rows = conn.execute(f'''
        SELECT
            id_viagem,
            COUNT(*) AS registros,
            COALESCE(SUM(quantidade), 0) AS qtd_devolvida,
            COUNT(DISTINCT codigo_barras) AS itens_unicos,
            MIN(data_hora) AS inicio,
            MAX(data_hora) AS fim
        FROM produtos_bipados
        WHERE COALESCE(fluxo, 'carregamento') = ?
          AND id_viagem IS NOT NULL
          AND trim(COALESCE(id_viagem,'')) != ''{filtro_dt}
        GROUP BY id_viagem
        ORDER BY MAX(data_hora) DESC
        LIMIT 30
        ''', (fluxo,) + params_dt).fetchall()

        itens_rows = conn.execute(f'''
        SELECT produto, codigo_barras, COALESCE(SUM(quantidade), 0) AS total
        FROM produtos_bipados
        WHERE COALESCE(fluxo, 'carregamento') = ?{filtro_dt}
        GROUP BY codigo_barras, produto
        ORDER BY total DESC
        LIMIT 20
        ''', (fluxo,) + params_dt).fetchall()

        veiculos_rows = conn.execute(f'''
        SELECT veiculo, COUNT(*) AS registros, COALESCE(SUM(quantidade), 0) AS total
        FROM produtos_bipados
        WHERE COALESCE(fluxo, 'carregamento') = ?
          AND veiculo IS NOT NULL
          AND trim(COALESCE(veiculo,'')) != ''{filtro_dt}
        GROUP BY veiculo
        ORDER BY total DESC, registros DESC
        LIMIT 20
        ''', (fluxo,) + params_dt).fetchall()

        docas_rows = conn.execute(f'''
        SELECT doca, COUNT(*) AS registros, COALESCE(SUM(quantidade), 0) AS total
        FROM produtos_bipados
        WHERE COALESCE(fluxo, 'carregamento') = ?
          AND doca IS NOT NULL
          AND trim(COALESCE(doca,'')) != ''{filtro_dt}
        GROUP BY doca
        ORDER BY total DESC, registros DESC
        ''', (fluxo,) + params_dt).fetchall()

        usuarios_rows = conn.execute(f'''
        SELECT usuario_bipagem, COUNT(*) AS registros, COALESCE(SUM(quantidade), 0) AS total
        FROM produtos_bipados
        WHERE COALESCE(fluxo, 'carregamento') = ?
          AND usuario_bipagem IS NOT NULL
          AND trim(COALESCE(usuario_bipagem,'')) != ''{filtro_dt}
        GROUP BY usuario_bipagem
        ORDER BY total DESC, registros DESC
        LIMIT 20
        ''', (fluxo,) + params_dt).fetchall()

        def _safe(r, key, default=0):
            try:
                val = r[key]
            except Exception:
                return default
            return default if val is None else val

        viagens = []
        for r in viagens_rows:
            inicio, fim = _safe(r, 'inicio', ''), _safe(r, 'fim', '')
            d_min = _calcular_duracao_minutos(inicio, fim)
            viagens.append({
                'id_viagem': _safe(r, 'id_viagem', ''),
                'registros': _safe(r, 'registros', 0),
                'qtd_devolvida': _safe(r, 'qtd_devolvida', 0),
                'itens_unicos': _safe(r, 'itens_unicos', 0),
                'inicio': _formatar_data_hora_periodo(inicio) or '',
                'fim': _formatar_data_hora_periodo(fim) or '',
                'duracao_minutos': d_min,
            })

        conn.close()
        return jsonify({
            'filtros': filtros_resp,
            'estatisticas': {
                'total_bipados': _safe(row_stats, 'total_bipados', 0),
                'soma_quantidades': _safe(row_stats, 'soma_quantidades', 0),
                'total_unicos': _safe(row_stats, 'total_unicos', 0),
                'total_viagens': _safe(row_stats, 'total_viagens', 0),
                'total_docas': _safe(row_stats, 'total_docas', 0),
                'total_usuarios': _safe(row_stats, 'total_usuarios', 0),
            },
            'viagens': viagens,
            'top_itens': [{'produto': _safe(r, 'produto', ''), 'codigo_barras': _safe(r, 'codigo_barras', ''), 'total': _safe(r, 'total', 0)} for r in itens_rows],
            'veiculos': [{'veiculo': _safe(r, 'veiculo', ''), 'registros': _safe(r, 'registros', 0), 'total': _safe(r, 'total', 0)} for r in veiculos_rows],
            'docas': [{'doca': _safe(r, 'doca', ''), 'registros': _safe(r, 'registros', 0), 'total': _safe(r, 'total', 0)} for r in docas_rows],
            'usuarios': [{'usuario': _safe(r, 'usuario_bipagem', ''), 'registros': _safe(r, 'registros', 0), 'total': _safe(r, 'total', 0)} for r in usuarios_rows],
        })
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({
            'estatisticas': {
                'total_bipados': 0,
                'soma_quantidades': 0,
                'total_unicos': 0,
                'total_viagens': 0,
                'total_docas': 0,
                'total_usuarios': 0,
            },
            'viagens': [],
            'top_itens': [],
            'veiculos': [],
            'docas': [],
            'usuarios': [],
            'erro': str(e),
        }), 200


def _agora_iso():
    return datetime.now().isoformat(timespec='seconds')


def _fmt_datahora_br(valor):
    if not valor:
        return ''
    if hasattr(valor, 'strftime'):
        return valor.strftime('%d/%m/%Y %H:%M:%S')
    dt = _parse_datetime(valor)
    if dt:
        return dt.strftime('%d/%m/%Y %H:%M:%S')
    return str(valor)


def _fmt_data_br(valor):
    if not valor:
        return ''
    dt = _parse_datetime(valor)
    if dt:
        return dt.strftime('%d/%m/%Y')
    s = str(valor).strip()
    return s[:10] if len(s) >= 10 else s


def _somente_digitos(valor):
    return re.sub(r'\D+', '', str(valor or ''))


def _fmt_cnpj_cpf_terceiros(valor):
    d = _somente_digitos(valor)
    if len(d) == 14:
        return '%s.%s.%s/%s-%s' % (d[0:2], d[2:5], d[5:8], d[8:12], d[12:14])
    if len(d) == 11:
        return '%s.%s.%s-%s' % (d[0:3], d[3:6], d[6:9], d[9:11])
    return d or '-'


def _fmt_chave_nfe_terceiros(chave):
    d = _somente_digitos(chave)
    if len(d) != 44:
        return chave or '-'
    return ' '.join(d[i:i + 4] for i in range(0, 44, 4))


def _fmt_moeda_br_terceiros(valor):
    try:
        v = float(str(valor or '').replace(',', '.'))
    except (TypeError, ValueError):
        return str(valor or '-')
    return ('{:,.2f}').format(v).replace(',', 'X').replace('.', ',').replace('X', '.')


def _render_danfe_html_terceiros(doc):
    from html import escape
    xml_texto = (doc.get('xml_conteudo') or '').strip()
    parsed = {}
    if xml_texto:
        try:
            parsed = _parse_nfe_xml(xml_texto)
        except Exception:
            parsed = {}
    chave = (doc.get('chave_nfe') or parsed.get('chave_nfe') or '').strip()
    numero_nf = (doc.get('numero_nf') or parsed.get('numero_nf') or '-').strip()
    serie_nf = (doc.get('serie_nf') or parsed.get('serie_nf') or '-').strip()
    data_emissao = _fmt_data_br(doc.get('data_emissao') or parsed.get('data_emissao') or '')
    remetente = (doc.get('remetente_nome') or parsed.get('remetente_nome') or '-').strip()
    rem_cnpj = _fmt_cnpj_cpf_terceiros(parsed.get('remetente_cnpj') or '')
    destinatario = (doc.get('destinatario_nome') or parsed.get('destinatario_nome') or '-').strip()
    dest_cnpj = _fmt_cnpj_cpf_terceiros(parsed.get('destinatario_cnpj') or '')
    dest_uf = (doc.get('destinatario_uf') or parsed.get('destinatario_uf') or '-').strip()
    pedido = (doc.get('numero_pedido') or parsed.get('numero_pedido') or '-').strip()
    valor_nf = parsed.get('valor_total_xml') or ''
    if valor_nf not in ('', None):
        valor_nf = _fmt_moeda_br_terceiros(valor_nf)
    else:
        valor_nf = '-'
    itens = doc.get('itens') or parsed.get('itens') or []
    linhas_itens = []
    for item in itens:
        if not isinstance(item, dict):
            continue
        linhas_itens.append(
            '<tr>'
            + '<td>' + escape(str(item.get('n_item') or '')) + '</td>'
            + '<td>' + escape(str(item.get('codigo_produto_xml') or item.get('codigo_ean') or '')) + '</td>'
            + '<td>' + escape(str(item.get('descricao_xml') or item.get('descricao_base') or '')) + '</td>'
            + '<td class="num">' + escape(str(item.get('unidade_xml') or '')) + '</td>'
            + '<td class="num">' + escape(str(item.get('quantidade_xml') or '')) + '</td>'
            + '</tr>'
        )
    if not linhas_itens:
        linhas_itens.append('<tr><td colspan="5" style="text-align:center;color:#666;">Sem itens no XML</td></tr>')
    url_sefaz = ''
    if len(_somente_digitos(chave)) == 44:
        url_sefaz = 'https://www.nfe.fazenda.gov.br/portal/consultaRecaptcha.aspx?tipoConsulta=resumo&tipoConteudo=7PhJ+gAVw2g=&nfe=' + escape(_somente_digitos(chave))
    bloco_sefaz = (
        '<p class="aviso"><a href="' + url_sefaz + '" target="_blank" rel="noopener">Consultar NF-e na Receita Federal</a></p>'
        if url_sefaz else ''
    )
    titulo_nf = escape('NF ' + numero_nf + ' / Série ' + serie_nf)
    return '''<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>''' + titulo_nf + ''' — DANFE</title>
<style>
  body { font-family: Arial, Helvetica, sans-serif; margin: 24px; color: #212121; font-size: 13px; }
  h1 { font-size: 18px; margin: 0 0 4px 0; color: #366092; }
  .sub { color: #607d8b; margin: 0 0 20px 0; font-size: 12px; }
  .grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }
  .box { border: 1px solid #cfd8dc; border-radius: 6px; padding: 12px; }
  .box h2 { margin: 0 0 8px 0; font-size: 12px; text-transform: uppercase; color: #546e7a; }
  .box p { margin: 4px 0; }
  table { width: 100%; border-collapse: collapse; margin-top: 8px; }
  th, td { border: 1px solid #b0bec5; padding: 6px 8px; text-align: left; }
  th { background: #eceff1; font-size: 11px; text-transform: uppercase; }
  td.num { text-align: right; white-space: nowrap; }
  .chave { font-family: Consolas, monospace; font-size: 11px; word-break: break-all; }
  .toolbar { margin-bottom: 16px; display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }
  .btn { background: #366092; color: #fff; border: none; padding: 8px 16px; border-radius: 4px; cursor: pointer; font-size: 14px; }
  .btn:hover { background: #2a4d73; }
  .aviso { font-size: 12px; color: #455a64; margin-top: 16px; }
  @media print { .toolbar { display: none; } body { margin: 12px; } }
</style>
</head>
<body>
<div class="toolbar">
  <button type="button" class="btn" onclick="window.print()">Imprimir / Salvar como PDF</button>
  <span class="aviso">Use a impressora «Salvar como PDF» se quiser arquivo PDF.</span>
</div>
<h1>DANFE — Documento auxiliar da NF-e</h1>
<p class="sub">''' + titulo_nf + ''' · Pedido: ''' + escape(pedido) + '''</p>
<div class="grid">
  <div class="box">
    <h2>Emitente</h2>
    <p><strong>''' + escape(remetente) + '''</strong></p>
    <p>CNPJ/CPF: ''' + escape(rem_cnpj) + '''</p>
  </div>
  <div class="box">
    <h2>Destinatário</h2>
    <p><strong>''' + escape(destinatario) + '''</strong></p>
    <p>CNPJ/CPF: ''' + escape(dest_cnpj) + ''' · UF: ''' + escape(dest_uf) + '''</p>
  </div>
</div>
<div class="box" style="margin-bottom: 20px;">
  <h2>Dados da nota</h2>
  <p><strong>Emissão:</strong> ''' + escape(data_emissao) + ''' &nbsp;·&nbsp; <strong>Valor NF:</strong> R$ ''' + escape(str(valor_nf)) + '''</p>
  <p class="chave"><strong>Chave de acesso:</strong> ''' + escape(_fmt_chave_nfe_terceiros(chave)) + '''</p>
</div>
<h2 style="font-size: 14px; margin: 0 0 8px 0;">Itens da nota</h2>
<table>
  <thead>
    <tr><th>#</th><th>Código</th><th>Descrição</th><th>Un.</th><th>Qtd</th></tr>
  </thead>
  <tbody>
    ''' + ''.join(linhas_itens) + '''
  </tbody>
</table>
''' + bloco_sefaz + '''
</body>
</html>'''


def _texto_xml(elem):
    return (elem.text or '').strip() if elem is not None and elem.text is not None else ''


def _tag_local(tag):
    if not tag:
        return ''
    return tag.split('}', 1)[-1] if '}' in tag else tag


def _find_child_text_by_local(parent, local_name):
    """Primeiro elemento com tag local (ignora namespace) e texto não vazio."""
    if parent is None:
        return ''
    alvo = (local_name or '').strip()
    if not alvo:
        return ''
    for el in parent.iter():
        if _tag_local(el.tag) == alvo:
            txt = _texto_xml(el)
            if txt:
                return txt
    return ''


_PEDIDO_INVALIDOS = frozenset({
    '', '-', '—', 'na', 'n/a', 'n.a', 'nao', 'não', 'none', 'null',
    's/n', 'sn', 'sem', 'sem pedido', '0',
})


def _numero_pedido_valido(valor):
    p = (valor or '').strip()
    if not p:
        return False
    pl = p.lower().replace('.', '').replace('/', '')
    if pl in _PEDIDO_INVALIDOS:
        return False
    if len(p) <= 2 and not any(ch.isdigit() for ch in p):
        return False
    return True


def _normalizar_numero_pedido_terceiros(valor):
    p = (valor or '').strip().strip('|').strip()
    return p if _numero_pedido_valido(p) else ''


def _find_first(parent, paths, ns):
    if parent is None:
        return ''
    for path in paths:
        el = parent.find(path, ns)
        if el is not None and _texto_xml(el):
            return _texto_xml(el)
    return ''


def _parse_decimal(valor):
    if valor is None:
        return 0.0
    s = str(valor).strip().replace(',', '.')
    try:
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def _status_bipagem_terceiros(qtd_xml, qtd_bipada):
    qtd_xml = float(qtd_xml or 0)
    qtd_bipada = float(qtd_bipada or 0)
    if qtd_bipada <= 0:
        return 'PENDENTE'
    if qtd_bipada < qtd_xml:
        return 'PARCIAL'
    if qtd_bipada == qtd_xml:
        return 'COMPLETO'
    return 'EXCEDENTE'


def _resolver_item_base_por_ean(codigo_ean):
    codigo_busca = str(codigo_ean or '').strip()
    if not codigo_busca:
        return None
    resultado = buscar_produto_na_planilha(codigo_busca)
    if not resultado:
        return None
    return {
        'codigo_produto_base': (resultado.get('codigo_produto') or '').strip(),
        'codigo_barras_base': (resultado.get('codigo_barras') or '').strip(),
        'descricao_base': (resultado.get('produto') or '').strip(),
    }


def _terceiros_criar_resolver_base(conn):
    """Resolver por EAN com cache na mesma conexão (evita abrir DB por item no upload)."""
    cache = {}

    def _from_row(r, codigo_busca):
        row = dict(r) if hasattr(r, 'keys') else r
        return {
            'codigo_produto_base': (row.get('codigo_interno') or '').strip(),
            'codigo_barras_base': codigo_busca,
            'descricao_base': (row.get('descricao') or '').strip(),
        }

    if _usa_banco_para_dados() and conn:
        ds = _get_latest_dataset_id(conn)
        if ds:

            def resolver_db(codigo_ean):
                codigo_busca = str(codigo_ean or '').strip()
                if not codigo_busca:
                    return None
                if codigo_busca in cache:
                    return cache[codigo_busca]
                base = None
                try:
                    row = conn.execute(
                        """SELECT codigo_interno, ean, dun, descricao FROM base_codigo_barras
                           WHERE dataset_id = ? AND (ean = ? OR dun = ?) LIMIT 1""",
                        (str(ds), codigo_busca, codigo_busca),
                    ).fetchone()
                    if row:
                        base = _from_row(row, codigo_busca)
                except Exception:
                    pass
                cache[codigo_busca] = base
                return base

            return resolver_db

    def resolver_legado(codigo_ean):
        codigo_busca = str(codigo_ean or '').strip()
        if not codigo_busca:
            return None
        if codigo_busca in cache:
            return cache[codigo_busca]
        base = _resolver_item_base_por_ean(codigo_busca)
        cache[codigo_busca] = base
        return base

    return resolver_legado


def _base_terceiros_para_bipagem(item_d, codigo_ean_solicitado):
    """Dados para gravar na linha do item na bipagem.

    Se o EAN não existir na planilha local, usa código/descrição do XML para não bloquear
    a contagem e não “zerar” o bipado ao recarregar a página.
    """
    codigo_ean_solicitado = str(codigo_ean_solicitado or '').strip()
    base = _resolver_item_base_por_ean(codigo_ean_solicitado)
    if base:
        return base
    ean_xml = (item_d.get('codigo_ean') or '').strip()
    return {
        'codigo_produto_base': (item_d.get('codigo_produto_xml') or '').strip(),
        'codigo_barras_base': ean_xml or codigo_ean_solicitado,
        'descricao_base': (item_d.get('descricao_xml') or '').strip(),
    }


def _extrair_uf_destinatario_nfe(dest, infnfe, ns):
    if dest is not None:
        for bloco in ('enderDest', 'entrega', 'enderEntrega'):
            if ns:
                sub = dest.find('nfe:' + bloco, ns)
            else:
                sub = dest.find(bloco)
            uf = _find_first(sub, ['nfe:UF', 'UF'], ns) if sub is not None else ''
            if not uf:
                uf = _find_child_text_by_local(sub, 'UF') if sub is not None else ''
            if uf:
                return uf.strip().upper()[:2]
        uf = _find_first(dest, ['nfe:UF', 'UF'], ns)
        if not uf:
            uf = _find_child_text_by_local(dest, 'UF')
        if uf:
            return uf.strip().upper()[:2]
    if infnfe is not None:
        for el in infnfe.iter():
            if _tag_local(el.tag) != 'enderDest':
                continue
            uf = _find_child_text_by_local(el, 'UF')
            if uf:
                return uf.strip().upper()[:2]
    return ''


def _extrair_pedido_obs_cont(infnfe, ns):
    info_adic = infnfe.find('nfe:infAdic', ns) if ns else infnfe.find('infAdic')
    if info_adic is None:
        return ''
    obs_list = info_adic.findall('nfe:obsCont', ns) if ns else info_adic.findall('obsCont')
    for obs in obs_list or []:
        campo = (_find_first(obs, ['nfe:xCampo', 'xCampo'], ns) or '').strip().lower()
        texto = _normalizar_numero_pedido_terceiros(_find_first(obs, ['nfe:xTexto', 'xTexto'], ns))
        if not texto:
            continue
        if 'ped' in campo or campo in ('xped', 'oc', 'ordem', 'pedido', 'n_ped', 'nped'):
            return texto
    return ''


def _extrair_numero_pedido_nfe(infnfe, ide, ns):
    candidatos = []
    dets = infnfe.findall('nfe:det', ns) if ns else infnfe.findall('det')
    for det in dets or []:
        prod = det.find('nfe:prod', ns) if ns else det.find('prod')
        if prod is None:
            continue
        for tag in ('xPed', 'nItemPed'):
            xp = _normalizar_numero_pedido_terceiros(_find_first(prod, ['nfe:' + tag, tag], ns))
            if xp:
                candidatos.append(xp)
    if ide is not None:
        compra = ide.find('nfe:compra', ns) if ns else ide.find('compra')
        if compra is not None:
            for tag in ('xPed', 'nPed'):
                xp = _normalizar_numero_pedido_terceiros(_find_first(compra, ['nfe:' + tag, tag], ns))
                if xp:
                    candidatos.append(xp)
    if candidatos:
        candidatos.sort(key=lambda x: (len(x), x), reverse=True)
        return candidatos[0]
    obs = _extrair_pedido_obs_cont(infnfe, ns)
    if obs:
        return obs
    info_adic = infnfe.find('nfe:infAdic', ns) if ns else infnfe.find('infAdic')
    for tag in ('infCpl', 'infAdFisco'):
        info_txt = _find_first(info_adic, ['nfe:' + tag, tag], ns)
        ped = _extrair_numero_pedido_inf_cpl(info_txt)
        if ped:
            return ped
    return ''


def _parse_nfe_xml(xml_texto, resolver_base=None):
    try:
        root = ET.fromstring(xml_texto)
    except Exception as e:
        raise ValueError('XML inválido: %s' % str(e))
    ns_uri = ''
    if root.tag.startswith('{') and '}' in root.tag:
        ns_uri = root.tag[1:].split('}', 1)[0]
    ns = {'nfe': ns_uri} if ns_uri else {}
    infnfe = root.find('.//nfe:infNFe', ns) if ns else root.find('.//infNFe')
    if infnfe is None:
        infnfe = root
    emit = infnfe.find('nfe:emit', ns) if ns else infnfe.find('emit')
    dest = infnfe.find('nfe:dest', ns) if ns else infnfe.find('dest')
    ide = infnfe.find('nfe:ide', ns) if ns else infnfe.find('ide')
    total = infnfe.find('nfe:total', ns) if ns else infnfe.find('total')

    itens = []
    dets = infnfe.findall('nfe:det', ns) if ns else infnfe.findall('det')
    for idx, det in enumerate(dets, 1):
        prod = det.find('nfe:prod', ns) if ns else det.find('prod')
        if prod is None:
            continue
        codigo_ean = _find_first(prod, ['nfe:cEAN', 'cEAN', 'nfe:cEANTrib', 'cEANTrib'], ns)
        codigo_produto_xml = _find_first(prod, ['nfe:cProd', 'cProd'], ns)
        descricao_xml = _find_first(prod, ['nfe:xProd', 'xProd'], ns)
        unidade_xml = _find_first(prod, ['nfe:uCom', 'uCom', 'nfe:uTrib', 'uTrib'], ns)
        quantidade_xml = _parse_decimal(_find_first(prod, ['nfe:qCom', 'qCom', 'nfe:qTrib', 'qTrib'], ns))
        if resolver_base:
            base = resolver_base(codigo_ean)
        else:
            base = _resolver_item_base_por_ean(codigo_ean)
        itens.append({
            'n_item': idx,
            'codigo_ean': codigo_ean,
            'codigo_produto_xml': codigo_produto_xml,
            'descricao_xml': descricao_xml,
            'unidade_xml': unidade_xml,
            'quantidade_xml': quantidade_xml,
            'codigo_produto_base': (base or {}).get('codigo_produto_base', ''),
            'codigo_barras_base': (base or {}).get('codigo_barras_base', ''),
            'descricao_base': (base or {}).get('descricao_base', ''),
        })

    numero_pedido = _extrair_numero_pedido_nfe(infnfe, ide, ns)
    destinatario_uf = _extrair_uf_destinatario_nfe(dest, infnfe, ns)

    return {
        'chave_nfe': (infnfe.attrib.get('Id') or '').replace('NFe', '').strip(),
        'numero_nf': _find_first(ide, ['nfe:nNF', 'nNF'], ns),
        'serie_nf': _find_first(ide, ['nfe:serie', 'serie'], ns),
        'data_emissao': _find_first(ide, ['nfe:dhEmi', 'dhEmi', 'nfe:dEmi', 'dEmi'], ns),
        'remetente_nome': _find_first(emit, ['nfe:xNome', 'xNome'], ns),
        'remetente_cnpj': _find_first(emit, ['nfe:CNPJ', 'CNPJ', 'nfe:CPF', 'CPF'], ns),
        'destinatario_nome': _find_first(dest, ['nfe:xNome', 'xNome'], ns),
        'destinatario_cnpj': _find_first(dest, ['nfe:CNPJ', 'CNPJ', 'nfe:CPF', 'CPF'], ns),
        'destinatario_uf': destinatario_uf,
        'numero_pedido': numero_pedido,
        'valor_total_xml': _find_first(total, ['nfe:ICMSTot/nfe:vNF', 'ICMSTot/vNF'], ns),
        'itens': itens,
    }


def _extrair_numero_pedido_inf_cpl(texto):
    if not texto:
        return ''
    patterns = (
        r'Numero do Pedido do Cliente:\s*([^\|\n\r;]+)',
        r'N[uú]mero\s+do\s+Pedido[^:]*:\s*([^\|\n\r;]+)',
        r'(?:ORDEM\s+DE\s+COMPRA|O\.?C\.?)\s*(?:n[º°\.]|N[º°]|No\.?|#)?\s*:?\s*([0-9][0-9A-Za-z\-\/]*)',
        r'(?:xPed|XPED)\s*[=:]\s*([0-9][0-9A-Za-z\-\/]*)',
        r'\bPED\.?\s*(?:CLIENTE|COMPRA)?\s*[=:#]?\s*([0-9][0-9A-Za-z\-\/]*)',
        r'Pedido\s*(?:do\s+Cliente\s*)?(?:n[º°\.]|N[º°]|No\.?|#)?\s*:+\s*([0-9][0-9A-Za-z\-\/]*)',
        r'\bOC\s*[=:#]?\s*([0-9][0-9A-Za-z\-\/]*)',
    )
    for pat in patterns:
        match = re.search(pat, texto, re.IGNORECASE)
        if match:
            valor = _normalizar_numero_pedido_terceiros(match.group(1))
            if valor:
                return valor
    return ''


def _terceiros_precisa_backfill_pedido(valor):
    p = (valor or '').strip()
    return not _numero_pedido_valido(p)


def _terceiros_precisa_backfill_uf(valor):
    uf = (valor or '').strip().upper()
    return len(uf) != 2


def _terceiros_backfill_campos_xml_pendentes(conn, limite=300):
    """Atualiza numero_pedido e destinatario_uf no banco a partir do XML (listagem não lê xml_conteudo)."""
    limite = max(1, min(int(limite or 300), 500))
    tbl = _tbl_terceiros_documentos(conn)
    rows = conn.execute(
        'SELECT id, numero_pedido, destinatario_uf, xml_conteudo FROM ' + tbl
        + ' WHERE xml_conteudo IS NOT NULL AND TRIM(xml_conteudo) != \'\''
        + ' ORDER BY id DESC LIMIT ?',
        (limite,)
    ).fetchall()
    atualizados = 0
    for r in rows or []:
        row = dict(r) if hasattr(r, 'keys') else {
            'id': r[0], 'numero_pedido': r[1], 'destinatario_uf': r[2], 'xml_conteudo': r[3],
        }
        doc_id = row.get('id')
        xml_texto = row.get('xml_conteudo') or ''
        if not doc_id or not xml_texto:
            continue
        precisa_ped = _terceiros_precisa_backfill_pedido(row.get('numero_pedido'))
        precisa_uf = _terceiros_precisa_backfill_uf(row.get('destinatario_uf'))
        if not precisa_ped and not precisa_uf:
            continue
        try:
            parsed = _parse_nfe_xml(xml_texto)
        except Exception:
            continue
        novo_ped = (parsed.get('numero_pedido') or '').strip() if precisa_ped else ''
        nova_uf = (parsed.get('destinatario_uf') or '').strip().upper() if precisa_uf else ''
        if precisa_ped and novo_ped:
            conn.execute(
                'UPDATE ' + tbl + ' SET numero_pedido = ? WHERE id = ?',
                (novo_ped, doc_id),
            )
            atualizados += 1
        if precisa_uf and nova_uf:
            conn.execute(
                'UPDATE ' + tbl + ' SET destinatario_uf = ? WHERE id = ?',
                (nova_uf, doc_id),
            )
            atualizados += 1
    if atualizados:
        conn.commit()
    return atualizados


def _terceiros_backfill_numero_pedido_pendentes(conn, limite=150):
    return _terceiros_backfill_campos_xml_pendentes(conn, limite=limite)


def _terceiros_enriquecer_campos_xml_listagem(conn, rows, limite=60):
    """Reparse XML só para linhas com pedido/UF inválidos na listagem (sem carregar xml de todas as NFs)."""
    if not rows:
        return
    limite = max(1, min(int(limite or 60), 120))
    tbl = _tbl_terceiros_documentos(conn)
    ids = []
    for r in rows:
        row = dict(r) if hasattr(r, 'keys') else {}
        if not row.get('id'):
            continue
        if _terceiros_precisa_backfill_pedido(row.get('numero_pedido')) or _terceiros_precisa_backfill_uf(row.get('destinatario_uf')):
            ids.append(int(row['id']))
    if not ids:
        return
    ids = ids[:limite]
    ph = ','.join(['?'] * len(ids))
    extra = conn.execute(
        'SELECT id, numero_pedido, destinatario_uf, xml_conteudo FROM ' + tbl
        + ' WHERE id IN (' + ph + ') AND xml_conteudo IS NOT NULL AND TRIM(xml_conteudo) != \'\'',
        tuple(ids),
    ).fetchall()
    alterou = False
    by_id = {int(dict(r)['id']): dict(r) for r in (extra or []) if dict(r).get('id')}
    for r in rows:
        row = dict(r) if hasattr(r, 'keys') else {}
        doc_id = row.get('id')
        if doc_id not in by_id:
            continue
        src = by_id[doc_id]
        xml_texto = src.get('xml_conteudo') or ''
        if not xml_texto:
            continue
        precisa_ped = _terceiros_precisa_backfill_pedido(row.get('numero_pedido'))
        precisa_uf = _terceiros_precisa_backfill_uf(row.get('destinatario_uf'))
        if not precisa_ped and not precisa_uf:
            continue
        try:
            parsed = _parse_nfe_xml(xml_texto)
        except Exception:
            continue
        if precisa_ped:
            novo_ped = (parsed.get('numero_pedido') or '').strip()
            if novo_ped:
                row['numero_pedido'] = novo_ped
                conn.execute('UPDATE ' + tbl + ' SET numero_pedido = ? WHERE id = ?', (novo_ped, doc_id))
                alterou = True
        if precisa_uf:
            nova_uf = (parsed.get('destinatario_uf') or '').strip().upper()
            if len(nova_uf) == 2:
                row['destinatario_uf'] = nova_uf
                conn.execute('UPDATE ' + tbl + ' SET destinatario_uf = ? WHERE id = ?', (nova_uf, doc_id))
                alterou = True
    if alterou:
        conn.commit()


def _terceiros_atualizar_campo_xml_doc(conn, documento_id, campo, valor):
    if not conn or not documento_id or not valor:
        return
    if campo not in ('numero_pedido', 'destinatario_uf'):
        return
    try:
        conn.execute(
            'UPDATE ' + _tbl_terceiros_documentos(conn) + ' SET ' + campo + ' = ? WHERE id = ?',
            (valor, documento_id),
        )
        conn.commit()
    except Exception:
        pass


def _numero_pedido_terceiros(doc, conn=None, documento_id=None):
    if not doc:
        return ''
    col = _normalizar_numero_pedido_terceiros(doc.get('numero_pedido'))
    if col:
        return col
    xml_texto = doc.get('xml_conteudo') or ''
    if not xml_texto:
        return ''
    try:
        ped = (_parse_nfe_xml(xml_texto).get('numero_pedido') or '').strip()
    except Exception:
        ped = ''
    doc_id = documento_id or doc.get('id')
    if ped and conn is not None and doc_id:
        _terceiros_atualizar_campo_xml_doc(conn, doc_id, 'numero_pedido', ped)
    return ped


def _uf_destinatario_terceiros(doc, conn=None, documento_id=None):
    if not doc:
        return ''
    uf = (doc.get('destinatario_uf') or '').strip().upper()
    if len(uf) == 2:
        return uf
    xml_texto = doc.get('xml_conteudo') or ''
    if not xml_texto:
        return uf
    try:
        parsed = _parse_nfe_xml(xml_texto)
        uf = (parsed.get('destinatario_uf') or '').strip().upper()
    except Exception:
        uf = ''
    doc_id = documento_id or doc.get('id')
    if len(uf) == 2 and conn is not None and doc_id:
        _terceiros_atualizar_campo_xml_doc(conn, doc_id, 'destinatario_uf', uf)
    return uf


def _valor_bool_texto(valor):
    txt = (valor or '').strip().lower()
    return txt if txt in ('sim', 'nao') else ''


def _terceiros_bool_sim_db(valor):
    if valor is True or valor == 1:
        return True
    txt = str(valor or '').strip().lower()
    return txt in ('sim', 's', 'true', '1', 'yes')


def _terceiros_doc_tem_bipagem_iniciada(conn, documento_id):
    tbl_i = _tbl_terceiros_documento_itens(conn)
    row = conn.execute(
        'SELECT COALESCE(SUM(quantidade_bipada), 0) AS total FROM ' + tbl_i + ' WHERE documento_id = ?',
        (documento_id,),
    ).fetchone()
    if not row:
        return False
    try:
        total = float(row['total'] if hasattr(row, 'keys') else row[0])
    except (TypeError, ValueError):
        total = 0.0
    return total > 1e-9


def _terceiros_pode_lancar_nota_sem_confirmacao_recebimento(conn, doc, documento_id):
    if _terceiros_bool_sim_db(doc.get('recebimento_concluido')):
        return True
    return _terceiros_doc_tem_bipagem_iniciada(conn, documento_id)


def _registrar_evento_terceiros(conn, documento_id, evento, valor_anterior='', valor_novo='', usuario='', detalhes=''):
    conn.execute(
        '''INSERT INTO ''' + _tbl_terceiros_documento_eventos(conn) + ''' (documento_id, evento, valor_anterior, valor_novo, usuario, criado_em, detalhes)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (documento_id, evento, valor_anterior or '', valor_novo or '', usuario or '', _agora_iso(), detalhes or '')
    )


def _criar_documento_terceiros(conn, area, previsao_chegada, arquivo_nome, xml_texto, xml_data, usuario, motorista_carreta_ini=None, placa_carreta_ini=None):
    mot = (motorista_carreta_ini or '').strip() or None
    plc = (placa_carreta_ini or '').strip().upper() or None
    mot_em = _agora_iso() if mot else None
    consumivel_sp = 'sim' if _valor_bool_texto(xml_data.get('consumivel_sp') or '') == 'sim' else ''
    recebedor_consumivel_sp = (xml_data.get('recebedor_consumivel_sp') or '').strip() if consumivel_sp == 'sim' else ''
    if getattr(conn, 'kind', None) == 'pg':
        row = conn.execute(
            '''INSERT INTO ''' + _tbl_terceiros_documentos(conn) + ''' (
                   area, chave_nfe, numero_nf, serie_nf, data_emissao, remetente_nome, remetente_cnpj,
                   destinatario_nome, destinatario_cnpj, destinatario_uf, numero_pedido, previsao_chegada, arquivo_nome, xml_conteudo,
                   motorista_carreta, motorista_carreta_em, placa_carreta,
                   consumivel_sp, recebedor_consumivel_sp,
                   criado_em, criado_por, atualizado_em, atualizado_por
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) RETURNING id''',
            (
                area, xml_data.get('chave_nfe') or '', xml_data.get('numero_nf') or '', xml_data.get('serie_nf') or '',
                xml_data.get('data_emissao') or '', xml_data.get('remetente_nome') or '', _somente_digitos(xml_data.get('remetente_cnpj') or ''),
                xml_data.get('destinatario_nome') or '', _somente_digitos(xml_data.get('destinatario_cnpj') or ''),
                (xml_data.get('destinatario_uf') or '').strip(),
                (xml_data.get('numero_pedido') or '').strip(),
                previsao_chegada or '', arquivo_nome or '', xml_texto,
                mot, mot_em, plc, consumivel_sp, recebedor_consumivel_sp,
                _agora_iso(), usuario or '', _agora_iso(), usuario or ''
            )
        ).fetchone()
        documento_id = int(row['id'])
    else:
        conn.execute(
            '''INSERT INTO ''' + _tbl_terceiros_documentos(conn) + ''' (
                   area, chave_nfe, numero_nf, serie_nf, data_emissao, remetente_nome, remetente_cnpj,
                   destinatario_nome, destinatario_cnpj, destinatario_uf, numero_pedido, previsao_chegada, arquivo_nome, xml_conteudo,
                   motorista_carreta, motorista_carreta_em, placa_carreta,
                   consumivel_sp, recebedor_consumivel_sp,
                   criado_em, criado_por, atualizado_em, atualizado_por
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (
                area, xml_data.get('chave_nfe') or '', xml_data.get('numero_nf') or '', xml_data.get('serie_nf') or '',
                xml_data.get('data_emissao') or '', xml_data.get('remetente_nome') or '', _somente_digitos(xml_data.get('remetente_cnpj') or ''),
                xml_data.get('destinatario_nome') or '', _somente_digitos(xml_data.get('destinatario_cnpj') or ''),
                (xml_data.get('destinatario_uf') or '').strip(),
                (xml_data.get('numero_pedido') or '').strip(),
                previsao_chegada or '', arquivo_nome or '', xml_texto,
                mot, mot_em, plc, consumivel_sp, recebedor_consumivel_sp,
                _agora_iso(), usuario or '', _agora_iso(), usuario or ''
            )
        )
        documento_id = int(conn.execute('SELECT last_insert_rowid() as id').fetchone()['id'])
    itens_nf = xml_data.get('itens') or []
    if itens_nf:
        agora_it = _agora_iso()
        tbl_it = _tbl_terceiros_documento_itens(conn)
        sql_it = (
            '''INSERT INTO ''' + tbl_it + ''' (
                   documento_id, n_item, codigo_ean, codigo_produto_xml, descricao_xml, unidade_xml,
                   quantidade_xml, codigo_produto_base, codigo_barras_base, descricao_base,
                   quantidade_bipada, status_bipagem, atualizado_em, atualizado_por
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
        )
        params_it = [
            (
                documento_id, item.get('n_item') or 0, item.get('codigo_ean') or '', item.get('codigo_produto_xml') or '',
                item.get('descricao_xml') or '', item.get('unidade_xml') or '', item.get('quantidade_xml') or 0,
                item.get('codigo_produto_base') or '', item.get('codigo_barras_base') or '', item.get('descricao_base') or '',
                0, 'PENDENTE', agora_it, usuario or ''
            )
            for item in itens_nf
        ]
        if len(params_it) == 1:
            conn.execute(sql_it, params_it[0])
        else:
            conn.executemany(sql_it, params_it)
    _registrar_evento_terceiros(conn, documento_id, 'upload_xml', '', area, usuario, arquivo_nome or '')
    return documento_id


def _sql_cols_terceiros_documento_itens():
    """Colunas dos itens para detalhe/bipagem — sem blobs desnecessários."""
    return (
        'id, documento_id, n_item, codigo_ean, codigo_produto_xml, descricao_xml, unidade_xml, '
        'quantidade_xml, codigo_produto_base, codigo_barras_base, descricao_base, quantidade_bipada, '
        'status_bipagem, motivo, ultimo_ean_bipado, atualizado_em, atualizado_por'
    )


def _carregar_documento_terceiros(conn, documento_id, incluir_eventos=False):
    tbl_d = _tbl_terceiros_documentos(conn)
    cols_d = _sql_cols_terceiros_documentos_listagem('d')
    row = conn.execute(
        'SELECT ' + cols_d + ' FROM ' + tbl_d + ' d WHERE d.id = ?',
        (documento_id,),
    ).fetchone()
    if not row:
        return None
    doc = dict(row) if hasattr(row, 'keys') else {}
    doc['numero_pedido'] = _normalizar_numero_pedido_terceiros(doc.get('numero_pedido'))
    doc['destinatario_uf'] = (doc.get('destinatario_uf') or '').strip().upper()
    tbl_i = _tbl_terceiros_documento_itens(conn)
    itens_rows = conn.execute(
        'SELECT ' + _sql_cols_terceiros_documento_itens() + ' FROM ' + tbl_i
        + ' WHERE documento_id = ? ORDER BY n_item, id',
        (documento_id,),
    ).fetchall()
    eventos_rows = []
    if incluir_eventos:
        eventos_rows = conn.execute(
            'SELECT * FROM ' + _tbl_terceiros_documento_eventos(conn)
            + ' WHERE documento_id = ? ORDER BY criado_em DESC, id DESC',
            (documento_id,),
        ).fetchall()
    itens = []
    for r in itens_rows or []:
        item = dict(r) if hasattr(r, 'keys') else {}
        qtd_xml = float(item.get('quantidade_xml') or 0)
        qtd_bipada = float(item.get('quantidade_bipada') or 0)
        item['quantidade_falta'] = max(0.0, qtd_xml - qtd_bipada)
        item['quantidade_sobra'] = max(0.0, qtd_bipada - qtd_xml)
        item['status_bipagem'] = _status_bipagem_terceiros(qtd_xml, qtd_bipada)
        itens.append(item)
    doc['itens'] = itens
    doc['eventos'] = [dict(r) if hasattr(r, 'keys') else {} for r in (eventos_rows or [])]
    doc['resumo'] = {
        'total_itens': len(itens),
        'quantidade_total_xml': round(sum(float(i.get('quantidade_xml') or 0) for i in itens), 3),
        'quantidade_total_bipada': round(sum(float(i.get('quantidade_bipada') or 0) for i in itens), 3),
        'itens_com_pendencia': sum(1 for i in itens if i.get('status_bipagem') != 'COMPLETO'),
    }
    doc.pop('xml_conteudo', None)
    return doc


_TERCEIROS_BIPAGEM_MSG_WMS = (
    'A bipagem de entrada é feita somente no módulo Endereçamento WMS. '
    'Use o WMS para registrar a recepção; o status será atualizado aqui automaticamente.'
)


def _terceiros_enriquecer_doc_wms(conn, doc, sincronizar=True, incluir_eventos=False):
    """Sincroniza quantidades do WMS e anexa metadados de recebimento WMS ao documento."""
    if not doc or not doc.get('id'):
        return doc
    try:
        from wms_enderecamento import (
            _sincronizar_terceiros_itens_desde_wms,
            _enriquecer_documento_nf_wms,
            _buscar_recebimento_wms_por_nf,
        )
    except ImportError:
        return doc
    doc_id = int(doc['id'])
    if sincronizar:
        _sincronizar_terceiros_itens_desde_wms(conn, doc_id)
        doc = _carregar_documento_terceiros(conn, doc_id, incluir_eventos=incluir_eventos)
        if not doc:
            return doc
    wms_itens = []
    for it in doc.get('itens') or []:
        wms_itens.append({
            'n_item': it.get('n_item'),
            'sku': (it.get('codigo_produto_base') or it.get('codigo_produto_xml') or '').strip(),
            'descricao': (it.get('descricao_base') or it.get('descricao_xml') or '').strip(),
            'quantidade_xml': it.get('quantidade_xml'),
            'quantidade_bipada': it.get('quantidade_bipada'),
            'codigo_ean': it.get('codigo_ean') or '',
            'unidade': it.get('unidade_xml') or '',
        })
    wms_doc = {
        'documento_id': doc_id,
        'numero_nf': doc.get('numero_nf') or '',
        'recebimento_concluido': doc.get('recebimento_concluido'),
        'itens': wms_itens,
    }
    wms_doc = _enriquecer_documento_nf_wms(conn, wms_doc) or wms_doc
    doc['recebimento_wms_id'] = wms_doc.get('recebimento_wms_id')
    doc['recebimento_wms_status'] = wms_doc.get('recebimento_wms_status')
    doc['wms_bloqueado'] = wms_doc.get('wms_bloqueado')
    doc['descarga_reabrivel'] = wms_doc.get('descarga_reabrivel')
    por_n = {}
    for wit in wms_doc.get('itens') or []:
        if wit.get('n_item') is not None:
            por_n[str(wit.get('n_item'))] = wit
    total_wms = 0.0
    pend_wms = 0
    for it in doc.get('itens') or []:
        w = por_n.get(str(it.get('n_item'))) or {}
        q_wms = float(w.get('quantidade_wms') or it.get('quantidade_bipada') or 0)
        q_xml = float(it.get('quantidade_xml') or 0)
        it['quantidade_wms'] = q_wms
        it['quantidade_armazenada'] = float(w.get('quantidade_armazenada') or 0)
        it['pendente_wms'] = float(w.get('pendente_wms') if w.get('pendente_wms') is not None else max(q_xml - q_wms, 0))
        it['status_wms'] = w.get('status_wms') or 'pendente'
        total_wms += q_wms
        if q_xml > 0 and abs(q_xml - q_wms) > 1e-6:
            pend_wms += 1
    doc['quantidade_total_wms'] = round(total_wms, 3)
    doc['itens_pendentes_wms'] = pend_wms
    rec = _buscar_recebimento_wms_por_nf(
        conn, documento_id=doc_id, numero_nf=doc.get('numero_nf'), somente_ativo=True,
    )
    if not rec:
        rec = _buscar_recebimento_wms_por_nf(
            conn, documento_id=doc_id, numero_nf=doc.get('numero_nf'), somente_ativo=False,
        )
    if rec:
        doc['recebimento_wms_id'] = rec.get('id')
        if not doc.get('recebimento_wms_status'):
            doc['recebimento_wms_status'] = rec.get('status')
    return doc


def _normalizar_texto_regra_terceiros(valor):
    return re.sub(r'[^a-z0-9]+', ' ', str(valor or '').lower()).strip()


def _terceiros_eh_area_carreta(doc):
    return str((doc or {}).get('area') or '').strip().lower() == 'carreta'


def _terceiros_eh_consumivel_sp(doc):
    return _terceiros_bool_sim_db((doc or {}).get('consumivel_sp'))


def _terceiros_usa_fluxo_mg(doc):
    return not _terceiros_eh_area_carreta(doc) and not _terceiros_eh_consumivel_sp(doc)


_TERCEIROS_CONSUMIVEL_SP_EMAILS_PADRAO = [
    'elder.tenorio@ultrapao.com.br',
    'jackeline.silva@ultrapao.com.br',
    'suellen.oliveira@ultrapao.com.br',
    'mariana.teixeira@ultrapao.com.br',
    'astro.santos@ultrapao.com.br',
]


def _terceiros_emails_consumivel_sp_destino():
    raw = (os.environ.get('TERCEIROS_CONSUMIVEL_SP_EMAILS') or '').strip()
    if raw:
        return [e.strip() for e in raw.split(',') if e.strip() and '@' in e.strip()]
    return list(_TERCEIROS_CONSUMIVEL_SP_EMAILS_PADRAO)


def _terceiros_nf_texto_email(doc):
    return '/'.join(filter(None, [
        str((doc or {}).get('numero_nf') or '').strip(),
        str((doc or {}).get('serie_nf') or '').strip(),
    ])) or '-'


def _terceiros_smtp_configurado():
    return bool((os.environ.get('SMTP_HOST') or '').strip())


def _terceiros_enviar_email(destinatarios, assunto, corpo_texto, corpo_html=None):
    """Retorna (ok: bool, motivo: str). motivo vazio se enviado com sucesso."""
    destinatarios = [d.strip() for d in (destinatarios or []) if d and '@' in str(d)]
    if not destinatarios:
        print('[terceiros-email] Nenhum destinatário válido.')
        return False, 'sem_destinatarios'
    smtp_host = (os.environ.get('SMTP_HOST') or '').strip()
    if not smtp_host:
        print('[terceiros-email] SMTP_HOST não configurado; e-mail não enviado.')
        return False, 'smtp_nao_configurado'
    smtp_port = int(os.environ.get('SMTP_PORT') or '587')
    smtp_user = (os.environ.get('SMTP_USER') or '').strip()
    smtp_pass = (os.environ.get('SMTP_PASSWORD') or '').strip().replace(' ', '')
    if smtp_host and not smtp_pass:
        print('[terceiros-email] SMTP_PASSWORD vazio; e-mail não enviado.')
        return False, 'smtp_sem_senha'
    smtp_from = (os.environ.get('SMTP_FROM') or smtp_user or 'noreply@ultrapao.com.br').strip()
    use_tls = str(os.environ.get('SMTP_USE_TLS', '1')).strip().lower() not in ('0', 'false', 'no')
    msg = MIMEMultipart('alternative')
    msg['Subject'] = assunto
    msg['From'] = smtp_from
    msg['To'] = ', '.join(destinatarios)
    msg.attach(MIMEText(corpo_texto or '', 'plain', 'utf-8'))
    if corpo_html:
        msg.attach(MIMEText(corpo_html, 'html', 'utf-8'))
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            if use_tls:
                server.starttls()
            if smtp_user and smtp_pass:
                server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_from, destinatarios, msg.as_string())
        print('[terceiros-email] Enviado para %s — %s' % (', '.join(destinatarios), assunto))
        return True, ''
    except Exception as e:
        print('[terceiros-email] Falha ao enviar: %s' % e)
        return False, 'erro_smtp: %s' % e


def _terceiros_notificar_consumivel_sp_pronto_lancamento(doc, usuario):
    """E-mail quando consumível SP conclui recebimento e segue para lançamento fiscal."""
    doc = doc or {}
    dest = _terceiros_emails_consumivel_sp_destino()
    nf = _terceiros_nf_texto_email(doc)
    pedido = (doc.get('numero_pedido') or '').strip() or '-'
    recebedor = (doc.get('recebedor_consumivel_sp') or '').strip() or '-'
    remetente = (doc.get('remetente_nome') or '').strip() or '-'
    concluido_por = (doc.get('recebimento_concluido_por') or usuario or '').strip() or '-'
    concluido_em = doc.get('recebimento_concluido_em') or _fmt_datahora_br(_agora_iso())
    assunto = 'Consumível SP — produto chegou (NF %s) — pronto para lançamento' % nf
    linhas = [
        'O produto (consumível SP) chegou e o recebimento foi concluído.',
        'A nota fiscal está pronta para lançamento.',
        '',
        'NF: %s' % nf,
        'Pedido: %s' % pedido,
        'Quem solicitou/irá receber: %s' % recebedor,
        'Remetente: %s' % remetente,
        'Recebimento registrado por: %s' % concluido_por,
        'Recebimento em: %s' % concluido_em,
        '',
        'Próximo passo: acesse Terceiros → NFs pendentes de lançamento e registre o lançamento fiscal.',
    ]
    corpo = '\n'.join(linhas)
    html = (
        '<p><strong>O produto chegou.</strong> Consumível SP com recebimento concluído — '
        '<strong>pronto para lançamento fiscal</strong>.</p>'
        '<ul>'
        '<li><strong>NF:</strong> %s</li>'
        '<li><strong>Pedido:</strong> %s</li>'
        '<li><strong>Quem solicitou/irá receber:</strong> %s</li>'
        '<li><strong>Remetente:</strong> %s</li>'
        '<li><strong>Recebimento registrado por:</strong> %s</li>'
        '<li><strong>Recebimento em:</strong> %s</li>'
        '</ul>'
        '<p><strong>Próximo passo:</strong> Terceiros → <em>NFs pendentes de lançamento</em>.</p>'
    ) % (nf, pedido, recebedor, remetente, concluido_por, concluido_em)
    return _terceiros_enviar_email(dest, assunto, corpo, html)


def _terceiros_email_consumivel_sp_apos_recebimento(doc, usuario, primeira_conclusao):
    """
    Dispara e-mail de consumível SP na primeira conclusão de recebimento.
    Envio síncrono (thread em background no Render/Gunicorn costuma ser interrompida).
    """
    info = {'enviado': False, 'motivo': 'nao_aplicavel'}
    if not primeira_conclusao:
        info['motivo'] = 'recebimento_ja_estava_concluido'
        return info
    if not _terceiros_eh_consumivel_sp(doc):
        info['motivo'] = 'nao_eh_consumivel_sp'
        return info
    try:
        ok, motivo = _terceiros_notificar_consumivel_sp_pronto_lancamento(doc, usuario)
        info['enviado'] = bool(ok)
        info['motivo'] = motivo or ('enviado' if ok else 'falha_desconhecida')
    except Exception as e:
        print('[terceiros-email] Exceção ao notificar consumível SP: %s' % e)
        info['motivo'] = 'erro: %s' % e
    return info


def _motorista_obrigatorio_terceiros(doc):
    remetente = _normalizar_texto_regra_terceiros((doc or {}).get('remetente_nome') or '')
    destinatario = _normalizar_texto_regra_terceiros((doc or {}).get('destinatario_nome') or '')
    return (
        'ultrapao de pouso alegre' in remetente
        and ('ultrapao de sp' in destinatario or 'ultrapao de sao paulo' in destinatario)
    )


def _identificador_duplicidade_terceiros(xml_data):
    chave = (xml_data or {}).get('chave_nfe') or ''
    if chave:
        return ('chave_nfe', chave.strip())
    numero_nf = str((xml_data or {}).get('numero_nf') or '').strip()
    serie_nf = str((xml_data or {}).get('serie_nf') or '').strip()
    emitente = _somente_digitos((xml_data or {}).get('remetente_cnpj') or '')
    destinatario = _somente_digitos((xml_data or {}).get('destinatario_cnpj') or '')
    return ('fallback', '|'.join([numero_nf, serie_nf, emitente, destinatario]))


def _documento_terceiros_ja_existe(conn, xml_data):
    tipo, valor = _identificador_duplicidade_terceiros(xml_data)
    if not valor:
        return False
    if tipo == 'chave_nfe':
        row = conn.execute(
            'SELECT id FROM ' + _tbl_terceiros_documentos(conn) + ' WHERE chave_nfe = ? LIMIT 1',
            (valor,)
        ).fetchone()
        return bool(row)
    row = conn.execute(
        'SELECT id FROM ' + _tbl_terceiros_documentos(conn) + ' WHERE numero_nf = ? AND serie_nf = ? AND remetente_cnpj = ? AND destinatario_cnpj = ? LIMIT 1',
        tuple(valor.split('|'))
    ).fetchone()
    return bool(row)


def _terceiros_periodo_request_args():
    de = (request.args.get('data_criacao_inicio') or request.args.get('data_inicio') or '').strip()
    ate = (request.args.get('data_criacao_fim') or request.args.get('data_fim') or '').strip()
    return de, ate


def _terceiros_sql_periodo_criado(alias, data_inicio, data_fim):
    parts = []
    params = []
    if data_inicio:
        d0 = (data_inicio[:10] + ' 00:00:00') if len(data_inicio) <= 10 else data_inicio
        parts.append(alias + '.criado_em >= ?')
        params.append(d0)
    if data_fim:
        d1 = (data_fim[:10] + ' 23:59:59') if len(data_fim) <= 10 else data_fim
        parts.append(alias + '.criado_em <= ?')
        params.append(d1)
    return parts, params


def _terceiros_workbook_simple(headers, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    return wb


def _terceiros_send_workbook(wb, download_name):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _terceiros_docs_com_resumo(conn, data_inicio='', data_fim=''):
    _ensure_terceiros_schema(conn)
    tbl_d = _tbl_terceiros_documentos(conn)
    tbl_i = _tbl_terceiros_documento_itens(conn)
    cols = _sql_cols_terceiros_documentos_listagem('d')
    where_parts = []
    params = []
    wp, pp = _terceiros_sql_periodo_criado('d', data_inicio, data_fim)
    where_parts.extend(wp)
    params.extend(pp)
    sql = (
        'SELECT ' + cols + ''',
                COALESCE(si.total_itens, 0) AS total_itens,
                COALESCE(si.quantidade_total_xml, 0) AS quantidade_total_xml,
                COALESCE(si.quantidade_total_bipada, 0) AS quantidade_total_bipada,
                COALESCE(si.itens_divergentes, 0) AS itens_divergentes,
                si.inicio_descarga
         FROM ''' + tbl_d + ''' d
         LEFT JOIN (
             SELECT documento_id,
                    COUNT(id) AS total_itens,
                    COALESCE(SUM(quantidade_xml), 0) AS quantidade_total_xml,
                    COALESCE(SUM(quantidade_bipada), 0) AS quantidade_total_bipada,
                    SUM(CASE WHEN ABS(COALESCE(quantidade_xml, 0) - COALESCE(quantidade_bipada, 0)) > 0.000001 THEN 1 ELSE 0 END) AS itens_divergentes,
                    MIN(CASE WHEN COALESCE(quantidade_bipada, 0) > 0 THEN atualizado_em ELSE NULL END) AS inicio_descarga
             FROM ''' + tbl_i + '''
             GROUP BY documento_id
         ) si ON si.documento_id = d.id
    ''')
    if where_parts:
        sql += ' WHERE ' + ' AND '.join(where_parts)
    sql += ' ORDER BY d.criado_em DESC, d.id DESC'
    rows = conn.execute(sql, params).fetchall()
    out = []
    for r in rows or []:
        row = dict(r) if hasattr(r, 'keys') else {}
        etapa = _terceiros_etapa_painel_row(row)
        out.append({
            'row': row,
            'etapa': _ETAPAS_PAINEL_TERCEIROS_LABELS.get(etapa, etapa),
        })
    return out


def _relatorio_terceiros_resumo_nf(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        docs = _terceiros_docs_com_resumo(conn, data_inicio, data_fim)
    finally:
        conn.close()
    linhas = []
    for item in docs:
        row = item['row']
        nf = '/'.join(filter(None, [str(row.get('numero_nf') or '').strip(), str(row.get('serie_nf') or '').strip()])) or '-'
        linhas.append([
            row.get('id'),
            nf,
            row.get('numero_pedido') or '',
            (row.get('area') or '').strip(),
            row.get('remetente_nome') or '',
            row.get('destinatario_nome') or '',
            row.get('destinatario_uf') or '',
            _fmt_datahora_br(row.get('previsao_chegada') or ''),
            _fmt_datahora_br(row.get('criado_em') or ''),
            _fmt_datahora_br(row.get('inicio_descarga') or ''),
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
            item['etapa'],
            float(row.get('quantidade_total_xml') or 0),
            float(row.get('quantidade_total_bipada') or 0),
            int(row.get('itens_divergentes') or 0),
            'Sim' if _terceiros_bool_sim_db(row.get('recebimento_concluido')) else 'Não',
            row.get('nota_lancada') or '',
            row.get('enviar_para_mg') or '',
            row.get('carga_recebida_mg') or '',
            row.get('recebedor_mg') or '',
            row.get('motorista_carreta') or '',
            row.get('placa_carreta') or '',
            row.get('motorista_saida_mg') or '',
            row.get('placa_saida_mg') or '',
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Área', 'Remetente', 'Destinatário', 'UF',
            'Previsão chegada', 'Cadastro NF', 'Início descarga', 'Recebimento concluído',
            'Etapa atual', 'Qtd. XML', 'Qtd. bipada', 'Itens divergentes',
            'Receb. concluído', 'Nota lançada', 'Enviar MG', 'Recebida MG',
            'Quem recebeu MG', 'Motorista que trouxe', 'Placa chegada', 'Motorista que levou MG', 'Placa saída MG',
        ],
        linhas,
    )
    wb.active.title = 'Resumo NFs'
    return wb


def _relatorio_terceiros_itens_bipados(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_i = _tbl_terceiros_documento_itens(conn)
        where_parts = ['COALESCE(i.quantidade_bipada, 0) > 0']
        params = []
        wp, pp = _terceiros_sql_periodo_criado('d', data_inicio, data_fim)
        where_parts.extend(wp)
        params.extend(pp)
        sql = (
            '''SELECT d.id AS documento_id, d.numero_nf, d.serie_nf, d.area, d.remetente_nome,
                      i.codigo_ean, i.descricao_xml, i.quantidade_xml, i.quantidade_bipada,
                      i.status_bipagem, i.atualizado_em, i.atualizado_por
               FROM ''' + tbl_i + ''' i
               INNER JOIN ''' + tbl_d + ''' d ON d.id = i.documento_id
               WHERE ''' + ' AND '.join(where_parts) + '''
               ORDER BY i.atualizado_em DESC, i.id DESC'''
        )
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    linhas = []
    for r in rows or []:
        rd = dict(r) if hasattr(r, 'keys') else {}
        nf = '/'.join(filter(None, [str(rd.get('numero_nf') or '').strip(), str(rd.get('serie_nf') or '').strip()])) or '-'
        linhas.append([
            rd.get('documento_id'),
            nf,
            rd.get('area') or '',
            rd.get('remetente_nome') or '',
            rd.get('codigo_ean') or '',
            (rd.get('descricao_xml') or '')[:120],
            float(rd.get('quantidade_xml') or 0),
            float(rd.get('quantidade_bipada') or 0),
            rd.get('status_bipagem') or '',
            _fmt_datahora_br(rd.get('atualizado_em') or ''),
            rd.get('atualizado_por') or '',
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID NF', 'NF', 'Área', 'Remetente', 'Código barras', 'Produto',
            'Qtd. XML', 'Qtd. bipada', 'Status', 'Última bipagem', 'Usuário',
        ],
        linhas,
    )
    wb.active.title = 'Itens bipados'
    return wb


def _relatorio_terceiros_itens_mais_bipados(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_i = _tbl_terceiros_documento_itens(conn)
        where_parts = ['COALESCE(i.quantidade_bipada, 0) > 0']
        params = []
        wp, pp = _terceiros_sql_periodo_criado('d', data_inicio, data_fim)
        where_parts.extend(wp)
        params.extend(pp)
        sql = (
            '''SELECT COALESCE(NULLIF(TRIM(i.descricao_xml), ''), NULLIF(TRIM(i.codigo_ean), ''), 'Item') AS produto,
                      COALESCE(NULLIF(TRIM(i.codigo_ean), ''), '-') AS codigo_ean,
                      COALESCE(SUM(i.quantidade_bipada), 0) AS total,
                      COUNT(DISTINCT i.documento_id) AS num_nfs
               FROM ''' + tbl_i + ''' i
               INNER JOIN ''' + tbl_d + ''' d ON d.id = i.documento_id
               WHERE ''' + ' AND '.join(where_parts) + '''
               GROUP BY i.descricao_xml, i.codigo_ean
               ORDER BY total DESC'''
        )
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    linhas = []
    for r in rows or []:
        rd = dict(r) if hasattr(r, 'keys') else {}
        linhas.append([
            rd.get('codigo_ean') or '',
            (rd.get('produto') or '')[:120],
            float(rd.get('total') or 0),
            int(rd.get('num_nfs') or 0),
        ])
    wb = _terceiros_workbook_simple(
        ['Código barras', 'Produto', 'Total bipado', 'Qtd. NFs'],
        linhas,
    )
    wb.active.title = 'Itens mais bipados'
    return wb


def _relatorio_terceiros_divergencias(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        docs = _terceiros_docs_com_resumo(conn, data_inicio, data_fim)
    finally:
        conn.close()
    linhas = []
    for item in docs:
        row = item['row']
        q_xml = float(row.get('quantidade_total_xml') or 0)
        q_bip = float(row.get('quantidade_total_bipada') or 0)
        div_it = int(row.get('itens_divergentes') or 0)
        if div_it == 0 and abs(q_xml - q_bip) <= 1e-6:
            continue
        if q_bip <= 1e-9 and div_it == 0:
            continue
        nf = '/'.join(filter(None, [str(row.get('numero_nf') or '').strip(), str(row.get('serie_nf') or '').strip()])) or '-'
        linhas.append([
            row.get('id'),
            nf,
            row.get('remetente_nome') or '',
            q_xml,
            q_bip,
            q_xml - q_bip,
            div_it,
            item['etapa'],
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Remetente', 'Qtd. XML', 'Qtd. bipada', 'Diferença',
            'Itens divergentes', 'Etapa', 'Recebimento concluído',
        ],
        linhas,
    )
    wb.active.title = 'NFs divergentes'
    return wb


def _relatorio_terceiros_carreta(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        docs = _terceiros_docs_com_resumo(conn, data_inicio, data_fim)
    finally:
        conn.close()
    linhas = []
    for item in docs:
        row = item['row']
        if (row.get('area') or '').strip().lower() != 'carreta':
            continue
        nf = '/'.join(filter(None, [str(row.get('numero_nf') or '').strip(), str(row.get('serie_nf') or '').strip()])) or '-'
        linhas.append([
            row.get('id'),
            nf,
            row.get('motorista_carreta') or '',
            (row.get('placa_carreta') or '').upper(),
            _fmt_datahora_br(row.get('previsao_chegada') or ''),
            _fmt_datahora_br(row.get('inicio_descarga') or ''),
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
            item['etapa'],
            float(row.get('quantidade_total_xml') or 0),
            float(row.get('quantidade_total_bipada') or 0),
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Motorista', 'Placa', 'Previsão chegada', 'Início descarga',
            'Recebimento concluído', 'Etapa', 'Qtd. XML', 'Qtd. bipada',
        ],
        linhas,
    )
    wb.active.title = 'Carreta'
    return wb


def _terceiros_nf_texto_rel(row):
    return '/'.join(filter(None, [str(row.get('numero_nf') or '').strip(), str(row.get('serie_nf') or '').strip()])) or '-'


def _terceiros_tipo_fluxo_rel(row):
    if _terceiros_eh_area_carreta(row):
        return 'Carreta'
    if _terceiros_eh_consumivel_sp(row):
        return 'Consumível SP'
    return 'Normal (MG)'


def _terceiros_parse_datahora_iso(s):
    if not s or not str(s).strip():
        return None
    txt = str(s).strip()[:19].replace('T', ' ')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            continue
    return None


def _terceiros_dias_entre(a, b):
    da = _terceiros_parse_datahora_iso(a)
    db = _terceiros_parse_datahora_iso(b)
    if not da or not db:
        return ''
    return round((db - da).total_seconds() / 86400.0, 1)


def _terceiros_ultima_acao_em(row):
    candidatos = [
        row.get('consumivel_sp_historico_em'), row.get('carga_recebida_mg_em'),
        row.get('enviar_para_mg_em'), row.get('nota_lancada_em'),
        row.get('recebimento_concluido_em'), row.get('atualizado_em'), row.get('criado_em'),
    ]
    datas = [_terceiros_parse_datahora_iso(c) for c in candidatos]
    datas = [d for d in datas if d]
    return max(datas) if datas else None


def _terceiros_dias_parados(row):
    ult = _terceiros_ultima_acao_em(row)
    if not ult:
        return ''
    return round((datetime.now() - ult).total_seconds() / 86400.0, 1)


def _terceiros_motivo_encerramento_rel(row):
    if _valor_bool_texto(row.get('nota_lancada') or '') == 'nao' and str(row.get('motivo_nao_lancada') or '').strip():
        return 'Não lançada: ' + str(row.get('motivo_nao_lancada') or '').strip()
    if _valor_bool_texto(row.get('enviar_para_mg') or '') == 'nao' and str(row.get('motivo_nao_enviar_mg') or '').strip():
        return 'Não enviou MG: ' + str(row.get('motivo_nao_enviar_mg') or '').strip()
    if _valor_bool_texto(row.get('carga_recebida_mg') or '') == 'nao' and str(row.get('motivo_nao_recebida_mg') or '').strip():
        return 'Não recebida MG: ' + str(row.get('motivo_nao_recebida_mg') or '').strip()
    if _terceiros_eh_consumivel_sp(row) and _valor_bool_texto(row.get('consumivel_sp_historico') or '') == 'sim':
        return 'Consumível SP enviado ao histórico'
    if _terceiros_eh_area_carreta(row) and _valor_bool_texto(row.get('carga_recebida_mg') or '') == 'sim':
        return 'Carreta registrada no histórico'
    if _valor_bool_texto(row.get('carga_recebida_mg') or '') == 'sim' and _valor_bool_texto(row.get('nota_lancada') or '') == 'sim':
        return 'Fluxo MG completo'
    return ''


def _terceiros_docs_filtrados(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        return _terceiros_docs_com_resumo(conn, data_inicio, data_fim)
    finally:
        conn.close()


def _relatorio_terceiros_notas_lancadas(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        nl = _valor_bool_texto(row.get('nota_lancada') or '')
        if nl != 'sim':
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('numero_pedido') or '',
            _terceiros_tipo_fluxo_rel(row),
            row.get('remetente_nome') or '',
            row.get('destinatario_nome') or '',
            row.get('destinatario_uf') or '',
            row.get('nota_lancada_por') or '',
            _fmt_datahora_br(row.get('nota_lancada_em') or ''),
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
            row.get('recebimento_concluido_por') or '',
            item['etapa'],
            row.get('enviar_para_mg') or '',
            row.get('carga_recebida_mg') or '',
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Tipo fluxo', 'Remetente', 'Destinatário', 'UF',
            'Quem lançou', 'Lançada em', 'Recebimento concluído em', 'Quem concluiu recebimento',
            'Etapa atual', 'Enviar MG', 'Recebida MG',
        ],
        linhas,
    )
    wb.active.title = 'Notas lançadas'
    return wb


def _relatorio_terceiros_historico(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        if _terceiros_etapa_painel_row(row) != 'historico':
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('numero_pedido') or '',
            _terceiros_tipo_fluxo_rel(row),
            row.get('recebedor_consumivel_sp') or '',
            row.get('remetente_nome') or '',
            row.get('destinatario_uf') or '',
            _terceiros_motivo_encerramento_rel(row),
            _fmt_datahora_br(row.get('criado_em') or ''),
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
            _fmt_datahora_br(row.get('nota_lancada_em') or ''),
            _fmt_datahora_br(row.get('enviar_para_mg_em') or ''),
            _fmt_datahora_br(row.get('carga_recebida_mg_em') or ''),
            _fmt_datahora_br(row.get('consumivel_sp_historico_em') or ''),
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Tipo fluxo', 'Quem irá receber (consumível)',
            'Remetente', 'UF', 'Motivo / forma de encerramento', 'Cadastro',
            'Recebimento concluído', 'Lançada em', 'Enviado MG em', 'Recebida MG em', 'Histórico consumível em',
        ],
        linhas,
    )
    wb.active.title = 'Histórico'
    return wb


def _relatorio_terceiros_consumivel_sp(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        if not _terceiros_eh_consumivel_sp(row):
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('numero_pedido') or '',
            row.get('recebedor_consumivel_sp') or '',
            item['etapa'],
            'Sim' if _terceiros_bool_sim_db(row.get('recebimento_concluido')) else 'Não',
            row.get('nota_lancada') or '',
            row.get('nota_lancada_por') or '',
            _fmt_datahora_br(row.get('nota_lancada_em') or ''),
            row.get('consumivel_sp_historico') or '',
            row.get('consumivel_sp_historico_por') or '',
            _fmt_datahora_br(row.get('consumivel_sp_historico_em') or ''),
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Quem solicitou/receberá', 'Etapa atual',
            'Receb. concluído', 'Lançada', 'Quem lançou', 'Lançada em',
            'No histórico', 'Enviado ao histórico por', 'Enviado ao histórico em',
        ],
        linhas,
    )
    wb.active.title = 'Consumível SP'
    return wb


def _relatorio_terceiros_pendencias_etapa(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        if _terceiros_etapa_painel_row(row) == 'historico':
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('numero_pedido') or '',
            _terceiros_tipo_fluxo_rel(row),
            item['etapa'],
            _terceiros_dias_parados(row),
            row.get('remetente_nome') or '',
            row.get('destinatario_uf') or '',
            _fmt_datahora_br(row.get('previsao_chegada') or ''),
            row.get('atualizado_por') or '',
            _fmt_datahora_br(row.get('atualizado_em') or ''),
        ])
    linhas.sort(key=lambda r: (r[4], -(float(r[5]) if r[5] != '' else 0)))
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Tipo fluxo', 'Etapa parada', 'Dias parados',
            'Remetente', 'UF', 'Previsão chegada', 'Último usuário', 'Última atualização',
        ],
        linhas,
    )
    wb.active.title = 'Pendências'
    return wb


def _relatorio_terceiros_encerradas_motivo(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        motivo = _terceiros_motivo_encerramento_rel(row)
        if not motivo or not motivo.startswith(('Não lançada', 'Não enviou', 'Não recebida')):
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('numero_pedido') or '',
            _terceiros_tipo_fluxo_rel(row),
            motivo,
            row.get('motivo_nao_lancada') or '',
            row.get('motivo_nao_enviar_mg') or '',
            row.get('motivo_nao_recebida_mg') or '',
            row.get('nota_lancada_por') or row.get('enviar_para_mg_por') or row.get('carga_recebida_mg_por') or '',
            _fmt_datahora_br(row.get('atualizado_em') or ''),
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Tipo fluxo', 'Resumo encerramento',
            'Motivo não lançada', 'Motivo não enviar MG', 'Motivo não recebida MG',
            'Registrado por', 'Atualizado em',
        ],
        linhas,
    )
    wb.active.title = 'Encerradas c motivo'
    return wb


def _relatorio_terceiros_fluxo_mg(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        if _terceiros_eh_area_carreta(row) or _terceiros_eh_consumivel_sp(row):
            continue
        emg = _valor_bool_texto(row.get('enviar_para_mg') or '')
        if emg not in ('sim', 'nao', 'pendente') and not str(row.get('enviar_para_mg') or '').strip():
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('destinatario_uf') or '',
            row.get('enviar_para_mg') or '',
            row.get('enviar_para_mg_por') or '',
            _fmt_datahora_br(row.get('enviar_para_mg_em') or ''),
            row.get('motorista_saida_mg') or '',
            row.get('placa_saida_mg') or '',
            row.get('carga_recebida_mg') or '',
            row.get('recebedor_mg') or '',
            row.get('carga_recebida_mg_por') or '',
            _fmt_datahora_br(row.get('carga_recebida_mg_em') or ''),
            row.get('nota_lancada') or '',
            item['etapa'],
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'UF', 'Enviar MG', 'Quem marcou envio', 'Enviado em',
            'Motorista saída MG', 'Placa saída', 'Recebida MG', 'Quem recebeu MG',
            'Recebida por', 'Recebida em', 'Nota lançada', 'Etapa atual',
        ],
        linhas,
    )
    wb.active.title = 'Fluxo MG'
    return wb


def _relatorio_terceiros_por_uf(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    agg = {}
    for item in docs:
        row = item['row']
        uf = (row.get('destinatario_uf') or '').strip().upper() or '—'
        if uf not in agg:
            agg[uf] = {'total': 0, 'historico': 0, 'pendentes': 0, 'q_xml': 0.0, 'q_bip': 0.0}
        agg[uf]['total'] += 1
        if _terceiros_etapa_painel_row(row) == 'historico':
            agg[uf]['historico'] += 1
        else:
            agg[uf]['pendentes'] += 1
        agg[uf]['q_xml'] += float(row.get('quantidade_total_xml') or 0)
        agg[uf]['q_bip'] += float(row.get('quantidade_total_bipada') or 0)
    linhas = []
    for uf in sorted(agg.keys(), key=lambda u: (-agg[u]['total'], u)):
        a = agg[uf]
        linhas.append([uf, a['total'], a['pendentes'], a['historico'], round(a['q_xml'], 3), round(a['q_bip'], 3)])
    wb = _terceiros_workbook_simple(
        ['UF destino', 'Total NFs', 'Em andamento', 'No histórico', 'Qtd. XML', 'Qtd. bipada'],
        linhas,
    )
    wb.active.title = 'Por UF'
    return wb


def _relatorio_terceiros_por_remetente(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    agg = {}
    for item in docs:
        row = item['row']
        rem = (row.get('remetente_nome') or '').strip() or 'Sem remetente'
        if rem not in agg:
            agg[rem] = {'total': 0, 'div': 0, 'dias_rec': []}
        agg[rem]['total'] += 1
        q_xml = float(row.get('quantidade_total_xml') or 0)
        q_bip = float(row.get('quantidade_total_bipada') or 0)
        div_it = int(row.get('itens_divergentes') or 0)
        if div_it > 0 or abs(q_xml - q_bip) > 1e-6:
            agg[rem]['div'] += 1
        dias = _terceiros_dias_entre(row.get('criado_em'), row.get('recebimento_concluido_em'))
        if dias != '':
            agg[rem]['dias_rec'].append(dias)
    linhas = []
    for rem in sorted(agg.keys(), key=lambda r: (-agg[r]['total'], r)):
        a = agg[rem]
        media = round(sum(a['dias_rec']) / len(a['dias_rec']), 1) if a['dias_rec'] else ''
        linhas.append([rem, a['total'], a['div'], media])
    wb = _terceiros_workbook_simple(
        ['Remetente', 'Total NFs', 'NFs com divergência', 'Média dias até recebimento'],
        linhas,
    )
    wb.active.title = 'Por remetente'
    return wb


def _relatorio_terceiros_previsao_chegada(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        prev = _terceiros_parse_datahora_iso(row.get('previsao_chegada'))
        ini = _terceiros_parse_datahora_iso(row.get('inicio_descarga'))
        rec = _terceiros_parse_datahora_iso(row.get('recebimento_concluido_em'))
        ref = ini or rec
        situacao = 'Sem descarga/recebimento'
        diff = ''
        if prev and ref:
            diff = round((ref - prev).total_seconds() / 86400.0, 1)
            situacao = 'No prazo' if diff <= 0 else 'Atrasada'
        elif prev and not ref:
            if prev.date() < datetime.now().date():
                situacao = 'Previsão vencida (sem descarga)'
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('remetente_nome') or '',
            _fmt_datahora_br(row.get('previsao_chegada') or ''),
            _fmt_datahora_br(row.get('inicio_descarga') or ''),
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
            situacao,
            diff,
            item['etapa'],
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Remetente', 'Previsão', 'Início descarga', 'Recebimento concluído',
            'Situação', 'Diferença dias (descarga vs previsão)', 'Etapa',
        ],
        linhas,
    )
    wb.active.title = 'Previsão x chegada'
    return wb


def _relatorio_terceiros_recebimentos_periodo(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        if not _terceiros_bool_sim_db(row.get('recebimento_concluido')):
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('numero_pedido') or '',
            _terceiros_tipo_fluxo_rel(row),
            row.get('recebimento_concluido_por') or '',
            _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
            float(row.get('quantidade_total_xml') or 0),
            float(row.get('quantidade_total_bipada') or 0),
            item['etapa'],
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Pedido', 'Tipo fluxo', 'Quem concluiu recebimento',
            'Recebimento concluído em', 'Qtd. XML', 'Qtd. bipada', 'Etapa atual',
        ],
        linhas,
    )
    wb.active.title = 'Recebimentos'
    return wb


def _relatorio_terceiros_divergencia_itens(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_i = _tbl_terceiros_documento_itens(conn)
        where_parts = ['ABS(COALESCE(i.quantidade_xml, 0) - COALESCE(i.quantidade_bipada, 0)) > 0.000001']
        params = []
        wp, pp = _terceiros_sql_periodo_criado('d', data_inicio, data_fim)
        where_parts.extend(wp)
        params.extend(pp)
        sql = (
            '''SELECT d.id AS documento_id, d.numero_nf, d.serie_nf, d.remetente_nome,
                      i.codigo_ean, i.descricao_xml, i.quantidade_xml, i.quantidade_bipada,
                      i.status_bipagem
               FROM ''' + tbl_i + ''' i
               INNER JOIN ''' + tbl_d + ''' d ON d.id = i.documento_id
               WHERE ''' + ' AND '.join(where_parts) + '''
               ORDER BY d.id DESC, i.n_item, i.id'''
        )
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    linhas = []
    for r in rows or []:
        rd = dict(r) if hasattr(r, 'keys') else {}
        q_xml = float(rd.get('quantidade_xml') or 0)
        q_bip = float(rd.get('quantidade_bipada') or 0)
        linhas.append([
            rd.get('documento_id'),
            _terceiros_nf_texto_rel(rd),
            rd.get('remetente_nome') or '',
            rd.get('codigo_ean') or '',
            (rd.get('descricao_xml') or '')[:120],
            q_xml,
            q_bip,
            round(q_xml - q_bip, 3),
            rd.get('status_bipagem') or '',
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID NF', 'NF', 'Remetente', 'Código barras', 'Produto',
            'Qtd. XML', 'Qtd. bipada', 'Diferença', 'Status',
        ],
        linhas,
    )
    wb.active.title = 'Divergência itens'
    return wb


def _relatorio_terceiros_sem_bipagem(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        q_bip = float(row.get('quantidade_total_bipada') or 0)
        if q_bip > 1e-9:
            continue
        if _terceiros_etapa_painel_row(row) == 'historico':
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('remetente_nome') or '',
            _fmt_datahora_br(row.get('criado_em') or ''),
            _fmt_datahora_br(row.get('previsao_chegada') or ''),
            _terceiros_dias_parados(row),
            item['etapa'],
        ])
    wb = _terceiros_workbook_simple(
        ['ID', 'NF', 'Remetente', 'Cadastro', 'Previsão', 'Dias parados', 'Etapa'],
        linhas,
    )
    wb.active.title = 'Sem bipagem'
    return wb


def _relatorio_terceiros_conferencia_incompleta(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        q_bip = float(row.get('quantidade_total_bipada') or 0)
        if q_bip <= 1e-9:
            continue
        if _terceiros_bool_sim_db(row.get('recebimento_concluido')):
            continue
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            row.get('remetente_nome') or '',
            float(row.get('quantidade_total_xml') or 0),
            q_bip,
            int(row.get('itens_divergentes') or 0),
            _fmt_datahora_br(row.get('inicio_descarga') or ''),
            item['etapa'],
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Remetente', 'Qtd. XML', 'Qtd. bipada', 'Itens divergentes',
            'Início descarga', 'Etapa',
        ],
        linhas,
    )
    wb.active.title = 'Conferência incompleta'
    return wb


def _relatorio_terceiros_eventos(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_e = _tbl_terceiros_documento_eventos(conn)
        where_parts = []
        params = []
        wp, pp = _terceiros_sql_periodo_criado('d', data_inicio, data_fim)
        where_parts.extend(wp)
        params.extend(pp)
        sql = (
            '''SELECT e.documento_id, d.numero_nf, d.serie_nf, e.evento, e.valor_anterior,
                      e.valor_novo, e.usuario, e.criado_em, e.detalhes
               FROM ''' + tbl_e + ''' e
               INNER JOIN ''' + tbl_d + ''' d ON d.id = e.documento_id'''
        )
        if where_parts:
            sql += ' WHERE ' + ' AND '.join(where_parts)
        sql += ' ORDER BY e.criado_em DESC, e.id DESC'
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    linhas = []
    for r in rows or []:
        rd = dict(r) if hasattr(r, 'keys') else {}
        linhas.append([
            rd.get('documento_id'),
            _terceiros_nf_texto_rel(rd),
            rd.get('evento') or '',
            rd.get('valor_anterior') or '',
            rd.get('valor_novo') or '',
            rd.get('usuario') or '',
            _fmt_datahora_br(rd.get('criado_em') or ''),
            (rd.get('detalhes') or '')[:200],
        ])
    wb = _terceiros_workbook_simple(
        ['ID NF', 'NF', 'Evento', 'Valor anterior', 'Valor novo', 'Usuário', 'Data/hora', 'Detalhes'],
        linhas,
    )
    wb.active.title = 'Eventos'
    return wb


def _relatorio_terceiros_auditoria_usuario(data_inicio='', data_fim=''):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_e = _tbl_terceiros_documento_eventos(conn)
        where_parts = ["COALESCE(TRIM(e.usuario), '') <> ''"]
        params = []
        wp, pp = _terceiros_sql_periodo_criado('d', data_inicio, data_fim)
        where_parts.extend(wp)
        params.extend(pp)
        sql = (
            '''SELECT COALESCE(NULLIF(TRIM(e.usuario), ''), 'Sem usuário') AS usuario,
                      e.evento,
                      COUNT(*) AS total,
                      COUNT(DISTINCT e.documento_id) AS num_nfs
               FROM ''' + tbl_e + ''' e
               INNER JOIN ''' + tbl_d + ''' d ON d.id = e.documento_id
               WHERE ''' + ' AND '.join(where_parts) + '''
               GROUP BY e.usuario, e.evento
               ORDER BY total DESC, usuario, e.evento'''
        )
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    linhas = []
    for r in rows or []:
        rd = dict(r) if hasattr(r, 'keys') else {}
        linhas.append([
            rd.get('usuario') or '',
            rd.get('evento') or '',
            int(rd.get('total') or 0),
            int(rd.get('num_nfs') or 0),
        ])
    wb = _terceiros_workbook_simple(
        ['Usuário', 'Tipo de evento', 'Quantidade', 'NFs distintas'],
        linhas,
    )
    wb.active.title = 'Auditoria usuário'
    return wb


def _relatorio_terceiros_sla_etapas(data_inicio='', data_fim=''):
    docs = _terceiros_docs_filtrados(data_inicio, data_fim)
    linhas = []
    for item in docs:
        row = item['row']
        linhas.append([
            row.get('id'),
            _terceiros_nf_texto_rel(row),
            _terceiros_tipo_fluxo_rel(row),
            _terceiros_dias_entre(row.get('criado_em'), row.get('recebimento_concluido_em')),
            _terceiros_dias_entre(row.get('recebimento_concluido_em'), row.get('nota_lancada_em')),
            _terceiros_dias_entre(row.get('nota_lancada_em'), row.get('enviar_para_mg_em')),
            _terceiros_dias_entre(row.get('enviar_para_mg_em'), row.get('carga_recebida_mg_em')),
            _terceiros_dias_entre(row.get('criado_em'), row.get('carga_recebida_mg_em')),
            item['etapa'],
        ])
    wb = _terceiros_workbook_simple(
        [
            'ID', 'NF', 'Tipo fluxo',
            'Dias cadastro → recebimento', 'Dias recebimento → lançamento',
            'Dias lançamento → envio MG', 'Dias envio → recebida MG',
            'Dias cadastro → recebida MG (total)', 'Etapa atual',
        ],
        linhas,
    )
    wb.active.title = 'SLA etapas'
    return wb


def _relatorio_terceiros_nf_detalhe(documento_id):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_i = _tbl_terceiros_documento_itens(conn)
        cols = _sql_cols_terceiros_documentos_listagem('d')
        doc_row = conn.execute(
            'SELECT ' + cols + ' FROM ' + tbl_d + ' d WHERE d.id = ?',
            (documento_id,),
        ).fetchone()
        if not doc_row:
            return None
        doc = dict(doc_row) if hasattr(doc_row, 'keys') else {}
        itens = conn.execute(
            '''SELECT n_item, codigo_ean, descricao_xml, quantidade_xml, quantidade_bipada,
                      status_bipagem, atualizado_em, atualizado_por
               FROM ''' + tbl_i + ''' WHERE documento_id = ? ORDER BY n_item, id''',
            (documento_id,),
        ).fetchall()
    finally:
        conn.close()
    from openpyxl import Workbook
    wb = Workbook()
    ws_nf = wb.active
    ws_nf.title = 'NF'
    hf = Font(bold=True)
    nf = '/'.join(filter(None, [str(doc.get('numero_nf') or '').strip(), str(doc.get('serie_nf') or '').strip()])) or '-'
    cabecalho = [
        ('ID', doc.get('id')),
        ('NF', nf),
        ('Área', doc.get('area') or ''),
        ('Remetente', doc.get('remetente_nome') or ''),
        ('Destinatário', doc.get('destinatario_nome') or ''),
        ('Previsão', _fmt_datahora_br(doc.get('previsao_chegada') or '')),
        ('Motorista que trouxe', doc.get('motorista_carreta') or ''),
        ('Placa chegada', doc.get('placa_carreta') or ''),
        ('Motorista que levou MG', doc.get('motorista_saida_mg') or ''),
        ('Placa saída MG', doc.get('placa_saida_mg') or ''),
        ('Quem recebeu MG', doc.get('recebedor_mg') or ''),
        ('Recebimento concluído', 'Sim' if _terceiros_bool_sim_db(doc.get('recebimento_concluido')) else 'Não'),
        ('Receb. concluído em', _fmt_datahora_br(doc.get('recebimento_concluido_em') or '')),
    ]
    for ri, (lb, val) in enumerate(cabecalho, 1):
        ws_nf.cell(row=ri, column=1, value=lb).font = hf
        ws_nf.cell(row=ri, column=2, value=val)
    ws_it = wb.create_sheet('Itens')
    headers = ['Item', 'Código', 'Produto', 'Qtd. XML', 'Qtd. bipada', 'Status', 'Atualizado em', 'Usuário']
    for col, h in enumerate(headers, 1):
        ws_it.cell(row=1, column=col, value=h).font = hf
    for row_idx, r in enumerate(itens or [], 2):
        rd = dict(r) if hasattr(r, 'keys') else {}
        ws_it.cell(row=row_idx, column=1, value=rd.get('n_item'))
        ws_it.cell(row=row_idx, column=2, value=rd.get('codigo_ean') or '')
        ws_it.cell(row=row_idx, column=3, value=(rd.get('descricao_xml') or '')[:120])
        ws_it.cell(row=row_idx, column=4, value=float(rd.get('quantidade_xml') or 0))
        ws_it.cell(row=row_idx, column=5, value=float(rd.get('quantidade_bipada') or 0))
        ws_it.cell(row=row_idx, column=6, value=rd.get('status_bipagem') or '')
        ws_it.cell(row=row_idx, column=7, value=_fmt_datahora_br(rd.get('atualizado_em') or ''))
        ws_it.cell(row=row_idx, column=8, value=rd.get('atualizado_por') or '')
    return wb


@app.route('/api/terceiros/relatorios/excel/resumo_nf', methods=['GET'])
def export_terceiros_resumo_nf():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_resumo_nf(de, ate), 'relatorio_terceiros_resumo_nf.xlsx')


@app.route('/api/terceiros/relatorios/excel/itens_bipados', methods=['GET'])
def export_terceiros_itens_bipados():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_itens_bipados(de, ate), 'relatorio_terceiros_itens_bipados.xlsx')


@app.route('/api/terceiros/relatorios/excel/itens_mais_bipados', methods=['GET'])
def export_terceiros_itens_mais_bipados():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_itens_mais_bipados(de, ate), 'relatorio_terceiros_itens_mais_bipados.xlsx')


@app.route('/api/terceiros/relatorios/excel/divergencias', methods=['GET'])
def export_terceiros_divergencias():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_divergencias(de, ate), 'relatorio_terceiros_divergencias.xlsx')


@app.route('/api/terceiros/relatorios/excel/carreta', methods=['GET'])
def export_terceiros_carreta():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_carreta(de, ate), 'relatorio_terceiros_carreta.xlsx')


@app.route('/api/terceiros/relatorios/excel/notas_lancadas', methods=['GET'])
def export_terceiros_notas_lancadas():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_notas_lancadas(de, ate), 'relatorio_terceiros_notas_lancadas.xlsx')


@app.route('/api/terceiros/relatorios/excel/historico', methods=['GET'])
def export_terceiros_historico():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_historico(de, ate), 'relatorio_terceiros_historico.xlsx')


@app.route('/api/terceiros/relatorios/excel/consumivel_sp', methods=['GET'])
def export_terceiros_consumivel_sp():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_consumivel_sp(de, ate), 'relatorio_terceiros_consumivel_sp.xlsx')


@app.route('/api/terceiros/relatorios/excel/pendencias_etapa', methods=['GET'])
def export_terceiros_pendencias_etapa():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_pendencias_etapa(de, ate), 'relatorio_terceiros_pendencias_etapa.xlsx')


@app.route('/api/terceiros/relatorios/excel/encerradas_motivo', methods=['GET'])
def export_terceiros_encerradas_motivo():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_encerradas_motivo(de, ate), 'relatorio_terceiros_encerradas_motivo.xlsx')


@app.route('/api/terceiros/relatorios/excel/fluxo_mg', methods=['GET'])
def export_terceiros_fluxo_mg():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_fluxo_mg(de, ate), 'relatorio_terceiros_fluxo_mg.xlsx')


@app.route('/api/terceiros/relatorios/excel/por_uf', methods=['GET'])
def export_terceiros_por_uf():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_por_uf(de, ate), 'relatorio_terceiros_por_uf.xlsx')


@app.route('/api/terceiros/relatorios/excel/por_remetente', methods=['GET'])
def export_terceiros_por_remetente():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_por_remetente(de, ate), 'relatorio_terceiros_por_remetente.xlsx')


@app.route('/api/terceiros/relatorios/excel/previsao_chegada', methods=['GET'])
def export_terceiros_previsao_chegada():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_previsao_chegada(de, ate), 'relatorio_terceiros_previsao_chegada.xlsx')


@app.route('/api/terceiros/relatorios/excel/recebimentos_periodo', methods=['GET'])
def export_terceiros_recebimentos_periodo():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_recebimentos_periodo(de, ate), 'relatorio_terceiros_recebimentos.xlsx')


@app.route('/api/terceiros/relatorios/excel/divergencia_itens', methods=['GET'])
def export_terceiros_divergencia_itens():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_divergencia_itens(de, ate), 'relatorio_terceiros_divergencia_itens.xlsx')


@app.route('/api/terceiros/relatorios/excel/sem_bipagem', methods=['GET'])
def export_terceiros_sem_bipagem():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_sem_bipagem(de, ate), 'relatorio_terceiros_sem_bipagem.xlsx')


@app.route('/api/terceiros/relatorios/excel/conferencia_incompleta', methods=['GET'])
def export_terceiros_conferencia_incompleta():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_conferencia_incompleta(de, ate), 'relatorio_terceiros_conferencia_incompleta.xlsx')


@app.route('/api/terceiros/relatorios/excel/eventos', methods=['GET'])
def export_terceiros_eventos():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_eventos(de, ate), 'relatorio_terceiros_eventos.xlsx')


@app.route('/api/terceiros/relatorios/excel/auditoria_usuario', methods=['GET'])
def export_terceiros_auditoria_usuario():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_auditoria_usuario(de, ate), 'relatorio_terceiros_auditoria_usuario.xlsx')


@app.route('/api/terceiros/relatorios/excel/sla_etapas', methods=['GET'])
def export_terceiros_sla_etapas():
    de, ate = _terceiros_periodo_request_args()
    return _terceiros_send_workbook(_relatorio_terceiros_sla_etapas(de, ate), 'relatorio_terceiros_sla_etapas.xlsx')


@app.route('/api/terceiros/relatorios/excel/nf', methods=['GET'])
def export_terceiros_nf_detalhe():
    doc_id = request.args.get('documento_id', type=int) or request.args.get('id', type=int)
    if not doc_id:
        return jsonify({'erro': 'Informe o ID da NF (documento_id).'}), 400
    wb = _relatorio_terceiros_nf_detalhe(doc_id)
    if wb is None:
        return jsonify({'erro': 'NF não encontrada.'}), 404
    return _terceiros_send_workbook(wb, 'relatorio_terceiros_nf_%s.xlsx' % doc_id)


@app.route('/api/terceiros/upload-xml', methods=['POST'])
def api_terceiros_upload_xml():
    area = (request.form.get('area') or '').strip().lower()
    previsao = (request.form.get('previsao_chegada') or '').strip()
    numero_pedido_form = _normalizar_numero_pedido_terceiros(request.form.get('numero_pedido') or '')
    consumivel_sp_form = _valor_bool_texto(request.form.get('consumivel_sp') or '') == 'sim'
    recebedor_consumivel_sp_form = (request.form.get('recebedor_consumivel_sp') or '').strip()
    if area not in ('recebimento', 'expedicao', 'carreta'):
        return jsonify({'ok': False, 'erro': 'Área inválida.'}), 400
    if not previsao:
        return jsonify({'ok': False, 'erro': 'Informe a previsão de chegada.'}), 400
    if area != 'carreta' and not numero_pedido_form:
        return jsonify({'ok': False, 'erro': 'Informe o número de pedido antes de subir XML de terceiros.'}), 400
    if area == 'carreta' and consumivel_sp_form:
        return jsonify({'ok': False, 'erro': 'XML de carreta não pode ser marcado como consumível SP.'}), 400
    if consumivel_sp_form and not recebedor_consumivel_sp_form:
        return jsonify({'ok': False, 'erro': 'Informe quem solicitou/irá receber o consumível SP.'}), 400
    motorista_carreta_form = (request.form.get('motorista_carreta') or '').strip()
    placa_carreta_form = (request.form.get('placa_carreta') or '').strip().upper()
    if area == 'carreta':
        if not motorista_carreta_form:
            return jsonify({'ok': False, 'erro': 'Informe o nome do motorista da carreta.'}), 400
        if not placa_carreta_form:
            return jsonify({'ok': False, 'erro': 'Informe a placa da carreta.'}), 400
    arquivos = request.files.getlist('files')
    if not arquivos:
        return jsonify({'ok': False, 'erro': 'Nenhum XML enviado.'}), 400
    conn = get_db()
    usuario = session.get('usuario', '')
    criados, erros, criados_rows = [], [], []
    vistos_lote = set()
    try:
        _ensure_terceiros_schema(conn, rodar_backfill=False)
        _pg_auditoria_sync_desligar(conn)
        resolver_base = _terceiros_criar_resolver_base(conn)
        chaves_db_cache = set()
        for arquivo in arquivos:
            nome = secure_filename(arquivo.filename or '')
            if not nome:
                erros.append('Arquivo sem nome.')
                continue
            if not allowed_xml_file(nome):
                erros.append('%s: formato não permitido.' % nome)
                continue
            xml_texto = arquivo.read().decode('utf-8', errors='ignore')
            if not xml_texto.strip():
                erros.append('%s: XML vazio.' % nome)
                continue
            try:
                xml_data = _parse_nfe_xml(xml_texto, resolver_base=resolver_base)
                if area != 'carreta':
                    xml_data['numero_pedido'] = numero_pedido_form
                    xml_data['consumivel_sp'] = 'sim' if consumivel_sp_form else ''
                    xml_data['recebedor_consumivel_sp'] = recebedor_consumivel_sp_form if consumivel_sp_form else ''
                identificador = _identificador_duplicidade_terceiros(xml_data)
                if identificador[1] and identificador in vistos_lote:
                    erros.append('%s: XML desta NF já foi incluído neste envio.' % nome)
                    continue
                chave_dup = (xml_data.get('chave_nfe') or '').strip()
                if chave_dup:
                    if _terceiros_chave_nfe_ja_existe(conn, chave_dup, chaves_db_cache):
                        erros.append('%s: XML desta NF já foi enviado anteriormente.' % nome)
                        continue
                elif _documento_terceiros_ja_existe(conn, xml_data):
                    erros.append('%s: XML desta NF já foi enviado anteriormente.' % nome)
                    continue
                if identificador[1]:
                    vistos_lote.add(identificador)
                doc_id = _criar_documento_terceiros(
                    conn, area, previsao, nome, xml_texto, xml_data, usuario,
                    motorista_carreta_form if area == 'carreta' else None,
                    placa_carreta_form if area == 'carreta' else None,
                )
                criados.append(doc_id)
                if chave_dup:
                    chaves_db_cache.add(chave_dup)
                criados_rows.append(_terceiros_row_listagem_apos_criar(
                    doc_id, area, previsao, xml_data, usuario,
                    motorista_carreta_form if area == 'carreta' else None,
                    placa_carreta_form if area == 'carreta' else None,
                ))
            except Exception as e:
                erros.append('%s: %s' % (nome, str(e)))
        conn.commit()
        return jsonify({
            'ok': True,
            'criados': criados,
            'criados_rows': criados_rows,
            'total_criados': len(criados),
            'erros': erros,
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'erro': str(e), 'criados': criados, 'erros': erros}), 500
    finally:
        conn.close()


def _terceiros_bool_sim_db(valor):
  if valor is True or valor == 1:
    return True
  txt = str(valor or '').strip().lower()
  return txt in ('sim', 's', 'true', '1', 'yes')


def _terceiros_etapa_painel_row(row):
  """Mesma lógica das abas do módulo (pendência → fornecedores → lançamento → MG ou histórico carreta)."""
  if _terceiros_eh_area_carreta(row):
    cmg_carreta = _valor_bool_texto(row.get('carga_recebida_mg') or '')
    if cmg_carreta == 'sim':
      return 'historico'
  if _terceiros_eh_consumivel_sp(row) and _valor_bool_texto(row.get('consumivel_sp_historico') or '') == 'sim':
    return 'historico'
  if _valor_bool_texto(row.get('nota_lancada') or '') == 'nao' and str(row.get('motivo_nao_lancada') or '').strip():
    return 'historico'
  if _valor_bool_texto(row.get('enviar_para_mg') or '') == 'nao' and str(row.get('motivo_nao_enviar_mg') or '').strip():
    return 'historico'
  if _valor_bool_texto(row.get('carga_recebida_mg') or '') == 'nao' and str(row.get('motivo_nao_recebida_mg') or '').strip():
    return 'historico'
  rc = (
    _terceiros_bool_sim_db(row.get('recebimento_concluido'))
    or bool(str(row.get('recebimento_concluido_em') or '').strip())
    or bool(str(row.get('nota_lancada') or '').strip())
    or bool(str(row.get('enviar_para_mg') or '').strip())
    or bool(str(row.get('carga_recebida_mg') or '').strip())
  )
  if not rc:
    return 'pendencia_recebimento'
  nl = _valor_bool_texto(row.get('nota_lancada') or '')
  if _terceiros_eh_area_carreta(row):
    if nl != 'sim':
      return 'fornecedores_recebidos'
    cmg = _valor_bool_texto(row.get('carga_recebida_mg') or '')
    if cmg != 'sim':
      return 'notas_lancadas'
    return 'historico'
  if _terceiros_eh_consumivel_sp(row):
    if nl == 'sim':
      return 'notas_lancadas'
    return 'pendentes_lancamento'
  emg_raw = str(row.get('enviar_para_mg') or '').strip().lower()
  emg = _valor_bool_texto(emg_raw)
  if emg == 'nao':
    return 'historico'
  if emg_raw == 'pendente' or not emg_raw:
    return 'pendencias_mg'
  cmg = _valor_bool_texto(row.get('carga_recebida_mg') or '')
  if emg == 'sim' and cmg != 'sim':
    return 'recebimentos_mg'
  if cmg == 'sim' and nl != 'sim':
    return 'pendentes_lancamento'
  if cmg == 'sim' and nl == 'sim':
    return 'historico'
  return 'pendencias_mg'


_ETAPAS_PAINEL_TERCEIROS_LABELS = {
  'pendencia_recebimento': 'Pendência de recebimento',
  'fornecedores_recebidos': 'Fornecedores recebidos',
  'pendentes_lancamento': 'Pendentes de lançamento',
  'notas_lancadas': 'Notas lançadas',
  'pendencias_mg': 'Pendências envio MG',
  'recebimentos_mg': 'Recebimentos MG',
  'historico': 'Histórico',
}


@app.route('/api/terceiros/painel', methods=['GET'])
def api_terceiros_painel():
  """Resumo e agregações para o painel de recebimento de terceiros."""
  conn = get_db()
  try:
    _ensure_terceiros_schema(conn, rodar_backfill=False)
    tbl_d = _tbl_terceiros_documentos(conn)
    tbl_i = _tbl_terceiros_documento_itens(conn)
    cols = _sql_cols_terceiros_documentos_listagem('d')
    data_inicio = (request.args.get('data_inicio') or request.args.get('data') or '').strip()[:10] or None
    data_fim = (request.args.get('data_fim') or request.args.get('data') or '').strip()[:10] or None
    hora_inicio = (request.args.get('hora_inicio') or '').strip()[:5] or None
    hora_fim = (request.args.get('hora_fim') or '').strip()[:5] or None
    wp, pp, janela_meta = _terceiros_sql_periodo_janela('d', conn, data_inicio, data_fim, hora_inicio, hora_fim)
    filtros_resp = {
      'data': janela_meta['data'],
      'data_iso': janela_meta['data_iso'],
      'data_inicio': janela_meta['data_inicio_iso'],
      'data_fim': janela_meta['data_fim_iso'],
      'hora_inicio': janela_meta['hora_inicio'],
      'hora_fim': janela_meta['hora_fim'],
      'legivel': janela_meta['legivel'],
    }
    where_sql = ' WHERE ' + ' AND '.join(wp)
    doc_rows = conn.execute(
      'SELECT ' + _terceiros_cols_select_listagem(cols, com_resumo=True)
      + ' FROM ' + tbl_d + ' d'
      + _terceiros_sql_join_resumo_itens(conn, tbl_i, 'd')
      + where_sql
      + ' ORDER BY d.criado_em DESC, d.id DESC',
      tuple(pp),
    ).fetchall()

    def _row_dict(r):
      return dict(r) if hasattr(r, 'keys') else {}

    docs = [_row_dict(r) for r in (doc_rows or [])]
    etapa_counts = {k: 0 for k in _ETAPAS_PAINEL_TERCEIROS_LABELS}
    remetente_agg = {}
    placa_agg = {}
    uf_agg = {}
    conferencia_ok = 0
    conferencia_div = 0
    total_qtd_xml = 0.0
    total_qtd_bip = 0.0
    total_itens_nf = 0
    nfs_carreta = 0
    recebimento_concluido = 0
    fornecedores_recebidos = 0

    ultimas = []
    for row in docs:
      q_xml = float(row.get('quantidade_total_xml') or 0)
      q_bip = float(row.get('quantidade_total_bipada') or 0)
      div_it = int(row.get('itens_divergentes') or 0)
      total_qtd_xml += q_xml
      total_qtd_bip += q_bip
      total_itens_nf += int(row.get('total_itens') or 0)
      if (row.get('area') or '').strip().lower() == 'carreta':
        nfs_carreta += 1
      if _terceiros_bool_sim_db(row.get('recebimento_concluido')):
        recebimento_concluido += 1
        fornecedores_recebidos += 1
      etapa = _terceiros_etapa_painel_row(row)
      etapa_counts[etapa] = etapa_counts.get(etapa, 0) + 1
      if div_it == 0 and abs(q_xml - q_bip) <= 1e-6 and q_xml > 1e-9:
        conferencia_ok += 1
      elif q_bip > 1e-9 or div_it > 0:
        conferencia_div += 1
      rem = (row.get('remetente_nome') or '').strip() or 'Sem remetente'
      remetente_agg[rem] = remetente_agg.get(rem, 0) + 1
      uf = (row.get('destinatario_uf') or '').strip() or '—'
      uf_agg[uf] = uf_agg.get(uf, 0) + 1
      plc = (row.get('placa_carreta') or '').strip().upper()
      if plc:
        placa_agg[plc] = placa_agg.get(plc, 0) + 1
      if len(ultimas) < 30:
        nf_txt = '/'.join(filter(None, [str(row.get('numero_nf') or '').strip(), str(row.get('serie_nf') or '').strip()])) or '-'
        ultimas.append({
          'id': row.get('id'),
          'nf': nf_txt,
          'remetente': rem,
          'destinatario': (row.get('destinatario_nome') or '').strip() or '-',
          'uf': uf,
          'previsao': _fmt_datahora_br(row.get('previsao_chegada') or ''),
          'etapa': _ETAPAS_PAINEL_TERCEIROS_LABELS.get(etapa, etapa),
          'qtd_xml': q_xml,
          'qtd_bipada': q_bip,
          'recebimento_concluido': _terceiros_bool_sim_db(row.get('recebimento_concluido')),
          'nota_lancada': row.get('nota_lancada') or '',
          'area': row.get('area') or '',
        })

    def _top_dict(agg, limit=12):
      return [{'nome': k, 'total': v} for k, v in sorted(agg.items(), key=lambda x: (-x[1], x[0]))[:limit]]

    itens_rows = conn.execute(
      '''SELECT COALESCE(NULLIF(TRIM(i.descricao_xml), ''), NULLIF(TRIM(i.codigo_ean), ''), 'Item') AS produto,
                COALESCE(NULLIF(TRIM(i.codigo_ean), ''), '-') AS codigo_ean,
                COALESCE(SUM(i.quantidade_bipada), 0) AS total
         FROM ''' + tbl_i + ''' i
         INNER JOIN ''' + tbl_d + ''' d ON d.id = i.documento_id
         WHERE COALESCE(i.quantidade_bipada, 0) > 0
           AND ''' + ' AND '.join(wp) + '''
         GROUP BY i.descricao_xml, i.codigo_ean
         ORDER BY total DESC
         LIMIT 20''',
      tuple(pp),
    ).fetchall()

    motorista_rows = conn.execute(
      '''SELECT COALESCE(NULLIF(TRIM(d.motorista_carreta), ''), 'Sem motorista') AS motorista,
                COUNT(*) AS nfs
         FROM ''' + tbl_d + ''' d
         WHERE d.motorista_carreta IS NOT NULL AND TRIM(COALESCE(d.motorista_carreta, '')) != ''
           AND ''' + ' AND '.join(wp) + '''
         GROUP BY d.motorista_carreta
         ORDER BY nfs DESC
         LIMIT 12''',
      tuple(pp),
    ).fetchall()

    chegadas_rows = conn.execute(
      '''SELECT d.id,
                d.numero_nf,
                d.serie_nf,
                COALESCE(NULLIF(TRIM(d.motorista_carreta), ''), '-') AS motorista,
                COALESCE(NULLIF(TRIM(d.placa_carreta), ''), '-') AS placa,
                MIN(i.atualizado_em) AS inicio_descarga
         FROM ''' + tbl_d + ''' d
         INNER JOIN ''' + tbl_i + ''' i ON i.documento_id = d.id
         WHERE LOWER(TRIM(COALESCE(d.area, ''))) = 'carreta'
           AND COALESCE(i.quantidade_bipada, 0) > 0
           AND ''' + ' AND '.join(wp) + '''
         GROUP BY d.id, d.numero_nf, d.serie_nf, d.motorista_carreta, d.placa_carreta
         ORDER BY inicio_descarga DESC
         LIMIT 20''',
      tuple(pp),
    ).fetchall()

    top_itens = []
    for r in itens_rows or []:
      rd = _row_dict(r)
      top_itens.append({
        'produto': rd.get('produto') or 'Item',
        'codigo_ean': rd.get('codigo_ean') or '-',
        'total': float(rd.get('total') or 0),
      })

    top_motoristas = []
    for r in motorista_rows or []:
      rd = _row_dict(r)
      top_motoristas.append({
        'motorista': rd.get('motorista') or '-',
        'nfs': int(rd.get('nfs') or 0),
      })

    chegadas_carreta = []
    for r in chegadas_rows or []:
      rd = _row_dict(r)
      nf_txt = '/'.join(filter(None, [str(rd.get('numero_nf') or '').strip(), str(rd.get('serie_nf') or '').strip()])) or '-'
      inicio_raw = rd.get('inicio_descarga')
      chegadas_carreta.append({
        'nf': nf_txt,
        'motorista': rd.get('motorista') or '-',
        'placa': (rd.get('placa') or '-').upper() if rd.get('placa') else '-',
        'inicio_descarga': _fmt_datahora_br(inicio_raw),
        'inicio_descarga_ordem': inicio_raw.isoformat() if hasattr(inicio_raw, 'isoformat') else str(inicio_raw or ''),
      })
    chegadas_carreta.reverse()

    etapas_chart = [
      {'etapa': k, 'label': _ETAPAS_PAINEL_TERCEIROS_LABELS[k], 'total': etapa_counts.get(k, 0)}
      for k in _ETAPAS_PAINEL_TERCEIROS_LABELS
      if etapa_counts.get(k, 0) > 0
    ]

    return jsonify({
      'filtros': filtros_resp,
      'estatisticas': {
        'total_nf': len(docs),
        'pendencia_recebimento': etapa_counts.get('pendencia_recebimento', 0),
        'fornecedores_recebidos': fornecedores_recebidos,
        'recebimento_concluido': recebimento_concluido,
        'pendentes_lancamento': etapa_counts.get('pendentes_lancamento', 0),
        'notas_lancadas': etapa_counts.get('notas_lancadas', 0),
        'pendencias_mg': etapa_counts.get('pendencias_mg', 0),
        'recebimentos_mg': etapa_counts.get('recebimentos_mg', 0),
        'total_itens': total_itens_nf,
        'quantidade_total_xml': total_qtd_xml,
        'quantidade_total_bipada': total_qtd_bip,
        'nfs_carreta': nfs_carreta,
        'conferencia_ok': conferencia_ok,
        'conferencia_divergente': conferencia_div,
      },
      'etapas': etapas_chart,
      'top_remetentes': _top_dict(remetente_agg),
      'top_placas': _top_dict(placa_agg),
      'por_uf': _top_dict(uf_agg, 10),
      'ultimas_nfs': ultimas,
      'top_itens': top_itens,
      'top_motoristas': top_motoristas,
      'chegadas_carreta': chegadas_carreta,
    })
  except Exception as e:
    return jsonify({'erro': str(e)}), 500
  finally:
    conn.close()


def _terceiros_row_listagem_apos_criar(documento_id, area, previsao_chegada, xml_data, usuario, motorista_carreta=None, placa_carreta=None):
    """Monta criados_rows sem SELECT pesado (upload)."""
    itens = xml_data.get('itens') or []
    q_xml = round(sum(float(i.get('quantidade_xml') or 0) for i in itens), 3)
    agora = _agora_iso()
    row = {
        'id': documento_id,
        'area': area,
        'chave_nfe': xml_data.get('chave_nfe') or '',
        'numero_nf': xml_data.get('numero_nf') or '',
        'serie_nf': xml_data.get('serie_nf') or '',
        'data_emissao': xml_data.get('data_emissao') or '',
        'remetente_nome': xml_data.get('remetente_nome') or '',
        'remetente_cnpj': _somente_digitos(xml_data.get('remetente_cnpj') or ''),
        'destinatario_nome': xml_data.get('destinatario_nome') or '',
        'destinatario_cnpj': _somente_digitos(xml_data.get('destinatario_cnpj') or ''),
        'destinatario_uf': (xml_data.get('destinatario_uf') or '').strip().upper(),
        'numero_pedido': (xml_data.get('numero_pedido') or '').strip(),
        'previsao_chegada': previsao_chegada or '',
        'recebimento_concluido': False,
        'nota_lancada': '',
        'enviar_para_mg': '',
        'motorista_carreta': (motorista_carreta or '').strip(),
        'placa_carreta': (placa_carreta or '').strip().upper(),
        'motorista_saida_mg': '',
        'placa_saida_mg': '',
        'carga_recebida_mg': '',
        'recebedor_mg': '',
        'consumivel_sp': 'sim' if _valor_bool_texto(xml_data.get('consumivel_sp') or '') == 'sim' else '',
        'recebedor_consumivel_sp': (xml_data.get('recebedor_consumivel_sp') or '').strip(),
        'consumivel_sp_historico': '',
        'criado_por': usuario or '',
        'atualizado_por': usuario or '',
        'total_itens': len(itens),
        'quantidade_total_xml': q_xml,
        'quantidade_total_bipada': 0.0,
        'itens_divergentes': len(itens),
        'criado_em': agora,
        'atualizado_em': agora,
    }
    return _terceiros_serializar_row_listagem(row)


def _terceiros_serializar_row_listagem(row):
    """Formata uma linha da listagem de NFs (mesmo formato do GET /documentos)."""
    rc = row.get('recebimento_concluido')
    if isinstance(rc, str) and rc.strip().lower() in ('sim', 's', 'true', '1'):
        rc_flag = True
    elif isinstance(rc, str) and rc.strip().lower() in ('nao', 'n', 'false', '0', ''):
        rc_flag = False
    else:
        rc_flag = bool(rc) if rc is not None else False
    if not rc_flag:
        rc_flag = any(str(row.get(campo) or '').strip() for campo in (
            'recebimento_concluido_em',
            'nota_lancada',
            'enviar_para_mg',
            'carga_recebida_mg',
        ))
    return {
        'id': row.get('id'),
        'area': row.get('area') or '',
        'numero_nf': row.get('numero_nf') or '',
        'serie_nf': row.get('serie_nf') or '',
        'numero_pedido': _normalizar_numero_pedido_terceiros(row.get('numero_pedido')),
        'chave_nfe': row.get('chave_nfe') or '',
        'data_emissao': _fmt_data_br(row.get('data_emissao') or ''),
        'remetente_nome': row.get('remetente_nome') or '',
        'destinatario_nome': row.get('destinatario_nome') or '',
        'destinatario_uf': (row.get('destinatario_uf') or '').strip().upper(),
        'previsao_chegada': _fmt_datahora_br(row.get('previsao_chegada') or ''),
        'recebimento_concluido': rc_flag,
        'nota_lancada': row.get('nota_lancada') or '',
        'enviar_para_mg': row.get('enviar_para_mg') or '',
        'motorista_carreta': row.get('motorista_carreta') or '',
        'placa_carreta': row.get('placa_carreta') or '',
        'motorista_saida_mg': row.get('motorista_saida_mg') or '',
        'placa_saida_mg': row.get('placa_saida_mg') or '',
        'motorista_obrigatorio': _motorista_obrigatorio_terceiros(row),
        'carga_recebida_mg': row.get('carga_recebida_mg') or '',
        'recebedor_mg': row.get('recebedor_mg') or '',
        'consumivel_sp': row.get('consumivel_sp') or '',
        'recebedor_consumivel_sp': row.get('recebedor_consumivel_sp') or '',
        'consumivel_sp_historico': row.get('consumivel_sp_historico') or '',
        'consumivel_sp_historico_por': row.get('consumivel_sp_historico_por') or '',
        'motivo_nao_lancada': row.get('motivo_nao_lancada') or '',
        'motivo_nao_enviar_mg': row.get('motivo_nao_enviar_mg') or '',
        'motivo_nao_recebida_mg': row.get('motivo_nao_recebida_mg') or '',
        'criado_por': row.get('criado_por') or '',
        'atualizado_por': row.get('atualizado_por') or '',
        'recebimento_concluido_por': row.get('recebimento_concluido_por') or '',
        'nota_lancada_por': row.get('nota_lancada_por') or '',
        'enviar_para_mg_por': row.get('enviar_para_mg_por') or '',
        'carga_recebida_mg_por': row.get('carga_recebida_mg_por') or '',
        'total_itens': int(row.get('total_itens') or 0),
        'quantidade_total_xml': float(row.get('quantidade_total_xml') or 0),
        'quantidade_total_bipada': float(row.get('quantidade_total_bipada') or 0),
        'itens_divergentes': int(row.get('itens_divergentes') or 0),
        'criado_em': _fmt_datahora_br(row.get('criado_em') or ''),
        'atualizado_em': _fmt_datahora_br(row.get('atualizado_em') or ''),
        'recebimento_concluido_em': _fmt_datahora_br(row.get('recebimento_concluido_em') or ''),
        'nota_lancada_em': _fmt_datahora_br(row.get('nota_lancada_em') or ''),
        'enviar_para_mg_em': _fmt_datahora_br(row.get('enviar_para_mg_em') or ''),
        'motorista_carreta_em': _fmt_datahora_br(row.get('motorista_carreta_em') or ''),
        'motorista_saida_mg_em': _fmt_datahora_br(row.get('motorista_saida_mg_em') or ''),
        'carga_recebida_mg_em': _fmt_datahora_br(row.get('carga_recebida_mg_em') or ''),
        'consumivel_sp_historico_em': _fmt_datahora_br(row.get('consumivel_sp_historico_em') or ''),
    }


def _terceiros_chave_nfe_ja_existe(conn, chave, cache_db):
    """Verifica chave no banco com cache por requisição (upload)."""
    chave = (chave or '').strip()
    if not chave:
        return False
    if chave in cache_db:
        return True
    row = conn.execute(
        'SELECT id FROM ' + _tbl_terceiros_documentos(conn) + ' WHERE chave_nfe = ? LIMIT 1',
        (chave,),
    ).fetchone()
    if row:
        cache_db.add(chave)
        return True
    return False


def _terceiros_listagem_rows_por_ids(conn, documento_ids, enriquecer=True):
    """Carrega só as NFs recém-criadas (evita GET de todas as notas após upload)."""
    ids = []
    for x in documento_ids or []:
        try:
            n = int(x)
            if n > 0:
                ids.append(n)
        except (TypeError, ValueError):
            pass
    if not ids:
        return []
    cols = _sql_cols_terceiros_documentos_listagem('d')
    tbl_d = _tbl_terceiros_documentos(conn)
    tbl_i = _tbl_terceiros_documento_itens(conn)
    ph = ','.join(['?'] * len(ids))
    rows = conn.execute(
        'SELECT ' + _terceiros_cols_select_listagem(cols, com_resumo=True)
        + ' FROM ' + tbl_d + ' d'
        + _terceiros_sql_join_resumo_itens(conn, tbl_i, 'd')
        + ' WHERE d.id IN (' + ph + ')'
        + ' ORDER BY d.criado_em DESC, d.id DESC',
        tuple(ids),
    ).fetchall()
    rows_list = [dict(r) if hasattr(r, 'keys') else {} for r in (rows or [])]
    if enriquecer:
        try:
            _terceiros_enriquecer_campos_xml_listagem(conn, rows_list)
        except Exception:
            pass
    return [_terceiros_serializar_row_listagem(row) for row in rows_list]


@app.route('/api/terceiros/documentos', methods=['GET'])
def api_terceiros_documentos():
    area = (request.args.get('area') or '').strip().lower()
    areas_unica = ('recebimento', 'expedicao', 'carreta')
    if area in ('todas', 'all'):
        where_area = 'd.area IN (?, ?, ?)'
        params_area = ('recebimento', 'expedicao', 'carreta')
    elif area in areas_unica:
        where_area = 'd.area = ?'
        params_area = (area,)
    else:
        return jsonify({'erro': 'Área inválida.', 'rows': []}), 400
    leve = (request.args.get('leve') or '').strip().lower() in ('1', 'sim', 'true', 'yes')
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn, rodar_backfill=False)
        cols = _sql_cols_terceiros_documentos_listagem('d')
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_i = _tbl_terceiros_documento_itens(conn)
        join_resumo = '' if leve else _terceiros_sql_join_resumo_itens(conn, tbl_i, 'd')
        rows = conn.execute(
            'SELECT ' + _terceiros_cols_select_listagem(cols, com_resumo=not leve)
            + ' FROM ' + tbl_d + ' d' + join_resumo
            + ' WHERE ' + where_area
            + ' ORDER BY d.criado_em DESC, d.id DESC',
            params_area
        ).fetchall()
        rows_list = [dict(r) if hasattr(r, 'keys') else {} for r in (rows or [])]
        out = [_terceiros_serializar_row_listagem(row) for row in rows_list]
        return jsonify({'rows': out})
    except Exception as e:
        return jsonify({'erro': str(e), 'rows': []}), 500
    finally:
        conn.close()


@app.route('/api/terceiros/documentos/<int:documento_id>', methods=['GET'])
def api_terceiros_documento_detalhe(documento_id):
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        incluir_eventos = (request.args.get('eventos') or '').strip().lower() in ('1', 'sim', 'true', 'yes')
        doc = _carregar_documento_terceiros(conn, documento_id, incluir_eventos=incluir_eventos)
        if not doc:
            return jsonify({'erro': 'Documento não encontrado.'}), 404
        doc = _terceiros_enriquecer_doc_wms(conn, doc, sincronizar=True, incluir_eventos=incluir_eventos)
        doc['motorista_obrigatorio'] = _motorista_obrigatorio_terceiros(doc)
        doc['previsao_chegada'] = _fmt_datahora_br(doc.get('previsao_chegada') or '')
        for campo in ('recebimento_concluido_em', 'nota_lancada_em', 'enviar_para_mg_em', 'motorista_carreta_em', 'motorista_saida_mg_em', 'carga_recebida_mg_em', 'consumivel_sp_historico_em', 'criado_em', 'atualizado_em'):
            doc[campo] = _fmt_datahora_br(doc.get(campo) or '')
        for ev in doc.get('eventos') or []:
            ev['criado_em'] = _fmt_datahora_br(ev.get('criado_em') or '')
        return jsonify(doc)
    finally:
        conn.close()


def _meu_danfe_api_key():
    return (
        os.environ.get('MEU_DANFE_API_KEY')
        or os.environ.get('MEUDANFE_API_KEY')
        or ''
    ).strip()


def _meu_danfe_converter_xml_para_pdf(xml_texto):
    """Converte XML NF-e/CT-e em PDF (DANFE) via API Meu Danfe. Retorna (bytes, nome_arquivo) ou (None, mensagem_erro)."""
    api_key = _meu_danfe_api_key()
    if not api_key:
        return None, 'Configure a variável MEU_DANFE_API_KEY no servidor (.env) para gerar o PDF pelo Meu Danfe.'
    if not http_requests:
        return None, 'Biblioteca requests não instalada no servidor.'
    xml_texto = (xml_texto or '').strip()
    if not xml_texto:
        return None, 'XML da nota fiscal está vazio.'
    url = (
        os.environ.get('MEU_DANFE_API_URL')
        or 'https://api.meudanfe.com.br/v2/fd/convert/xml-to-da'
    ).strip()
    try:
        resp = http_requests.post(
            url,
            data=xml_texto.encode('utf-8'),
            headers={
                'Api-Key': api_key,
                'Content-Type': 'text/plain; charset=utf-8',
                'Accept': 'application/json',
            },
            timeout=int(os.environ.get('MEU_DANFE_TIMEOUT', '60') or 60),
        )
    except Exception as exc:
        return None, 'Falha ao contactar a API Meu Danfe: %s' % exc
    if resp.status_code == 401:
        return None, 'Api-Key Meu Danfe inválida ou não informada.'
    if resp.status_code == 403:
        return None, 'Api-Key Meu Danfe foi substituída. Atualize MEU_DANFE_API_KEY no servidor.'
    if resp.status_code == 400:
        return None, 'XML inválido ou vazio para o Meu Danfe. Confira o arquivo XML desta NF.'
    if resp.status_code >= 500:
        return None, 'Meu Danfe indisponível no momento. Tente novamente mais tarde.'
    if not resp.ok:
        det = ''
        try:
            det = (resp.json() or {}).get('message') or (resp.json() or {}).get('error') or ''
        except Exception:
            det = (resp.text or '')[:300]
        return None, 'Meu Danfe retornou erro %s%s' % (resp.status_code, (': ' + det) if det else '')
    try:
        payload = resp.json()
    except Exception:
        return None, 'Resposta inválida da API Meu Danfe.'
    b64_pdf = (payload.get('data') or '').strip()
    if not b64_pdf:
        return None, 'Meu Danfe não retornou o PDF (campo data vazio).'
    try:
        pdf_bytes = base64.b64decode(b64_pdf, validate=False)
    except Exception:
        return None, 'PDF retornado pelo Meu Danfe está corrompido (Base64 inválido).'
    if not pdf_bytes:
        return None, 'PDF retornado pelo Meu Danfe está vazio.'
    nome = secure_filename((payload.get('name') or 'danfe.pdf').strip()) or 'danfe.pdf'
    if not nome.lower().endswith('.pdf'):
        nome += '.pdf'
    return pdf_bytes, nome


@app.route('/api/terceiros/documentos/<int:documento_id>/danfe', methods=['GET'])
def api_terceiros_documento_danfe(documento_id):
    """DANFE em PDF a partir do XML (API Meu Danfe). Sem Api-Key, fallback HTML imprimível."""
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        doc = _carregar_documento_terceiros(conn, documento_id)
        if not doc:
            return Response('Documento não encontrado.', status=404, mimetype='text/plain; charset=utf-8')
        row_xml = conn.execute(
            'SELECT xml_conteudo FROM ' + _tbl_terceiros_documentos(conn) + ' WHERE id = ?',
            (documento_id,),
        ).fetchone()
        xml_texto = ''
        if row_xml:
            xml_texto = (row_xml['xml_conteudo'] if hasattr(row_xml, 'keys') else row_xml[0]) or ''
        xml_texto = str(xml_texto).strip()
        if not xml_texto:
            return Response(
                'XML da nota fiscal não disponível para esta NF.',
                status=404,
                mimetype='text/plain; charset=utf-8',
            )
        pdf_bytes, pdf_ou_erro = _meu_danfe_converter_xml_para_pdf(xml_texto)
        if pdf_bytes is not None:
            return Response(
                pdf_bytes,
                mimetype='application/pdf',
                headers={
                    'Content-Disposition': 'inline; filename="%s"' % pdf_ou_erro,
                    'Cache-Control': 'no-store',
                },
            )
        if _meu_danfe_api_key():
            return Response(pdf_ou_erro, status=502, mimetype='text/plain; charset=utf-8')
        html = _render_danfe_html_terceiros(doc)
        return Response(html, mimetype='text/html; charset=utf-8')
    except Exception as e:
        return Response('Erro ao gerar visualização da NF: %s' % str(e), status=500, mimetype='text/plain; charset=utf-8')
    finally:
        conn.close()


def _nome_arquivo_xml_terceiros(doc):
    nome = (doc.get('arquivo_nome') or '').strip()
    if nome and not nome.lower().endswith('.xml'):
        nome = nome + '.xml'
    if nome:
        return secure_filename(nome) or 'nota_fiscal.xml'
    chave = (doc.get('chave_nfe') or '').strip()
    if chave:
        return secure_filename(chave[:60] + '.xml') or 'nota_fiscal.xml'
    nf = (doc.get('numero_nf') or 'nf').strip()
    serie = (doc.get('serie_nf') or '').strip()
    base = 'NFe_%s' % nf
    if serie:
        base += '_S%s' % serie
    return secure_filename(base + '.xml') or 'nota_fiscal.xml'


@app.route('/api/terceiros/documentos/<int:documento_id>/xml', methods=['GET'])
def api_terceiros_documento_xml(documento_id):
    """Download do XML original da NF armazenado no cadastro."""
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        row = conn.execute(
            'SELECT xml_conteudo, arquivo_nome, chave_nfe, numero_nf, serie_nf FROM '
            + _tbl_terceiros_documentos(conn) + ' WHERE id = ?',
            (documento_id,),
        ).fetchone()
        if not row:
            return Response('Documento não encontrado.', status=404, mimetype='text/plain; charset=utf-8')
        doc = dict(row) if hasattr(row, 'keys') else {
            'xml_conteudo': row[0],
            'arquivo_nome': row[1],
            'chave_nfe': row[2],
            'numero_nf': row[3],
            'serie_nf': row[4],
        }
        xml_texto = (doc.get('xml_conteudo') or '').strip()
        if not xml_texto:
            return Response(
                'XML da nota fiscal não disponível para esta NF.',
                status=404,
                mimetype='text/plain; charset=utf-8',
            )
        nome_arquivo = _nome_arquivo_xml_terceiros(doc)
        return Response(
            xml_texto.encode('utf-8'),
            mimetype='application/xml; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename="%s"' % nome_arquivo},
        )
    except Exception as e:
        return Response('Erro ao baixar XML: %s' % str(e), status=500, mimetype='text/plain; charset=utf-8')
    finally:
        conn.close()


@app.route('/api/terceiros/documentos/<int:documento_id>', methods=['DELETE'])
def api_terceiros_excluir_documento(documento_id):
    conn = get_db()
    usuario = session.get('usuario', '')
    try:
        _ensure_terceiros_schema(conn)
        tbl_d = _tbl_terceiros_documentos(conn)
        tbl_i = _tbl_terceiros_documento_itens(conn)
        tbl_e = _tbl_terceiros_documento_eventos(conn)
        numero_nf = ''
        serie_nf = ''
        row = conn.execute(
            'SELECT numero_nf, serie_nf FROM ' + tbl_d + ' WHERE id = ?',
            (documento_id,),
        ).fetchone()
        if not row:
            return jsonify({'ok': False, 'erro': 'Documento não encontrado.'}), 404
        numero_nf = (row['numero_nf'] if hasattr(row, 'keys') else row[0]) or ''
        serie_nf = (row['serie_nf'] if hasattr(row, 'keys') else row[1]) or ''
        _pg_auditoria_sync_desligar(conn)
        if getattr(conn, 'kind', None) == 'pg':
            rid = str(documento_id)
            for fonte in (tbl_d, 'public.terceiros_documentos'):
                try:
                    conn.execute(
                        'DELETE FROM public.tabela_geral_snapshot WHERE fonte_tabela = ? AND row_id = ?',
                        (fonte, rid),
                    )
                except Exception:
                    pass
        conn.execute('DELETE FROM ' + tbl_i + ' WHERE documento_id = ?', (documento_id,))
        conn.execute('DELETE FROM ' + tbl_e + ' WHERE documento_id = ?', (documento_id,))
        if getattr(conn, 'kind', None) == 'pg':
            deleted = conn.execute(
                'DELETE FROM ' + tbl_d + ' WHERE id = ? RETURNING id',
                (documento_id,),
            ).fetchone()
            if not deleted:
                conn.rollback()
                return jsonify({'ok': False, 'erro': 'Documento não encontrado.'}), 404
        else:
            conn.execute('DELETE FROM ' + tbl_d + ' WHERE id = ?', (documento_id,))
        conn.commit()
        return jsonify({
            'ok': True,
            'mensagem': 'NF excluída com sucesso.',
            'documento_id': documento_id,
            'numero_nf': numero_nf,
            'serie_nf': serie_nf,
            'usuario': usuario
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'erro': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/terceiros/documentos/<int:documento_id>/bipar', methods=['POST'])
def api_terceiros_bipar_item(documento_id):
    return jsonify({'ok': False, 'erro': _TERCEIROS_BIPAGEM_MSG_WMS}), 403


@app.route('/api/terceiros/documentos/<int:documento_id>/item-motivo', methods=['POST'])
def api_terceiros_item_motivo(documento_id):
    """Motivo livre por linha (item da NF). Body: item_id, motivo."""
    data = request.get_json() or {}
    item_id = int(data.get('item_id') or 0)
    motivo = data.get('motivo')
    if motivo is None:
        motivo = ''
    motivo = (str(motivo) or '').strip()
    if len(motivo) > 2000:
        motivo = motivo[:2000]
    if not item_id:
        return jsonify({'ok': False, 'erro': 'item_id obrigatório.'}), 400
    usuario = session.get('usuario', '')
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        tbl = _tbl_terceiros_documento_itens(conn)
        row = conn.execute(
            'SELECT id FROM ' + tbl + ' WHERE id = ? AND documento_id = ?', (item_id, documento_id)
        ).fetchone()
        if not row:
            return jsonify({'ok': False, 'erro': 'Item não encontrado.'}), 404
        conn.execute(
            '''UPDATE ''' + tbl + ''' SET motivo = ?, atualizado_em = ?, atualizado_por = ? WHERE id = ? AND documento_id = ?''',
            (motivo, _agora_iso(), usuario, item_id, documento_id),
        )
        conn.execute('UPDATE ' + _tbl_terceiros_documentos(conn) + ' SET atualizado_em = ?, atualizado_por = ? WHERE id = ?', (_agora_iso(), usuario, documento_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'erro': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/terceiros/documentos/<int:documento_id>/desbipar', methods=['POST'])
def api_terceiros_desbipar_item(documento_id):
    return jsonify({'ok': False, 'erro': _TERCEIROS_BIPAGEM_MSG_WMS}), 403


@app.route('/api/terceiros/documentos/<int:documento_id>/zerar-bipagem', methods=['POST'])
def api_terceiros_zerar_bipagem(documento_id):
    return jsonify({'ok': False, 'erro': _TERCEIROS_BIPAGEM_MSG_WMS}), 403


@app.route('/api/terceiros/documentos/<int:documento_id>/status', methods=['POST'])
def api_terceiros_status(documento_id):
    data = request.get_json() or {}
    campo = (data.get('campo') or '').strip()
    valor = (data.get('valor') or '').strip()
    forcar_lancamento_sem_recebimento = bool(data.get('forcar_lancamento_sem_recebimento'))
    forcar_fluxo_carreta = bool(data.get('forcar_fluxo_carreta'))
    motivo = (data.get('motivo') or '').strip()
    recebedor_mg = (data.get('recebedor_mg') or '').strip()
    usuario = session.get('usuario', '')
    _MOTIVO_COL_TERCEIROS = {
        'nota_lancada': 'motivo_nao_lancada',
        'enviar_para_mg': 'motivo_nao_enviar_mg',
        'carga_recebida_mg': 'motivo_nao_recebida_mg',
    }
    if campo not in ('recebimento_concluido', 'nota_lancada', 'enviar_para_mg', 'carga_recebida_mg', 'consumivel_sp_historico'):
        return jsonify({'ok': False, 'erro': 'Campo inválido.'}), 400
    conn = get_db()
    try:
        if getattr(conn, 'kind', None) == 'pg':
            _ensure_pg_tabela_geral_dados(conn)
        _ensure_terceiros_schema(conn, rodar_backfill=False)
        tbl_d = _tbl_terceiros_documentos(conn)
        cols_d = _sql_cols_terceiros_documentos_listagem('d')
        row = conn.execute(
            'SELECT ' + cols_d + ' FROM ' + tbl_d + ' d WHERE d.id = ?',
            (documento_id,),
        ).fetchone()
        if not row:
            return jsonify({'ok': False, 'erro': 'Documento não encontrado.'}), 404
        doc = dict(row) if hasattr(row, 'keys') else {}
        if campo == 'consumivel_sp_historico' and not _terceiros_eh_consumivel_sp(doc):
            return jsonify({'ok': False, 'erro': 'Somente NF marcada como consumível SP pode usar este envio ao Histórico.'}), 400
        if campo == 'consumivel_sp_historico' and _valor_bool_texto(doc.get('nota_lancada') or '') != 'sim':
            return jsonify({'ok': False, 'erro': 'Consumível SP só pode ir para o Histórico depois de Lançada = Sim.'}), 400
        if campo in ('enviar_para_mg', 'carga_recebida_mg') and _terceiros_eh_consumivel_sp(doc):
            return jsonify({
                'ok': False,
                'erro': 'Consumível SP não utiliza fluxo MG. Depois de lançado, envie ao Histórico.'
            }), 400
        if campo == 'enviar_para_mg' and _terceiros_eh_area_carreta(doc):
            return jsonify({
                'ok': False,
                'erro': 'NF de carreta não utiliza envio para MG. Conclua na 5ª aba — Notas lançadas.'
            }), 400
        agora = _agora_iso()
        if campo == 'recebimento_concluido':
            ja_concluido = _terceiros_bool_sim_db(doc.get('recebimento_concluido'))
            novo_bool = str(valor).strip().lower() in ('1', 'true', 'sim', 's', 'yes')
            _pg_auditoria_sync_desligar(conn)
            conn.execute(
                'UPDATE ' + _tbl_terceiros_documentos(conn) + ' SET recebimento_concluido = ?, recebimento_concluido_em = ?, recebimento_concluido_por = ?, atualizado_em = ?, atualizado_por = ? WHERE id = ?',
                (novo_bool, agora, usuario, agora, usuario, documento_id)
            )
            _registrar_evento_terceiros(conn, documento_id, campo, str(doc.get('recebimento_concluido') or 0), str(novo_bool), usuario)
        else:
            valor_norm = _valor_bool_texto(valor)
            if campo == 'enviar_para_mg' and (valor or '').strip().lower() == 'pendente':
                valor_norm = 'pendente'
            if campo == 'consumivel_sp_historico' and valor_norm != 'sim':
                return jsonify({'ok': False, 'erro': 'Use Sim para enviar o consumível SP ao Histórico.'}), 400
            if not valor_norm:
                return jsonify({'ok': False, 'erro': 'Use sim, nao ou pendente.'}), 400
            if campo == 'nota_lancada' and valor_norm == 'sim' and not forcar_lancamento_sem_recebimento and not _terceiros_pode_lancar_nota_sem_confirmacao_recebimento(conn, doc, documento_id):
                return jsonify({
                    'ok': False,
                    'erro': 'A nota fiscal precisa ser recebida antes de ser lançada.',
                    'confirmacao_necessaria': True,
                    'codigo': 'confirmar_lancamento_sem_recebimento'
                }), 409
            if campo == 'nota_lancada' and valor_norm == 'sim' and _terceiros_usa_fluxo_mg(doc) and _valor_bool_texto(doc.get('carga_recebida_mg') or '') != 'sim':
                return jsonify({
                    'ok': False,
                    'erro': 'NF de MG só pode ser lançada depois de Recebida MG = Sim.',
                    'codigo': 'recebida_mg_obrigatoria_para_lancamento'
                }), 400
            if campo in ('enviar_para_mg', 'carga_recebida_mg') and valor_norm == 'sim' and _motorista_obrigatorio_terceiros(doc) and not (doc.get('motorista_saida_mg') or '').strip():
                return jsonify({
                    'ok': False,
                    'erro': 'Para esta rota, informe o motorista que vai levar para MG antes de continuar.',
                    'codigo': 'motorista_obrigatorio_terceiros'
                }), 400
            if campo == 'carga_recebida_mg' and valor_norm == 'sim' and not _terceiros_eh_area_carreta(doc) and not recebedor_mg:
                return jsonify({
                    'ok': False,
                    'erro': 'Informe quem recebeu em MG.',
                    'codigo': 'recebedor_mg_obrigatorio'
                }), 400
            campo_em = campo + '_em'
            campo_por = campo + '_por'
            col_motivo = _MOTIVO_COL_TERCEIROS.get(campo)
            if valor_norm == 'nao' and col_motivo:
                if not motivo:
                    return jsonify({
                        'ok': False,
                        'erro': 'Informe o motivo para registrar «Não».',
                        'motivo_obrigatorio': True,
                    }), 400
                _pg_auditoria_sync_desligar(conn)
                sql_recebedor_nao = ', recebedor_mg = NULL' if campo == 'carga_recebida_mg' else ''
                conn.execute(
                    'UPDATE ' + _tbl_terceiros_documentos(conn) + ' SET '
                    + campo + ' = ?, ' + campo_em + ' = ?, ' + campo_por + ' = ?, '
                    + col_motivo + ' = ?' + sql_recebedor_nao + ', atualizado_em = ?, atualizado_por = ? WHERE id = ?',
                    (valor_norm, agora, usuario, motivo, agora, usuario, documento_id),
                )
            else:
                sql_clear_motivo = ''
                params_extra = ()
                if col_motivo:
                    sql_clear_motivo = ', ' + col_motivo + ' = NULL'
                if campo == 'carga_recebida_mg':
                    sql_clear_motivo += ', recebedor_mg = ?'
                    params_extra = (recebedor_mg if valor_norm == 'sim' else None,)
                _pg_auditoria_sync_desligar(conn)
                conn.execute(
                    'UPDATE ' + _tbl_terceiros_documentos(conn) + ' SET '
                    + campo + ' = ?, ' + campo_em + ' = ?, ' + campo_por + ' = ?'
                    + sql_clear_motivo + ', atualizado_em = ?, atualizado_por = ? WHERE id = ?',
                    (valor_norm, agora, usuario) + params_extra + (agora, usuario, documento_id),
                )
            _registrar_evento_terceiros(conn, documento_id, campo, str(doc.get(campo) or ''), valor_norm, usuario)
        conn.commit()
        if campo == 'recebimento_concluido':
            doc_resp = dict(doc)
            doc_resp.pop('xml_conteudo', None)
            doc_resp['recebimento_concluido'] = bool(novo_bool)
            doc_resp['recebimento_concluido_em'] = _fmt_datahora_br(agora)
            doc_resp['recebimento_concluido_por'] = usuario
            doc_resp['atualizado_em'] = _fmt_datahora_br(agora)
            doc_resp['atualizado_por'] = usuario
            email_info = {'enviado': False, 'motivo': 'nao_aplicavel'}
            if novo_bool:
                email_info = _terceiros_email_consumivel_sp_apos_recebimento(
                    doc_resp, usuario, primeira_conclusao=not ja_concluido
                )
            return jsonify({
                'ok': True,
                'documento': doc_resp,
                'email_consumivel_sp': email_info,
            })
        return jsonify({'ok': True, 'documento': _carregar_documento_terceiros(conn, documento_id)})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'erro': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/terceiros/email/diagnostico', methods=['GET'])
def api_terceiros_email_diagnostico():
    """Verifica se SMTP está configurado no servidor (Render). Não expõe senhas."""
    dest = _terceiros_emails_consumivel_sp_destino()
    return jsonify({
        'ok': True,
        'smtp_configurado': _terceiros_smtp_configurado(),
        'smtp_host': (os.environ.get('SMTP_HOST') or '').strip() or None,
        'smtp_user_definido': bool((os.environ.get('SMTP_USER') or '').strip()),
        'smtp_senha_definida': bool((os.environ.get('SMTP_PASSWORD') or '').strip()),
        'destinatarios_consumivel_sp': dest,
        'dica': 'No Render: SMTP_HOST, SMTP_PORT=587, SMTP_USER, SMTP_PASSWORD (senha de app Gmail), SMTP_FROM, SMTP_USE_TLS=1',
    })


@app.route('/api/terceiros/email/teste', methods=['POST'])
def api_terceiros_email_teste():
    """Envia e-mail de teste para SMTP_USER (ou primeiro destinatário consumível SP)."""
    usuario = session.get('usuario', '') or 'sistema'
    dest = [(os.environ.get('SMTP_USER') or '').strip()]
    if not dest[0] or '@' not in dest[0]:
        dest = _terceiros_emails_consumivel_sp_destino()[:1]
    if not dest:
        return jsonify({'ok': False, 'erro': 'Nenhum destinatário para teste.'}), 400
    assunto = 'Teste — Controle de Carregamento / Terceiros'
    corpo = 'E-mail de teste enviado por %s em %s.' % (usuario, _fmt_datahora_br(_agora_iso()))
    ok, motivo = _terceiros_enviar_email(dest, assunto, corpo)
    if ok:
        return jsonify({'ok': True, 'enviado_para': dest[0]})
    return jsonify({'ok': False, 'erro': motivo or 'Falha ao enviar'}), 500


@app.route('/api/terceiros/documentos/<int:documento_id>/motorista', methods=['POST'])
def api_terceiros_motorista(documento_id):
    data = request.get_json() or {}
    motorista = (data.get('motorista') or '').strip()
    placa_informada = 'placa' in data
    placa = (data.get('placa') or '').strip().upper() if placa_informada else None
    tipo = (data.get('tipo') or 'chegada').strip().lower()
    saida_mg = tipo in ('saida_mg', 'envio_mg', 'levar_mg')
    if not motorista:
        return jsonify({'ok': False, 'erro': 'Informe o motorista da carreta.'}), 400
    usuario = session.get('usuario', '')
    conn = get_db()
    try:
        _ensure_terceiros_schema(conn)
        row = conn.execute(
            'SELECT motorista_carreta, placa_carreta, motorista_saida_mg, placa_saida_mg, remetente_nome, destinatario_nome FROM '
            + _tbl_terceiros_documentos(conn) + ' WHERE id = ?',
            (documento_id,)
        ).fetchone()
        if not row:
            return jsonify({'ok': False, 'erro': 'Documento não encontrado.'}), 404
        row_d = dict(row) if hasattr(row, 'keys') else {
            'motorista_carreta': row[0] if len(row) > 0 else '',
            'placa_carreta': row[1] if len(row) > 1 else '',
            'motorista_saida_mg': row[2] if len(row) > 2 else '',
            'placa_saida_mg': row[3] if len(row) > 3 else '',
            'remetente_nome': row[4] if len(row) > 4 else '',
            'destinatario_nome': row[5] if len(row) > 5 else '',
        }
        campo_motorista = 'motorista_saida_mg' if saida_mg else 'motorista_carreta'
        campo_motorista_em = 'motorista_saida_mg_em' if saida_mg else 'motorista_carreta_em'
        campo_placa = 'placa_saida_mg' if saida_mg else 'placa_carreta'
        valor_antigo = (row_d.get(campo_motorista) or '')
        if _motorista_obrigatorio_terceiros(row_d) and not motorista:
            return jsonify({'ok': False, 'erro': 'Para esta rota, informe o motorista.'}), 400
        agora = _agora_iso()
        if placa_informada:
            placa_antiga = (row_d.get(campo_placa) or '')
            conn.execute(
                'UPDATE ' + _tbl_terceiros_documentos(conn)
                + ' SET ' + campo_motorista + ' = ?, ' + campo_motorista_em + ' = ?, ' + campo_placa + ' = ?, atualizado_em = ?, atualizado_por = ? WHERE id = ?',
                (motorista, agora, placa or None, agora, usuario, documento_id)
            )
            if placa != placa_antiga:
                _registrar_evento_terceiros(conn, documento_id, campo_placa, placa_antiga, placa, usuario)
        else:
            conn.execute(
                'UPDATE ' + _tbl_terceiros_documentos(conn) + ' SET ' + campo_motorista + ' = ?, ' + campo_motorista_em + ' = ?, atualizado_em = ?, atualizado_por = ? WHERE id = ?',
                (motorista, agora, agora, usuario, documento_id)
            )
        _registrar_evento_terceiros(conn, documento_id, campo_motorista, valor_antigo, motorista, usuario)
        conn.commit()
        return jsonify({'ok': True, 'documento': _carregar_documento_terceiros(conn, documento_id)})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'erro': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/estatisticas', methods=['GET'])
def get_estatisticas():
    """Retorna estatísticas para o painel"""
    conn = get_db()
    
    total_bipados = conn.execute('SELECT COUNT(*) as total FROM produtos_bipados').fetchone()['total']
    total_carregados = conn.execute(
        'SELECT COUNT(*) as total FROM produtos_bipados WHERE status = ?',
        ('CARREGADO',)
    ).fetchone()['total']
    
    total_unicos = conn.execute('''
        SELECT COUNT(DISTINCT codigo_barras) as total FROM produtos_bipados
    ''').fetchone()['total']
    
    # Total de viagens com bipagem
    total_viagens = conn.execute('''
        SELECT COUNT(DISTINCT id_viagem) as total FROM produtos_bipados
        WHERE id_viagem IS NOT NULL AND id_viagem != ''
    ''').fetchone()['total']
    
    # Soma total de quantidades bipadas (não só contagem de linhas)
    soma_quantidades = conn.execute('SELECT COALESCE(SUM(quantidade), 0) as total FROM produtos_bipados').fetchone()['total']
    
    # Contar divergências nos roteiros mais recentes (evita timeout)
    try:
        total_divergencias = len(_coletar_divergencias('carregamento', limit_viagens=40))
    except Exception:
        total_divergencias = 0
    
    # Estatísticas por veículo
    veiculos = conn.execute('''
        SELECT veiculo, COUNT(*) as total
        FROM produtos_bipados
        WHERE status = 'CARREGADO' AND veiculo != ''
        GROUP BY veiculo
    ''').fetchall()
    
    conn.close()
    
    return jsonify({
        'total_bipados': total_bipados,
        'total_carregados': total_carregados,
        'total_unicos': total_unicos,
        'total_divergencias': total_divergencias,
        'total_viagens': total_viagens,
        'soma_quantidades': soma_quantidades,
        'veiculos': [dict(row) for row in veiculos]
    })

def _init_db_em_background():
    """Não bloqueia o boot do Gunicorn (evita 502 no health check do Render)."""
    try:
        init_db()
        sync_usuarios_from_config()
    except Exception as e:
        import traceback
        try:
            print("[controle-carregamento] init_db/sync_usuarios falhou:", e, flush=True)
            traceback.print_exc()
        except Exception:
            pass


# Inicializar banco ao carregar o app (gunicorn não executa __main__)
threading.Thread(target=_init_db_em_background, daemon=True, name='init_db').start()

if __name__ == '__main__':
    init_db()
    sync_usuarios_from_config()
    port = int(os.environ.get('PORT', '5000'))
    print("=" * 60)
    print(SYSTEM_NAME.upper() + " — WMS")
    print("=" * 60)
    print(f"\nServidor iniciado em: http://127.0.0.1:{port}")
    print("\nPressione Ctrl+C para parar o servidor")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=port, threaded=True)
