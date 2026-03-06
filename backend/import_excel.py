from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

import openpyxl
import psycopg
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.startswith("="):
        return ""
    return s


def _as_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        s = str(v).strip()
        if not s or s.startswith("="):
            return None
        return int(float(s))
    except Exception:
        return None


def _header_list(ws) -> list[str]:
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row:
        return []
    raw = header_row[0] or []
    headers = []
    for i, h in enumerate(raw):
        if h is None:
            headers.append(f"Col_{i}")
        else:
            hs = str(h).strip()
            headers.append(hs if hs else f"Col_{i}")
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _row_to_json(headers: list[str], row_values: tuple[Any, ...]) -> dict[str, Any]:
    d: dict[str, Any] = {}
    for i, h in enumerate(headers):
        if i >= len(row_values):
            d[h] = None
            continue
        v = row_values[i]
        if isinstance(v, datetime):
            d[h] = v.isoformat()
        else:
            d[h] = v
    return d


def _coluna_base_eh_ean(header_str: str) -> bool:
    h = (header_str or "").upper().strip()
    return "EAN" in h and "13" in h


def _coluna_base_eh_dun(header_str: str) -> bool:
    h = (header_str or "").upper().strip()
    return "DUN" in h and "14" in h


def _coluna_base_eh_codigo_interno(header_str: str) -> bool:
    h = (header_str or "").upper().strip()
    if h in ("CODIGO", "CÓDIGO"):
        return True
    if "INTERNO" in h and ("CODIGO" in h or "CÓDIGO" in h):
        return True
    if (
        "PRODUTO" in h
        and ("CODIGO" in h or "CÓDIGO" in h)
        and "BARRAS" not in h
        and "EAN" not in h
        and "DUN" not in h
    ):
        return True
    return False


def _coluna_base_eh_descricao(header_str: str) -> bool:
    h = (header_str or "").upper().strip()
    if "DESCRI" in h or "DESCRIÇÃO" in h:
        return True
    if h in ("PRODUTO", "NOME", "NOME DO PRODUTO"):
        return True
    return False


def _find_sheet(wb, exact_upper: str | None = None, contains_all: list[str] | None = None) -> str | None:
    if exact_upper:
        for s in wb.sheetnames:
            if s.upper().strip() == exact_upper:
                return s
    if contains_all:
        needles = [n.upper() for n in contains_all]
        for s in wb.sheetnames:
            up = s.upper()
            if all(n in up for n in needles):
                return s
    return None


@dataclass
class ImportResult:
    dataset_id: str
    base_rows: int
    romaneio_rows: int
    colaboradores_rows: int
    placas_rows: int


def criar_dataset(conn: psycopg.Connection, arquivo_nome: str | None) -> str:
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute(
            "insert into public.excel_datasets (arquivo_nome) values (%s) returning dataset_id",
            (arquivo_nome,),
        )
        row = cur.fetchone()
        if not row or not row.get("dataset_id"):
            raise RuntimeError("Falha ao criar dataset")
        return str(row["dataset_id"])


def importar_excel_para_supabase(
    conn: psycopg.Connection,
    caminho_arquivo: str,
    arquivo_nome: str | None = None,
) -> ImportResult:
    try:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=False)

    dataset_id = criar_dataset(conn, arquivo_nome)
    agora = datetime.now(timezone.utc)

    base_rows = 0
    romaneio_rows = 0
    colaboradores_rows = 0
    placas_rows = 0

    # BASE
    nome_base = _find_sheet(wb, exact_upper="BASE")
    if nome_base:
        ws = wb[nome_base]
        headers = _header_list(ws)
        col_codigo = next((i for i, h in enumerate(headers) if _coluna_base_eh_codigo_interno(h)), None)
        col_descricao = next((i for i, h in enumerate(headers) if _coluna_base_eh_descricao(h)), None)
        col_ean = next((i for i, h in enumerate(headers) if _coluna_base_eh_ean(h)), None)
        col_dun = next((i for i, h in enumerate(headers) if _coluna_base_eh_dun(h)), None)

        valores = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            codigo_interno = _as_str(row[col_codigo]) if (col_codigo is not None and col_codigo < len(row)) else ""
            if not codigo_interno:
                continue
            ean = _as_str(row[col_ean]) if (col_ean is not None and col_ean < len(row)) else ""
            dun = _as_str(row[col_dun]) if (col_dun is not None and col_dun < len(row)) else ""
            descricao = _as_str(row[col_descricao]) if (col_descricao is not None and col_descricao < len(row)) else ""

            data = _row_to_json(headers, row)
            valores.append(
                (
                    dataset_id,
                    idx,
                    agora,
                    codigo_interno,
                    ean,
                    dun,
                    descricao,
                    _as_str(data.get("Unidade") or data.get("UNIDADE") or ""),
                    _as_str(data.get("Peso") or data.get("PESO") or ""),
                    Jsonb(data),
                )
            )

        if valores:
            with conn.cursor() as cur:
                cur.executemany(
                    """
                    insert into public.base_codigo_barras
                      (dataset_id, row_index, importado_em, codigo_interno, ean, dun, descricao, unidade, peso, data)
                    values
                      (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    valores,
                )
            base_rows = len(valores)

    # ROMANEIO POR ITEM
    nome_rom = _find_sheet(wb, exact_upper="ROMANEIO POR ITEM") or _find_sheet(wb, contains_all=["ROMANEIO", "ITEM"])
    if nome_rom:
        ws = wb[nome_rom]
        headers = _header_list(ws)

        def find_col(pred):
            return next((i for i, h in enumerate(headers) if pred((h or "").upper().strip())), None)

        col_id_roteiro = find_col(lambda h: "ID" in h and "ROTEIRO" in h) or (1 if len(headers) > 1 else None)
        col_id_viagem = find_col(lambda h: "ID" in h and "VIAGEM" in h)
        col_codigo_produto = find_col(lambda h: ("CODIGO" in h or "CÓDIGO" in h) and "PRODUTO" in h and "BARRAS" not in h)
        col_descricao = find_col(lambda h: ("DESCRI" in h or "DESCRIÇÃO" in h) and "PRODUTO" in h)
        col_quantidade = find_col(lambda h: "QUANTIDADE" in h and "PRODUTO" in h)
        col_unidade = find_col(lambda h: "UNIDADE" in h and "MEDIDA" in h) or find_col(lambda h: h == "UNIDADE")
        col_peso = find_col(lambda h: "PESO" in h and "BRUTO" in h)
        col_codigo_cliente = find_col(lambda h: ("CODIGO" in h or "CÓDIGO" in h) and "CLIENTE" in h)
        col_endereco = find_col(lambda h: "ENDERE" in h)
        col_cidade = find_col(lambda h: "CIDADE" in h)

        # Fallback do mapeamento “fixo” usado no sistema atual
        if len(headers) >= 17:
            col_id_roteiro = 1
            col_codigo_produto = 14
            col_descricao = 15
            col_quantidade = 16

        valores = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            id_roteiro = _as_str(row[col_id_roteiro]) if (col_id_roteiro is not None and col_id_roteiro < len(row)) else ""
            codigo_produto = _as_str(row[col_codigo_produto]) if (col_codigo_produto is not None and col_codigo_produto < len(row)) else ""
            if not id_roteiro or not codigo_produto:
                continue
            descricao = _as_str(row[col_descricao]) if (col_descricao is not None and col_descricao < len(row)) else ""
            quantidade = _as_int(row[col_quantidade]) if (col_quantidade is not None and col_quantidade < len(row)) else None
            unidade = _as_str(row[col_unidade]) if (col_unidade is not None and col_unidade < len(row)) else ""
            peso_bruto = _as_str(row[col_peso]) if (col_peso is not None and col_peso < len(row)) else ""
            codigo_cliente = _as_str(row[col_codigo_cliente]) if (col_codigo_cliente is not None and col_codigo_cliente < len(row)) else ""
            endereco = _as_str(row[col_endereco]) if (col_endereco is not None and col_endereco < len(row)) else ""
            cidade = _as_str(row[col_cidade]) if (col_cidade is not None and col_cidade < len(row)) else ""
            id_viagem = _as_str(row[col_id_viagem]) if (col_id_viagem is not None and col_id_viagem < len(row)) else ""

            data = _row_to_json(headers, row)
            valores.append(
                (
                    dataset_id,
                    idx,
                    agora,
                    id_roteiro,
                    id_viagem,
                    codigo_produto,
                    descricao,
                    quantidade,
                    unidade,
                    peso_bruto,
                    codigo_cliente,
                    endereco,
                    cidade,
                    Jsonb(data),
                )
            )

        if valores:
            with conn.cursor() as cur:
                cur.executemany(
                    """
                    insert into public.excel_romaneio_por_item
                      (dataset_id, row_index, importado_em,
                       id_roteiro, id_viagem, codigo_produto, descricao, quantidade, unidade, peso_bruto,
                       codigo_cliente, endereco, cidade, data)
                    values
                      (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    valores,
                )
            romaneio_rows = len(valores)

    # COLABORADORES: tabela excel_colaboradores removida; usar tabelas colaboradores e motoristas
    colaboradores_rows = 0

    # PLACAS: tabela excel_placas removida; usar tabela placas
    placas_rows = 0

    wb.close()
    return ImportResult(
        dataset_id=dataset_id,
        base_rows=base_rows,
        romaneio_rows=romaneio_rows,
        colaboradores_rows=colaboradores_rows,
        placas_rows=placas_rows,
    )

