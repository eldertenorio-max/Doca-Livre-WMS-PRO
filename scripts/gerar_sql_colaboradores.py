#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Lê a aba COLABORADORES da planilha Excel e gera SQL para inserir no Supabase.
"""
import openpyxl
import os

def main():
    # Procurar a planilha
    planilha = 'CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx'
    if not os.path.isfile(planilha):
        print(f"Erro: Planilha não encontrada: {planilha}")
        return
    
    print(f"Lendo planilha: {planilha}")
    wb = openpyxl.load_workbook(planilha, data_only=True)
    
    if 'COLABORADORES' not in wb.sheetnames:
        print("Erro: Aba COLABORADORES não encontrada")
        wb.close()
        return
    
    ws = wb['COLABORADORES']
    print(f"Total de linhas: {ws.max_row}")
    
    colaboradores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        
        centro_custo = str(row[0] or '').strip()
        funcao = str(row[1] or '').strip()
        nome = str(row[2] or '').strip()
        
        if not nome:
            continue
        
        # Identificar tipo baseado no centro de custo ou função
        cc_upper = centro_custo.upper()
        fn_upper = funcao.upper()
        tipo = ''
        
        if 'TRANSPORTE' in cc_upper or 'MOTORISTA' in fn_upper:
            tipo = 'MOTORISTA'
        elif 'CONFERENTE' in fn_upper:
            tipo = 'CONFERENTE'
        elif 'AJUDANTE' in fn_upper or 'AUXILIAR' in fn_upper or 'AUX' in fn_upper.replace('AUXILIAR', 'AUX'):
            tipo = 'AJUDANTE'
        elif 'COORDENADOR' in fn_upper:
            tipo = 'COORDENADOR'
        
        # PULAR MOTORISTAS (motoristas não vão na tabela colaboradores)
        if tipo == 'MOTORISTA':
            continue
        
        colaboradores.append({
            'nome': nome.replace("'", "''"),
            'funcao': funcao.replace("'", "''"),
            'centro_custo': centro_custo.replace("'", "''"),
            'tipo': tipo or 'OUTRO'
        })
    
    wb.close()
    
    print(f"Total de colaboradores extraídos: {len(colaboradores)}")
    
    # Gerar SQL
    output_file = os.path.join('supabase', 'insert_colaboradores_data.sql')
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('-- ============================================================\n')
        f.write('-- DADOS: Colaboradores (da planilha Excel)\n')
        f.write('-- ============================================================\n')
        f.write(f'-- Total: {len(colaboradores)} colaboradores\n')
        f.write('-- Execute no SQL Editor do Supabase\n')
        f.write('-- ============================================================\n\n')
        
        # Agrupar por tipo para organizar
        tipos = {}
        for c in colaboradores:
            t = c['tipo'] or 'OUTRO'
            if t not in tipos:
                tipos[t] = []
            tipos[t].append(c)
        
        for tipo, lista in sorted(tipos.items()):
            f.write(f'-- {tipo} ({len(lista)} colaboradores)\n')
            f.write('insert into public.colaboradores (nome, funcao, centro_custo, tipo, ativo) values\n')
            for i, c in enumerate(lista):
                virgula = ',' if i < len(lista) - 1 else ''
                linha = f"  ('{c['nome']}', '{c['funcao']}', '{c['centro_custo']}', '{c['tipo']}', true){virgula}\n"
                f.write(linha)
            f.write('on conflict do nothing;\n\n')
    
    print(f'Arquivo gerado: {output_file}')
    print(f'Cole o conteúdo no SQL Editor do Supabase.')

if __name__ == '__main__':
    main()
