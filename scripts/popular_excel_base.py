#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Popula a tabela base_codigo_barras no Supabase com dados da aba BASE da planilha.
Conecta direto no Supabase (mais eficiente que gerar SQL).
"""
import openpyxl
import os
import json
from datetime import datetime

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:
    print("Erro: psycopg nao instalado. Execute: pip install psycopg[binary]")
    exit(1)

# Carrega .env da raiz do projeto (se existir)
try:
    from dotenv import load_dotenv
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    load_dotenv(os.path.join(base_dir, '.env'))
except ImportError:
    pass

def main():
    # Verificar DATABASE_URL
    db_url = os.environ.get('DATABASE_URL', '').strip()
    if not db_url:
        print("Erro: Configure DATABASE_URL antes de executar")
        print("Exemplo (PowerShell): $env:DATABASE_URL = \"postgresql://...\"")
        print("Ou crie um arquivo .env na raiz do projeto com DATABASE_URL e PGSSLMODE")
        return
    
    planilha = 'CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx'
    if not os.path.isfile(planilha):
        print(f"Erro: Planilha não encontrada: {planilha}")
        return
    
    print(f"Lendo planilha: {planilha}")
    wb = openpyxl.load_workbook(planilha, data_only=True)
    
    if 'BASE' not in wb.sheetnames:
        print("Erro: Aba BASE não encontrada")
        wb.close()
        return
    
    ws = wb['BASE']
    print(f"Total de linhas: {ws.max_row}")
    
    # Ler cabeçalho
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h or f'col_{i}') for i, h in enumerate(header_row)]
    
    print(f"Colunas: {', '.join(headers[:6])}...")
    
    # Conectar no Supabase
    print("Conectando no Supabase...")
    sslmode = os.environ.get('PGSSLMODE', 'require')
    
    with psycopg.connect(db_url, sslmode=sslmode, row_factory=dict_row) as conn:
        # Criar dataset (versão da importação)
        print("Criando dataset...")
        row = conn.execute(
            """INSERT INTO public.excel_datasets (arquivo_nome, importado_em, ativo)
               VALUES (%s, NOW(), TRUE)
               RETURNING dataset_id""",
            (os.path.basename(planilha),)
        ).fetchone()
        dataset_id = row['dataset_id']
        print(f"Dataset criado: {dataset_id}")
        
        # Desativar datasets antigos
        conn.execute(
            "UPDATE public.excel_datasets SET ativo = FALSE WHERE dataset_id != %s",
            (dataset_id,)
        )
        
        # Inserir produtos da aba BASE
        print("Inserindo produtos da aba BASE...")
        inseridos = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            
            # Montar dict com todos os dados
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    # Converter datetime para string
                    if isinstance(val, datetime):
                        val = val.strftime('%Y-%m-%d %H:%M:%S')
                    row_dict[headers[i]] = val
            
            # Extrair campos principais para índice
            codigo_interno = str(row_dict.get('Codigo') or row_dict.get('codigo') or '').strip()
            descricao = str(row_dict.get('Descricao') or row_dict.get('descricao') or '').strip()
            unidade = str(row_dict.get('Unidade') or row_dict.get('unidade') or '').strip()
            
            # EAN e DUN (podem estar em colunas diferentes dependendo da planilha)
            ean = ''
            dun = ''
            for key, val in row_dict.items():
                key_upper = str(key).upper()
                if 'EAN' in key_upper and not 'DUN' in key_upper:
                    ean = str(val or '').strip()
                elif 'DUN' in key_upper:
                    dun = str(val or '').strip()
            
            # Peso
            peso = str(row_dict.get('Peso') or row_dict.get('peso') or '').strip()
            
            # Inserir no Supabase
            conn.execute(
                """INSERT INTO public.base_codigo_barras 
                   (dataset_id, row_index, codigo_interno, ean, dun, descricao, unidade, peso, data)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    dataset_id,
                    row_idx,
                    codigo_interno or None,
                    ean or None,
                    dun or None,
                    descricao or None,
                    unidade or None,
                    peso or None,
                    json.dumps(row_dict, ensure_ascii=False, default=str)
                )
            )
            inseridos += 1
            
            if inseridos % 100 == 0:
                print(f"  Inseridos: {inseridos}...")
                conn.commit()
        
        conn.commit()
        print(f"\nTotal inserido: {inseridos} produtos")
        print("✓ Aba BASE importada com sucesso!")
    
    wb.close()

if __name__ == '__main__':
    main()
