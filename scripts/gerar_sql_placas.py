#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Lê a aba TODAS AS PLACAS da planilha e gera SQL para inserir no Supabase.
"""
import openpyxl
import os

def main():
    planilha = 'CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx'
    if not os.path.isfile(planilha):
        print(f"Erro: Planilha não encontrada: {planilha}")
        return
    
    print(f"Lendo planilha: {planilha}")
    wb = openpyxl.load_workbook(planilha, data_only=True)
    
    # Procurar aba "Todas as placas" (pode ter variações no nome)
    nome_aba = next((s for s in wb.sheetnames if 'TODAS' in s.upper() and 'PLACAS' in s.upper()), None)
    if not nome_aba:
        print("Erro: Aba TODAS AS PLACAS não encontrada")
        wb.close()
        return
    
    ws = wb[nome_aba]
    print(f"Aba encontrada: {nome_aba}")
    print(f"Total de linhas: {ws.max_row}")
    
    placas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        
        # Coluna B (índice 1) = Placa
        placa = str(row[1] or '').strip()
        
        if not placa:
            continue
        
        # Coluna C (índice 2) = Descrição (se existir)
        descricao = str(row[2] or '').strip() if len(row) > 2 else ''
        
        placas.append({
            'placa': placa.replace("'", "''"),
            'descricao': descricao.replace("'", "''")
        })
    
    wb.close()
    
    print(f"Total de placas extraídas: {len(placas)}")
    
    # Gerar SQL
    output_file = os.path.join('supabase', 'insert_placas_data.sql')
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('-- ============================================================\n')
        f.write('-- DADOS: Placas (da planilha Excel)\n')
        f.write('-- ============================================================\n')
        f.write(f'-- Total: {len(placas)} placas\n')
        f.write('-- Execute no SQL Editor do Supabase\n')
        f.write('-- ============================================================\n\n')
        
        f.write('insert into public.placas (placa, descricao, ativo) values\n')
        for i, p in enumerate(placas):
            virgula = ',' if i < len(placas) - 1 else ''
            if p['descricao']:
                linha = f"  ('{p['placa']}', '{p['descricao']}', true){virgula}\n"
            else:
                linha = f"  ('{p['placa']}', null, true){virgula}\n"
            f.write(linha)
        f.write('on conflict (placa) do nothing;\n')
    
    print(f'Arquivo gerado: {output_file}')
    print(f'Cole o conteúdo no SQL Editor do Supabase.')

if __name__ == '__main__':
    main()
