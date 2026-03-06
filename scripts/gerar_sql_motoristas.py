#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Lê a aba COLABORADORES da planilha e extrai apenas os MOTORISTAS.
"""
import openpyxl
import os

def main():
    planilha = 'CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx'
    if not os.path.isfile(planilha):
        print(f"Erro: Planilha não encontrada: {planilha}")
        return
    
    print(f"Lendo planilha: {planilha}")
    wb = openpyxl.load_workbook(planilha, data_only=True)
    
    if 'COLABORADORES' not in wb.sheetnames:
        print("Erro: Aba COLABORADORES não encontrada")
        wb.close()
        return
    
    ws = wb['COLABORADORES']
    print(f"Total de linhas: {ws.max_row}")
    
    motoristas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        
        centro_custo = str(row[0] or '').strip()
        funcao = str(row[1] or '').strip()
        nome = str(row[2] or '').strip()
        
        if not nome:
            continue
        
        # Filtrar APENAS motoristas (TRANSPORTE)
        cc_upper = centro_custo.upper()
        fn_upper = funcao.upper()
        
        if 'TRANSPORTE' not in cc_upper and 'MOTORISTA' not in fn_upper:
            continue
        
        motoristas.append({
            'nome': nome.replace("'", "''"),
            'centro_custo': centro_custo.replace("'", "''")
        })
    
    wb.close()
    
    print(f"Total de motoristas extraídos: {len(motoristas)}")
    
    # Gerar SQL
    output_file = os.path.join('supabase', 'insert_motoristas_data.sql')
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('-- ============================================================\n')
        f.write('-- DADOS: Motoristas (da planilha Excel)\n')
        f.write('-- ============================================================\n')
        f.write(f'-- Total: {len(motoristas)} motoristas\n')
        f.write('-- Execute no SQL Editor do Supabase\n')
        f.write('-- ============================================================\n\n')
        
        f.write('insert into public.motoristas (nome, centro_custo, ativo) values\n')
        for i, m in enumerate(motoristas):
            virgula = ',' if i < len(motoristas) - 1 else ''
            linha = f"  ('{m['nome']}', '{m['centro_custo']}', true){virgula}\n"
            f.write(linha)
        f.write('on conflict (nome) do nothing;\n')
    
    print(f'Arquivo gerado: {output_file}')
    print(f'Cole o conteúdo no SQL Editor do Supabase.')

if __name__ == '__main__':
    main()
