import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime

# Criar workbook
wb = Workbook()

# Remover sheet padrão
wb.remove(wb.active)

# Definir estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
# Cores para formatação condicional
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# ========== ABA BASE ==========
ws_base = wb.create_sheet("BASE")
ws_base.append(["CÓDIGO DE BARRAS", "PRODUTO", "QUANTIDADE", "DATA/HORA", "VEÍCULO", "STATUS"])

# Formatar cabeçalho BASE
for cell in ws_base[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar largura das colunas BASE
ws_base.column_dimensions['A'].width = 20
ws_base.column_dimensions['B'].width = 40
ws_base.column_dimensions['C'].width = 15
ws_base.column_dimensions['D'].width = 20
ws_base.column_dimensions['E'].width = 20
ws_base.column_dimensions['F'].width = 15

# Adicionar validação de dados na coluna STATUS (coluna F)
status_validation = DataValidation(type="list", formula1='"CARREGADO,PENDENTE,CANCELADO"', allow_blank=True)
status_validation.add(f'F2:F10000')
ws_base.add_data_validation(status_validation)

# Adicionar comentário na célula F1
comment = Comment("Selecione o status do item. Use 'CARREGADO' para itens que foram carregados no veículo.", "Sistema")
ws_base['F1'].comment = comment

# ========== ABA CONFERENCIA ==========
ws_conferencia = wb.create_sheet("CONFERENCIA")
ws_conferencia.append(["CÓDIGO DE BARRAS", "PRODUTO", "QUANTIDADE TOTAL", "ÚLTIMA DATA/HORA", "VEÍCULO"])

# Formatar cabeçalho CONFERENCIA
for cell in ws_conferencia[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar largura das colunas CONFERENCIA
ws_conferencia.column_dimensions['A'].width = 20
ws_conferencia.column_dimensions['B'].width = 40
ws_conferencia.column_dimensions['C'].width = 15
ws_conferencia.column_dimensions['D'].width = 20
ws_conferencia.column_dimensions['E'].width = 20

# Fórmula para consolidar dados da BASE (sem repetir código de barras)
# Usando fórmula de array para valores únicos (será inserida como fórmula de array no Excel)
# Para Excel, precisamos usar uma abordagem diferente - vamos criar uma fórmula que funciona linha por linha
ws_conferencia['A2'] = '=IFERROR(INDEX(BASE!$A$2:$A$10000, MATCH(0, COUNTIF($A$1:A1, BASE!$A$2:$A$10000&""), 0)), "")'
ws_conferencia['B2'] = '=IF(A2="", "", VLOOKUP(A2, BASE!$A$2:$B$10000, 2, FALSE))'
ws_conferencia['C2'] = '=IF(A2="", "", SUMIF(BASE!$A$2:$A$10000, A2, BASE!$C$2:$C$10000))'
ws_conferencia['D2'] = '=IF(A2="", "", INDEX(BASE!$D$2:$D$10000, MATCH(A2, BASE!$A$2:$A$10000, 0)))'
ws_conferencia['E2'] = '=IF(A2="", "", INDEX(BASE!$E$2:$E$10000, MATCH(A2, BASE!$A$2:$A$10000, 0)))'

# Copiar fórmulas para várias linhas (até linha 1000)
for row in range(3, 1001):
    ws_conferencia[f'A{row}'] = f'=IFERROR(INDEX(BASE!$A$2:$A$10000, MATCH(0, COUNTIF($A$1:A{row-1}, BASE!$A$2:$A$10000&""), 0)), "")'
    ws_conferencia[f'B{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, BASE!$A$2:$B$10000, 2, FALSE))'
    ws_conferencia[f'C{row}'] = f'=IF(A{row}="", "", SUMIF(BASE!$A$2:$A$10000, A{row}, BASE!$C$2:$C$10000))'
    ws_conferencia[f'D{row}'] = f'=IF(A{row}="", "", INDEX(BASE!$D$2:$D$10000, MATCH(A{row}, BASE!$A$2:$A$10000, 0)))'
    ws_conferencia[f'E{row}'] = f'=IF(A{row}="", "", INDEX(BASE!$E$2:$E$10000, MATCH(A{row}, BASE!$A$2:$A$10000, 0)))'

# ========== ABA EXTRATO ==========
ws_extrato = wb.create_sheet("EXTRATO")
ws_extrato.append(["CÓDIGO DE BARRAS", "PRODUTO", "QUANTIDADE", "DATA/HORA", "VEÍCULO", "STATUS"])

# Formatar cabeçalho EXTRATO
for cell in ws_extrato[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar largura das colunas EXTRATO
ws_extrato.column_dimensions['A'].width = 20
ws_extrato.column_dimensions['B'].width = 40
ws_extrato.column_dimensions['C'].width = 15
ws_extrato.column_dimensions['D'].width = 20
ws_extrato.column_dimensions['E'].width = 20
ws_extrato.column_dimensions['F'].width = 15

# Fórmula para filtrar apenas itens com STATUS = "CARREGADO"
# Usando fórmula de array (será inserida como fórmula de array no Excel)
ws_extrato['A2'] = '=IFERROR(INDEX(BASE!$A$2:$A$10000, SMALL(IF(BASE!$F$2:$F$10000="CARREGADO", ROW(BASE!$A$2:$A$10000)-ROW(BASE!$A$2)+1), ROW(A1))), "")'
ws_extrato['B2'] = '=IF(A2="", "", VLOOKUP(A2, BASE!$A$2:$B$10000, 2, FALSE))'
ws_extrato['C2'] = '=IF(A2="", "", VLOOKUP(A2, BASE!$A$2:$C$10000, 3, FALSE))'
ws_extrato['D2'] = '=IF(A2="", "", VLOOKUP(A2, BASE!$A$2:$D$10000, 4, FALSE))'
ws_extrato['E2'] = '=IF(A2="", "", VLOOKUP(A2, BASE!$A$2:$E$10000, 5, FALSE))'
ws_extrato['F2'] = '=IF(A2="", "", "CARREGADO")'

# Copiar fórmulas para várias linhas
for row in range(3, 1001):
    ws_extrato[f'A{row}'] = f'=IFERROR(INDEX(BASE!$A$2:$A$10000, SMALL(IF(BASE!$F$2:$F$10000="CARREGADO", ROW(BASE!$A$2:$A$10000)-ROW(BASE!$A$2)+1), ROW(A{row-1}))), "")'
    ws_extrato[f'B{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, BASE!$A$2:$B$10000, 2, FALSE))'
    ws_extrato[f'C{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, BASE!$A$2:$C$10000, 3, FALSE))'
    ws_extrato[f'D{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, BASE!$A$2:$D$10000, 4, FALSE))'
    ws_extrato[f'E{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, BASE!$A$2:$E$10000, 5, FALSE))'
    ws_extrato[f'F{row}'] = f'=IF(A{row}="", "", "CARREGADO")'

# ========== ABA ROMANEIO POR ITEM ==========
ws_romaneio = wb.create_sheet("ROMANEIO POR ITEM")
ws_romaneio.append(["CÓDIGO DE BARRAS", "PRODUTO", "QUANTIDADE BIPADA", "QUANTIDADE ROMANEIO", "DIFERENÇA", "STATUS"])

# Formatar cabeçalho ROMANEIO
for cell in ws_romaneio[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar largura das colunas ROMANEIO
ws_romaneio.column_dimensions['A'].width = 20
ws_romaneio.column_dimensions['B'].width = 40
ws_romaneio.column_dimensions['C'].width = 18
ws_romaneio.column_dimensions['D'].width = 20
ws_romaneio.column_dimensions['E'].width = 15
ws_romaneio.column_dimensions['F'].width = 15

# Adicionar formatação condicional na aba ROMANEIO

# Formatação para OK (diferença = 0)
ok_rule = CellIsRule(operator='equal', formula=['0'], fill=green_fill)
ws_romaneio.conditional_formatting.add('E2:E10000', ok_rule)

# Formatação para SOBRA
sobra_rule_romaneio = CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill)
ws_romaneio.conditional_formatting.add('E2:E10000', sobra_rule_romaneio)

# Formatação para FALTA
falta_rule_romaneio = CellIsRule(operator='lessThan', formula=['0'], fill=yellow_fill)
ws_romaneio.conditional_formatting.add('E2:E10000', falta_rule_romaneio)

# Fórmulas para comparar CONFERENCIA com ROMANEIO
# Assumindo que o usuário vai inserir manualmente a quantidade do romaneio na coluna D
ws_romaneio['A2'] = '=IF(CONFERENCIA!A2="", "", CONFERENCIA!A2)'
ws_romaneio['B2'] = '=IF(A2="", "", CONFERENCIA!B2)'
ws_romaneio['C2'] = '=IF(A2="", "", CONFERENCIA!C2)'
ws_romaneio['D2'] = ''  # Campo manual para inserir quantidade do romaneio
ws_romaneio['E2'] = '=IF(A2="", "", C2-D2)'
ws_romaneio['F2'] = '=IF(A2="", "", IF(E2=0, "OK", IF(E2>0, "SOBRA", "FALTA")))'

# Copiar fórmulas para várias linhas
for row in range(3, 1001):
    ws_romaneio[f'A{row}'] = f'=IF(CONFERENCIA!A{row}="", "", CONFERENCIA!A{row})'
    ws_romaneio[f'B{row}'] = f'=IF(A{row}="", "", CONFERENCIA!B{row})'
    ws_romaneio[f'C{row}'] = f'=IF(A{row}="", "", CONFERENCIA!C{row})'
    ws_romaneio[f'D{row}'] = ''  # Campo manual
    ws_romaneio[f'E{row}'] = f'=IF(A{row}="", "", C{row}-D{row})'
    ws_romaneio[f'F{row}'] = f'=IF(A{row}="", "", IF(E{row}=0, "OK", IF(E{row}>0, "SOBRA", "FALTA")))'

# ========== ABA DIVERGÊNCIAS ==========
ws_divergencias = wb.create_sheet("DIVERGÊNCIAS")
ws_divergencias.append(["CÓDIGO DE BARRAS", "PRODUTO", "QUANTIDADE BIPADA", "QUANTIDADE ROMANEIO", "DIFERENÇA", "TIPO"])

# Formatar cabeçalho DIVERGÊNCIAS
for cell in ws_divergencias[1]:
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar largura das colunas DIVERGÊNCIAS
ws_divergencias.column_dimensions['A'].width = 20
ws_divergencias.column_dimensions['B'].width = 40
ws_divergencias.column_dimensions['C'].width = 18
ws_divergencias.column_dimensions['D'].width = 20
ws_divergencias.column_dimensions['E'].width = 15
ws_divergencias.column_dimensions['F'].width = 15

# Adicionar formatação condicional para destacar divergências

# Formatação para SOBRA (valores positivos na coluna E)
sobra_rule = CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill)
ws_divergencias.conditional_formatting.add('E2:E10000', sobra_rule)

# Formatação para FALTA (valores negativos na coluna E)  
falta_rule = CellIsRule(operator='lessThan', formula=['0'], fill=yellow_fill)
ws_divergencias.conditional_formatting.add('E2:E10000', falta_rule)

# Fórmula para filtrar apenas divergências (diferença diferente de zero)
ws_divergencias['A2'] = '=IFERROR(INDEX(\'ROMANEIO POR ITEM\'!$A$2:$A$10000, SMALL(IF(\'ROMANEIO POR ITEM\'!$E$2:$E$10000<>0, ROW(\'ROMANEIO POR ITEM\'!$A$2:$A$10000)-ROW(\'ROMANEIO POR ITEM\'!$A$2)+1), ROW(A1))), "")'
ws_divergencias['B2'] = '=IF(A2="", "", VLOOKUP(A2, \'ROMANEIO POR ITEM\'!$A$2:$B$10000, 2, FALSE))'
ws_divergencias['C2'] = '=IF(A2="", "", VLOOKUP(A2, \'ROMANEIO POR ITEM\'!$A$2:$C$10000, 3, FALSE))'
ws_divergencias['D2'] = '=IF(A2="", "", VLOOKUP(A2, \'ROMANEIO POR ITEM\'!$A$2:$D$10000, 4, FALSE))'
ws_divergencias['E2'] = '=IF(A2="", "", VLOOKUP(A2, \'ROMANEIO POR ITEM\'!$A$2:$E$10000, 5, FALSE))'
ws_divergencias['F2'] = '=IF(A2="", "", VLOOKUP(A2, \'ROMANEIO POR ITEM\'!$A$2:$F$10000, 6, FALSE))'

# Copiar fórmulas para várias linhas
for row in range(3, 1001):
    ws_divergencias[f'A{row}'] = f'=IFERROR(INDEX(\'ROMANEIO POR ITEM\'!$A$2:$A$10000, SMALL(IF(\'ROMANEIO POR ITEM\'!$E$2:$E$10000<>0, ROW(\'ROMANEIO POR ITEM\'!$A$2:$A$10000)-ROW(\'ROMANEIO POR ITEM\'!$A$2)+1), ROW(A{row-1}))), "")'
    ws_divergencias[f'B{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, \'ROMANEIO POR ITEM\'!$A$2:$B$10000, 2, FALSE))'
    ws_divergencias[f'C{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, \'ROMANEIO POR ITEM\'!$A$2:$C$10000, 3, FALSE))'
    ws_divergencias[f'D{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, \'ROMANEIO POR ITEM\'!$A$2:$D$10000, 4, FALSE))'
    ws_divergencias[f'E{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, \'ROMANEIO POR ITEM\'!$A$2:$E$10000, 5, FALSE))'
    ws_divergencias[f'F{row}'] = f'=IF(A{row}="", "", VLOOKUP(A{row}, \'ROMANEIO POR ITEM\'!$A$2:$F$10000, 6, FALSE))'

# ========== ABA PAINEL ==========
ws_painel = wb.create_sheet("PAINEL", 0)  # Criar como primeira aba

# Título principal
ws_painel.merge_cells('A1:F1')
title_cell = ws_painel['A1']
title_cell.value = "CONTROLE DE CARREGAMENTO ULTRAPÃO"
title_cell.font = Font(bold=True, size=18, color="FFFFFF")
title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_painel.row_dimensions[1].height = 30

# Seção de Resumo Geral
ws_painel['A3'] = "RESUMO GERAL"
ws_painel['A3'].font = title_font
ws_painel['A3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws_painel['A4'] = "Total de Itens Bipados:"
ws_painel['B4'] = '=COUNTA(BASE!A:A)-1'
ws_painel['A5'] = "Total de Itens Carregados:"
ws_painel['B5'] = '=COUNTIF(BASE!F:F, "CARREGADO")'
ws_painel['A6'] = "Total de Itens Únicos:"
ws_painel['B6'] = '=COUNTA(CONFERENCIA!A:A)-1'
ws_painel['A7'] = "Total de Divergências:"
ws_painel['B7'] = '=COUNTA(DIVERGÊNCIAS!A:A)-1'

# Formatar células de resumo
for row in range(4, 8):
    ws_painel[f'A{row}'].font = Font(bold=True)
    ws_painel[f'B{row}'].font = Font(size=12)
    ws_painel[f'A{row}'].border = border
    ws_painel[f'B{row}'].border = border

# Seção de Estatísticas por Veículo
ws_painel['D3'] = "ESTATÍSTICAS POR VEÍCULO"
ws_painel['D3'].font = title_font
ws_painel['D3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws_painel['D4'] = "Veículo"
ws_painel['E4'] = "Itens Carregados"
ws_painel['D4'].font = header_font
ws_painel['E4'].font = header_font
ws_painel['D4'].fill = header_fill
ws_painel['E4'].fill = header_fill
ws_painel['D4'].border = border
ws_painel['E4'].border = border

# Ajustar larguras
ws_painel.column_dimensions['A'].width = 25
ws_painel.column_dimensions['B'].width = 20
ws_painel.column_dimensions['D'].width = 20
ws_painel.column_dimensions['E'].width = 20

# Seção de Acesso Rápido
ws_painel['A10'] = "ACESSO RÁPIDO ÀS ABAS"
ws_painel['A10'].font = title_font
ws_painel['A10'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

abas = [
    ("BASE", "Cadastro de produtos bipados"),
    ("CONFERENCIA", "Conferência sem repetir códigos"),
    ("EXTRATO", "Relação de itens carregados"),
    ("ROMANEIO POR ITEM", "Conferência com romaneio"),
    ("DIVERGÊNCIAS", "Lista de divergências encontradas")
]

for idx, (aba, descricao) in enumerate(abas, start=11):
    ws_painel[f'A{idx}'] = aba
    ws_painel[f'B{idx}'] = descricao
    ws_painel[f'A{idx}'].font = Font(bold=True, color="366092")
    ws_painel[f'A{idx}'].border = border
    ws_painel[f'B{idx}'].border = border

# Instruções
ws_painel['A17'] = "INSTRUÇÕES DE USO"
ws_painel['A17'].font = title_font
ws_painel['A17'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

instrucoes = [
    "1. Na aba BASE, registre todos os produtos bipados com código de barras, produto, quantidade, data/hora, veículo e status",
    "2. A aba CONFERENCIA consolida automaticamente os itens sem repetir códigos de barras",
    "3. A aba EXTRATO mostra apenas os itens com status 'CARREGADO'",
    "4. Na aba ROMANEIO POR ITEM, insira manualmente a quantidade do romaneio na coluna D para comparar",
    "5. A aba DIVERGÊNCIAS mostra automaticamente todas as diferenças encontradas"
]

for idx, instrucao in enumerate(instrucoes, start=18):
    ws_painel[f'A{idx}'] = instrucao
    ws_painel[f'A{idx}'].font = Font(size=10)
    ws_painel.merge_cells(f'A{idx}:F{idx}')

# Congelar primeira linha em todas as abas (exceto PAINEL)
for ws in wb.worksheets:
    if ws.title != "PAINEL":
        ws.freeze_panes = 'A2'

# Salvar arquivo
import os
arquivo = "CONTROLE DE CARREGAMENTO ULTRAPAO.xlsx"
arquivo_path = os.path.join(os.getcwd(), arquivo)

# Tentar remover arquivo existente se houver
if os.path.exists(arquivo_path):
    try:
        os.remove(arquivo_path)
    except PermissionError:
        print(f"AVISO: Não foi possível remover o arquivo existente. Por favor, feche o arquivo '{arquivo}' se estiver aberto.")
        arquivo = "CONTROLE DE CARREGAMENTO ULTRAPAO_NOVO.xlsx"
        arquivo_path = os.path.join(os.getcwd(), arquivo)

wb.save(arquivo_path)
print(f"Planilha '{arquivo}' criada com sucesso!")
print(f"\nAbas criadas:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
